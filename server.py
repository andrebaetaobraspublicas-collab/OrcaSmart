#!/usr/bin/env python3
"""
server.py — Servidor principal do Sistema de Orçamentação de Obras
Backend: Python + Flask + SQLite (built-in)

Para executar:  python server.py
Acesse:         http://localhost:3000
"""

import os, sqlite3, json, sys, re, io, unicodedata, shutil, uuid, time
from pathlib import Path
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, session, redirect
from flask_cors import CORS
from datetime import date
import json as _json
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.exceptions import HTTPException
import traceback

try:
    import stripe
except Exception:
    stripe = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# ── Custom JSON encoder: converte tipos numpy/pandas para Python nativo ────────
class _NumpyEncoder(_json.JSONEncoder):
    def default(self, obj):
        try:
            import numpy as np
            if isinstance(obj, (np.integer,)):   return int(obj)
            if isinstance(obj, (np.floating,)):  return float(obj)
            if isinstance(obj, (np.bool_,)):     return bool(obj)
            if isinstance(obj, np.ndarray):      return obj.tolist()
        except ImportError:
            pass
        return super().default(obj)

# ─── CONFIGURAÇÃO DO BANCO ────────────────────────────────────────────────────
BASE_DIR = os.environ.get('ORCASMART_SAAS_BASE_DIR') or os.path.dirname(os.path.abspath(__file__))
MASTER_DB_PATH = os.path.join(BASE_DIR, 'saas_master.db')
TENANT_DB_DIR = os.path.join(BASE_DIR, 'tenant_dbs')
DB_TEMPLATE_PATH = os.path.join(BASE_DIR, 'database', 'orcamento_obras_template.db')
DB_PATH  = DB_TEMPLATE_PATH
PORT     = int(os.environ.get('PORT', 3000))
PUBLIC_DOMAIN = os.environ.get('PUBLIC_DOMAIN', 'https://www.calculoobras.com.br').rstrip('/')
STRIPE_PRICE_ID = os.environ.get('STRIPE_PRICE_ID', '')
STRIPE_SECRET_KEY = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')

if not os.path.exists(DB_TEMPLATE_PATH):
    print(f"\n⚠️  Banco-template não encontrado em:\n    {DB_TEMPLATE_PATH}")
    print("    O app inicia, mas novos tenants exigem o upload desse arquivo.\n")

if stripe and STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY

def get_db():
    db_path = session.get('tenant_db_path') or getattr(request, 'tenant_db_path', None)
    if not db_path:
        raise RuntimeError('Tenant não definido para esta requisição.')
    if not os.path.abspath(db_path).startswith(os.path.abspath(TENANT_DB_DIR)):
        raise RuntimeError('Caminho de banco de dados de tenant inválido.')
    conn = sqlite3.connect(db_path, timeout=60)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA ignore_check_constraints = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    return conn

def rows_to_list(rows):
    return [dict(r) for r in rows]

def _table_exists(db, table_name):
    return db.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        [table_name]
    ).fetchone() is not None

def _ensure_columns(db, table_name, columns):
    if not _table_exists(db, table_name):
        return
    existentes = {r['name'] for r in db.execute(f"PRAGMA table_info({table_name})").fetchall()}
    for nome, ddl in columns.items():
        if nome not in existentes:
            db.execute(f"ALTER TABLE {table_name} ADD COLUMN {nome} {ddl}")

def get_master_db():
    conn = sqlite3.connect(MASTER_DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_saas_master():
    os.makedirs(os.path.dirname(MASTER_DB_PATH), exist_ok=True)
    os.makedirs(TENANT_DB_DIR, exist_ok=True)
    db = get_master_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS tenants (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            db_path TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            created_at INTEGER NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tenant_id TEXT NOT NULL REFERENCES tenants(id) ON DELETE CASCADE,
            email TEXT NOT NULL UNIQUE,
            nome TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'owner',
            is_active INTEGER NOT NULL DEFAULT 0,
            stripe_customer_id TEXT,
            created_at INTEGER NOT NULL
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            stripe_subscription_id TEXT UNIQUE,
            stripe_customer_id TEXT,
            status TEXT NOT NULL DEFAULT 'incomplete',
            current_period_end INTEGER,
            created_at INTEGER NOT NULL,
            updated_at INTEGER NOT NULL
        )
    """)
    admin_email = os.environ.get('ORCASMART_ADMIN_EMAIL')
    admin_password = os.environ.get('ORCASMART_ADMIN_PASSWORD')
    if admin_email and admin_password and os.path.exists(DB_TEMPLATE_PATH):
        row = db.execute("SELECT id FROM users WHERE email=?", [admin_email.lower()]).fetchone()
        if not row:
            tenant_id, tenant_db = create_tenant_database('admin')
            db.execute("INSERT INTO tenants (id,nome,db_path,status,created_at) VALUES (?,?,?,?,?)",
                       [tenant_id, 'Administrador', tenant_db, 'active', int(time.time())])
            db.execute("""INSERT INTO users
                (tenant_id,email,nome,password_hash,role,is_active,created_at)
                VALUES (?,?,?,?,?,?,?)""",
                [tenant_id, admin_email.lower(), 'Administrador',
                 generate_password_hash(admin_password), 'admin', 1, int(time.time())])
    elif admin_email and admin_password:
        print('⚠️  Admin inicial não criado porque o banco-template ainda não foi enviado.')
    db.commit()
    db.close()

def create_tenant_database(prefix='tenant'):
    if not os.path.exists(DB_TEMPLATE_PATH):
        raise FileNotFoundError(
            f'Banco-template não encontrado: {DB_TEMPLATE_PATH}. '
            'Envie database/orcamento_obras_template.db para o servidor antes de criar usuários.'
        )
    tenant_id = uuid.uuid4().hex
    safe_prefix = re.sub(r'[^a-zA-Z0-9_-]+', '_', prefix or 'tenant')[:32]
    db_path = os.path.join(TENANT_DB_DIR, f'{safe_prefix}_{tenant_id}.db')
    shutil.copyfile(DB_TEMPLATE_PATH, db_path)
    return tenant_id, db_path

def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    db = get_master_db()
    row = db.execute("""SELECT u.*, t.db_path, t.status AS tenant_status
                        FROM users u JOIN tenants t ON t.id=u.tenant_id
                        WHERE u.id=?""", [uid]).fetchone()
    db.close()
    return dict(row) if row else None

def subscription_allows_access(user):
    return bool(user and (user.get('role') == 'admin' or user.get('is_active') or user.get('tenant_status') == 'active'))

def login_user(user):
    session.clear()
    session['user_id'] = user['id']
    session['tenant_id'] = user['tenant_id']
    session['tenant_db_path'] = user['db_path']
    session['role'] = user['role']

def api_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user:
            return jsonify({'erro': 'Autenticação necessária.'}), 401
        if not subscription_allows_access(user):
            return jsonify({'erro': 'Assinatura inativa.', 'code': 'subscription_required'}), 402
        request.tenant_db_path = user['db_path']
        return fn(*args, **kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user or user.get('role') != 'admin':
            return jsonify({'erro': 'Acesso administrativo necessário.'}), 403
        return fn(*args, **kwargs)
    return wrapper

def ensure_municipio_aliquotas_table(db):
    db.execute("""
        CREATE TABLE IF NOT EXISTS municipio_aliquotas_anuais (
            id_aliquota INTEGER PRIMARY KEY AUTOINCREMENT,
            id_municipio INTEGER NOT NULL,
            ano INTEGER NOT NULL,
            iva_percentual REAL NOT NULL DEFAULT 0.0,
            aliquota_cbs REAL NOT NULL DEFAULT 0.0,
            aliquota_ibs REAL NOT NULL DEFAULT 0.0,
            aliquota_iss REAL NOT NULL DEFAULT 0.0,
            data_atualizacao TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (id_municipio) REFERENCES municipios(id_municipio) ON DELETE CASCADE,
            UNIQUE(id_municipio, ano)
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_municipio_aliquotas_anuais_ano ON municipio_aliquotas_anuais(ano)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_municipio_aliquotas_anuais_municipio ON municipio_aliquotas_anuais(id_municipio)")

def ensure_obras_reforma_fields(db):
    cols = {r['name'] for r in db.execute("PRAGMA table_info(obras)").fetchall()}
    if 'ano_realizacao' not in cols:
        db.execute("ALTER TABLE obras ADD COLUMN ano_realizacao INTEGER")
    if 'fator_setorial' not in cols:
        db.execute("ALTER TABLE obras ADD COLUMN fator_setorial REAL NOT NULL DEFAULT 0.0")
    if 'redutor_compras_governamentais' not in cols:
        db.execute("ALTER TABLE obras ADD COLUMN redutor_compras_governamentais REAL NOT NULL DEFAULT 0.0")

def ensure_orcamentos_reforma_fields(db):
    cols = {r['name'] for r in db.execute("PRAGMA table_info(orcamentos)").fetchall()}
    if 'regime_previdenciario' not in cols:
        db.execute("ALTER TABLE orcamentos ADD COLUMN regime_previdenciario TEXT NOT NULL DEFAULT 'Onerado'")

def ensure_orcamento_sintetico_bdi_linha(db):
    cols = {r['name'] for r in db.execute("PRAGMA table_info(orcamento_sintetico)").fetchall()}
    if 'bdi_percentual_linha' not in cols:
        db.execute("ALTER TABLE orcamento_sintetico ADD COLUMN bdi_percentual_linha REAL")

def ensure_eventograma_schema(db):
    db.executescript("""
        CREATE TABLE IF NOT EXISTS eventogramas (
            id_eventograma   INTEGER PRIMARY KEY AUTOINCREMENT,
            id_orcamento     INTEGER NOT NULL REFERENCES orcamentos(id_orcamento) ON DELETE CASCADE,
            nome             TEXT    NOT NULL,
            descricao        TEXT,
            modo_geracao     TEXT    DEFAULT 'manual',
            status           TEXT    DEFAULT 'Rascunho',
            valor_total_ref  REAL    DEFAULT 0,
            observacoes      TEXT,
            data_criacao     TEXT    DEFAULT (date('now')),
            data_atualizacao TEXT    DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS ev_eventos (
            id_evento        INTEGER PRIMARY KEY AUTOINCREMENT,
            id_eventograma   INTEGER NOT NULL REFERENCES eventogramas(id_eventograma) ON DELETE CASCADE,
            id_evento_pai    INTEGER REFERENCES ev_eventos(id_evento) ON DELETE CASCADE,
            numero_evento    TEXT    NOT NULL,
            descricao        TEXT    NOT NULL,
            grupo            TEXT,
            criterio_medicao TEXT,
            condicao_pagamento TEXT,
            prazo_marco      TEXT,
            docs_comprobatorios TEXT,
            observacoes      TEXT,
            valor_calculado  REAL    DEFAULT 0,
            ordem            INTEGER DEFAULT 0
        );

        CREATE INDEX IF NOT EXISTS idx_ev_eventos_evgrama ON ev_eventos(id_eventograma);
        CREATE INDEX IF NOT EXISTS idx_ev_eventos_pai     ON ev_eventos(id_evento_pai);

        CREATE TABLE IF NOT EXISTS ev_evento_itens (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            id_evento INTEGER NOT NULL REFERENCES ev_eventos(id_evento) ON DELETE CASCADE,
            id_item   INTEGER NOT NULL REFERENCES orcamento_sintetico(id_item) ON DELETE CASCADE,
            UNIQUE(id_evento, id_item)
        );

        CREATE INDEX IF NOT EXISTS idx_ev_itens_evento ON ev_evento_itens(id_evento);
        CREATE INDEX IF NOT EXISTS idx_ev_itens_item   ON ev_evento_itens(id_item);
    """)
    _ensure_columns(db, 'eventogramas', {
        'descricao': 'TEXT',
        'modo_geracao': "TEXT DEFAULT 'manual'",
        'status': "TEXT DEFAULT 'Rascunho'",
        'valor_total_ref': 'REAL DEFAULT 0',
        'observacoes': 'TEXT',
        'data_criacao': "TEXT DEFAULT (date('now'))",
        'data_atualizacao': "TEXT DEFAULT (datetime('now'))",
    })
    _ensure_columns(db, 'ev_eventos', {
        'id_evento_pai': 'INTEGER REFERENCES ev_eventos(id_evento) ON DELETE CASCADE',
        'grupo': 'TEXT',
        'criterio_medicao': 'TEXT',
        'condicao_pagamento': 'TEXT',
        'prazo_marco': 'TEXT',
        'docs_comprobatorios': 'TEXT',
        'observacoes': 'TEXT',
        'valor_calculado': 'REAL DEFAULT 0',
        'ordem': 'INTEGER DEFAULT 0',
    })

def ensure_pem_schema(db):
    db.executescript("""
        CREATE TABLE IF NOT EXISTS pem_servicos (
            id_pem          INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo          TEXT NOT NULL UNIQUE,
            servico         TEXT NOT NULL,
            producao_equipe REAL,
            unidade         TEXT,
            observacoes     TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_pem_codigo ON pem_servicos(codigo);

        CREATE TABLE IF NOT EXISTS pem_equipamentos (
            id_pem_equip           INTEGER PRIMARY KEY AUTOINCREMENT,
            id_pem                 INTEGER NOT NULL REFERENCES pem_servicos(id_pem) ON DELETE CASCADE,
            codigo_equip           TEXT,
            descricao_equip        TEXT,
            formula                TEXT,
            producao_horaria       REAL,
            num_unidades           REAL DEFAULT 1.0,
            utilizacao_operativa   REAL DEFAULT 1.0,
            utilizacao_improdutiva REAL DEFAULT 0.0,
            ordem                  INTEGER DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_pem_eq_pem ON pem_equipamentos(id_pem);

        CREATE TABLE IF NOT EXISTS pem_variaveis (
            id_var        INTEGER PRIMARY KEY AUTOINCREMENT,
            id_pem_equip  INTEGER NOT NULL REFERENCES pem_equipamentos(id_pem_equip) ON DELETE CASCADE,
            letra         TEXT NOT NULL,
            nome_variavel TEXT NOT NULL,
            unidade       TEXT,
            valor         REAL,
            UNIQUE(id_pem_equip, letra)
        );
        CREATE INDEX IF NOT EXISTS idx_pem_var_equip ON pem_variaveis(id_pem_equip);
    """)

def ensure_encargos_schema(db):
    db.executescript("""
        CREATE TABLE IF NOT EXISTS perfis_encargos (
            id_perfil          INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_perfil        TEXT NOT NULL,
            categoria          TEXT NOT NULL DEFAULT 'Horista',
            regime             TEXT NOT NULL DEFAULT 'Normal',
            uf_referencia      TEXT,
            id_data_base       INTEGER REFERENCES datas_base(id_data_base),
            descricao          TEXT,
            total_grupo_a      REAL DEFAULT 0,
            total_grupo_b      REAL DEFAULT 0,
            total_grupo_c      REAL DEFAULT 0,
            total_grupo_d      REAL DEFAULT 0,
            encargo_total      REAL DEFAULT 0,
            observacoes        TEXT,
            situacao           TEXT DEFAULT 'Ativo'
        );

        CREATE TABLE IF NOT EXISTS grupos_encargos (
            id_grupo_enc    INTEGER PRIMARY KEY AUTOINCREMENT,
            id_perfil       INTEGER NOT NULL REFERENCES perfis_encargos(id_perfil) ON DELETE CASCADE,
            letra           TEXT NOT NULL,
            descricao       TEXT,
            total_grupo     REAL DEFAULT 0,
            UNIQUE(id_perfil, letra)
        );

        CREATE TABLE IF NOT EXISTS itens_encargo (
            id_item         INTEGER PRIMARY KEY AUTOINCREMENT,
            id_grupo_enc    INTEGER NOT NULL REFERENCES grupos_encargos(id_grupo_enc) ON DELETE CASCADE,
            descricao       TEXT NOT NULL,
            base_legal      TEXT,
            percentual      REAL NOT NULL DEFAULT 0,
            observacoes     TEXT,
            ordem           INTEGER DEFAULT 0
        );

        CREATE INDEX IF NOT EXISTS idx_itens_grupo ON itens_encargo(id_grupo_enc);
    """)
    cols = {r['name'] for r in db.execute("PRAGMA table_info(perfis_encargos)").fetchall()}
    if 'fonte_referencia' not in cols:
        db.execute("ALTER TABLE perfis_encargos ADD COLUMN fonte_referencia TEXT NOT NULL DEFAULT 'SINAPI'")
    if 'vigencia_inicio' not in cols:
        db.execute("ALTER TABLE perfis_encargos ADD COLUMN vigencia_inicio TEXT")
    if 'vigencia_fim' not in cols:
        db.execute("ALTER TABLE perfis_encargos ADD COLUMN vigencia_fim TEXT")
    if 'encargo_original_percentual' not in cols:
        db.execute("ALTER TABLE perfis_encargos ADD COLUMN encargo_original_percentual REAL")
    db.execute("""
        UPDATE perfis_encargos
           SET fonte_referencia = COALESCE(NULLIF(fonte_referencia,''), 'SINAPI'),
               vigencia_inicio = COALESCE(vigencia_inicio, '2026-01-01'),
               vigencia_fim = COALESCE(vigencia_fim, '2026-12-31'),
               vigencia = CASE
                   WHEN vigencia IS NULL OR vigencia='' OR vigencia='01/2026' THEN '01/2026 a 12/2026'
                   ELSE vigencia
               END,
               encargo_original_percentual = COALESCE(encargo_original_percentual, encargo_total)
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS encargos_orcamento_aplicacoes (
            id_aplicacao INTEGER PRIMARY KEY AUTOINCREMENT,
            id_orcamento INTEGER NOT NULL,
            id_perfil INTEGER NOT NULL,
            encargo_novo_percentual REAL NOT NULL DEFAULT 0,
            itens_atualizados INTEGER NOT NULL DEFAULT 0,
            custo_antes REAL NOT NULL DEFAULT 0,
            custo_depois REAL NOT NULL DEFAULT 0,
            data_aplicacao TEXT DEFAULT (datetime('now')),
            observacoes TEXT,
            FOREIGN KEY (id_orcamento) REFERENCES orcamentos(id_orcamento) ON DELETE CASCADE,
            FOREIGN KEY (id_perfil) REFERENCES perfis_encargos(id_perfil) ON DELETE RESTRICT
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_encargos_aplic_orc ON encargos_orcamento_aplicacoes(id_orcamento)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_encargos_aplic_perfil ON encargos_orcamento_aplicacoes(id_perfil)")
    db.execute("""
        CREATE TABLE IF NOT EXISTS encargos_sicro_profissionais (
            id_profissional_enc INTEGER PRIMARY KEY AUTOINCREMENT,
            id_perfil INTEGER NOT NULL,
            codigo_profissional TEXT NOT NULL,
            descricao TEXT NOT NULL,
            unidade TEXT,
            total_grupo_a REAL NOT NULL DEFAULT 0,
            total_grupo_b REAL NOT NULL DEFAULT 0,
            total_grupo_c REAL NOT NULL DEFAULT 0,
            total_grupo_d REAL NOT NULL DEFAULT 0,
            encargo_total REAL NOT NULL DEFAULT 0,
            parcelas_json TEXT,
            FOREIGN KEY (id_perfil) REFERENCES perfis_encargos(id_perfil) ON DELETE CASCADE,
            UNIQUE (id_perfil, codigo_profissional)
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_enc_sicro_perfil ON encargos_sicro_profissionais(id_perfil)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_enc_sicro_codigo ON encargos_sicro_profissionais(codigo_profissional)")
    db.execute("""
        CREATE TABLE IF NOT EXISTS encargos_goinfra_profissionais (
            id_profissional_enc INTEGER PRIMARY KEY AUTOINCREMENT,
            id_perfil INTEGER NOT NULL,
            codigo_profissional TEXT NOT NULL,
            descricao TEXT NOT NULL,
            unidade TEXT,
            total_grupo_a REAL NOT NULL DEFAULT 0,
            total_grupo_b REAL NOT NULL DEFAULT 0,
            total_grupo_c REAL NOT NULL DEFAULT 0,
            total_grupo_d REAL NOT NULL DEFAULT 0,
            encargo_total REAL NOT NULL DEFAULT 0,
            parcelas_json TEXT,
            FOREIGN KEY (id_perfil) REFERENCES perfis_encargos(id_perfil) ON DELETE CASCADE,
            UNIQUE (id_perfil, codigo_profissional)
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_enc_goinfra_perfil ON encargos_goinfra_profissionais(id_perfil)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_enc_goinfra_codigo ON encargos_goinfra_profissionais(codigo_profissional)")

def ensure_insumos_encargos_schema(db):
    cols_ins = {r['name'] for r in db.execute("PRAGMA table_info(insumos)").fetchall()}
    if 'encargos_sociais_percentual' not in cols_ins:
        db.execute("ALTER TABLE insumos ADD COLUMN encargos_sociais_percentual REAL")
    cols_precos = {r['name'] for r in db.execute("PRAGMA table_info(precos_insumos)").fetchall()}
    if 'encargos_sociais_percentual' not in cols_precos:
        db.execute("ALTER TABLE precos_insumos ADD COLUMN encargos_sociais_percentual REAL")
    db.execute("CREATE INDEX IF NOT EXISTS idx_precos_insumos_enc_soc ON precos_insumos(id_insumo, uf_referencia, id_data_base)")

def _insumo_categoria_mao_obra(row):
    desc = _norm_ascii((row.get('descricao') if hasattr(row, 'get') else row['descricao']) or '')
    sigla = _norm_ascii((row.get('sigla_unidade') if hasattr(row, 'get') else row['sigla_unidade']) or '')
    if 'MENSALISTA' in desc or sigla in ('MES', 'MENSAL', 'MENSALISTA'):
        return 'Mensalista'
    return 'Horista'

def _preco_regime_encargos(row):
    pref = float(row['preco_referencia'] or 0)
    pdes = float(row['preco_desonerado'] or 0)
    pon = float(row['preco_nao_desonerado'] or 0)
    if pdes > 0 and (pon <= 0 or abs(pref - pdes) < 0.0001):
        return 'Desonerado'
    return 'Normal'

def _fonte_insumo_encargos(origem):
    fonte = _norm_ascii(origem or '')
    if fonte.startswith('SEINFRA'):
        return 'SEINFRA'
    if fonte.startswith('SUDECAP'):
        return 'SUDECAP'
    if fonte.startswith('SICRO'):
        return 'SICRO'
    if fonte.startswith('GOINFRA'):
        return 'GOINFRA'
    if fonte.startswith('CDHU'):
        return 'CDHU'
    if fonte.startswith('SINAPI'):
        return 'SINAPI'
    return fonte

def _codigo_profissional_referencial(codigo):
    txt = (codigo or '').strip().upper()
    m = re.search(r'\b(P\d+)\b', txt)
    if m:
        return m.group(1)
    return (
        txt.replace('SINAPI.', '')
           .replace('SICRO.', '')
           .replace('SEINFRA.', '')
           .replace('SUDECAP.', '')
           .replace('GOINFRA.', '')
           .replace('CDHU.', '')
           .replace('USUARIO.', '')
           .strip()
    )

def _codigo_sicro_profissional(codigo):
    return _codigo_profissional_referencial(codigo)

def _buscar_perfil_encargos(db, fonte, uf, ano, mes, categoria, regime):
    if not fonte or not ano or not mes:
        return None
    data_ref = f"{int(ano):04d}-{int(mes):02d}-01"
    fonte_like = f"{fonte}%"
    perfil = db.execute("""
        SELECT * FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,'')) LIKE ?
           AND UPPER(COALESCE(uf_referencia,'')) = ?
           AND categoria = ?
           AND regime = ?
           AND COALESCE(vigencia_inicio, '1900-01-01') <= ?
           AND COALESCE(vigencia_fim, '2999-12-31') >= ?
         ORDER BY id_perfil DESC
         LIMIT 1
    """, [fonte_like, (uf or '').upper(), categoria, regime, data_ref, data_ref]).fetchone()
    if perfil or fonte != 'SICRO':
        return perfil
    return db.execute("""
        SELECT * FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,'')) LIKE ?
           AND categoria = ?
           AND regime = ?
           AND COALESCE(vigencia_inicio, '1900-01-01') <= ?
           AND COALESCE(vigencia_fim, '2999-12-31') >= ?
         ORDER BY
           CASE WHEN UPPER(COALESCE(uf_referencia,''))=? THEN 0 ELSE 1 END,
           id_perfil DESC
         LIMIT 1
    """, [fonte_like, categoria, regime, data_ref, data_ref, (uf or '').upper()]).fetchone()

def sincronizar_encargos_sociais_insumos(db, fontes=None):
    ensure_encargos_schema(db)
    ensure_insumos_encargos_schema(db)
    fontes_norm = {_fonte_insumo_encargos(f) for f in (fontes or []) if f}
    rows = db.execute("""
        SELECT p.id_preco, i.id_insumo, i.codigo_insumo, i.descricao, i.tipo_insumo, i.origem,
               um.sigla AS sigla_unidade, p.uf_referencia, p.preco_referencia,
               p.preco_desonerado, p.preco_nao_desonerado, db2.mes, db2.ano
          FROM precos_insumos p
          JOIN insumos i ON i.id_insumo = p.id_insumo
          LEFT JOIN unidades_medida um ON um.id_unidade = i.id_unidade
          LEFT JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
         WHERE UPPER(i.tipo_insumo) LIKE '%OBRA%'
    """).fetchall()
    atualizados = 0
    cache_perfis = {}
    for row in rows:
        fonte = _fonte_insumo_encargos(row['origem'])
        if fontes_norm and fonte not in fontes_norm:
            continue
        uf = (row['uf_referencia'] or '').upper()
        categoria = _insumo_categoria_mao_obra(row)
        regime = _preco_regime_encargos(row)
        key = (fonte, uf, row['ano'], row['mes'], categoria, regime)
        if key not in cache_perfis:
            cache_perfis[key] = _buscar_perfil_encargos(db, fonte, uf, row['ano'], row['mes'], categoria, regime)
        perfil = cache_perfis[key]
        if not perfil:
            continue
        encargo = None
        if fonte in ('SICRO', 'GOINFRA'):
            tabela_prof = 'encargos_sicro_profissionais' if fonte == 'SICRO' else 'encargos_goinfra_profissionais'
            codigo_prof = _codigo_profissional_referencial(row['codigo_insumo'])
            prof = db.execute("""
                SELECT encargo_total
                  FROM """ + tabela_prof + """
                 WHERE id_perfil=? AND UPPER(codigo_profissional)=?
                 LIMIT 1
            """, [perfil['id_perfil'], codigo_prof]).fetchone()
            if prof:
                encargo = float(prof['encargo_total'] or 0)
        else:
            encargo = float(perfil['encargo_total'] or 0)
        if encargo is None:
            continue
        db.execute("UPDATE precos_insumos SET encargos_sociais_percentual=? WHERE id_preco=?", [encargo, row['id_preco']])
        atualizados += 1
    db.execute("""
        UPDATE insumos
           SET encargos_sociais_percentual = (
               SELECT p.encargos_sociais_percentual
                 FROM precos_insumos p
                WHERE p.id_insumo = insumos.id_insumo
                  AND p.encargos_sociais_percentual IS NOT NULL
                ORDER BY p.id_preco DESC
                LIMIT 1
           )
         WHERE UPPER(tipo_insumo) LIKE '%OBRA%'
           AND EXISTS (
               SELECT 1 FROM precos_insumos p
                WHERE p.id_insumo = insumos.id_insumo
                  AND p.encargos_sociais_percentual IS NOT NULL
           )
    """)
    return atualizados

def ensure_bdi_reforma_fields(db):
    cols = {r['name'] for r in db.execute("PRAGMA table_info(perfis_bdi)").fetchall()}
    novos = {
        'ano_orcamento': 'INTEGER',
        'quartil': 'TEXT',
        'cbs_percentual': 'REAL NOT NULL DEFAULT 0.0',
        'ibs_percentual': 'REAL NOT NULL DEFAULT 0.0',
        'fator_efetivo_ivaeq': 'REAL NOT NULL DEFAULT 0.5',
        'percentual_mat_ivaeq': 'REAL NOT NULL DEFAULT 0.4',
        'credito_bdi_ivaeq': 'REAL NOT NULL DEFAULT 0.0',
        'ivaeq_percentual': 'REAL NOT NULL DEFAULT 0.0',
        'iss_percentual_manual': 'REAL',
        'id_orcamento_ivaeq': 'INTEGER',
        'regime_previdenciario': "TEXT NOT NULL DEFAULT 'Onerado'",
        'simples_faixa': 'INTEGER',
        'simples_faixa_label': 'TEXT',
        'simples_receita_limite': 'REAL',
        'simples_aliquota_efetiva': 'REAL',
        'simples_irpj_percentual': 'REAL NOT NULL DEFAULT 0.0',
        'simples_csll_percentual': 'REAL NOT NULL DEFAULT 0.0',
    }
    for col, ddl in novos.items():
        if col not in cols:
            db.execute(f"ALTER TABLE perfis_bdi ADD COLUMN {col} {ddl}")

def _float_or_zero(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0

def _ano_tributario_orcamento(orc):
    ano = orc.get('ano_realizacao') or orc.get('data_base_ano')
    try:
        return int(ano)
    except (TypeError, ValueError):
        return 2026

# ─── APP ───────────────────────────────────────────────────────────────────────
app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')
# Flask 3.x: configurar JSON para não escapar Unicode e suportar acentos
app.json.ensure_ascii = False
app.json.sort_keys    = False
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('SECRET_KEY') or uuid.uuid4().hex
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE', '1') != '0'
CORS(app, supports_credentials=True)
init_saas_master()

@app.errorhandler(Exception)
def api_exception_handler(exc):
    if request.path.startswith('/api/'):
        if isinstance(exc, HTTPException):
            return jsonify({
                'erro': exc.description,
                'tipo': exc.__class__.__name__,
                'rota': request.path,
            }), exc.code
        app.logger.exception("Erro na API %s %s", request.method, request.path)
        detalhe = traceback.format_exc(limit=8)
        return jsonify({
            'erro': str(exc) or exc.__class__.__name__,
            'tipo': exc.__class__.__name__,
            'rota': request.path,
            'detalhe': detalhe[-1600:],
        }), 500
    raise exc

def inicializar_schemas_tenant_runtime():
    db = get_db()
    try:
        ensure_municipio_aliquotas_table(db)
        ensure_obras_reforma_fields(db)
        ensure_orcamentos_reforma_fields(db)
        ensure_orcamento_sintetico_bdi_linha(db)
        ensure_encargos_schema(db)
        ensure_insumos_encargos_schema(db)
        ensure_eventograma_schema(db)
        ensure_pem_schema(db)
        db.commit()
    finally:
        db.close()

@app.before_request
def saas_auth_gate():
    public_api_prefixes = (
        '/api/auth/',
        '/api/billing/',
        '/api/stripe/webhook',
        '/api/status',
    )
    public_paths = (
        '/login.html',
        '/css/',
        '/js/',
        '/img/',
        '/favicon.ico',
    )
    if request.path == '/' or any(request.path.startswith(p) for p in public_paths):
        return None
    if request.path.startswith('/api/'):
        if any(request.path.startswith(p) for p in public_api_prefixes):
            return None
        user = current_user()
        if not user:
            return jsonify({'erro': 'Autenticação necessária.'}), 401
        if not subscription_allows_access(user):
            return jsonify({'erro': 'Assinatura inativa.', 'code': 'subscription_required'}), 402
        request.tenant_db_path = user['db_path']
        inicializar_schemas_tenant_runtime()
    return None

@app.after_request
def add_no_cache_headers(resp):
    if request.path == '/' or request.path.startswith(('/js/', '/css/')):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
    return resp

# ─── STATIC + SPA ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if not session.get('user_id'):
        return redirect('/login.html')
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/login.html')
def login_page():
    return send_from_directory(BASE_DIR, 'login.html')

@app.route('/css/<path:p>')
def css(p): return send_from_directory(os.path.join(BASE_DIR,'css'), p)

@app.route('/img/<path:p>')
def img(p): return send_from_directory(os.path.join(BASE_DIR,'img'), p)

@app.route('/js/<path:p>')
def js(p):  return send_from_directory(os.path.join(BASE_DIR,'js'), p)

# ─── STATUS + DASHBOARD ───────────────────────────────────────────────────────
@app.route('/api/status')
def status():
    return jsonify({'status':'ok','version':'1.0.0-saas','domain':PUBLIC_DOMAIN})

@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    user = current_user()
    if not user:
        return jsonify({'authenticated': False}), 200
    return jsonify({
        'authenticated': True,
        'id': user['id'],
        'email': user['email'],
        'nome': user.get('nome'),
        'role': user.get('role'),
        'tenant_id': user.get('tenant_id'),
        'active': subscription_allows_access(user),
        'subscription_required': not subscription_allows_access(user),
    })

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    d = request.json or {}
    email = (d.get('email') or '').strip().lower()
    password = d.get('password') or ''
    db = get_master_db()
    row = db.execute("""SELECT u.*, t.db_path, t.status AS tenant_status
                        FROM users u JOIN tenants t ON t.id=u.tenant_id
                        WHERE u.email=?""", [email]).fetchone()
    db.close()
    if not row or not check_password_hash(row['password_hash'], password):
        return jsonify({'erro': 'E-mail ou senha inválidos.'}), 401
    user = dict(row)
    login_user(user)
    if not subscription_allows_access(user):
        return jsonify({'ok': True, 'subscription_required': True}), 402
    return jsonify({'ok': True, 'user': {'email': user['email'], 'role': user['role'], 'nome': user.get('nome')}})

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    d = request.json or {}
    email = (d.get('email') or '').strip().lower()
    password = d.get('password') or ''
    nome = (d.get('nome') or email).strip()
    if not email or '@' not in email:
        return jsonify({'erro': 'Informe um e-mail válido.'}), 400
    if len(password) < 8:
        return jsonify({'erro': 'A senha deve ter pelo menos 8 caracteres.'}), 400
    db = get_master_db()
    if db.execute("SELECT id FROM users WHERE email=?", [email]).fetchone():
        db.close()
        return jsonify({'erro': 'E-mail já cadastrado.'}), 409
    tenant_id, tenant_db = create_tenant_database(email.split('@')[0])
    now = int(time.time())
    db.execute("INSERT INTO tenants (id,nome,db_path,status,created_at) VALUES (?,?,?,?,?)",
               [tenant_id, nome, tenant_db, 'pending', now])
    cur = db.execute("""INSERT INTO users
        (tenant_id,email,nome,password_hash,role,is_active,created_at)
        VALUES (?,?,?,?,?,?,?)""",
        [tenant_id, email, nome, generate_password_hash(password), 'owner', 0, now])
    user_id = cur.lastrowid
    db.commit()
    user = dict(db.execute("""SELECT u.*, t.db_path, t.status AS tenant_status
                              FROM users u JOIN tenants t ON t.id=u.tenant_id
                              WHERE u.id=?""", [user_id]).fetchone())
    db.close()
    login_user(user)
    checkout = create_stripe_checkout_for_user(user)
    return jsonify({
        'ok': True,
        'user_id': user_id,
        'subscription_required': True,
        'checkout_url': checkout.get('url'),
        'stripe_configured': checkout.get('configured'),
    }), 201

def create_stripe_checkout_for_user(user):
    if not (stripe and STRIPE_SECRET_KEY and STRIPE_PRICE_ID):
        return {'configured': False, 'url': None}
    session_obj = stripe.checkout.Session.create(
        mode='subscription',
        customer_email=user['email'],
        client_reference_id=str(user['id']),
        line_items=[{'price': STRIPE_PRICE_ID, 'quantity': 1}],
        success_url=f'{PUBLIC_DOMAIN}/login.html?checkout=success',
        cancel_url=f'{PUBLIC_DOMAIN}/login.html?checkout=cancel',
        metadata={'user_id': str(user['id']), 'tenant_id': user['tenant_id']},
    )
    return {'configured': True, 'url': session_obj.url}

@app.route('/api/billing/create-checkout-session', methods=['POST'])
def billing_checkout():
    user = current_user()
    if not user:
        return jsonify({'erro': 'Autenticação necessária.'}), 401
    checkout = create_stripe_checkout_for_user(user)
    if not checkout.get('configured'):
        return jsonify({'erro': 'Stripe não configurado no servidor.'}), 503
    return jsonify({'url': checkout['url']})

@app.route('/api/billing/create-portal-session', methods=['POST'])
def billing_portal():
    user = current_user()
    if not user:
        return jsonify({'erro': 'Autenticação necessária.'}), 401
    if not (stripe and STRIPE_SECRET_KEY):
        return jsonify({'erro': 'Stripe não configurado no servidor.'}), 503
    if not user.get('stripe_customer_id'):
        return jsonify({'erro': 'Cliente Stripe ainda não vinculado.'}), 400
    portal = stripe.billing_portal.Session.create(
        customer=user['stripe_customer_id'],
        return_url=f'{PUBLIC_DOMAIN}/',
    )
    return jsonify({'url': portal.url})

@app.route('/api/stripe/webhook', methods=['POST'])
def stripe_webhook():
    if not (stripe and STRIPE_WEBHOOK_SECRET):
        return jsonify({'erro': 'Webhook Stripe não configurado.'}), 503
    payload = request.get_data()
    sig_header = request.headers.get('Stripe-Signature')
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        return jsonify({'erro': str(e)}), 400
    handle_stripe_event(event)
    return jsonify({'received': True})

def handle_stripe_event(event):
    typ = event.get('type')
    obj = event['data']['object']
    db = get_master_db()
    now = int(time.time())
    try:
        if typ == 'checkout.session.completed':
            user_id = int(obj.get('client_reference_id') or obj.get('metadata', {}).get('user_id') or 0)
            customer = obj.get('customer')
            sub = obj.get('subscription')
            if user_id:
                db.execute("UPDATE users SET stripe_customer_id=?, is_active=1 WHERE id=?", [customer, user_id])
                db.execute("""UPDATE tenants SET status='active'
                              WHERE id=(SELECT tenant_id FROM users WHERE id=?)""", [user_id])
                db.execute("""INSERT OR IGNORE INTO subscriptions
                    (user_id,stripe_subscription_id,stripe_customer_id,status,created_at,updated_at)
                    VALUES (?,?,?,?,?,?)""", [user_id, sub, customer, 'active', now, now])
        elif typ in ('customer.subscription.updated', 'customer.subscription.deleted'):
            sub = obj.get('id')
            customer = obj.get('customer')
            status = obj.get('status') or 'inactive'
            active = 1 if status in ('active', 'trialing') else 0
            user = db.execute("SELECT id, tenant_id FROM users WHERE stripe_customer_id=?", [customer]).fetchone()
            if user:
                db.execute("UPDATE users SET is_active=? WHERE id=?", [active, user['id']])
                db.execute("UPDATE tenants SET status=? WHERE id=?", ['active' if active else 'inactive', user['tenant_id']])
                db.execute("""INSERT INTO subscriptions
                    (user_id,stripe_subscription_id,stripe_customer_id,status,current_period_end,created_at,updated_at)
                    VALUES (?,?,?,?,?,?,?)
                    ON CONFLICT(stripe_subscription_id) DO UPDATE SET
                      status=excluded.status,
                      current_period_end=excluded.current_period_end,
                      updated_at=excluded.updated_at""",
                    [user['id'], sub, customer, status, obj.get('current_period_end'), now, now])
        db.commit()
    finally:
        db.close()

@app.route('/api/admin/users', methods=['GET'])
@admin_required
def admin_users():
    db = get_master_db()
    rows = rows_to_list(db.execute("""SELECT u.id,u.email,u.nome,u.role,u.is_active,u.created_at,
                                             t.id AS tenant_id,t.status AS tenant_status,t.db_path
                                      FROM users u JOIN tenants t ON t.id=u.tenant_id
                                      ORDER BY u.created_at DESC""").fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/admin/users', methods=['POST'])
@admin_required
def admin_create_user():
    d = request.json or {}
    email = (d.get('email') or '').strip().lower()
    password = d.get('password') or ''
    nome = d.get('nome') or email
    role = d.get('role') or 'owner'
    active = 1 if d.get('is_active', True) else 0
    if not email or len(password) < 8:
        return jsonify({'erro': 'Informe e-mail e senha com pelo menos 8 caracteres.'}), 400
    db = get_master_db()
    if db.execute("SELECT id FROM users WHERE email=?", [email]).fetchone():
        db.close()
        return jsonify({'erro': 'E-mail já cadastrado.'}), 409
    tenant_id, tenant_db = create_tenant_database(email.split('@')[0])
    now = int(time.time())
    db.execute("INSERT INTO tenants (id,nome,db_path,status,created_at) VALUES (?,?,?,?,?)",
               [tenant_id, nome, tenant_db, 'active' if active else 'pending', now])
    cur = db.execute("""INSERT INTO users
        (tenant_id,email,nome,password_hash,role,is_active,created_at)
        VALUES (?,?,?,?,?,?,?)""",
        [tenant_id, email, nome, generate_password_hash(password), role, active, now])
    db.commit()
    db.close()
    return jsonify({'id': cur.lastrowid, 'email': email, 'tenant_id': tenant_id}), 201

@app.route('/api/dashboard')
def dashboard():
    db = get_db()
    def safe_count(sql):
        try: return db.execute(sql).fetchone()[0]
        except: return 0
    result = {
        'totalObras':       safe_count('SELECT COUNT(*) FROM obras'),
        'totalOrcamentos':  safe_count('SELECT COUNT(*) FROM orcamentos'),
        'totalUnidades':    safe_count('SELECT COUNT(*) FROM unidades_medida'),
        'totalFontes':      safe_count('SELECT COUNT(*) FROM fontes_referencia'),
        'totalInsumos':     safe_count('SELECT COUNT(*) FROM insumos'),
        'totalCompSINAPI':  safe_count("SELECT COUNT(*) FROM composicoes WHERE fonte='SINAPI'"),
        'totalCompSICRO':   safe_count("SELECT COUNT(*) FROM composicoes WHERE fonte='SICRO'"),
        'totalCompUsuario': safe_count("SELECT COUNT(*) FROM composicoes WHERE fonte='USUARIO'"),
        'totalComposicoes':   safe_count('SELECT COUNT(*) FROM composicoes'),
        'totalEventogramas':  safe_count('SELECT COUNT(*) FROM eventogramas'),
        'ultimosOrcamentos': rows_to_list(db.execute("""
            SELECT o.id_orcamento, o.nome_orcamento, o.status, o.data_criacao,
                   o.valor_total, ob.nome_obra
            FROM orcamentos o LEFT JOIN obras ob ON o.id_obra = ob.id_obra
            ORDER BY o.data_criacao DESC LIMIT 5""").fetchall()),
    }
    db.close()
    return jsonify(result)

# ─── OBRAS ────────────────────────────────────────────────────────────────────
# ── Municípios / Estados ──────────────────────────────────────────────────────

@app.route('/api/estados', methods=['GET'])
def estados_list():
    db = get_db()
    rows = rows_to_list(db.execute(
        "SELECT id_estado, codigo_ibge, uf, nome_estado FROM estados ORDER BY uf"
    ).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/municipios', methods=['GET'])
def municipios_list():
    uf     = request.args.get('uf','').upper().strip()
    busca  = request.args.get('busca','').strip()
    ano_raw = request.args.get('ano','').strip()
    try:
        ano = int(ano_raw) if ano_raw else 2026
    except ValueError:
        return jsonify({'erro':'Ano de referencia invalido.'}), 400
    db = get_db()
    ensure_municipio_aliquotas_table(db)
    params = []
    where  = []
    if uf:
        where.append("m.uf=?"); params.append(uf)
    if busca:
        where.append("(m.nome_municipio LIKE ? OR CAST(m.codigo_ibge_municipio AS TEXT) LIKE ?)")
        params += [f'%{busca}%', f'%{busca}%']
    w = ('WHERE ' + ' AND '.join(where)) if where else ''
    rows = rows_to_list(db.execute(
        f"""SELECT m.id_municipio, m.codigo_ibge_municipio, m.nome_municipio, m.uf,
                   COALESCE(ma.aliquota_ibs, m.aliquota_ibs) AS aliquota_ibs,
                   COALESCE(ma.aliquota_cbs, m.aliquota_cbs) AS aliquota_cbs,
                   COALESCE(ma.aliquota_iss, m.aliquota_iss) AS aliquota_iss,
                   COALESCE(ma.ano, m.ano_aliquota, ?) AS ano_aliquota,
                   COALESCE(ma.iva_percentual, COALESCE(m.aliquota_ibs,0) + COALESCE(m.aliquota_cbs,0)) AS iva_percentual,
                   e.nome_estado
            FROM municipios m
            LEFT JOIN estados e ON m.id_estado=e.id_estado
            LEFT JOIN municipio_aliquotas_anuais ma ON ma.id_municipio=m.id_municipio AND ma.ano=?
            {w} ORDER BY m.uf, m.nome_municipio""", [ano, ano] + params
    ).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/municipios/estados', methods=['GET'])
def municipios_estados():
    db = get_db()
    rows = rows_to_list(db.execute(
        "SELECT id_estado, codigo_ibge, uf, nome_estado FROM estados ORDER BY uf"
    ).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/municipios/<int:id>', methods=['GET'])
def municipios_get(id):
    db = get_db()
    ensure_municipio_aliquotas_table(db)
    row = db.execute(
        """SELECT m.*, e.nome_estado FROM municipios m
           LEFT JOIN estados e ON m.id_estado=e.id_estado
           WHERE m.id_municipio=?""", [id]
    ).fetchone()
    if not row:
        db.close()
        return jsonify({'erro':'Municipio nao encontrado.'}), 404
    aliquotas = rows_to_list(db.execute(
        """SELECT ano, iva_percentual, aliquota_cbs, aliquota_ibs, aliquota_iss, data_atualizacao
           FROM municipio_aliquotas_anuais
           WHERE id_municipio=?
           ORDER BY ano""", [id]
    ).fetchall())
    data = dict(row)
    data['aliquotas_anuais'] = aliquotas
    db.close()
    return jsonify(data)
    row = db.execute(
        """SELECT m.*, e.nome_estado FROM municipios m
           LEFT JOIN estados e ON m.id_estado=e.id_estado
           WHERE m.id_municipio=?""", [id]
    ).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Município não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/municipios/<int:id>', methods=['PUT'])
def municipios_update(id):
    d = request.json or {}
    db = get_db()
    ensure_municipio_aliquotas_table(db)
    exists = db.execute('SELECT 1 FROM municipios WHERE id_municipio=?', [id]).fetchone()
    if not exists:
        db.close()
        return jsonify({'erro':'Municipio nao encontrado.'}), 404

    try:
        ano = int(d.get('ano_aliquota') or 2026)
    except (TypeError, ValueError):
        db.close()
        return jsonify({'erro':'Ano de referencia invalido.'}), 400

    ibs = float(d.get('aliquota_ibs', 0.0) or 0.0)
    cbs = float(d.get('aliquota_cbs', 0.0) or 0.0)
    iss = float(d.get('aliquota_iss', 0.0) or 0.0)
    iva = float(d.get('iva_percentual', ibs + cbs) or 0.0)

    db.execute(
        """INSERT INTO municipio_aliquotas_anuais
              (id_municipio, ano, iva_percentual, aliquota_cbs, aliquota_ibs, aliquota_iss, data_atualizacao)
           VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
           ON CONFLICT(id_municipio, ano) DO UPDATE SET
              iva_percentual=excluded.iva_percentual,
              aliquota_cbs=excluded.aliquota_cbs,
              aliquota_ibs=excluded.aliquota_ibs,
              aliquota_iss=excluded.aliquota_iss,
              data_atualizacao=datetime('now')""",
        [id, ano, iva, cbs, ibs, iss]
    )

    if ano == 2026:
        db.execute(
            """UPDATE municipios
               SET aliquota_ibs=?, aliquota_cbs=?, aliquota_iss=?, ano_aliquota=?
               WHERE id_municipio=?""",
            [ibs, cbs, iss, ano, id]
        )
    db.commit()
    row = dict(db.execute(
        """SELECT m.id_municipio, m.codigo_ibge_municipio, m.nome_municipio, m.uf,
                  ma.aliquota_ibs, ma.aliquota_cbs, ma.aliquota_iss, ma.ano AS ano_aliquota,
                  ma.iva_percentual
           FROM municipios m
           JOIN municipio_aliquotas_anuais ma ON ma.id_municipio=m.id_municipio AND ma.ano=?
           WHERE m.id_municipio=?""",
        [ano, id]
    ).fetchone())
    db.close()
    return jsonify(row)
    cur = db.execute(
        """UPDATE municipios
           SET aliquota_ibs=?, aliquota_cbs=?, aliquota_iss=?, ano_aliquota=?
           WHERE id_municipio=?""",
        [d.get('aliquota_ibs', 0.0), d.get('aliquota_cbs', 0.0),
         d.get('aliquota_iss', 0.0), d.get('ano_aliquota'), id]
    )
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Município não encontrado.'}), 404
    row = dict(db.execute('SELECT * FROM municipios WHERE id_municipio=?',[id]).fetchone())
    db.close()
    return jsonify(row)

@app.route('/api/municipios/importar-aliquotas', methods=['POST'])
def municipios_importar_aliquotas():
    """Importa alíquotas ISS/IBS/CBS de um arquivo Excel para todos os municípios."""
    import tempfile, pandas as pd
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Arquivo não enviado.'}), 400
    f = request.files['arquivo']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx','.xls','.xlsm','.ods'):
        return jsonify({'erro': 'Formato inválido. Use .xlsx, .xls ou .ods.'}), 400

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        # Detectar aba correta
        xl = pd.ExcelFile(tmp_path)
        aba = next((s for s in xl.sheet_names if 'munic' in s.lower() or 'estado' in s.lower()), xl.sheet_names[0])
        df = xl.parse(aba)

        # Identificar colunas por posição ou nome
        col_ibge = next((c for c in df.columns if 'código município' in str(c).lower() or 'codigo_municipio' in str(c).lower()), df.columns[3])
        anos_iss = {int(c.replace('ISS','').strip()): c for c in df.columns if str(c).startswith('ISS ') and df[c].notna().any()}
        anos_ibs = {int(c.replace('IBS','').strip()): c for c in df.columns if str(c).startswith('IBS ') and df[c].notna().any()}
        anos_cbs = {int(c.replace('CBS','').strip()): c for c in df.columns if str(c).startswith('CBS ') and df[c].notna().any()}

        todos_anos = sorted(set(list(anos_ibs.keys()) + list(anos_cbs.keys()) + list(anos_iss.keys())))
        if not todos_anos:
            return jsonify({'erro': 'Nenhuma coluna ISS/IBS/CBS encontrada na planilha.'}), 400

        db  = get_db()
        ensure_municipio_aliquotas_table(db)
        ok  = 0
        err = 0
        for ano in todos_anos:
            col_i = anos_iss.get(ano)
            col_b = anos_ibs.get(ano)
            col_c = anos_cbs.get(ano)
            sub = df[
                (df[col_b].notna() if col_b else False) |
                (df[col_c].notna() if col_c else False) |
                (df[col_i].notna() if col_i else False)
            ]
            for _, row in sub.iterrows():
                try:
                    cod = int(row[col_ibge])
                except:
                    err += 1; continue
                iss_v = float(row[col_i]) if col_i and pd.notna(row[col_i]) else None
                ibs_v = float(row[col_b]) if col_b and pd.notna(row[col_b]) else None
                cbs_v = float(row[col_c]) if col_c and pd.notna(row[col_c]) else None
                mun = db.execute(
                    "SELECT id_municipio, aliquota_iss, aliquota_ibs, aliquota_cbs FROM municipios WHERE codigo_ibge_municipio=?",
                    [cod]
                ).fetchone()
                if not mun:
                    err += 1
                    continue
                iss_final = iss_v if iss_v is not None else (mun['aliquota_iss'] or 0.0)
                ibs_final = ibs_v if ibs_v is not None else (mun['aliquota_ibs'] or 0.0)
                cbs_final = cbs_v if cbs_v is not None else (mun['aliquota_cbs'] or 0.0)
                db.execute(
                    """INSERT INTO municipio_aliquotas_anuais
                          (id_municipio, ano, iva_percentual, aliquota_cbs, aliquota_ibs, aliquota_iss, data_atualizacao)
                       VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                       ON CONFLICT(id_municipio, ano) DO UPDATE SET
                          iva_percentual=excluded.iva_percentual,
                          aliquota_cbs=excluded.aliquota_cbs,
                          aliquota_ibs=excluded.aliquota_ibs,
                          aliquota_iss=excluded.aliquota_iss,
                          data_atualizacao=datetime('now')""",
                    [mun['id_municipio'], ano, ibs_final + cbs_final, cbs_final, ibs_final, iss_final]
                )
                if ano == 2026:
                    db.execute(
                        """UPDATE municipios
                           SET aliquota_iss=?, aliquota_ibs=?, aliquota_cbs=?, ano_aliquota=?
                           WHERE id_municipio=?""",
                        [iss_final, ibs_final, cbs_final, ano, mun['id_municipio']]
                    )
                ok += 1
        db.commit()
        db.close()
        return jsonify({'status':'ok', 'atualizados': ok, 'nao_encontrados': err,
                        'anos': todos_anos,
                        'mensagem': f'{ok} municípios atualizados para os anos {todos_anos}.'})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass

# ── Obras ─────────────────────────────────────────────────────────────────────

@app.route('/api/obras', methods=['GET'])
def obras_list():
    q = request.args.get('q','')
    situacao = request.args.get('situacao','')
    sql = """SELECT o.*, (SELECT COUNT(*) FROM orcamentos WHERE id_obra=o.id_obra) AS qtd_orcamentos
             FROM obras o WHERE 1=1"""
    params = []
    if q:
        sql += " AND (o.nome_obra LIKE ? OR o.codigo_obra LIKE ? OR o.contratante LIKE ? OR o.municipio LIKE ?)"
        like = f'%{q}%'
        params += [like,like,like,like]
    if situacao:
        sql += " AND o.situacao=?"; params.append(situacao)
    sql += " ORDER BY o.id_obra DESC"
    db = get_db()
    ensure_obras_reforma_fields(db)
    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/obras/<int:id>', methods=['GET'])
def obras_get(id):
    db = get_db()
    ensure_obras_reforma_fields(db)
    row = db.execute("""SELECT o.*,(SELECT COUNT(*) FROM orcamentos WHERE id_obra=o.id_obra) AS qtd_orcamentos
                        FROM obras o WHERE o.id_obra=?""", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Obra não encontrada.'}), 404
    return jsonify(dict(row))

@app.route('/api/obras', methods=['POST'])
def obras_create():
    d = request.json or {}
    if not d.get('nome_obra','').strip():
        return jsonify({'erro':'Nome da obra é obrigatório.'}), 400
    if d.get('uf') and len(d['uf']) != 2:
        return jsonify({'erro':'UF deve ter exatamente 2 caracteres.'}), 400
    db = get_db()
    ensure_obras_reforma_fields(db)
    cur = db.execute("""INSERT INTO obras (codigo_obra,nome_obra,descricao,tipo_obra,contratante,
        uf,municipio,id_municipio,cib,endereco,area_construida_m2,situacao,
        ano_realizacao,fator_setorial,redutor_compras_governamentais)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [d.get('codigo_obra'), d['nome_obra'].strip(), d.get('descricao'), d.get('tipo_obra'),
         d.get('contratante'), d.get('uf'), d.get('municipio'), d.get('id_municipio'),
         d.get('cib'), d.get('endereco'), d.get('area_construida_m2'), d.get('situacao','Ativa'),
         d.get('ano_realizacao'), _float_or_zero(d.get('fator_setorial')),
         _float_or_zero(d.get('redutor_compras_governamentais'))])
    db.commit()
    row = dict(db.execute('SELECT * FROM obras WHERE id_obra=?',[cur.lastrowid]).fetchone())
    db.close()
    return jsonify(row), 201

@app.route('/api/obras/<int:id>', methods=['PUT'])
def obras_update(id):
    d = request.json or {}
    if not d.get('nome_obra','').strip():
        return jsonify({'erro':'Nome da obra é obrigatório.'}), 400
    if d.get('uf') and len(d['uf']) != 2:
        return jsonify({'erro':'UF deve ter exatamente 2 caracteres.'}), 400
    db = get_db()
    ensure_obras_reforma_fields(db)
    cur = db.execute("""UPDATE obras SET codigo_obra=?,nome_obra=?,descricao=?,tipo_obra=?,
        contratante=?,uf=?,municipio=?,id_municipio=?,cib=?,endereco=?,area_construida_m2=?,situacao=?,
        ano_realizacao=?,fator_setorial=?,redutor_compras_governamentais=?
        WHERE id_obra=?""",
        [d.get('codigo_obra'), d['nome_obra'].strip(), d.get('descricao'), d.get('tipo_obra'),
         d.get('contratante'), d.get('uf'), d.get('municipio'), d.get('id_municipio'),
         d.get('cib'), d.get('endereco'), d.get('area_construida_m2'), d.get('situacao','Ativa'),
         d.get('ano_realizacao'), _float_or_zero(d.get('fator_setorial')),
         _float_or_zero(d.get('redutor_compras_governamentais')), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Obra não encontrada.'}), 404
    row = dict(db.execute('SELECT * FROM obras WHERE id_obra=?',[id]).fetchone())
    db.close()
    return jsonify(row)

@app.route('/api/obras/<int:id>', methods=['DELETE'])
def obras_delete(id):
    db = get_db()
    total = db.execute('SELECT COUNT(*) FROM orcamentos WHERE id_obra=?',[id]).fetchone()[0]
    if total > 0:
        db.close()
        return jsonify({'erro':f'Não é possível excluir: obra possui {total} orçamento(s) vinculado(s).'}), 409
    cur = db.execute('DELETE FROM obras WHERE id_obra=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Obra não encontrada.'}), 404
    return jsonify({'mensagem':'Obra excluída com sucesso.'})

@app.route('/api/obras/<int:id>/duplicar', methods=['POST'])
def obras_duplicate(id):
    db = get_db()
    ensure_obras_reforma_fields(db)
    row = db.execute('SELECT * FROM obras WHERE id_obra=?',[id]).fetchone()
    if not row: db.close(); return jsonify({'erro':'Obra não encontrada.'}), 404
    r = dict(row)
    cur = db.execute("""INSERT INTO obras (codigo_obra,nome_obra,descricao,tipo_obra,contratante,
        uf,municipio,id_municipio,cib,endereco,area_construida_m2,situacao,
        ano_realizacao,fator_setorial,redutor_compras_governamentais)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [(r['codigo_obra']+'-COPIA') if r['codigo_obra'] else None,
         'Cópia de '+r['nome_obra'], r['descricao'], r['tipo_obra'],
         r['contratante'], r['uf'], r['municipio'], r['id_municipio'],
         r.get('cib'), r['endereco'], r['area_construida_m2'], 'Ativa',
         r.get('ano_realizacao'), r.get('fator_setorial', 0),
         r.get('redutor_compras_governamentais', 0)])
    db.commit()
    novo = dict(db.execute('SELECT * FROM obras WHERE id_obra=?',[cur.lastrowid]).fetchone())
    db.close()
    return jsonify(novo), 201

@app.route('/api/obras/<int:id>/orcamentos', methods=['GET'])
def obras_orcamentos(id):
    db = get_db()
    rows = rows_to_list(db.execute("""
        SELECT o.*, db.mes, db.ano FROM orcamentos o
        LEFT JOIN datas_base db ON o.id_data_base=db.id_data_base
        WHERE o.id_obra=? ORDER BY o.id_orcamento DESC""",[id]).fetchall())
    db.close()
    return jsonify(rows)

# ─── ORÇAMENTOS ──────────────────────────────────────────────────────────────
SEL_ORC = """SELECT o.*, ob.nome_obra, ob.uf AS obra_uf, ob.id_municipio,
                ob.ano_realizacao, ob.fator_setorial, ob.redutor_compras_governamentais,
                db.mes AS data_base_mes, db.ano AS data_base_ano
             FROM orcamentos o
             LEFT JOIN obras ob ON o.id_obra=ob.id_obra
             LEFT JOIN datas_base db ON o.id_data_base=db.id_data_base"""

def _orc_credito_iva_totais(db, id_orcamento):
    servicos = db.execute("""
        SELECT quantidade, custo_unitario, codigo, tipo_item, id_composicao
        FROM orcamento_sintetico
        WHERE id_orcamento=? AND tipo_linha='item'
        ORDER BY ordem, id_item
    """, [id_orcamento]).fetchall()
    comp_rows = db.execute("SELECT id_composicao, codigo FROM composicoes").fetchall()
    comp_cache = {}
    for r in comp_rows:
        raw = (r['codigo'] or '').strip().upper()
        if not raw:
            continue
        bare = raw.replace('SINAPI.', '').replace('SICRO.', '').replace('SEINFRA.', '').replace('SUDECAP.', '').replace('GOINFRA.', '').replace('CDHU.', '').replace('USUARIO.', '').strip()
        for k in (bare, 'SINAPI.' + bare, 'SICRO.' + bare, 'SEINFRA.' + bare, 'SUDECAP.' + bare, 'GOINFRA.' + bare, 'CDHU.' + bare, raw):
            comp_cache.setdefault(k, r['id_composicao'])

    sql_itens = """
        SELECT codigo_item, descricao, unidade, coeficiente, tipo_item, preco_unitario
        FROM itens_composicao
        WHERE id_composicao=?
        ORDER BY ordem
    """
    sql_preco = """
        SELECT COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia) AS preco,
               COALESCE(p.ibs_percentual, 0) AS ibs_percentual,
               COALESCE(p.cbs_percentual, 0) AS cbs_percentual
        FROM precos_insumos p
        JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
        WHERE p.id_insumo = (
            SELECT id_insumo FROM insumos
            WHERE codigo_insumo = ? LIMIT 1
        )
        ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1
    """
    total_ibs = 0.0
    total_cbs = 0.0

    def resolver_preco(codigo_item, preco_armazenado):
        if codigo_item:
            r = db.execute(sql_preco, [codigo_item.strip()]).fetchone()
            if r and r['preco']:
                return float(r['preco']), float(r['ibs_percentual'] or 0), float(r['cbs_percentual'] or 0)
        return float(preco_armazenado or 0), 0.0, 0.0

    def add(codigo, qtd, preco_armazenado):
        nonlocal total_ibs, total_cbs
        preco, ibs_pct, cbs_pct = resolver_preco(codigo, preco_armazenado)
        custo = qtd * preco
        total_ibs += custo * (ibs_pct / 100.0)
        total_cbs += custo * (cbs_pct / 100.0)

    def expandir(id_comp, fator, visitados=None):
        if visitados is None:
            visitados = set()
        if id_comp in visitados:
            return
        visitados = visitados | {id_comp}
        for row in db.execute(sql_itens, [id_comp]).fetchall():
            coef = float(row['coeficiente'] or 0)
            if coef <= 0:
                continue
            cod = (row['codigo_item'] or '').strip()
            tipo = (row['tipo_item'] or 'INSUMO').upper()
            if tipo == 'COMPOSICAO':
                cod_up = cod.upper()
                bare = cod_up.replace('SINAPI.','').replace('SICRO.','').replace('SEINFRA.','').replace('SUDECAP.','').replace('GOINFRA.','').replace('CDHU.','').replace('USUARIO.','').strip()
                sub_id = comp_cache.get(bare) or comp_cache.get('SINAPI.' + bare) or comp_cache.get('SICRO.' + bare) or comp_cache.get('SEINFRA.' + bare) or comp_cache.get('SUDECAP.' + bare) or comp_cache.get('GOINFRA.' + bare) or comp_cache.get('CDHU.' + bare) or comp_cache.get(cod_up)
                if sub_id:
                    expandir(sub_id, coef * fator, visitados)
                else:
                    add(cod, coef * fator, row['preco_unitario'])
            else:
                add(cod, coef * fator, row['preco_unitario'])

    for row in servicos:
        qtd = float(row['quantidade'] or 0)
        if qtd <= 0:
            continue
        id_comp = row['id_composicao']
        if not id_comp:
            cod_item = (row['codigo'] or '').strip().upper()
            if cod_item:
                id_comp = comp_cache.get(cod_item) or comp_cache.get('SINAPI.' + cod_item) or comp_cache.get('SICRO.' + cod_item) or comp_cache.get('SEINFRA.' + cod_item) or comp_cache.get('SUDECAP.' + cod_item) or comp_cache.get('GOINFRA.' + cod_item) or comp_cache.get('CDHU.' + cod_item)
        if id_comp:
            expandir(id_comp, qtd)
        else:
            add(row['codigo'] or '', qtd, row['custo_unitario'])
    return round(total_ibs, 2), round(total_cbs, 2)

def _enriquecer_orcamento_reforma(db, orc):
    ensure_municipio_aliquotas_table(db)
    ensure_obras_reforma_fields(db)
    total_ibs, total_cbs = _orc_credito_iva_totais(db, orc['id_orcamento'])
    credito_iva = total_ibs + total_cbs
    ano = _ano_tributario_orcamento(orc)
    aliq = None
    if orc.get('id_municipio'):
        aliq = db.execute("""
            SELECT iva_percentual, aliquota_cbs, aliquota_ibs
            FROM municipio_aliquotas_anuais
            WHERE id_municipio=? AND ano=?
        """, [orc.get('id_municipio'), ano]).fetchone()
    iva_nominal = float(aliq['iva_percentual']) if aliq else 0.0
    fator = _float_or_zero(orc.get('fator_setorial'))
    redutor = _float_or_zero(orc.get('redutor_compras_governamentais'))
    fator_efetivo = max(0.0, (1 - fator) * (1 - redutor))
    custo_direto = _float_or_zero(orc.get('valor_custo_direto'))
    credito_pct = (credito_iva / custo_direto) if custo_direto > 0 else 0.0
    ivaeq = max(0.0, (iva_nominal * fator_efetivo) - credito_pct)
    orc['credito_iva'] = round(credito_iva, 2)
    orc['credito_ibs'] = round(total_ibs, 2)
    orc['credito_cbs'] = round(total_cbs, 2)
    orc['credito_iva_percentual'] = round(credito_pct * 100, 4)
    orc['ivaeq_percentual'] = round(ivaeq * 100, 4)
    orc['iva_nominal_percentual'] = round(iva_nominal * 100, 4)
    orc['ano_tributario'] = ano
    return orc

@app.route('/api/orcamentos', methods=['GET'])
def orc_list():
    id_obra = request.args.get('id_obra','')
    status  = request.args.get('status','')
    q       = request.args.get('q','')
    sql = SEL_ORC + " WHERE 1=1"; params = []
    if id_obra: sql += " AND o.id_obra=?"; params.append(id_obra)
    if status:  sql += " AND o.status=?";  params.append(status)
    if q:
        sql += " AND (o.nome_orcamento LIKE ? OR ob.nome_obra LIKE ?)"; like=f'%{q}%'; params+=[like,like]
    sql += " ORDER BY o.id_orcamento DESC"
    db = get_db()
    ensure_obras_reforma_fields(db)
    ensure_orcamentos_reforma_fields(db)
    rows = [_enriquecer_orcamento_reforma(db, dict(r)) for r in db.execute(sql,params).fetchall()]
    db.close(); return jsonify(rows)

@app.route('/api/orcamentos/<int:id>', methods=['GET'])
def orc_get(id):
    db = get_db()
    ensure_obras_reforma_fields(db)
    ensure_orcamentos_reforma_fields(db)
    row = db.execute(SEL_ORC+" WHERE o.id_orcamento=?",[id]).fetchone()
    if not row:
        db.close()
        return jsonify({'erro':'Orcamento nao encontrado.'}), 404
    data = _enriquecer_orcamento_reforma(db, dict(row))
    db.close()
    return jsonify(data)

@app.route('/api/orcamentos', methods=['POST'])
def orc_create():
    d = request.json or {}
    if not d.get('id_obra'): return jsonify({'erro':'Obra é obrigatória.'}), 400
    if not d.get('nome_orcamento','').strip(): return jsonify({'erro':'Nome do orçamento é obrigatório.'}), 400
    db = get_db()
    ensure_orcamentos_reforma_fields(db)
    if not db.execute('SELECT id_obra FROM obras WHERE id_obra=?',[d['id_obra']]).fetchone():
        db.close(); return jsonify({'erro':'Obra não encontrada.'}), 400
    cur = db.execute("""INSERT INTO orcamentos (id_obra,nome_orcamento,descricao,id_data_base,
        uf_referencia,versao,status,observacoes,regime_previdenciario) VALUES (?,?,?,?,?,?,?,?,?)""",
        [d['id_obra'], d['nome_orcamento'].strip(), d.get('descricao'), d.get('id_data_base'),
         d.get('uf_referencia'), d.get('versao','1.0'), d.get('status','Em elaboração'), d.get('observacoes'),
         d.get('regime_previdenciario','Onerado')])
    db.commit()
    row = db.execute(SEL_ORC+" WHERE o.id_orcamento=?",[cur.lastrowid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/orcamentos/<int:id>', methods=['PUT'])
def orc_update(id):
    d = request.json or {}
    if not d.get('nome_orcamento','').strip(): return jsonify({'erro':'Nome do orçamento é obrigatório.'}), 400
    cd  = float(d.get('valor_custo_direto') or 0)
    bdi = float(d.get('valor_bdi') or 0)
    # Converter strings vazias em NULL para campos com CHECK/FK constraints
    def _nv(v): return v if v not in (None,'','0',0) else None
    id_obra    = _nv(d.get('id_obra'))
    id_db      = _nv(d.get('id_data_base'))
    uf_ref     = d.get('uf_referencia','').strip() or None  # '' → NULL
    db = get_db()
    ensure_orcamentos_reforma_fields(db)
    try:
        cur = db.execute("""UPDATE orcamentos SET id_obra=?,nome_orcamento=?,descricao=?,id_data_base=?,
            uf_referencia=?,versao=?,status=?,valor_custo_direto=?,valor_bdi=?,valor_total=?,observacoes=?,
            regime_previdenciario=?
            WHERE id_orcamento=?""",
            [id_obra, d['nome_orcamento'].strip(), d.get('descricao') or None,
             id_db, uf_ref, d.get('versao','1.0') or '1.0',
             d.get('status','Em elaboração'),
             cd, bdi, cd+bdi, d.get('observacoes') or None,
             d.get('regime_previdenciario','Onerado'), id])
        db.commit()
        if cur.rowcount == 0: return jsonify({'erro':'Orçamento não encontrado.'}), 404
        row = db.execute(SEL_ORC+" WHERE o.id_orcamento=?",[id]).fetchone()
        return jsonify(dict(row))
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-500:]}), 500
    finally:
        db.close()

@app.route('/api/orcamentos/<int:id>', methods=['DELETE'])
def orc_delete(id):
    db = get_db()
    cur = db.execute('DELETE FROM orcamentos WHERE id_orcamento=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Orçamento não encontrado.'}), 404
    return jsonify({'mensagem':'Orçamento excluído com sucesso.'})

@app.route('/api/orcamentos/<int:id>/duplicar', methods=['POST'])
def orc_duplicate(id):
    db = get_db()
    ensure_orcamentos_reforma_fields(db)
    row = db.execute('SELECT * FROM orcamentos WHERE id_orcamento=?',[id]).fetchone()
    if not row: db.close(); return jsonify({'erro':'Orçamento não encontrado.'}), 404
    ensure_orcamento_sintetico_bdi_linha(db)
    r = dict(row)
    import re as _re
    _versao_raw = r.get('versao') or '1.0'
    _partes = _versao_raw.split('.')
    _minor_raw = _partes[1] if len(_partes) > 1 else '0'
    _minor_digits = _re.sub(r'\D.*', '', _minor_raw)   # remove sufixos como '-IA'
    _minor = int(_minor_digits) if _minor_digits else 0
    nova_versao = _partes[0] + '.' + str(_minor + 1)
    cur = db.execute("""INSERT INTO orcamentos (id_obra,nome_orcamento,descricao,id_data_base,
        uf_referencia,versao,status,observacoes,
        bdi_percentual,id_bdi_perfil,
        valor_custo_direto,valor_bdi,valor_total,regime_previdenciario)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [r['id_obra'], 'Cópia de '+r['nome_orcamento'], r.get('descricao'),
         r.get('id_data_base'), r.get('uf_referencia'), nova_versao, 'Em elaboração', r.get('observacoes'),
         r.get('bdi_percentual', 0), r.get('id_bdi_perfil'),
         r.get('valor_custo_direto', 0), r.get('valor_bdi', 0), r.get('valor_total', 0),
         r.get('regime_previdenciario', 'Onerado')])
    novo_id = cur.lastrowid

    # Copiar todos os itens do orçamento sintético
    itens_orig = db.execute(
        'SELECT * FROM orcamento_sintetico WHERE id_orcamento=? ORDER BY ordem, id_item', [id]
    ).fetchall()
    for it in itens_orig:
        it = dict(it)
        db.execute("""INSERT INTO orcamento_sintetico
            (id_orcamento,item_num,tipo_linha,profundidade,ordem,tipo_item,
             id_composicao,id_insumo,codigo,fonte,descricao,unidade,quantidade,custo_unitario,bdi_percentual_linha)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [novo_id, it.get('item_num'), it.get('tipo_linha'), it.get('profundidade'),
             it.get('ordem'), it.get('tipo_item'), it.get('id_composicao'), it.get('id_insumo'),
             it.get('codigo',''), it.get('fonte',''), it.get('descricao',''),
             it.get('unidade',''), it.get('quantidade',0), it.get('custo_unitario',0),
             it.get('bdi_percentual_linha')])

    db.commit()
    novo = db.execute(SEL_ORC+" WHERE o.id_orcamento=?",[novo_id]).fetchone()
    db.close()
    return jsonify(dict(novo)), 201

# ─── UNIDADES ────────────────────────────────────────────────────────────────
@app.route('/api/unidades', methods=['GET'])
def un_list():
    db = get_db()
    rows = rows_to_list(db.execute('SELECT * FROM unidades_medida ORDER BY sigla').fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/unidades/<int:id>', methods=['GET'])
def un_get(id):
    db = get_db()
    row = db.execute('SELECT * FROM unidades_medida WHERE id_unidade=?',[id]).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Unidade não encontrada.'}), 404
    return jsonify(dict(row))

@app.route('/api/unidades', methods=['POST'])
def un_create():
    d = request.json or {}
    if not d.get('sigla','').strip(): return jsonify({'erro':'Sigla é obrigatória.'}), 400
    db = get_db()
    try:
        cur = db.execute('INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)',
            [d['sigla'].strip(), d.get('descricao'), d.get('tipo_unidade')])
        db.commit()
        row = dict(db.execute('SELECT * FROM unidades_medida WHERE id_unidade=?',[cur.lastrowid]).fetchone())
        db.close(); return jsonify(row), 201
    except sqlite3.IntegrityError:
        db.close(); return jsonify({'erro':f'Sigla "{d["sigla"]}" já existe.'}), 409

@app.route('/api/unidades/<int:id>', methods=['PUT'])
def un_update(id):
    d = request.json or {}
    if not d.get('sigla','').strip(): return jsonify({'erro':'Sigla é obrigatória.'}), 400
    db = get_db()
    try:
        cur = db.execute('UPDATE unidades_medida SET sigla=?,descricao=?,tipo_unidade=? WHERE id_unidade=?',
            [d['sigla'].strip(), d.get('descricao'), d.get('tipo_unidade'), id])
        db.commit()
        if cur.rowcount == 0: db.close(); return jsonify({'erro':'Unidade não encontrada.'}), 404
        row = dict(db.execute('SELECT * FROM unidades_medida WHERE id_unidade=?',[id]).fetchone())
        db.close(); return jsonify(row)
    except sqlite3.IntegrityError:
        db.close(); return jsonify({'erro':f'Sigla "{d["sigla"]}" já existe.'}), 409

@app.route('/api/unidades/<int:id>', methods=['DELETE'])
def un_delete(id):
    db = get_db()
    cur = db.execute('DELETE FROM unidades_medida WHERE id_unidade=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Unidade não encontrada.'}), 404
    return jsonify({'mensagem':'Unidade excluída com sucesso.'})

# ─── FONTES ──────────────────────────────────────────────────────────────────
@app.route('/api/fontes', methods=['GET'])
def fontes_list():
    db = get_db()
    rows = rows_to_list(db.execute('SELECT * FROM fontes_referencia ORDER BY nome_fonte').fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/fontes/<int:id>', methods=['GET'])
def fontes_get(id):
    db = get_db()
    row = db.execute('SELECT * FROM fontes_referencia WHERE id_fonte=?',[id]).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Fonte não encontrada.'}), 404
    return jsonify(dict(row))

@app.route('/api/fontes', methods=['POST'])
def fontes_create():
    d = request.json or {}
    if not d.get('nome_fonte','').strip(): return jsonify({'erro':'Nome da fonte é obrigatório.'}), 400
    db = get_db()
    cur = db.execute('INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes) VALUES (?,?,?,?,?)',
        [d['nome_fonte'].strip(), d.get('tipo_fonte'), d.get('orgao_responsavel'), d.get('abrangencia'), d.get('observacoes')])
    db.commit()
    row = dict(db.execute('SELECT * FROM fontes_referencia WHERE id_fonte=?',[cur.lastrowid]).fetchone())
    db.close(); return jsonify(row), 201

@app.route('/api/fontes/<int:id>', methods=['PUT'])
def fontes_update(id):
    d = request.json or {}
    if not d.get('nome_fonte','').strip(): return jsonify({'erro':'Nome da fonte é obrigatório.'}), 400
    db = get_db()
    cur = db.execute('UPDATE fontes_referencia SET nome_fonte=?,tipo_fonte=?,orgao_responsavel=?,abrangencia=?,observacoes=? WHERE id_fonte=?',
        [d['nome_fonte'].strip(), d.get('tipo_fonte'), d.get('orgao_responsavel'), d.get('abrangencia'), d.get('observacoes'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Fonte não encontrada.'}), 404
    row = dict(db.execute('SELECT * FROM fontes_referencia WHERE id_fonte=?',[id]).fetchone())
    db.close(); return jsonify(row)

@app.route('/api/fontes/<int:id>', methods=['DELETE'])
def fontes_delete(id):
    db = get_db()
    cur = db.execute('DELETE FROM fontes_referencia WHERE id_fonte=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Fonte não encontrada.'}), 404
    return jsonify({'mensagem':'Fonte excluída com sucesso.'})

# ─── DATAS-BASE ──────────────────────────────────────────────────────────────
MESES = ['','Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']

@app.route('/api/datas-base', methods=['GET'])
def db_list():
    db = get_db()
    rows = rows_to_list(db.execute('SELECT * FROM datas_base ORDER BY ano DESC, mes DESC').fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/datas-base/<int:id>', methods=['GET'])
def db_get(id):
    db = get_db()
    row = db.execute('SELECT * FROM datas_base WHERE id_data_base=?',[id]).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Data-base não encontrada.'}), 404
    return jsonify(dict(row))

@app.route('/api/datas-base', methods=['POST'])
def db_create():
    d = request.json or {}
    try: m, a = int(d.get('mes',0)), int(d.get('ano',0))
    except: return jsonify({'erro':'Mês e ano inválidos.'}), 400
    if m < 1 or m > 12: return jsonify({'erro':'Mês inválido (1–12).'}), 400
    if len(str(a)) != 4: return jsonify({'erro':'Ano deve ter 4 dígitos.'}), 400
    data_ref = f'{str(m).zfill(2)}/{a}'
    desc = d.get('descricao') or f'{MESES[m]}/{a}'
    db = get_db()
    try:
        cur = db.execute('INSERT INTO datas_base (mes,ano,data_referencia,descricao) VALUES (?,?,?,?)',
            [m, a, data_ref, desc])
        db.commit()
        row = dict(db.execute('SELECT * FROM datas_base WHERE id_data_base=?',[cur.lastrowid]).fetchone())
        db.close(); return jsonify(row), 201
    except sqlite3.IntegrityError:
        db.close(); return jsonify({'erro':f'Data-base {m}/{a} já existe.'}), 409

@app.route('/api/datas-base/<int:id>', methods=['PUT'])
def db_update(id):
    d = request.json or {}
    try: m, a = int(d.get('mes',0)), int(d.get('ano',0))
    except: return jsonify({'erro':'Mês e ano inválidos.'}), 400
    if m < 1 or m > 12: return jsonify({'erro':'Mês inválido.'}), 400
    if len(str(a)) != 4: return jsonify({'erro':'Ano deve ter 4 dígitos.'}), 400
    db = get_db()
    try:
        cur = db.execute('UPDATE datas_base SET mes=?,ano=?,data_referencia=?,descricao=? WHERE id_data_base=?',
            [m, a, f'{str(m).zfill(2)}/{a}', d.get('descricao'), id])
        db.commit()
        if cur.rowcount == 0: db.close(); return jsonify({'erro':'Data-base não encontrada.'}), 404
        row = dict(db.execute('SELECT * FROM datas_base WHERE id_data_base=?',[id]).fetchone())
        db.close(); return jsonify(row)
    except sqlite3.IntegrityError:
        db.close(); return jsonify({'erro':f'Data-base {m}/{a} já existe.'}), 409

@app.route('/api/datas-base/<int:id>', methods=['DELETE'])
def db_delete(id):
    db = get_db()
    total = db.execute('SELECT COUNT(*) FROM orcamentos WHERE id_data_base=?',[id]).fetchone()[0]
    if total > 0:
        db.close()
        return jsonify({'erro':f'Não é possível excluir: data-base está vinculada a {total} orçamento(s).'}), 409
    cur = db.execute('DELETE FROM datas_base WHERE id_data_base=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Data-base não encontrada.'}), 404
    return jsonify({'mensagem':'Data-base excluída com sucesso.'})

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 2 — INSUMOS
# ═══════════════════════════════════════════════════════════════════════════════

SEL_INS = """
    SELECT i.*,
           um.sigla        AS sigla_unidade,
           um.descricao    AS desc_unidade,
           gi.nome_grupo   AS nome_grupo,
           p.id_preco, p.id_data_base AS preco_id_data_base,
           p.preco_referencia, p.preco_desonerado, p.preco_nao_desonerado,
           p.uf_referencia AS preco_uf, p.iva_equivalente,
           p.cbs_percentual, p.ibs_percentual, p.is_percentual, p.preco_sem_tributos,
           p.encargos_sociais_percentual AS preco_encargos_sociais_percentual,
           COALESCE(p.encargos_sociais_percentual, i.encargos_sociais_percentual) AS encargos_sociais_calculado,
           db2.mes AS preco_mes, db2.ano AS preco_ano,
           fr.nome_fonte   AS nome_fonte
    FROM insumos i
    LEFT JOIN unidades_medida um ON i.id_unidade = um.id_unidade
    LEFT JOIN grupos_insumos  gi ON i.id_grupo   = gi.id_grupo
    LEFT JOIN precos_insumos  p  ON p.id_preco   = (
        SELECT id_preco FROM precos_insumos
        WHERE id_insumo = i.id_insumo
        ORDER BY id_preco DESC LIMIT 1
    )
    LEFT JOIN datas_base      db2 ON p.id_data_base = db2.id_data_base
    LEFT JOIN fontes_referencia fr ON p.id_fonte    = fr.id_fonte
"""

def _build_sel_ins(uf='', mes='', ano='', regime=''):
    """Builds a dynamic SEL_INS with optional subquery filters for uf/mes/ano."""
    sub_where = "WHERE id_insumo = i.id_insumo"
    sub_params = []
    if uf:
        sub_where += " AND uf_referencia = ?"
        sub_params.append(uf)
    if mes and ano:
        sub_where += (" AND id_data_base IN "
                      "(SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?)")
        sub_params += [int(mes), int(ano)]
    regime_norm = (regime or '').strip().lower()
    if regime_norm == 'onerado':
        sub_where += " AND COALESCE(preco_nao_desonerado, 0) > 0"
    elif regime_norm == 'desonerado':
        sub_where += " AND COALESCE(preco_desonerado, 0) > 0"
    preco_regime_expr = "p.preco_referencia"
    if regime_norm == 'onerado':
        preco_regime_expr = "COALESCE(NULLIF(p.preco_nao_desonerado,0), p.preco_referencia)"
    elif regime_norm == 'desonerado':
        preco_regime_expr = "COALESCE(NULLIF(p.preco_desonerado,0), p.preco_referencia)"
    sel = f"""
        SELECT i.*,
               um.sigla AS sigla_unidade, um.descricao AS desc_unidade,
               gi.nome_grupo AS nome_grupo,
               p.id_preco, p.id_data_base AS preco_id_data_base,
               p.preco_referencia, p.preco_desonerado, p.preco_nao_desonerado,
               {preco_regime_expr} AS preco_regime,
               p.uf_referencia AS preco_uf, p.iva_equivalente,
               p.cbs_percentual, p.ibs_percentual, p.is_percentual, p.preco_sem_tributos,
               p.encargos_sociais_percentual AS preco_encargos_sociais_percentual,
               COALESCE(p.encargos_sociais_percentual, i.encargos_sociais_percentual) AS encargos_sociais_calculado,
               db2.mes AS preco_mes, db2.ano AS preco_ano,
               fr.nome_fonte AS nome_fonte
        FROM insumos i
        LEFT JOIN unidades_medida um ON i.id_unidade = um.id_unidade
        LEFT JOIN grupos_insumos  gi ON i.id_grupo   = gi.id_grupo
        LEFT JOIN precos_insumos  p  ON p.id_preco = (
            SELECT id_preco FROM precos_insumos
            {sub_where}
            ORDER BY id_preco DESC LIMIT 1
        )
        LEFT JOIN datas_base db2 ON p.id_data_base = db2.id_data_base
        LEFT JOIN fontes_referencia fr ON p.id_fonte = fr.id_fonte
    """
    return sel, sub_params

def _save_preco_principal(db, id_insumo, d):
    """Cria ou atualiza o registro de preço principal ao salvar o insumo."""
    pref = float(d.get('preco_referencia') or 0)
    if pref <= 0:
        return
    cbs = float(d.get('cbs_percentual') or 0)
    ibs = float(d.get('ibs_percentual') or 0)
    isp = float(d.get('is_percentual') or 0)
    enc_soc = d.get('encargos_sociais_percentual')
    enc_soc = float(enc_soc) if enc_soc not in (None, '') else None
    iva = round(cbs + ibs + isp, 6)
    psem = round(pref / (1 + iva / 100), 6) if iva > 0 else pref
    existing = db.execute(
        "SELECT id_preco FROM precos_insumos WHERE id_insumo=? ORDER BY id_preco DESC LIMIT 1",
        [id_insumo]
    ).fetchone()
    if existing:
        db.execute("""UPDATE precos_insumos SET
            id_data_base=?, uf_referencia=?,
            preco_desonerado=?, preco_nao_desonerado=?, preco_referencia=?,
            cbs_percentual=?, ibs_percentual=?, is_percentual=?,
            iva_equivalente=?, preco_sem_tributos=?, encargos_sociais_percentual=?
            WHERE id_preco=?""",
            [d.get('id_data_base') or None, d.get('uf_referencia') or None,
             float(d.get('preco_desonerado') or 0), float(d.get('preco_nao_desonerado') or 0),
             pref, cbs, ibs, isp, iva, psem, enc_soc, existing[0]])
    else:
        db.execute("""INSERT INTO precos_insumos
            (id_insumo, id_data_base, uf_referencia,
             preco_desonerado, preco_nao_desonerado, preco_referencia,
             cbs_percentual, ibs_percentual, is_percentual, iva_equivalente, preco_sem_tributos,
             encargos_sociais_percentual)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            [id_insumo, d.get('id_data_base') or None, d.get('uf_referencia') or None,
             float(d.get('preco_desonerado') or 0), float(d.get('preco_nao_desonerado') or 0),
             pref, cbs, ibs, isp, iva, psem, enc_soc])

def _insumo_codigo_variantes(codigo):
    raw = (codigo or '').strip()
    if not raw:
        return []
    bare = raw.split('.', 1)[1] if '.' in raw else raw
    vals = {raw, bare}
    for prefix in ('SINAPI', 'SICRO', 'SEINFRA', 'SUDECAP', 'GOINFRA', 'CDHU'):
        vals.add(f'{prefix}.{bare}')
    return [v for v in vals if v]

def _insumo_item_tipo(tipo_insumo):
    t = (tipo_insumo or '').strip().lower()
    if 'mão' in t or 'mao' in t:
        return 'MO'
    if 'equip' in t:
        return 'EQUIPAMENTO'
    if 'serv' in t:
        return 'SERVICO'
    return 'INSUMO'

def _insumo_preco_payload(d):
    return float(
        d.get('preco_referencia')
        or d.get('preco_nao_desonerado')
        or d.get('preco_desonerado')
        or 0
    )

def _novo_codigo_preservado(db, codigo):
    base = (codigo or '').strip() or 'INSUMO'
    for i in range(1, 1000):
        cand = f'{base}.REV{i:03d}'
        if not db.execute("SELECT 1 FROM insumos WHERE codigo_insumo=? LIMIT 1", [cand]).fetchone():
            return cand
    return f'{base}.REV'

def _impacto_insumo(db, id_insumo):
    ins = db.execute("SELECT * FROM insumos WHERE id_insumo=?", [id_insumo]).fetchone()
    if not ins:
        return None
    variantes = _insumo_codigo_variantes(ins['codigo_insumo'])
    ph = ','.join('?' * len(variantes)) if variantes else "''"

    comps = {}
    if variantes:
        rows = db.execute(f"""
            SELECT DISTINCT c.id_composicao, c.codigo, c.descricao, c.fonte, c.custo_unitario
            FROM itens_composicao ic
            JOIN composicoes c ON c.id_composicao = ic.id_composicao
            WHERE ic.codigo_item IN ({ph})
              AND COALESCE(ic.tipo_item,'') <> 'COMPOSICAO'
        """, variantes).fetchall()
        for r in rows:
            comps[r['id_composicao']] = dict(r)

    # Sobe a cadeia de composições auxiliares até encontrar composições principais.
    mudou = True
    while mudou and comps:
        mudou = False
        cods = set()
        for c in comps.values():
            cods.update(_insumo_codigo_variantes(c.get('codigo')))
        cods = [c for c in cods if c]
        if not cods:
            break
        phc = ','.join('?' * len(cods))
        rows = db.execute(f"""
            SELECT DISTINCT c.id_composicao, c.codigo, c.descricao, c.fonte, c.custo_unitario
            FROM itens_composicao ic
            JOIN composicoes c ON c.id_composicao = ic.id_composicao
            WHERE ic.tipo_item='COMPOSICAO'
              AND ic.codigo_item IN ({phc})
        """, cods).fetchall()
        for r in rows:
            if r['id_composicao'] not in comps:
                comps[r['id_composicao']] = dict(r)
                mudou = True

    comp_ids = list(comps.keys())
    direct_orc = []
    if variantes:
        direct_orc = rows_to_list(db.execute(f"""
            SELECT DISTINCT os.id_item, os.id_orcamento, o.nome_orcamento, ob.nome_obra,
                   os.codigo, os.descricao, os.custo_unitario
            FROM orcamento_sintetico os
            JOIN orcamentos o ON o.id_orcamento = os.id_orcamento
            LEFT JOIN obras ob ON ob.id_obra = o.id_obra
            WHERE os.id_insumo = ?
               OR os.codigo IN ({ph})
        """, [id_insumo] + variantes).fetchall())

    indirect_orc = []
    if comp_ids:
        phids = ','.join('?' * len(comp_ids))
        indirect_orc = rows_to_list(db.execute(f"""
            SELECT DISTINCT os.id_item, os.id_orcamento, o.nome_orcamento, ob.nome_obra,
                   os.id_composicao, os.codigo, os.descricao, os.custo_unitario
            FROM orcamento_sintetico os
            JOIN orcamentos o ON o.id_orcamento = os.id_orcamento
            LEFT JOIN obras ob ON ob.id_obra = o.id_obra
            WHERE os.id_composicao IN ({phids})
        """, comp_ids).fetchall())

    comps_list = list(comps.values())
    return {
        'insumo': dict(ins),
        'composicoes': comps_list,
        'orcamentos_diretos': direct_orc,
        'orcamentos_indiretos': indirect_orc,
        'total_composicoes': len(comps_list),
        'total_orcamentos_diretos': len({r['id_item'] for r in direct_orc}),
        'total_orcamentos_indiretos': len({r['id_item'] for r in indirect_orc}),
        'tem_impacto': bool(comps_list or direct_orc or indirect_orc),
    }

def _propagar_insumo_em_composicoes(db, insumo_antigo, d, comp_ids):
    if not comp_ids:
        return 0
    variantes = _insumo_codigo_variantes(insumo_antigo['codigo_insumo'])
    if not variantes:
        return 0
    unidade = ''
    if d.get('id_unidade'):
        r = db.execute("SELECT sigla FROM unidades_medida WHERE id_unidade=?", [d.get('id_unidade')]).fetchone()
        unidade = r['sigla'] if r else ''
    preco = _insumo_preco_payload(d)
    tipo_item = _insumo_item_tipo(d.get('tipo_insumo'))
    ph = ','.join('?' * len(variantes))
    cur = db.execute(f"""
        UPDATE itens_composicao
        SET descricao=?,
            unidade=COALESCE(NULLIF(?,''), unidade),
            tipo_item=?,
            preco_unitario=?,
            custo_parcial=COALESCE(coeficiente,0) * ?
        WHERE codigo_item IN ({ph})
          AND COALESCE(tipo_item,'') <> 'COMPOSICAO'
    """, [d.get('descricao', '').strip(), unidade, tipo_item, preco, preco] + variantes)

    # Recalcula somente as composições impactadas a partir dos itens já gravados.
    for cid in comp_ids:
        total = db.execute("""
            SELECT COALESCE(SUM(COALESCE(custo_parcial, COALESCE(coeficiente,0) * COALESCE(preco_unitario,0))),0)
            FROM itens_composicao WHERE id_composicao=?
        """, [cid]).fetchone()[0]
        db.execute("UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?", [round(total or 0, 4), cid])
    return cur.rowcount

def _propagar_insumo_em_orcamentos(db, insumo_antigo, d, comp_ids):
    variantes = _insumo_codigo_variantes(insumo_antigo['codigo_insumo'])
    unidade = ''
    if d.get('id_unidade'):
        r = db.execute("SELECT sigla FROM unidades_medida WHERE id_unidade=?", [d.get('id_unidade')]).fetchone()
        unidade = r['sigla'] if r else ''
    preco = _insumo_preco_payload(d)
    atualizados = 0
    if variantes:
        ph = ','.join('?' * len(variantes))
        cur = db.execute(f"""
            UPDATE orcamento_sintetico
            SET descricao=?, unidade=COALESCE(NULLIF(?,''), unidade), custo_unitario=?
            WHERE id_insumo=? OR codigo IN ({ph})
        """, [d.get('descricao','').strip(), unidade, preco, insumo_antigo['id_insumo']] + variantes)
        atualizados += cur.rowcount
    if comp_ids:
        phids = ','.join('?' * len(comp_ids))
        cur = db.execute(f"""
            UPDATE orcamento_sintetico
            SET custo_unitario = COALESCE((
                SELECT c.custo_unitario FROM composicoes c
                WHERE c.id_composicao = orcamento_sintetico.id_composicao
            ), custo_unitario)
            WHERE id_composicao IN ({phids})
        """, comp_ids)
        atualizados += cur.rowcount
    return atualizados

# ── Grupos ─────────────────────────────────────────────────────────────────────
@app.route('/api/grupos-insumos', methods=['GET'])
def grupos_list():
    db = get_db()
    rows = rows_to_list(db.execute('SELECT * FROM grupos_insumos ORDER BY nome_grupo').fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/grupos-insumos', methods=['POST'])
def grupos_create():
    d = request.json or {}
    if not d.get('nome_grupo','').strip():
        return jsonify({'erro':'Nome do grupo é obrigatório.'}), 400
    db = get_db()
    cur = db.execute('INSERT INTO grupos_insumos (nome_grupo,descricao) VALUES (?,?)',
                     [d['nome_grupo'].strip(), d.get('descricao')])
    db.commit()
    row = dict(db.execute('SELECT * FROM grupos_insumos WHERE id_grupo=?',[cur.lastrowid]).fetchone())
    db.close(); return jsonify(row), 201

@app.route('/api/grupos-insumos/<int:id>', methods=['PUT'])
def grupos_update(id):
    d = request.json or {}
    db = get_db()
    db.execute('UPDATE grupos_insumos SET nome_grupo=?,descricao=? WHERE id_grupo=?',
               [d.get('nome_grupo','').strip(), d.get('descricao'), id])
    db.commit(); db.close(); return jsonify({'ok': True})

@app.route('/api/grupos-insumos/<int:id>', methods=['DELETE'])
def grupos_delete(id):
    db = get_db()
    db.execute('DELETE FROM grupos_insumos WHERE id_grupo=?',[id])
    db.commit(); db.close(); return jsonify({'mensagem':'Grupo excluído.'})

# ── Insumos — CRUD ─────────────────────────────────────────────────────────────
@app.route('/api/insumos', methods=['GET'])
def ins_list():
    db   = get_db()
    ensure_insumos_encargos_schema(db)
    tipo     = request.args.get('tipo','')
    origem   = request.args.get('origem','')
    situacao = request.args.get('situacao','')
    q        = request.args.get('q','')
    uf       = request.args.get('uf','')
    mes      = request.args.get('mes','')
    ano      = request.args.get('ano','')
    regime   = request.args.get('regime','')

    sel, sub_params = _build_sel_ins(uf, mes, ano, regime)
    sql    = sel + " WHERE 1=1"
    params = list(sub_params)          # sub_params posicionais ANTES do WHERE principal

    if tipo:     sql += " AND i.tipo_insumo=?";  params.append(tipo)
    if origem:   sql += " AND i.origem=?";        params.append(origem)
    if situacao: sql += " AND i.situacao=?";      params.append(situacao)
    if q:
        sql += " AND (i.descricao LIKE ? OR i.codigo_insumo LIKE ?)"
        like = f'%{q}%'; params += [like, like]
    if uf or (mes and ano) or regime:  # se filtro ativo, mostrar só quem tem preço correspondente
        sql += " AND p.id_preco IS NOT NULL"
    sql += """ ORDER BY
        CASE i.tipo_insumo
            WHEN 'Material' THEN 0
            WHEN 'MÃ£o de Obra' THEN 1
            WHEN 'Equipamento' THEN 2
            WHEN 'ServiÃ§o Auxiliar' THEN 3
            ELSE 4
        END,
        i.descricao"""

    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/insumos/stats', methods=['GET'])
def ins_stats():
    db = get_db()
    ensure_insumos_encargos_schema(db)
    tipos = ['Material','Mão de Obra','Equipamento','Serviço Auxiliar']
    result = {'total': db.execute('SELECT COUNT(*) FROM insumos').fetchone()[0]}
    for t in tipos:
        key = t.lower().replace(' ','_').replace('ã','a').replace('ç','c')
        result[key] = db.execute('SELECT COUNT(*) FROM insumos WHERE tipo_insumo=?',[t]).fetchone()[0]
    result['com_preco'] = db.execute(
        'SELECT COUNT(DISTINCT id_insumo) FROM precos_insumos').fetchone()[0]
    db.close(); return jsonify(result)

@app.route('/api/insumos/<int:id>', methods=['GET'])
def ins_get(id):
    db = get_db()
    ensure_insumos_encargos_schema(db)
    row = db.execute(SEL_INS + " WHERE i.id_insumo=?", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro':'Insumo não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/insumos/<int:id>/impacto', methods=['GET'])
def ins_impacto(id):
    db = get_db()
    ensure_insumos_encargos_schema(db)
    impacto = _impacto_insumo(db, id)
    db.close()
    if not impacto:
        return jsonify({'erro':'Insumo não encontrado.'}), 404
    # Listas resumidas para a confirmação da interface.
    for k in ('composicoes', 'orcamentos_diretos', 'orcamentos_indiretos'):
        impacto[k] = impacto[k][:12]
    return jsonify(impacto)

@app.route('/api/insumos', methods=['POST'])
def ins_create():
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro':'Descrição é obrigatória.'}), 400
    db = get_db()
    ensure_insumos_encargos_schema(db)
    cur = db.execute("""INSERT INTO insumos
        (codigo_insumo,descricao,tipo_insumo,id_unidade,id_grupo,
         origem,encargos_aplicaveis,encargos_sociais_percentual,situacao,observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        [d.get('codigo_insumo'), d['descricao'].strip(), d.get('tipo_insumo'),
         d.get('id_unidade') or None, d.get('id_grupo') or None,
         d.get('origem'), d.get('encargos_aplicaveis','Sim'),
         float(d.get('encargos_sociais_percentual') or 0) if d.get('encargos_sociais_percentual') not in (None, '') else None,
         d.get('situacao','Ativo'), d.get('observacoes')])
    db.commit()
    new_id = cur.lastrowid
    _save_preco_principal(db, new_id, d)
    db.commit()
    row = db.execute(SEL_INS+' WHERE i.id_insumo=?',[new_id]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/insumos/<int:id>', methods=['PUT'])
def ins_update(id):
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro':'Descrição é obrigatória.'}), 400
    db = get_db()
    ensure_insumos_encargos_schema(db)
    atual = db.execute("SELECT * FROM insumos WHERE id_insumo=?", [id]).fetchone()
    if not atual:
        db.close()
        return jsonify({'erro':'Insumo não encontrado.'}), 404
    modo_impacto = (d.get('modo_impacto') or '').strip()
    impacto = _impacto_insumo(db, id)
    if impacto and impacto.get('tem_impacto') and not modo_impacto:
        db.close()
        return jsonify({'erro':'Escolha como tratar as composições e orçamentos impactados antes de salvar.'}), 409

    if modo_impacto == 'preservar':
        d_novo = dict(d)
        d_novo['codigo_insumo'] = _novo_codigo_preservado(db, d.get('codigo_insumo') or atual['codigo_insumo'])
        cur = db.execute("""INSERT INTO insumos
            (codigo_insumo,descricao,tipo_insumo,id_unidade,id_grupo,
             origem,encargos_aplicaveis,encargos_sociais_percentual,situacao,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            [d_novo.get('codigo_insumo'), d_novo['descricao'].strip(), d_novo.get('tipo_insumo'),
             d_novo.get('id_unidade') or None, d_novo.get('id_grupo') or None,
             d_novo.get('origem'), d_novo.get('encargos_aplicaveis','Sim'),
             float(d_novo.get('encargos_sociais_percentual') or 0) if d_novo.get('encargos_sociais_percentual') not in (None, '') else None,
             d_novo.get('situacao','Ativo'), d_novo.get('observacoes')])
        novo_id = cur.lastrowid
        _save_preco_principal(db, novo_id, d_novo)
        db.commit()
        row = db.execute(SEL_INS+' WHERE i.id_insumo=?',[novo_id]).fetchone()
        db.close()
        resp = dict(row)
        resp['mensagem'] = 'Novo insumo criado; composições e orçamentos existentes foram preservados.'
        resp['id_insumo_original_preservado'] = id
        return jsonify(resp), 201

    cur = db.execute("""UPDATE insumos SET
        codigo_insumo=?,descricao=?,tipo_insumo=?,id_unidade=?,id_grupo=?,
        origem=?,encargos_aplicaveis=?,encargos_sociais_percentual=?,situacao=?,observacoes=?
        WHERE id_insumo=?""",
        [d.get('codigo_insumo'), d['descricao'].strip(), d.get('tipo_insumo'),
         d.get('id_unidade') or None, d.get('id_grupo') or None,
         d.get('origem'), d.get('encargos_aplicaveis','Sim'),
         float(d.get('encargos_sociais_percentual') or 0) if d.get('encargos_sociais_percentual') not in (None, '') else None,
         d.get('situacao','Ativo'), d.get('observacoes'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Insumo não encontrado.'}), 404
    _save_preco_principal(db, id, d)
    comp_ids = [c['id_composicao'] for c in (impacto or {}).get('composicoes', [])]
    itens_atualizados = 0
    orc_atualizados = 0
    if modo_impacto in ('alterar_composicoes', 'alterar_composicoes_orcamentos'):
        itens_atualizados = _propagar_insumo_em_composicoes(db, atual, d, comp_ids)
    if modo_impacto == 'alterar_composicoes_orcamentos':
        orc_atualizados = _propagar_insumo_em_orcamentos(db, atual, d, comp_ids)
    db.commit()
    row = db.execute(SEL_INS+' WHERE i.id_insumo=?',[id]).fetchone()
    resp = dict(row)
    resp['itens_composicao_atualizados'] = itens_atualizados
    resp['itens_orcamento_atualizados'] = orc_atualizados
    db.close(); return jsonify(resp)

@app.route('/api/insumos/<int:id>', methods=['DELETE'])
def ins_delete(id):
    db = get_db()
    ensure_insumos_encargos_schema(db)
    impacto = _impacto_insumo(db, id)
    modo = (request.args.get('modo') or 'preservar').strip()
    if not impacto:
        db.close()
        return jsonify({'erro':'Insumo não encontrado.'}), 404
    if impacto.get('tem_impacto') and modo == 'preservar':
        db.execute("UPDATE insumos SET situacao='Inativo' WHERE id_insumo=?", [id])
        db.commit()
        db.close()
        return jsonify({
            'mensagem':'Insumo inativado. Composições e orçamentos existentes foram preservados.',
            'inativado': True,
            'impacto': {
                'total_composicoes': impacto.get('total_composicoes', 0),
                'total_orcamentos_diretos': impacto.get('total_orcamentos_diretos', 0),
                'total_orcamentos_indiretos': impacto.get('total_orcamentos_indiretos', 0),
            }
        })
    try:
        cur = db.execute('DELETE FROM insumos WHERE id_insumo=?',[id])
        db.commit()
    except sqlite3.IntegrityError:
        db.rollback(); db.close()
        return jsonify({'erro':'Não foi possível excluir: insumo vinculado a composição ou orçamento.'}), 409
    db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Insumo não encontrado.'}), 404
    return jsonify({'mensagem':'Insumo excluído com sucesso.'})

@app.route('/api/insumos/excluir-lote', methods=['POST'])
def ins_excluir_lote():
    d = request.json or {}
    tipo     = d.get('tipo', '')
    origem   = d.get('origem', '')
    situacao = d.get('situacao', '')
    uf       = d.get('uf', '')
    mes      = d.get('mes', '')
    ano      = d.get('ano', '')
    regime   = (d.get('regime') or '').strip().lower()
    id_grupo = d.get('id_grupo', '')
    q        = (d.get('q') or '').strip()
    dry_run  = bool(d.get('dry_run', False))

    if not any([tipo, origem, situacao, uf, mes, ano, regime, id_grupo, q]):
        return jsonify({'erro': 'Informe pelo menos um critério de seleção para excluir.'}), 400

    joins = ""
    where = " WHERE 1=1"
    params = []
    precisa_preco = bool(uf or (mes and ano) or regime)
    if precisa_preco:
        joins += " JOIN precos_insumos p ON p.id_insumo = i.id_insumo"
    if tipo:
        where += " AND i.tipo_insumo=?"; params.append(tipo)
    if origem:
        where += " AND i.origem=?"; params.append(origem)
    if situacao:
        where += " AND i.situacao=?"; params.append(situacao)
    if id_grupo:
        where += " AND i.id_grupo=?"; params.append(int(id_grupo))
    if q:
        where += " AND (i.descricao LIKE ? OR i.codigo_insumo LIKE ?)"
        like = f"%{q}%"; params += [like, like]
    if uf:
        where += " AND p.uf_referencia=?"; params.append(uf)
    if mes and ano:
        where += " AND p.id_data_base IN (SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?)"
        params += [int(mes), int(ano)]
    if regime == 'onerado':
        where += " AND COALESCE(p.preco_nao_desonerado, 0) > 0"
    elif regime == 'desonerado':
        where += " AND COALESCE(p.preco_desonerado, 0) > 0"

    db = get_db()
    ensure_insumos_encargos_schema(db)
    select_ids = "SELECT DISTINCT i.id_insumo FROM insumos i" + joins + where
    total = db.execute("SELECT COUNT(*) FROM (" + select_ids + ")", params).fetchone()[0]
    if dry_run:
        db.close()
        return jsonify({'total': total})
    try:
        cur = db.execute("DELETE FROM insumos WHERE id_insumo IN (" + select_ids + ")", params)
        db.commit()
    except sqlite3.IntegrityError:
        db.rollback(); db.close()
        return jsonify({'erro':'Não foi possível excluir todos os insumos selecionados porque há vínculos em composições ou orçamentos.'}), 409
    excluidos = cur.rowcount
    db.close()
    return jsonify({'excluidos': excluidos, 'mensagem': f'{excluidos} insumo(s) excluído(s) com sucesso.'})

# ── Preços de insumos ──────────────────────────────────────────────────────────
SEL_PRECO = """
    SELECT p.*, db2.mes, db2.ano, db2.descricao AS desc_data_base,
           fr.nome_fonte, um.sigla AS sigla_unidade
    FROM precos_insumos p
    LEFT JOIN datas_base db2 ON p.id_data_base = db2.id_data_base
    LEFT JOIN fontes_referencia fr ON p.id_fonte = fr.id_fonte
    LEFT JOIN insumos i ON p.id_insumo = i.id_insumo
    LEFT JOIN unidades_medida um ON i.id_unidade = um.id_unidade
"""

@app.route('/api/insumos/<int:id_ins>/precos', methods=['GET'])
def precos_list(id_ins):
    db = get_db()
    ensure_insumos_encargos_schema(db)
    rows = rows_to_list(db.execute(
        SEL_PRECO + " WHERE p.id_insumo=? ORDER BY p.id_preco DESC", [id_ins]
    ).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/insumos/<int:id_ins>/precos', methods=['POST'])
def precos_create(id_ins):
    d = request.json or {}
    cbs = float(d.get('cbs_percentual') or 0)
    ibs = float(d.get('ibs_percentual') or 0)
    isp = float(d.get('is_percentual') or 0)
    enc_soc = d.get('encargos_sociais_percentual')
    enc_soc = float(enc_soc) if enc_soc not in (None, '') else None
    iva = round(cbs + ibs + isp, 6)
    pref = float(d.get('preco_referencia') or 0)
    psem = round(pref / (1 + iva/100), 6) if iva > 0 and pref > 0 else pref
    db = get_db()
    ensure_insumos_encargos_schema(db)
    cur = db.execute("""INSERT INTO precos_insumos
        (id_insumo,id_data_base,id_fonte,uf_referencia,
         preco_desonerado,preco_nao_desonerado,preco_referencia,
         cbs_percentual,ibs_percentual,is_percentual,iva_equivalente,preco_sem_tributos,
         encargos_sociais_percentual,data_coleta,observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [id_ins, d.get('id_data_base') or None, d.get('id_fonte') or None,
         d.get('uf_referencia') or None,
         float(d.get('preco_desonerado') or 0), float(d.get('preco_nao_desonerado') or 0),
         pref, cbs, ibs, isp, iva, psem,
         enc_soc, d.get('data_coleta'), d.get('observacoes')])
    db.commit()
    row = db.execute(SEL_PRECO+" WHERE p.id_preco=?",[cur.lastrowid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/precos-insumos/<int:id>', methods=['PUT'])
def precos_update(id):
    d = request.json or {}
    cbs = float(d.get('cbs_percentual') or 0)
    ibs = float(d.get('ibs_percentual') or 0)
    isp = float(d.get('is_percentual') or 0)
    enc_soc = d.get('encargos_sociais_percentual')
    enc_soc = float(enc_soc) if enc_soc not in (None, '') else None
    iva = round(cbs + ibs + isp, 6)
    pref = float(d.get('preco_referencia') or 0)
    psem = round(pref / (1 + iva/100), 6) if iva > 0 and pref > 0 else pref
    db = get_db()
    ensure_insumos_encargos_schema(db)
    cur = db.execute("""UPDATE precos_insumos SET
        id_data_base=?,id_fonte=?,uf_referencia=?,
        preco_desonerado=?,preco_nao_desonerado=?,preco_referencia=?,
        cbs_percentual=?,ibs_percentual=?,is_percentual=?,iva_equivalente=?,preco_sem_tributos=?,
        encargos_sociais_percentual=?,data_coleta=?,observacoes=?
        WHERE id_preco=?""",
        [d.get('id_data_base') or None, d.get('id_fonte') or None, d.get('uf_referencia') or None,
         float(d.get('preco_desonerado') or 0), float(d.get('preco_nao_desonerado') or 0),
         pref, cbs, ibs, isp, iva, psem,
         enc_soc, d.get('data_coleta'), d.get('observacoes'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Preço não encontrado.'}), 404
    row = db.execute(SEL_PRECO+" WHERE p.id_preco=?",[id]).fetchone()
    db.close(); return jsonify(dict(row))

@app.route('/api/precos-insumos/<int:id>', methods=['DELETE'])
def precos_delete(id):
    db = get_db()
    cur = db.execute('DELETE FROM precos_insumos WHERE id_preco=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Preço não encontrado.'}), 404
    return jsonify({'mensagem':'Preço excluído.'})

# ── Pesquisa de mercado ───────────────────────────────────────────────────────
ALIQUOTAS_TRANSICAO_RT = {
    2026: {'cbs': 0.90, 'ibs': 0.10},
    2027: {'cbs': 8.70, 'ibs': 0.10},
    2028: {'cbs': 8.70, 'ibs': 0.10},
    2029: {'cbs': 8.80, 'ibs': 1.77},
    2030: {'cbs': 8.80, 'ibs': 3.54},
    2031: {'cbs': 8.80, 'ibs': 5.31},
    2032: {'cbs': 8.80, 'ibs': 7.08},
    2033: {'cbs': 8.80, 'ibs': 17.70},
}

def _market_float(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).strip().replace('R$', '').replace('%', '').replace(' ', '')
    if ',' in txt and '.' in txt:
        txt = txt.replace('.', '').replace(',', '.')
    elif ',' in txt:
        txt = txt.replace(',', '.')
    try:
        return float(txt)
    except Exception:
        return default

def _norm_text(value):
    import unicodedata
    txt = unicodedata.normalize('NFKD', str(value or ''))
    txt = ''.join(ch for ch in txt if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', txt).strip().lower()

def _read_env_file_value(name):
    """Le OPENAI_API_KEY de .env local quando existir."""
    for filename in ('.env', 'orcasmart.env'):
        path = os.path.join(BASE_DIR, filename)
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as fh:
                for line in fh:
                    line = line.strip()
                    if not line or line.startswith('#') or '=' not in line:
                        continue
                    k, v = line.split('=', 1)
                    if k.strip() == name:
                        return v.strip().strip('"').strip("'")
        except Exception:
            pass
    return ''

def _read_windows_env_value(name):
    """Le variaveis persistidas com setx sem depender do processo atual."""
    if os.name != 'nt':
        return ''
    try:
        import winreg
        locations = [
            (winreg.HKEY_CURRENT_USER, r'Environment'),
            (winreg.HKEY_LOCAL_MACHINE, r'SYSTEM\CurrentControlSet\Control\Session Manager\Environment'),
        ]
        for root, subkey in locations:
            try:
                with winreg.OpenKey(root, subkey) as key:
                    value, _ = winreg.QueryValueEx(key, name)
                    if value:
                        return str(value).strip()
            except OSError:
                continue
    except Exception:
        return ''
    return ''

def _get_config_value(name, default=''):
    return (
        os.environ.get(name, '').strip()
        or _read_windows_env_value(name)
        or _read_env_file_value(name)
        or default
    )

def _get_openai_api_key():
    return _get_config_value('OPENAI_API_KEY')

def _ensure_data_base(db, mes, ano, descricao=None):
    mes = int(mes or date.today().month)
    ano = int(ano or date.today().year)
    row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
    if row:
        return row['id_data_base']
    return db.execute(
        "INSERT INTO datas_base (mes,ano,data_referencia,descricao) VALUES (?,?,?,?)",
        [mes, ano, f"{ano:04d}-{mes:02d}-01", descricao or f"Pesquisa de mercado {mes:02d}/{ano}"]
    ).lastrowid

def _ensure_fonte_cotacao(db):
    row = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='Cotação de Mercado'").fetchone()
    if row:
        return row['id_fonte']
    return db.execute("""INSERT INTO fontes_referencia
        (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes)
        VALUES (?,?,?,?,?)""",
        ['Cotação de Mercado', 'Cotação', 'Pesquisa de mercado do usuário', 'Variável',
         'Fonte criada automaticamente pelo módulo Pesquisa de mercado.']
    ).lastrowid

def _ensure_fonte_compras_gov(db):
    row = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='Compras Governamentais'").fetchone()
    if row:
        return row['id_fonte']
    return db.execute("""INSERT INTO fontes_referencia
        (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes)
        VALUES (?,?,?,?,?)""",
        ['Compras Governamentais', 'Cotação', 'Dados Abertos Compras.gov.br', 'Nacional',
         'Fonte criada automaticamente pelo módulo Pesquisa em Compras Governamentais.']
    ).lastrowid

def _ensure_unidade(db, sigla):
    sigla = (sigla or 'un').strip()[:20] or 'un'
    row = db.execute("SELECT id_unidade FROM unidades_medida WHERE lower(sigla)=lower(?)", [sigla]).fetchone()
    if row:
        return row['id_unidade']
    return db.execute(
        "INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)",
        [sigla, sigla.upper(), 'Pesquisa de mercado']
    ).lastrowid

def _aliquotas_por_ano(ano):
    ano = int(ano or date.today().year)
    if ano <= 2025:
        return {'cbs': 0.0, 'ibs': 0.0}
    if ano >= 2033:
        return ALIQUOTAS_TRANSICAO_RT[2033].copy()
    return ALIQUOTAS_TRANSICAO_RT.get(ano, ALIQUOTAS_TRANSICAO_RT[2026]).copy()

COMPRAS_GOV_BASE = 'https://dadosabertos.compras.gov.br'

def _compras_gov_get(path, params=None, timeout=45):
    import urllib.request, urllib.error, urllib.parse
    query = urllib.parse.urlencode({k: v for k, v in (params or {}).items() if v not in (None, '')})
    url = COMPRAS_GOV_BASE + path + (('?' + query) if query else '')
    req = urllib.request.Request(url, headers={
        'Accept': 'application/json',
        'User-Agent': 'OrcaSmart/1.0 pesquisa-compras-governamentais',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode('utf-8', errors='ignore')
            return _json.loads(raw or '{}')
    except urllib.error.HTTPError as e:
        corpo = e.read().decode(errors='ignore')[:500]
        raise RuntimeError(f"Erro HTTP {e.code} na API Compras.gov.br: {corpo}")
    except urllib.error.URLError as e:
        raise RuntimeError(f"Erro de rede ao consultar Compras.gov.br: {getattr(e, 'reason', e)}")

def _compras_num(v, default=0.0):
    return _market_float(v, default)

def _compras_date_parts(value):
    txt = str(value or '').strip()
    if len(txt) >= 10 and re.match(r'^\d{4}-\d{2}-\d{2}', txt):
        return txt[:10], int(txt[:4]), int(txt[5:7])
    hoje = date.today()
    return hoje.isoformat(), hoje.year, hoje.month

def _compras_normalizar_resultado(item, tipo_catalogo):
    preco = _compras_num(item.get('precoUnitario') or item.get('valorUnitarioHomologado') or item.get('valorUnitarioResultado'))
    data_ref, ano, mes = _compras_date_parts(item.get('dataResultado') or item.get('dataCompra'))
    codigo = item.get('codigoItemCatalogo') or item.get('codigoItem') or item.get('codigoServico') or ''
    unidade = item.get('siglaUnidadeFornecimento') or item.get('siglaUnidadeMedida') or item.get('nomeUnidadeFornecimento') or 'un'
    descricao = item.get('descricaoItem') or item.get('descricaoDetalhadaItem') or item.get('descricaoServico') or item.get('nomePdm') or ''
    detalhe = item.get('descricaoDetalhadaItem') or descricao
    return {
        'id': f"{tipo_catalogo}-{codigo}-{item.get('idCompra') or ''}-{item.get('idItemCompra') or item.get('numeroItemCompra') or ''}",
        'tipo_catalogo': tipo_catalogo,
        'tipo_insumo': 'Serviço Auxiliar' if tipo_catalogo == 'CATSER' else 'Material',
        'codigo_catalogo': str(codigo or ''),
        'descricao': descricao,
        'descricao_detalhada': detalhe,
        'unidade': str(unidade or 'un').strip()[:20],
        'preco': preco,
        'quantidade': _compras_num(item.get('quantidade')),
        'fornecedor': item.get('nomeFornecedor') or '',
        'marca': item.get('marca') or '',
        'uasg': item.get('codigoUasg') or '',
        'orgao': item.get('nomeOrgao') or item.get('nomeUasg') or '',
        'municipio': item.get('municipio') or '',
        'uf': item.get('estado') or '',
        'data_resultado': data_ref,
        'mes': mes,
        'ano': ano,
        'id_compra': item.get('idCompra') or '',
        'id_item_compra': item.get('idItemCompra') or '',
        'objeto_compra': item.get('objetoCompra') or '',
        'fonte_url': COMPRAS_GOV_BASE,
    }

def _compras_buscar_precos_por_codigo(codigo, tipo='material', uf='', data_inicio='', data_fim='', limite=20):
    tipo = (tipo or 'material').lower()
    if tipo.startswith('serv'):
        path = '/modulo-pesquisa-preco/3_consultarServico'
        tipo_catalogo = 'CATSER'
    else:
        path = '/modulo-pesquisa-preco/1_consultarMaterial'
        tipo_catalogo = 'CATMAT'
    data = _compras_gov_get(path, {
        'pagina': 1,
        'tamanhoPagina': max(10, min(100, int(limite or 20))),
        'codigoItemCatalogo': codigo,
        'estado': (uf or '').strip().upper(),
        'dataCompraInicio': data_inicio,
        'dataCompraFim': data_fim,
    })
    resultados = data.get('resultado') or []
    return [_compras_normalizar_resultado(r, tipo_catalogo) for r in resultados][:limite]

def _compras_catalogo_material_por_descricao(termo, limite=12):
    # O endpoint oficial de catálogo aceita descricaoItem em algumas instalações.
    # Quando o filtro remoto não encontra, filtramos a primeira página localmente como fallback.
    params = {'pagina': 1, 'tamanhoPagina': max(10, min(100, int(limite or 12))), 'descricaoItem': termo}
    data = _compras_gov_get('/modulo-material/4_consultarItemMaterial', params)
    rows = data.get('resultado') or []
    if not rows:
        data = _compras_gov_get('/modulo-material/4_consultarItemMaterial', {'pagina': 1, 'tamanhoPagina': 100})
        termo_norm = _norm_text(termo)
        rows = [r for r in (data.get('resultado') or []) if termo_norm in _norm_text(r.get('descricaoItem') or r.get('nomePdm') or '')]
    out = []
    for r in rows[:limite]:
        out.append({
            'tipo_catalogo': 'CATMAT',
            'codigo_catalogo': str(r.get('codigoItem') or ''),
            'descricao': r.get('descricaoItem') or r.get('nomePdm') or '',
            'descricao_detalhada': r.get('descricaoItem') or '',
            'unidade': 'un',
            'preco': 0,
            'quantidade': 0,
            'fornecedor': '',
            'marca': '',
            'uasg': '',
            'orgao': 'Catálogo de Materiais Compras.gov.br',
            'municipio': '',
            'uf': '',
            'data_resultado': '',
            'mes': '',
            'ano': '',
            'objeto_compra': r.get('nomeClasse') or r.get('nomeGrupo') or '',
            'fonte_url': COMPRAS_GOV_BASE,
            'catalogo_sem_preco': True,
        })
    return out

def _compras_resultados_por_termo(termo, tipo='todos', uf='', data_inicio='', data_fim='', limite=20):
    termo = (termo or '').strip()
    if not termo:
        return [], ['Informe uma descrição ou um código CATMAT/CATSER.']
    limite = max(1, min(50, int(limite or 20)))
    uf = (uf or '').strip().upper()
    avisos = []
    resultados = []

    m = re.search(r'\d{4,}', termo)
    if m:
        codigo = m.group(0)
        tipos = ['material', 'servico'] if tipo in ('todos', '') else [tipo]
        for tp in tipos:
            try:
                resultados.extend(_compras_buscar_precos_por_codigo(codigo, tp, uf, data_inicio, data_fim, limite))
            except Exception as exc:
                avisos.append(str(exc))
        if resultados:
            return resultados[:limite], avisos

    if tipo in ('todos', 'material', ''):
        try:
            catalogo = _compras_catalogo_material_por_descricao(termo, min(10, limite))
            for c in catalogo:
                if c.get('codigo_catalogo'):
                    precos = _compras_buscar_precos_por_codigo(c['codigo_catalogo'], 'material', uf, data_inicio, data_fim, 8)
                    resultados.extend(precos or [c])
        except Exception as exc:
            avisos.append(f"Catálogo de materiais: {exc}")

    # A API pública atual expõe preços de serviços por código CATSER; busca textual
    # de CATSER não está estável no servidor real. Mantemos o aviso para orientar o usuário.
    if tipo in ('todos', 'servico'):
        avisos.append('Para serviços, informe o código CATSER quando disponível; a busca textual pública ainda não retorna catálogo de serviços de forma consistente.')

    termo_norm = _norm_text(termo)
    filtrados = []
    vistos = set()
    for r in resultados:
        hay = _norm_text(' '.join(str(r.get(k) or '') for k in ('descricao', 'descricao_detalhada', 'objeto_compra', 'fornecedor', 'orgao')))
        if termo_norm and termo_norm not in hay and not str(r.get('codigo_catalogo', '')).startswith(termo):
            continue
        key = r.get('id') or (r.get('codigo_catalogo'), r.get('preco'), r.get('fornecedor'), r.get('data_resultado'))
        if key in vistos:
            continue
        vistos.add(key)
        filtrados.append(r)
    return filtrados[:limite], avisos

@app.route('/api/compras-gov/pesquisar', methods=['POST'])
def compras_gov_pesquisar():
    d = request.json or {}
    termo = (d.get('termo') or '').strip()
    if not termo:
        return jsonify({'erro': 'Informe uma descrição ou código CATMAT/CATSER.'}), 400
    try:
        resultados, avisos = _compras_resultados_por_termo(
            termo,
            d.get('tipo') or 'todos',
            d.get('uf') or '',
            d.get('data_inicio') or '',
            d.get('data_fim') or '',
            int(d.get('limite') or 20),
        )
        if not resultados and not avisos:
            avisos.append('Nenhum preço público encontrado. Tente informar um código CATMAT/CATSER ou ampliar o período/UF.')
        return jsonify({
            'termo': termo,
            'fonte': 'Dados Abertos Compras.gov.br',
            'resultados': resultados,
            'avisos': avisos,
        })
    except Exception as exc:
        return jsonify({'erro': str(exc)}), 502

@app.route('/api/compras-gov/importar', methods=['POST'])
def compras_gov_importar():
    d = request.json or {}
    descricao = (d.get('descricao') or d.get('descricao_detalhada') or '').strip()
    if not descricao:
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    preco = _market_float(d.get('preco'))
    if preco <= 0:
        return jsonify({'erro': 'Selecione ou informe um preço unitário maior que zero.'}), 400
    tipo = d.get('tipo_insumo') or ('Serviço Auxiliar' if d.get('tipo_catalogo') == 'CATSER' else 'Material')
    if tipo not in ('Material', 'Mão de Obra', 'Equipamento', 'Serviço Auxiliar'):
        tipo = 'Material'

    data_ref, ano, mes = _compras_date_parts(d.get('data_resultado') or d.get('data_pesquisa'))
    aliq = _aliquotas_por_ano(ano)
    cbs = _market_float(d.get('cbs_percentual'), aliq['cbs'])
    ibs = _market_float(d.get('ibs_percentual'), aliq['ibs'])
    isp = _market_float(d.get('is_percentual'), 0)
    iva = round(cbs + ibs + isp, 6)
    psem = round(preco / (1 + iva/100), 6) if iva > 0 else preco
    regime = (d.get('regime') or 'Onerado').strip()

    codigo_catalogo = (d.get('codigo_catalogo') or '').strip()
    codigo = (d.get('codigo_insumo') or '').strip()
    if not codigo:
        prefixo = d.get('tipo_catalogo') or 'CG'
        codigo = f"{prefixo}-{codigo_catalogo or abs(hash(descricao)) % 100000}"

    observacoes = '\n'.join(x for x in [
        'Importado pelo módulo Pesquisa em Compras Governamentais.',
        f"Fonte: Dados Abertos Compras.gov.br",
        f"Catálogo: {d.get('tipo_catalogo') or ''} {codigo_catalogo}".strip(),
        f"Fornecedor: {d.get('fornecedor') or ''}".strip(),
        f"Marca: {d.get('marca') or ''}".strip(),
        f"Órgão/UASG: {d.get('orgao') or ''} {d.get('uasg') or ''}".strip(),
        f"Município/UF: {d.get('municipio') or ''}/{d.get('uf') or ''}".strip('/'),
        f"Compra: {d.get('id_compra') or ''} - Item: {d.get('id_item_compra') or ''}".strip(),
        f"Objeto: {d.get('objeto_compra') or ''}".strip(),
    ] if x and not x.endswith(': '))

    db = get_db()
    try:
        id_data_base = _ensure_data_base(db, mes, ano, f"Compras Governamentais {mes:02d}/{ano}")
        id_fonte = _ensure_fonte_compras_gov(db)
        id_unidade = _ensure_unidade(db, d.get('unidade') or 'un')
        cur = db.execute("""INSERT INTO insumos
            (codigo_insumo,descricao,tipo_insumo,id_unidade,id_grupo,
             origem,encargos_aplicaveis,situacao,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            [codigo, descricao, tipo, id_unidade, d.get('id_grupo') or None,
             'Cotação', 'Sim' if tipo == 'Mão de Obra' else 'Não',
             'Ativo', observacoes])
        id_insumo = cur.lastrowid
        preco_des = preco if regime.lower().startswith('des') else 0
        preco_on = preco if not regime.lower().startswith('des') else 0
        db.execute("""INSERT INTO precos_insumos
            (id_insumo,id_data_base,id_fonte,uf_referencia,
             preco_desonerado,preco_nao_desonerado,preco_referencia,
             cbs_percentual,ibs_percentual,is_percentual,iva_equivalente,preco_sem_tributos,
             data_coleta,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [id_insumo, id_data_base, id_fonte, d.get('uf_referencia') or d.get('uf') or None,
             preco_des, preco_on, preco,
             cbs, ibs, isp, iva, psem,
             data_ref, observacoes])
        db.commit()
        row = db.execute(SEL_INS + " WHERE i.id_insumo=?", [id_insumo]).fetchone()
        return jsonify(dict(row)), 201
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

def _extract_openai_text(data):
    if isinstance(data, dict) and data.get('output_text'):
        return data.get('output_text') or ''
    parts = []
    for out in (data.get('output') or []):
        for c in (out.get('content') or []):
            if isinstance(c, dict):
                if c.get('type') in ('output_text', 'text'):
                    parts.append(c.get('text') or '')
    return '\n'.join(p for p in parts if p).strip()

def _call_openai_market_research(termo, tipo, uf, mes, ano):
    """Pesquisa assistida por IA. Usa OPENAI_API_KEY quando configurada."""
    import urllib.request, urllib.error
    api_key = _get_openai_api_key()
    if not api_key:
        raise EnvironmentError(
            "OPENAI_API_KEY não configurada. A pesquisa assistida por IA fica disponível "
            "quando a variável de ambiente for definida."
        )

    model = _get_config_value('OPENAI_MODEL', 'gpt-4o-mini')
    prompt = f"""
Pesquise preços de mercado no Brasil para cadastrar insumo de orçamento público.
Termo pesquisado: {termo}
Tipo pretendido: {tipo or 'a definir'}
UF de referência: {uf or 'não informada'}
Data-base: {int(mes):02d}/{int(ano)}

Retorne somente JSON, sem markdown, neste formato:
{{
  "resultados": [
    {{
      "nome": "descrição curta do bem/serviço",
      "descricao": "descrição técnica para cadastro",
      "tipo_sugerido": "Material|Equipamento|Mão de Obra|Serviço Auxiliar",
      "unidade": "un",
      "preco": 0.0,
      "moeda": "BRL",
      "fornecedor": "nome do fornecedor ou marketplace",
      "marca_modelo": "marca/modelo quando houver",
      "uf": "UF quando houver",
      "url": "URL pública consultada",
      "imagem_url": "URL de foto pública quando houver",
      "especificacoes": ["até 8 especificações objetivas"],
      "observacoes": "nota curta sobre validade, frete, impostos ou incerteza",
      "confianca": "Alta|Média|Baixa"
    }}
  ],
  "avisos": ["limitações relevantes da pesquisa"]
}}

Priorize fontes com preço visível, fornecedor identificável, URL e imagem. Não invente preço:
se o preço não estiver claro, deixe preco=0 e explique em observacoes.
"""
    body_base = {
        'model': model,
        'input': [
            {'role': 'system', 'content': 'Você é um pesquisador técnico de preços para orçamento de obras públicas.'},
            {'role': 'user', 'content': prompt},
        ],
    }

    def do_request(body):
        req = urllib.request.Request(
            'https://api.openai.com/v1/responses',
            data=_json.dumps(body).encode('utf-8'),
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            }
        )
        with urllib.request.urlopen(req, timeout=180) as resp:
            return _json.loads(resp.read())

    try:
        try:
            data = do_request({**body_base, 'tools': [{'type': 'web_search_preview'}]})
        except urllib.error.HTTPError as e:
            if e.code not in (400, 404):
                raise
            data = do_request(body_base)
        text = _extract_openai_text(data)
        parsed = _clean_json(text)
        resultados = parsed.get('resultados') if isinstance(parsed, dict) else []
        if not isinstance(resultados, list):
            resultados = []
        return {
            'provedor': 'OpenAI',
            'busca_web': True,
            'resultados': resultados[:12],
            'avisos': parsed.get('avisos', []) if isinstance(parsed, dict) else [],
        }
    except urllib.error.HTTPError as e:
        corpo = e.read().decode(errors='ignore')[:600]
        raise RuntimeError(f"Erro HTTP {e.code} na API OpenAI: {corpo}")
    except urllib.error.URLError as e:
        reason = getattr(e, "reason", e)
        reason_text = str(reason)
        if "WinError 10013" in reason_text:
            raise RuntimeError(
                "A chamada externa para a API OpenAI foi bloqueada pelo Windows ou pelo ambiente "
                "em que o servidor foi iniciado (WinError 10013). Reinicie o servidor com permissão de rede "
                "e verifique firewall/antivírus/proxy."
            )
        raise RuntimeError(f"Erro de rede ao chamar a API OpenAI: {reason_text}")

def _call_anthropic_market_research(termo, tipo, uf, mes, ano):
    """Fallback para a IA ja configurada no sistema via ANTHROPIC_API_KEY."""
    if not os.environ.get('ANTHROPIC_API_KEY', '').strip():
        raise EnvironmentError(
            "Nenhuma chave de IA configurada. Defina OPENAI_API_KEY para pesquisa com busca web "
            "ou ANTHROPIC_API_KEY para pesquisa assistida."
        )

    prompt = f"""
Voce esta no modulo Pesquisa de mercado do OrcaSmart. Estruture opcoes para cotacao de um insumo.

Termo pesquisado: {termo}
Tipo pretendido: {tipo or 'a definir'}
UF de referencia: {uf or 'nao informada'}
Data-base: {int(mes):02d}/{int(ano)}

Retorne somente JSON, sem markdown:
{{
  "resultados": [
    {{
      "nome": "descricao curta",
      "descricao": "descricao tecnica para cadastro",
      "tipo_sugerido": "Material|Equipamento|Mão de Obra|Serviço Auxiliar",
      "unidade": "un",
      "preco": 0.0,
      "moeda": "BRL",
      "fornecedor": "",
      "marca_modelo": "",
      "uf": "{uf or ''}",
      "url": "",
      "imagem_url": "",
      "especificacoes": ["especificacao objetiva"],
      "observacoes": "informe claramente se o preco precisa ser confirmado em fonte externa",
      "confianca": "Baixa"
    }}
  ],
  "avisos": [
    "Fallback via Claude sem ferramenta de busca web no servidor; confirme preco, fornecedor e URL antes de importar."
  ]
}}

Regras:
- Nao invente URL.
- Se nao houver preco verificavel, use preco 0.
- Sugira ate 6 alternativas tecnicas realistas, com especificacoes uteis ao comprador.
- Se souber faixas historicas aproximadas, coloque preco 0 e descreva a faixa em observacoes, deixando claro que precisa confirmacao.
"""
    text = _call_claude_ia([
        {'role': 'user', 'content': prompt}
    ], max_tokens=4000)
    parsed = _clean_json(text)
    resultados = parsed.get('resultados') if isinstance(parsed, dict) else []
    if not isinstance(resultados, list):
        resultados = []
    avisos = parsed.get('avisos', []) if isinstance(parsed, dict) else []
    if not avisos:
        avisos = ['Fallback via Claude sem busca web; confirme preço e fonte antes de importar.']
    return {
        'provedor': 'Claude',
        'busca_web': False,
        'resultados': resultados[:12],
        'avisos': avisos,
    }

def _call_market_research(termo, tipo, uf, mes, ano):
    if _get_openai_api_key():
        return _call_openai_market_research(termo, tipo, uf, mes, ano)
    if os.environ.get('ANTHROPIC_API_KEY', '').strip():
        return _call_anthropic_market_research(termo, tipo, uf, mes, ano)
    raise EnvironmentError(
        "Nenhuma chave de IA configurada. Defina OPENAI_API_KEY para pesquisa com busca web "
        "ou ANTHROPIC_API_KEY para pesquisa assistida."
    )

@app.route('/api/pesquisa-mercado/parametros', methods=['GET'])
def pesquisa_mercado_parametros():
    hoje = date.today()
    aliq = _aliquotas_por_ano(hoje.year)
    return jsonify({
        'data_pesquisa': hoje.isoformat(),
        'mes': hoje.month,
        'ano': hoje.year,
        'cbs_percentual': aliq['cbs'],
        'ibs_percentual': aliq['ibs'],
        'openai_configurado': bool(_get_openai_api_key()),
        'anthropic_configurado': bool(os.environ.get('ANTHROPIC_API_KEY', '').strip()),
        'provedor_ia': 'OpenAI' if _get_openai_api_key()
                      else ('Claude' if os.environ.get('ANTHROPIC_API_KEY', '').strip() else None),
        'busca_web_configurada': bool(_get_openai_api_key()),
    })

@app.route('/api/pesquisa-mercado/pesquisar', methods=['POST'])
def pesquisa_mercado_pesquisar():
    d = request.json or {}
    termo = (d.get('termo') or '').strip()
    if not termo:
        return jsonify({'erro': 'Informe o bem ou serviço a pesquisar.'}), 400
    hoje = date.today()
    mes = int(d.get('mes') or hoje.month)
    ano = int(d.get('ano') or hoje.year)
    try:
        res = _call_market_research(
            termo,
            d.get('tipo') or '',
            d.get('uf') or '',
            mes,
            ano,
        )
        return jsonify({
            'modo': 'ia',
            'termo': termo,
            'mes': mes,
            'ano': ano,
            **res,
        })
    except Exception as e:
        return jsonify({
            'modo': 'manual',
            'termo': termo,
            'mes': mes,
            'ano': ano,
            'resultados': [],
            'avisos': [str(e)],
            'mensagem': (
                'Pesquisa por IA indisponível neste momento. Você ainda pode cadastrar a cotação '
                'manualmente na mesma tela.'
            )
        })

@app.route('/api/pesquisa-mercado/importar', methods=['POST'])
def pesquisa_mercado_importar():
    d = request.json or {}
    descricao = (d.get('descricao') or d.get('nome') or '').strip()
    if not descricao:
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    tipo = d.get('tipo_insumo') or d.get('tipo_sugerido') or 'Material'
    if tipo not in ('Material', 'Mão de Obra', 'Equipamento', 'Serviço Auxiliar'):
        tipo = 'Material'
    preco = _market_float(d.get('preco_referencia') or d.get('preco'))
    if preco <= 0:
        return jsonify({'erro': 'Informe um preço válido para importar a cotação.'}), 400

    hoje = date.today()
    data_pesquisa = d.get('data_pesquisa') or hoje.isoformat()
    mes = int(d.get('mes') or (int(data_pesquisa[5:7]) if len(data_pesquisa) >= 7 else hoje.month))
    ano = int(d.get('ano') or (int(data_pesquisa[:4]) if len(data_pesquisa) >= 4 else hoje.year))
    aliq = _aliquotas_por_ano(ano)
    cbs = _market_float(d.get('cbs_percentual'), aliq['cbs'])
    ibs = _market_float(d.get('ibs_percentual'), aliq['ibs'])
    isp = _market_float(d.get('is_percentual'), 0)
    iva = round(cbs + ibs + isp, 6)
    psem = round(preco / (1 + iva/100), 6) if iva > 0 else preco
    regime = (d.get('regime') or 'Onerado').strip()

    specs = d.get('especificacoes') or []
    if isinstance(specs, list):
        specs_txt = '; '.join(str(x) for x in specs if str(x).strip())
    else:
        specs_txt = str(specs)
    obs_partes = [
        'Importado pelo módulo Pesquisa de mercado.',
        f"Termo pesquisado: {d.get('termo') or ''}".strip(),
        f"Fornecedor: {d.get('fornecedor') or ''}".strip(),
        f"Marca/modelo: {d.get('marca_modelo') or ''}".strip(),
        f"URL: {d.get('url') or ''}".strip(),
        f"Imagem: {d.get('imagem_url') or ''}".strip(),
        f"Especificações: {specs_txt}".strip(),
        f"Observações da pesquisa: {d.get('observacoes') or ''}".strip(),
    ]
    observacoes = '\n'.join(p for p in obs_partes if not p.endswith(': '))

    db = get_db()
    try:
        id_data_base = _ensure_data_base(db, mes, ano)
        id_fonte = _ensure_fonte_cotacao(db)
        id_unidade = _ensure_unidade(db, d.get('unidade') or 'un')
        codigo = (d.get('codigo_insumo') or '').strip()
        if not codigo:
            codigo = f"COT-{ano}{mes:02d}-{date.today().strftime('%d')}-{abs(hash(descricao)) % 100000:05d}"

        cur = db.execute("""INSERT INTO insumos
            (codigo_insumo,descricao,tipo_insumo,id_unidade,id_grupo,
             origem,encargos_aplicaveis,situacao,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            [codigo, descricao, tipo, id_unidade, d.get('id_grupo') or None,
             'Cotação', 'Sim' if tipo == 'Mão de Obra' else 'Não',
             'Ativo', observacoes])
        id_insumo = cur.lastrowid

        preco_des = preco if regime.lower().startswith('des') else 0
        preco_on = preco if not regime.lower().startswith('des') else 0
        db.execute("""INSERT INTO precos_insumos
            (id_insumo,id_data_base,id_fonte,uf_referencia,
             preco_desonerado,preco_nao_desonerado,preco_referencia,
             cbs_percentual,ibs_percentual,is_percentual,iva_equivalente,preco_sem_tributos,
             data_coleta,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [id_insumo, id_data_base, id_fonte, d.get('uf_referencia') or d.get('uf') or None,
             preco_des, preco_on, preco,
             cbs, ibs, isp, iva, psem,
             data_pesquisa, observacoes])
        db.commit()
        row = db.execute(SEL_INS + " WHERE i.id_insumo=?", [id_insumo]).fetchone()
        return jsonify(dict(row)), 201
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 3 — ENCARGOS SOCIAIS
# ═══════════════════════════════════════════════════════════════════════════════

def _calc_encargos(db, pid, recalc_d=False):
    """
    Atualiza os totais de um perfil.
    Por padrão apenas soma os itens armazenados (incluindo D1/D2 do SINAPI).
    Se recalc_d=True, recalcula D usando a fórmula simplificada A×(B+C).
    """
    def soma_grupo(letra):
        row = db.execute("""
            SELECT COALESCE(SUM(ie.percentual),0)
            FROM itens_encargo ie
            JOIN grupos_encargos ge ON ie.id_grupo_enc = ge.id_grupo_enc
            WHERE ge.id_perfil=? AND ge.letra=?""", [pid, letra]).fetchone()
        return row[0] if row else 0.0

    A = soma_grupo('A')
    B = soma_grupo('B')
    C = soma_grupo('C')

    if recalc_d:
        # Fórmula simplificada: D = (1+A/100)×B - B + (1+A/100)×C - C
        fator = 1 + A / 100.0
        d_sobre_b = round(fator * B - B, 6)
        d_sobre_c = round(fator * C - C, 6)
        D = round(d_sobre_b + d_sobre_c, 6)
        # Atualizar itens D1 e D2
        gd = db.execute("SELECT id_grupo_enc FROM grupos_encargos WHERE id_perfil=? AND letra='D'",
                        [pid]).fetchone()
        if gd:
            itens_d = db.execute(
                "SELECT id_item FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem",
                [gd[0]]).fetchall()
            if len(itens_d) >= 1:
                db.execute("UPDATE itens_encargo SET percentual=? WHERE id_item=?",
                           [round(d_sobre_b, 6), itens_d[0][0]])
            if len(itens_d) >= 2:
                db.execute("UPDATE itens_encargo SET percentual=? WHERE id_item=?",
                           [round(d_sobre_c, 6), itens_d[1][0]])
            db.execute("UPDATE grupos_encargos SET total_grupo=? WHERE id_grupo_enc=?", [D, gd[0]])
    else:
        D = soma_grupo('D')

    total = round(A + B + C + D, 6)

    # Persiste totais no perfil e grupos
    db.execute("""UPDATE perfis_encargos
        SET total_grupo_a=?,total_grupo_b=?,total_grupo_c=?,
            total_grupo_d=?,encargo_total=?
        WHERE id_perfil=?""", [round(A,6), round(B,6), round(C,6), round(D,6), total, pid])
    for letra, val in [('A', A), ('B', B), ('C', C)]:
        db.execute("UPDATE grupos_encargos SET total_grupo=? WHERE id_perfil=? AND letra=?",
                   [round(val,6), pid, letra])
    db.commit()
    return {'A': round(A,4), 'B': round(B,4), 'C': round(C,4),
            'D': round(D,4), 'total': round(total,4)}

SEL_PERFIL = """
    SELECT pe.*,
           db2.mes AS db_mes, db2.ano AS db_ano
    FROM perfis_encargos pe
    LEFT JOIN datas_base db2 ON pe.id_data_base = db2.id_data_base
"""

# ── Perfis ─────────────────────────────────────────────────────────────────────
@app.route('/api/encargos/perfis', methods=['GET'])
def enc_perfis_list():
    uf       = request.args.get('uf','')
    fonte    = request.args.get('fonte','')
    cat      = request.args.get('categoria','')
    reg      = request.args.get('regime','')
    situacao = request.args.get('situacao','')
    vig_ini_mes = request.args.get('vigencia_inicio_mes','')
    vig_fim_mes = request.args.get('vigencia_fim_mes','')
    q        = request.args.get('q','')

    sql = SEL_PERFIL + " WHERE 1=1"
    params = []
    if fonte:    sql += " AND UPPER(COALESCE(pe.fonte_referencia,''))=?"; params.append(fonte.upper())
    if uf:       sql += " AND pe.uf_referencia=?";  params.append(uf)
    if cat and not cat.startswith('Profissional'):
        sql += " AND pe.categoria=?";       params.append(cat)
    if reg:      sql += " AND pe.regime=?";          params.append(reg)
    if situacao: sql += " AND pe.situacao=?";        params.append(situacao)
    if vig_ini_mes:
        sql += " AND substr(COALESCE(pe.vigencia_inicio,''),1,7)=?"
        params.append(vig_ini_mes)
    if vig_fim_mes:
        sql += " AND substr(COALESCE(pe.vigencia_fim,''),1,7)=?"
        params.append(vig_fim_mes)
    if q:
        sql += " AND pe.nome_perfil LIKE ?";         params.append(f'%{q}%')
    sql += " ORDER BY pe.fonte_referencia, pe.uf_referencia, pe.categoria, pe.regime, pe.vigencia_inicio"

    db = get_db()
    ensure_encargos_schema(db)
    # Recalcula totais (sem tocar em D) para manter consistência
    pids = [r[0] for r in db.execute("SELECT id_perfil FROM perfis_encargos").fetchall()]
    for pid in pids:
        _calc_encargos(db, pid, recalc_d=False)
    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close()
    return jsonify(rows)

ENCARGOS_REF_VALORES = {
    'Desonerado': {
        'Horista': {
            'totais': {'A': 16.80, 'B': 48.36, 'C': 10.70, 'D': 8.58},
            'itens': {
                'A': [('A1','INSS',0.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00)],
                'B': [('B1','Descanso Semanal Remunerado',17.85),('B2','Feriados',3.71),('B3','Auxílio Enfermidade',0.87),('B4','13º Salário',11.03),('B5','Licença Paternidade',0.07),('B6','Faltas Justificadas',0.74),('B7','Dias de Chuvas',1.59),('B8','Auxílio Acidente de Trabalho',0.11),('B9','Férias Gozadas',12.35),('B10','Salário Maternidade',0.04)],
                'C': [('C1','Aviso Prévio Indenizado',5.52),('C2','Aviso Prévio Trabalhado',0.13),('C3','Férias Indenizadas',1.72),('C4','Depósito de rescisão sem justa causa',2.87),('C5','Indenização Adicional',0.46)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',8.12),('D2','Reincidência de Grupo A sobre aviso prévio trabalhado e FGTS sobre aviso prévio indenizado',0.46)],
            },
        },
        'Mensalista': {
            'totais': {'A': 16.80, 'B': 19.04, 'C': 8.09, 'D': 3.55},
            'itens': {
                'A': [('A1','INSS',0.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00)],
                'B': [('B1','Descanso Semanal Remunerado',0.00),('B2','Feriados',0.00),('B3','Auxílio Enfermidade',0.66),('B4','13º Salário',8.33),('B5','Licença Paternidade',0.05),('B6','Faltas Justificadas',0.56),('B7','Dias de Chuvas',0.00),('B8','Auxílio Acidente de Trabalho',0.08),('B9','Férias Gozadas',9.33),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',4.17),('C2','Aviso Prévio Trabalhado',0.10),('C3','Férias Indenizadas',1.30),('C4','Depósito de rescisão sem justa causa',2.17),('C5','Indenização Adicional',0.35)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',3.20),('D2','Reincidência de Grupo A sobre aviso prévio trabalhado e FGTS sobre aviso prévio indenizado',0.35)],
            },
        },
    },
    'Normal': {
        'Horista': {
            'totais': {'A': 36.80, 'B': 48.36, 'C': 10.70, 'D': 18.29},
            'itens': {
                'A': [('A1','INSS',20.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00)],
                'B': [('B1','Descanso Semanal Remunerado',17.85),('B2','Feriados',3.71),('B3','Auxílio Enfermidade',0.87),('B4','13º Salário',11.03),('B5','Licença Paternidade',0.07),('B6','Faltas Justificadas',0.74),('B7','Dias de Chuvas',1.59),('B8','Auxílio Acidente de Trabalho',0.11),('B9','Férias Gozadas',12.35),('B10','Salário Maternidade',0.04)],
                'C': [('C1','Aviso Prévio Indenizado',5.52),('C2','Aviso Prévio Trabalhado',0.13),('C3','Férias Indenizadas',1.72),('C4','Depósito de rescisão sem justa causa',2.87),('C5','Indenização Adicional',0.46)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',17.80),('D2','Reincidência de Grupo A sobre aviso prévio trabalhado e FGTS sobre aviso prévio indenizado',0.49)],
            },
        },
        'Mensalista': {
            'totais': {'A': 36.80, 'B': 19.04, 'C': 8.09, 'D': 7.38},
            'itens': {
                'A': [('A1','INSS',20.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00)],
                'B': [('B1','Descanso Semanal Remunerado',0.00),('B2','Feriados',0.00),('B3','Auxílio Enfermidade',0.66),('B4','13º Salário',8.33),('B5','Licença Paternidade',0.05),('B6','Faltas Justificadas',0.56),('B7','Dias de Chuvas',0.00),('B8','Auxílio Acidente de Trabalho',0.08),('B9','Férias Gozadas',9.33),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',4.17),('C2','Aviso Prévio Trabalhado',0.10),('C3','Férias Indenizadas',1.30),('C4','Depósito de rescisão sem justa causa',2.17),('C5','Indenização Adicional',0.35)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',7.01),('D2','Reincidência de Grupo A sobre aviso prévio trabalhado e FGTS sobre aviso prévio indenizado',0.37)],
            },
        },
    },
}

ENCARGOS_REF_GRUPOS = {
    'A': 'Encargos Básicos',
    'B': 'Encargos sobre Tempo Trabalhado',
    'C': 'Encargos Rescisórios',
    'D': 'Reincidências',
}

SUDECAP_ENCARGOS_2025 = {
    'Desonerado': {
        'Horista': {
            'totais': {'A': 23.00, 'B': 49.23, 'C': 10.32, 'D': 11.28},
            'itens': {
                'A': [('A1','INSS',5.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro Contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00),('A9','SECONCI',1.20)],
                'B': [('B1','Repouso Semanal Remunerado',17.76),('B2','Feriados',3.68),('B3','Auxílio - Enfermidade',0.86),('B4','13º Salário',11.06),('B5','Licença Paternidade',0.07),('B6','Faltas Justificadas',0.74),('B7','Dias de Chuvas',1.09),('B8','Auxílio Acidente de Trabalho',0.10),('B9','Férias Gozadas',13.84),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',5.97),('C2','Aviso Prévio Trabalhado',0.14),('C3','Férias Indenizadas',0.93),('C4','Depósito Rescisão Sem Justa Causa',2.78),('C5','Indenização Adicional',0.50)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',10.77),('D2','Reincidência de Grupo A sobre Aviso Prévio Trabalhado e Reincidência do FGTS sobre Aviso Prévio Indenizado',0.51)],
            },
        },
        'Mensalista': {
            'totais': {'A': 23.00, 'B': 20.11, 'C': 7.78, 'D': 4.60},
            'itens': {
                'A': [('A1','INSS',5.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro Contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00),('A9','SECONCI',1.20)],
                'B': [('B1','Repouso Semanal Remunerado',0.00),('B2','Feriados',0.00),('B3','Auxílio - Enfermidade',0.65),('B4','13º Salário',8.33),('B5','Licença Paternidade',0.05),('B6','Faltas Justificadas',0.56),('B7','Dias de Chuvas',0.00),('B8','Auxílio Acidente de Trabalho',0.07),('B9','Férias Gozadas',10.42),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',4.50),('C2','Aviso Prévio Trabalhado',0.11),('C3','Férias Indenizadas',0.70),('C4','Depósito Rescisão Sem Justa Causa',2.09),('C5','Indenização Adicional',0.38)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',4.21),('D2','Reincidência de Grupo A sobre Aviso Prévio Trabalhado e Reincidência do FGTS sobre Aviso Prévio Indenizado',0.39)],
            },
        },
    },
    'Normal': {
        'Horista': {
            'totais': {'A': 38.00, 'B': 49.23, 'C': 10.32, 'D': 19.24},
            'itens': {
                'A': [('A1','INSS',20.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro Contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00),('A9','SECONCI',1.20)],
                'B': [('B1','Repouso Semanal Remunerado',17.76),('B2','Feriados',3.68),('B3','Auxílio - Enfermidade',0.86),('B4','13º Salário',11.06),('B5','Licença Paternidade',0.07),('B6','Faltas Justificadas',0.74),('B7','Dias de Chuvas',1.09),('B8','Auxílio Acidente de Trabalho',0.10),('B9','Férias Gozadas',13.84),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',5.97),('C2','Aviso Prévio Trabalhado',0.14),('C3','Férias Indenizadas',0.93),('C4','Depósito Rescisão Sem Justa Causa',2.78),('C5','Indenização Adicional',0.50)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',18.71),('D2','Reincidência de Grupo A sobre Aviso Prévio Trabalhado e Reincidência do FGTS sobre Aviso Prévio Indenizado',0.53)],
            },
        },
        'Mensalista': {
            'totais': {'A': 38.00, 'B': 20.11, 'C': 7.78, 'D': 8.04},
            'itens': {
                'A': [('A1','INSS',20.00),('A2','SESI',1.50),('A3','SENAI',1.00),('A4','INCRA',0.20),('A5','SEBRAE',0.60),('A6','Salário Educação',2.50),('A7','Seguro Contra Acidentes de Trabalho',3.00),('A8','FGTS',8.00),('A9','SECONCI',1.20)],
                'B': [('B1','Repouso Semanal Remunerado',0.00),('B2','Feriados',0.00),('B3','Auxílio - Enfermidade',0.65),('B4','13º Salário',8.33),('B5','Licença Paternidade',0.05),('B6','Faltas Justificadas',0.56),('B7','Dias de Chuvas',0.00),('B8','Auxílio Acidente de Trabalho',0.07),('B9','Férias Gozadas',10.42),('B10','Salário Maternidade',0.03)],
                'C': [('C1','Aviso Prévio Indenizado',4.50),('C2','Aviso Prévio Trabalhado',0.11),('C3','Férias Indenizadas',0.70),('C4','Depósito Rescisão Sem Justa Causa',2.09),('C5','Indenização Adicional',0.38)],
                'D': [('D1','Reincidência de Grupo A sobre Grupo B',7.64),('D2','Reincidência de Grupo A sobre Aviso Prévio Trabalhado e Reincidência do FGTS sobre Aviso Prévio Indenizado',0.40)],
            },
        },
    },
}

def _upsert_encargo_referencial(db, fonte, uf, regime, categoria, vig_ini, vig_fim, vigencia, descricao, valores=None):
    dados = (valores or ENCARGOS_REF_VALORES)[regime][categoria]
    total = round(sum(dados['totais'].values()), 4)
    nome_regime = 'Com Desoneração' if regime == 'Desonerado' else 'Sem Desoneração'
    nome = f"{fonte}/{uf} — {uf} — {categoria} — {nome_regime}"
    row = db.execute("""
        SELECT id_perfil FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,''))=?
           AND COALESCE(uf_referencia,'')=?
           AND categoria=?
           AND regime=?
           AND COALESCE(vigencia_inicio,'')=COALESCE(?, '')
           AND COALESCE(vigencia_fim,'')=COALESCE(?, '')
    """, [fonte.upper(), uf, categoria, regime, vig_ini, vig_fim]).fetchone()
    if row:
        pid = row['id_perfil']
        db.execute("""
            UPDATE perfis_encargos
               SET nome_perfil=?, descricao=?, observacoes=?, situacao='Ativo',
                   fonte_referencia=?, vigencia=?, encargo_original_percentual=?
             WHERE id_perfil=?
        """, [nome, descricao, 'Perfil importado de fonte referencial externa.', fonte.upper(), vigencia, total, pid])
        gids = [r['id_grupo_enc'] for r in db.execute("SELECT id_grupo_enc FROM grupos_encargos WHERE id_perfil=?", [pid]).fetchall()]
        if gids:
            db.executemany("DELETE FROM itens_encargo WHERE id_grupo_enc=?", [(gid,) for gid in gids])
        db.execute("DELETE FROM grupos_encargos WHERE id_perfil=?", [pid])
    else:
        cur = db.execute("""INSERT INTO perfis_encargos
            (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
             fonte_referencia,vigencia,vigencia_inicio,vigencia_fim,encargo_original_percentual)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [nome, categoria, regime, uf, None, descricao, 'Perfil importado de fonte referencial externa.',
             'Ativo', fonte.upper(), vigencia, vig_ini, vig_fim, total])
        pid = cur.lastrowid
    for letra in ('A','B','C','D'):
        gcur = db.execute("""INSERT INTO grupos_encargos
            (id_perfil,letra,descricao,total_grupo) VALUES (?,?,?,?)""",
            [pid, letra, ENCARGOS_REF_GRUPOS[letra], dados['totais'][letra]])
        gid = gcur.lastrowid
        for ordem, (codigo, desc, pct) in enumerate(dados['itens'][letra], start=1):
            db.execute("""INSERT INTO itens_encargo
                (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem)
                VALUES (?,?,?,?,?,?)""",
                [gid, f"{codigo} - {desc}", 'Tabela referencial de encargos sociais', pct, None, ordem])
    _calc_encargos(db, pid, recalc_d=False)
    db.execute("UPDATE perfis_encargos SET encargo_original_percentual=? WHERE id_perfil=?", [total, pid])
    return pid

def _clone_encargo_perfil(db, source_pid, fonte, uf, vig_ini, vig_fim, vigencia, descricao):
    source = db.execute("SELECT * FROM perfis_encargos WHERE id_perfil=?", [source_pid]).fetchone()
    if not source:
        raise ValueError('Perfil de origem não encontrado.')
    source = dict(source)
    nome_regime = 'Com Desoneração' if source['regime'] == 'Desonerado' else 'Sem Desoneração'
    nome = f"{fonte}/{uf} — {uf} — {source['categoria']} — {nome_regime}"
    row = db.execute("""
        SELECT id_perfil FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,''))=?
           AND COALESCE(uf_referencia,'')=?
           AND categoria=?
           AND regime=?
           AND COALESCE(vigencia_inicio,'')=COALESCE(?, '')
           AND COALESCE(vigencia_fim,'')=COALESCE(?, '')
    """, [fonte.upper(), uf, source['categoria'], source['regime'], vig_ini, vig_fim]).fetchone()
    if row:
        pid = row['id_perfil']
        db.execute("""
            UPDATE perfis_encargos
               SET nome_perfil=?, descricao=?, observacoes=?, situacao='Ativo',
                   fonte_referencia=?, vigencia=?, encargo_original_percentual=?
             WHERE id_perfil=?
        """, [nome, descricao, 'Perfil importado de fonte referencial externa.',
              fonte.upper(), vigencia, source.get('encargo_total') or 0, pid])
        gids = [r['id_grupo_enc'] for r in db.execute("SELECT id_grupo_enc FROM grupos_encargos WHERE id_perfil=?", [pid]).fetchall()]
        if gids:
            db.executemany("DELETE FROM itens_encargo WHERE id_grupo_enc=?", [(gid,) for gid in gids])
        db.execute("DELETE FROM grupos_encargos WHERE id_perfil=?", [pid])
    else:
        cur = db.execute("""INSERT INTO perfis_encargos
            (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
             fonte_referencia,vigencia,vigencia_inicio,vigencia_fim,encargo_original_percentual)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [nome, source['categoria'], source['regime'], uf, None, descricao,
             'Perfil importado de fonte referencial externa.', 'Ativo', fonte.upper(),
             vigencia, vig_ini, vig_fim, source.get('encargo_total') or 0])
        pid = cur.lastrowid
    grupos = db.execute("SELECT * FROM grupos_encargos WHERE id_perfil=? ORDER BY letra", [source_pid]).fetchall()
    for g in grupos:
        g = dict(g)
        gc = db.execute("""INSERT INTO grupos_encargos
            (id_perfil,letra,descricao,total_grupo) VALUES (?,?,?,?)""",
            [pid, g['letra'], g['descricao'], g.get('total_grupo') or 0])
        novo_gid = gc.lastrowid
        itens = db.execute("SELECT * FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem", [g['id_grupo_enc']]).fetchall()
        for it in itens:
            it = dict(it)
            db.execute("""INSERT INTO itens_encargo
                (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem)
                VALUES (?,?,?,?,?,?)""",
                [novo_gid, it['descricao'], it.get('base_legal') or 'Tabela referencial de encargos sociais',
                 it.get('percentual') or 0, it.get('observacoes'), it.get('ordem') or 0])
    _calc_encargos(db, pid, recalc_d=False)
    return pid

def _encargos_import_form(prefix):
    pdf = request.files.get('arquivo_pdf') if request.files else None
    vig_ini = (request.form.get('vigencia_inicio') if request.form else None) or '2026-01-01'
    vig_fim = (request.form.get('vigencia_fim') if request.form else None) or '2026-12-31'
    vigencia = (request.form.get('vigencia') if request.form else None) or f"{vig_ini[:7]} a {vig_fim[:7]}"
    if request.content_type and 'multipart/form-data' in request.content_type:
        if not pdf or not pdf.filename:
            raise ValueError(f'Envie o PDF de encargos sociais da {prefix}.')
        if not pdf.filename.lower().endswith('.pdf'):
            raise ValueError('Arquivo inválido. Use um PDF.')
        payload = pdf.read()
        if not payload.startswith(b'%PDF'):
            raise ValueError('O arquivo enviado não parece ser um PDF válido.')
    return vig_ini, vig_fim, vigencia

def _encargos_import_pdf_payload(prefix):
    pdf = request.files.get('arquivo_pdf') if request.files else None
    if not pdf or not pdf.filename:
        raise ValueError(f'Envie o PDF de encargos sociais do {prefix}.')
    if not pdf.filename.lower().endswith('.pdf'):
        raise ValueError('Arquivo inválido. Use um PDF.')
    payload = pdf.read()
    if not payload.startswith(b'%PDF'):
        raise ValueError('O arquivo enviado não parece ser um PDF válido.')
    vig_ini = (request.form.get('vigencia_inicio') if request.form else None) or ''
    vig_fim = (request.form.get('vigencia_fim') if request.form else None) or ''
    vigencia = (request.form.get('vigencia') if request.form else None) or ''
    return payload, vig_ini, vig_fim, vigencia

def _encargos_import_pypdf_reader(payload):
    try:
        from pypdf import PdfReader
    except Exception:
        dep = Path.home() / '.cache' / 'codex-runtimes' / 'codex-primary-runtime' / 'dependencies' / 'python' / 'Lib' / 'site-packages'
        if dep.exists() and str(dep) not in sys.path:
            sys.path.append(str(dep))
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise ValueError('A leitura do PDF SINAPI exige a biblioteca pypdf no ambiente Python.') from exc
    return PdfReader(io.BytesIO(payload))

UF_NOME_TO_SIGLA = {
    'ACRE': 'AC', 'ALAGOAS': 'AL', 'AMAPA': 'AP', 'AMAZONAS': 'AM', 'BAHIA': 'BA',
    'CEARA': 'CE', 'DISTRITO FEDERAL': 'DF', 'ESPIRITO SANTO': 'ES', 'GOIAS': 'GO',
    'MARANHAO': 'MA', 'MATO GROSSO': 'MT', 'MATO GROSSO DO SUL': 'MS', 'MINAS GERAIS': 'MG',
    'PARA': 'PA', 'PARAIBA': 'PB', 'PARANA': 'PR', 'PERNAMBUCO': 'PE', 'PIAUI': 'PI',
    'RIO DE JANEIRO': 'RJ', 'RIO GRANDE DO NORTE': 'RN', 'RIO GRANDE DO SUL': 'RS',
    'RONDONIA': 'RO', 'RORAIMA': 'RR', 'SANTA CATARINA': 'SC', 'SAO PAULO': 'SP',
    'SERGIPE': 'SE', 'TOCANTINS': 'TO',
}

ENCARGOS_ITEM_DESCRICOES = {
    'A1': 'INSS', 'A2': 'SESI', 'A3': 'SENAI', 'A4': 'INCRA', 'A5': 'SEBRAE',
    'A6': 'Salário Educação', 'A7': 'Seguro Contra Acidentes de Trabalho',
    'A8': 'FGTS', 'A9': 'SECONCI',
    'B1': 'Repouso Semanal Remunerado', 'B2': 'Feriados', 'B3': 'Auxílio - Enfermidade',
    'B4': '13º Salário', 'B5': 'Licença Paternidade', 'B6': 'Faltas Justificadas',
    'B7': 'Dias de Chuvas', 'B8': 'Auxílio Acidente de Trabalho',
    'B9': 'Férias Gozadas', 'B10': 'Salário Maternidade',
    'C1': 'Aviso Prévio Indenizado', 'C2': 'Aviso Prévio Trabalhado',
    'C3': 'Férias Indenizadas', 'C4': 'Depósito Rescisão Sem Justa Causa',
    'C5': 'Indenização Adicional',
    'D1': 'Reincidência de Grupo A sobre Grupo B',
    'D2': 'Reincidência de Grupo A sobre Aviso Prévio Trabalhado e Reincidência do FGTS sobre Aviso Prévio Indenizado',
}

def _norm_ascii(txt):
    txt = unicodedata.normalize('NFKD', txt or '')
    return ''.join(ch for ch in txt if not unicodedata.combining(ch)).upper().strip()

def _encargos_pct_token_to_float(token):
    token = (token or '').strip()
    if re.search(r'n[aã]o\s+inc', token, flags=re.I):
        return 0.0
    token = token.replace(' ', '').replace('%', '').replace(',', '.')
    try:
        return round(float(token), 6)
    except Exception:
        return 0.0

def _encargos_extract_four_values(segment):
    tokens = re.findall(r'(?:\d+,\s*\d+\s*%|N[aã]o\s+inc\s*ide)', segment or '', flags=re.I)
    return [_encargos_pct_token_to_float(t) for t in tokens[:4]] if len(tokens) >= 4 else None

def _encargos_parse_sinapi_page(text):
    text = (text or '').replace('\xa0', ' ')
    m_uf = re.search(r'Encargos\s+Sociais\s+.{0,6}?\s*([A-Za-zÀ-ÿ ]+?)\s*\n\s*\d{2}/\d{4}', text, flags=re.I)
    if not m_uf:
        return None
    nome = _norm_ascii(m_uf.group(1))
    uf = UF_NOME_TO_SIGLA.get(nome)
    if not uf:
        return None
    m_ref = re.search(r'\b(\d{2})/(\d{4})\b', text)
    mes, ano = (m_ref.group(1), m_ref.group(2)) if m_ref else ('01', '2025')

    grupos = {'A': [], 'B': [], 'C': [], 'D': []}
    codigos = ['A1','A2','A3','A4','A5','A6','A7','A8','A9',
               'B1','B2','B3','B4','B5','B6','B7','B8','B9','B10',
               'C1','C2','C3','C4','C5','D1','D2']
    for idx, codigo in enumerate(codigos):
        prox = codigos[idx + 1] if idx + 1 < len(codigos) else None
        if prox:
            pat = rf'\b{codigo}\b\s+(.*?)(?=\b{prox}\b)'
        else:
            pat = rf'\b{codigo}\b\s+(.*?)(?=\bD\s+Total\b|TOTAL\(A\+B\+C\+D\)|$)'
        m = re.search(pat, text, flags=re.S)
        if not m:
            continue
        vals = _encargos_extract_four_values(m.group(1))
        if not vals:
            continue
        grupos[codigo[0]].append((codigo, ENCARGOS_ITEM_DESCRICOES.get(codigo, codigo), vals))
    if not all(grupos[g] for g in ('A','B','C','D')):
        return None

    valores = {'Desonerado': {'Horista': {}, 'Mensalista': {}}, 'Normal': {'Horista': {}, 'Mensalista': {}}}
    colunas = [
        ('Desonerado', 'Horista', 0),
        ('Desonerado', 'Mensalista', 1),
        ('Normal', 'Horista', 2),
        ('Normal', 'Mensalista', 3),
    ]
    for regime, categoria, col in colunas:
        itens = {}
        totais = {}
        for letra in ('A','B','C','D'):
            itens[letra] = [(cod, desc, vals[col]) for cod, desc, vals in grupos[letra]]
            totais[letra] = round(sum(v for _, _, v in itens[letra]), 6)
        valores[regime][categoria] = {'totais': totais, 'itens': itens}
    return {'uf': uf, 'mes': mes, 'ano': ano, 'valores': valores}

def _importar_encargos_sinapi_pdf(db, payload, vig_ini='', vig_fim='', vigencia=''):
    reader = _encargos_import_pypdf_reader(payload)
    paginas = []
    for page in reader.pages:
        parsed = _encargos_parse_sinapi_page(page.extract_text() or '')
        if parsed:
            paginas.append(parsed)
    if not paginas:
        raise ValueError('Não foi possível localizar tabelas de encargos SINAPI no PDF enviado.')
    criados = []
    for pg in paginas:
        ini = vig_ini or f"{pg['ano']}-{pg['mes']}-01"
        fim = vig_fim or f"{pg['ano']}-12-31"
        vig = vigencia or f"SINAPI {pg['mes']}/{pg['ano']} a {fim[:7]}"
        descricao = 'Encargos sociais SINAPI importados do PDF oficial informado pelo usuário.'
        for regime in ('Normal', 'Desonerado'):
            for categoria in ('Horista', 'Mensalista'):
                criados.append(_upsert_encargo_referencial(
                    db, 'SINAPI', pg['uf'], regime, categoria,
                    ini, fim, vig, descricao,
                    valores=pg['valores']
                ))
    return criados

def _sicro_detect_regime(filename, fallback):
    txt = _norm_ascii(filename or '')
    if 'DESONER' in txt:
        return 'Desonerado'
    return fallback or 'Normal'

def _sicro_parse_xlsx(payload):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    if 'Encargos' not in wb.sheetnames:
        raise ValueError('A planilha SICRO deve conter a aba "Encargos".')
    ws = wb['Encargos']
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    header2 = next(rows, None) or []
    codigos = [(idx, str(codigo).strip()) for idx, codigo in enumerate(header2) if idx >= 3 and idx < 29 and codigo]
    out = []
    for row_values in rows:
        codigo_prof = row_values[0] if len(row_values) > 0 else None
        desc = row_values[1] if len(row_values) > 1 else None
        if not codigo_prof or not str(codigo_prof).strip().upper().startswith('P'):
            continue
        parcelas = {}
        grupos = {'A': 0.0, 'B': 0.0, 'C': 0.0, 'D': 0.0}
        for idx, codigo_parcela in codigos:
            val = _market_float(row_values[idx] if idx < len(row_values) else 0, 0.0)
            parcelas[codigo_parcela] = val
            letra = codigo_parcela[:1].upper()
            if letra in grupos:
                grupos[letra] += val
        total = _market_float(row_values[29] if len(row_values) > 29 else None, sum(grupos.values()))
        out.append({
            'codigo': str(codigo_prof).strip(),
            'descricao': str(desc or '').strip(),
            'unidade': str(row_values[2] if len(row_values) > 2 and row_values[2] is not None else '').strip(),
            'A': round(grupos['A'], 6),
            'B': round(grupos['B'], 6),
            'C': round(grupos['C'], 6),
            'D': round(grupos['D'], 6),
            'total': round(total, 6),
            'parcelas': parcelas,
        })
    if not out:
        raise ValueError('Nenhum profissional SICRO foi encontrado na aba Encargos.')
    return out

def _sicro_categoria_profissional(prof):
    unidade = _norm_ascii(prof.get('unidade') or '')
    desc = _norm_ascii(prof.get('descricao') or '')
    if unidade == 'H' or unidade.startswith('HORA'):
        return 'Horista'
    if 'MENSALISTA' in desc or unidade:
        return 'Mensalista'
    return 'Horista'

def _sicro_upsert_perfil(db, uf, regime, categoria, vig_ini, vig_fim, vigencia, profissionais):
    uf = (uf or 'DF').strip().upper()
    regime = 'Desonerado' if regime == 'Desonerado' else 'Normal'
    categoria = 'Mensalista' if categoria == 'Mensalista' else 'Horista'
    nome_regime = 'Com Desoneração' if regime == 'Desonerado' else 'Sem Desoneração'
    nome = f"SICRO/{uf} — {uf} — {categoria} por Profissional — {nome_regime}"
    desc = (
        'Encargos sociais e trabalhistas SICRO importados por profissional da mão de obra. '
        'Diferentemente do SINAPI/SEINFRA/SUDECAP, o percentual aplicável varia conforme o código profissional.'
    )
    medias = {
        'A': round(sum(p['A'] for p in profissionais) / len(profissionais), 6),
        'B': round(sum(p['B'] for p in profissionais) / len(profissionais), 6),
        'C': round(sum(p['C'] for p in profissionais) / len(profissionais), 6),
        'D': round(sum(p['D'] for p in profissionais) / len(profissionais), 6),
        'total': round(sum(p['total'] for p in profissionais) / len(profissionais), 6),
    }
    row = db.execute("""
        SELECT id_perfil FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,''))='SICRO'
           AND COALESCE(uf_referencia,'')=?
           AND categoria=?
           AND regime=?
           AND COALESCE(vigencia_inicio,'')=COALESCE(?, '')
           AND COALESCE(vigencia_fim,'')=COALESCE(?, '')
    """, [uf, categoria, regime, vig_ini, vig_fim]).fetchone()
    if row:
        pid = row['id_perfil']
        db.execute("""
            UPDATE perfis_encargos
               SET nome_perfil=?, descricao=?, observacoes=?, situacao='Ativo',
                   fonte_referencia='SICRO', vigencia=?, encargo_original_percentual=?,
                   total_grupo_a=?, total_grupo_b=?, total_grupo_c=?, total_grupo_d=?, encargo_total=?
             WHERE id_perfil=?
        """, [nome, desc, f'{len(profissionais)} profissionais SICRO importados.', vigencia, medias['total'],
              medias['A'], medias['B'], medias['C'], medias['D'], medias['total'], pid])
        gids = [r['id_grupo_enc'] for r in db.execute("SELECT id_grupo_enc FROM grupos_encargos WHERE id_perfil=?", [pid]).fetchall()]
        if gids:
            db.executemany("DELETE FROM itens_encargo WHERE id_grupo_enc=?", [(gid,) for gid in gids])
        db.execute("DELETE FROM grupos_encargos WHERE id_perfil=?", [pid])
        db.execute("DELETE FROM encargos_sicro_profissionais WHERE id_perfil=?", [pid])
    else:
        pid = db.execute("""INSERT INTO perfis_encargos
            (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
             fonte_referencia,vigencia,vigencia_inicio,vigencia_fim,encargo_original_percentual,
             total_grupo_a,total_grupo_b,total_grupo_c,total_grupo_d,encargo_total)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [nome, categoria, regime, uf, None, desc, f'{len(profissionais)} profissionais SICRO importados.',
             'Ativo', 'SICRO', vigencia, vig_ini, vig_fim, medias['total'],
             medias['A'], medias['B'], medias['C'], medias['D'], medias['total']]).lastrowid
    for letra, titulo in [('A','Encargos Sociais'), ('B','Encargos Trabalhistas'), ('C','Verbas Rescisórias'), ('D','Reincidências')]:
        gid = db.execute("""INSERT INTO grupos_encargos
            (id_perfil,letra,descricao,total_grupo) VALUES (?,?,?,?)""",
            [pid, letra, titulo, medias[letra]]).lastrowid
        db.execute("""INSERT INTO itens_encargo
            (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem)
            VALUES (?,?,?,?,?,?)""",
            [gid, f"Média dos profissionais SICRO - Grupo {letra}", 'SICRO - Relatório Analítico de Encargos Sociais e Trabalhistas',
             medias[letra], 'Resumo estatístico; use a tabela de profissionais para aplicação por código.', 1])
    for p in profissionais:
        db.execute("""INSERT INTO encargos_sicro_profissionais
            (id_perfil,codigo_profissional,descricao,unidade,total_grupo_a,total_grupo_b,total_grupo_c,total_grupo_d,encargo_total,parcelas_json)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            [pid, p['codigo'], p['descricao'], p['unidade'], p['A'], p['B'], p['C'], p['D'], p['total'],
             _json.dumps(p['parcelas'], ensure_ascii=False)])
    return pid

def _importar_encargos_sicro(db):
    ensure_encargos_schema(db)
    files = request.files or {}
    uploads = [
        ('Normal', files.get('arquivo_onerado')),
        ('Desonerado', files.get('arquivo_desonerado')),
    ]
    if not uploads[0][1] or not uploads[1][1]:
        raise ValueError('Envie as duas planilhas SICRO: onerada e desonerada. A tabela analítica exige os dois regimes.')
    uf = (request.form.get('uf') if request.form else None) or 'DF'
    vig_ini = (request.form.get('vigencia_inicio') if request.form else None) or '2026-01-01'
    vig_fim = (request.form.get('vigencia_fim') if request.form else None) or '2026-12-31'
    vigencia = (request.form.get('vigencia') if request.form else None) or f"SICRO {uf.upper()} {vig_ini[:7]} a {vig_fim[:7]}"
    criados = []
    for fallback_regime, upload in uploads:
        if not upload or not upload.filename:
            continue
        if not upload.filename.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError('Use arquivos SICRO em formato .xlsx.')
        regime = _sicro_detect_regime(upload.filename, fallback_regime)
        profissionais = _sicro_parse_xlsx(upload.read())
        for categoria in ('Horista', 'Mensalista'):
            subset = [p for p in profissionais if _sicro_categoria_profissional(p) == categoria]
            if subset:
                criados.append(_sicro_upsert_perfil(db, uf, regime, categoria, vig_ini, vig_fim, vigencia, subset))
    return criados

def _encargos_valor_percentual_planilha(v):
    val = _market_float(v, 0.0)
    if abs(val) <= 3:
        return val * 100
    return val

def _goinfra_detect_regime(filename, fallback):
    txt = _norm_ascii(filename or '')
    if 'SEM' in txt and 'DESONER' in txt:
        return 'Normal'
    if 'COM' in txt and 'DESONER' in txt:
        return 'Desonerado'
    if 'DESONER' in txt:
        return 'Desonerado'
    return fallback or 'Normal'

def _goinfra_detect_referencia(payload):
    meses = {
        'JANEIRO': 1, 'FEVEREIRO': 2, 'MARCO': 3, 'MARÇO': 3, 'ABRIL': 4,
        'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9,
        'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
    }
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
        ws = wb['Encargos'] if 'Encargos' in wb.sheetnames else wb.active
        texto = _norm_ascii(ws.cell(3, 1).value or '')
        ano_m = re.search(r'(20\d{2})', texto)
        mes = None
        for nome, num in meses.items():
            if nome in texto:
                mes = num
                break
        if ano_m and mes:
            import calendar
            ano = int(ano_m.group(1))
            return f'{ano}-{mes:02d}-01', f'{ano}-{mes:02d}-{calendar.monthrange(ano, mes)[1]:02d}', f'GOINFRA/GO {mes:02d}/{ano}'
    except Exception:
        pass
    return None, None, None

def _goinfra_parse_xlsx(payload):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    if 'Encargos' not in wb.sheetnames:
        raise ValueError('A planilha GOINFRA deve conter a aba "Encargos".')
    ws = wb['Encargos']
    codigos = []
    for col in range(4, ws.max_column + 1):
        codigo = ws.cell(6, col).value
        if codigo and re.match(r'^[ABCD]\d+$', str(codigo).strip(), flags=re.I):
            codigos.append((col, str(codigo).strip().upper()))
    if not codigos:
        raise ValueError('Não foi possível localizar as parcelas A/B/C/D na planilha GOINFRA.')
    out = []
    for row in range(7, ws.max_row + 1):
        codigo_prof = ws.cell(row, 1).value
        desc = ws.cell(row, 2).value
        unidade = ws.cell(row, 3).value
        if codigo_prof in (None, '') or not str(desc or '').strip():
            continue
        if not re.match(r'^\d+$', str(codigo_prof).strip()):
            continue
        parcelas = {}
        grupos = {'A': 0.0, 'B': 0.0, 'C': 0.0, 'D': 0.0}
        for col, codigo_parcela in codigos:
            val = _encargos_valor_percentual_planilha(ws.cell(row, col).value)
            parcelas[codigo_parcela] = round(val, 6)
            letra = codigo_parcela[:1].upper()
            if letra in grupos:
                grupos[letra] += val
        total_planilha = _encargos_valor_percentual_planilha(ws.cell(row, 30).value)
        total = total_planilha if total_planilha else sum(grupos.values())
        out.append({
            'codigo': str(codigo_prof).strip(),
            'descricao': str(desc or '').strip(),
            'unidade': str(unidade or '').strip(),
            'A': round(grupos['A'], 6),
            'B': round(grupos['B'], 6),
            'C': round(grupos['C'], 6),
            'D': round(grupos['D'], 6),
            'total': round(total, 6),
            'parcelas': parcelas,
        })
    if not out:
        raise ValueError('Nenhum profissional GOINFRA foi encontrado na aba Encargos.')
    return out

def _goinfra_categoria_profissional(prof):
    unidade = _norm_ascii(prof.get('unidade') or '')
    desc = _norm_ascii(prof.get('descricao') or '')
    if unidade == 'H' or unidade.startswith('HORA'):
        return 'Horista'
    if unidade in ('MES', 'MÊS') or 'MENSALISTA' in desc:
        return 'Mensalista'
    return 'Horista'

def _goinfra_upsert_perfil(db, uf, regime, categoria, vig_ini, vig_fim, vigencia, profissionais):
    uf = (uf or 'GO').strip().upper()
    regime = 'Desonerado' if regime == 'Desonerado' else 'Normal'
    categoria = 'Mensalista' if categoria == 'Mensalista' else 'Horista'
    nome_regime = 'Com Desoneração' if regime == 'Desonerado' else 'Sem Desoneração'
    nome = f"GOINFRA/{uf} — {uf} — {categoria} por Profissional — {nome_regime}"
    desc = (
        'Encargos sociais e trabalhistas GOINFRA/GO importados por profissional da mão de obra. '
        'Assim como no SICRO, o percentual aplicável varia conforme o código profissional.'
    )
    medias = {
        'A': round(sum(p['A'] for p in profissionais) / len(profissionais), 6),
        'B': round(sum(p['B'] for p in profissionais) / len(profissionais), 6),
        'C': round(sum(p['C'] for p in profissionais) / len(profissionais), 6),
        'D': round(sum(p['D'] for p in profissionais) / len(profissionais), 6),
        'total': round(sum(p['total'] for p in profissionais) / len(profissionais), 6),
    }
    row = db.execute("""
        SELECT id_perfil FROM perfis_encargos
         WHERE UPPER(COALESCE(fonte_referencia,''))='GOINFRA'
           AND COALESCE(uf_referencia,'')=?
           AND categoria=?
           AND regime=?
           AND COALESCE(vigencia_inicio,'')=COALESCE(?, '')
           AND COALESCE(vigencia_fim,'')=COALESCE(?, '')
    """, [uf, categoria, regime, vig_ini, vig_fim]).fetchone()
    if row:
        pid = row['id_perfil']
        db.execute("""
            UPDATE perfis_encargos
               SET nome_perfil=?, descricao=?, observacoes=?, situacao='Ativo',
                   fonte_referencia='GOINFRA', vigencia=?, encargo_original_percentual=?,
                   total_grupo_a=?, total_grupo_b=?, total_grupo_c=?, total_grupo_d=?, encargo_total=?
             WHERE id_perfil=?
        """, [nome, desc, f'{len(profissionais)} profissionais GOINFRA importados.', vigencia, medias['total'],
              medias['A'], medias['B'], medias['C'], medias['D'], medias['total'], pid])
        gids = [r['id_grupo_enc'] for r in db.execute("SELECT id_grupo_enc FROM grupos_encargos WHERE id_perfil=?", [pid]).fetchall()]
        if gids:
            db.executemany("DELETE FROM itens_encargo WHERE id_grupo_enc=?", [(gid,) for gid in gids])
        db.execute("DELETE FROM grupos_encargos WHERE id_perfil=?", [pid])
        db.execute("DELETE FROM encargos_goinfra_profissionais WHERE id_perfil=?", [pid])
    else:
        pid = db.execute("""INSERT INTO perfis_encargos
            (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
             fonte_referencia,vigencia,vigencia_inicio,vigencia_fim,encargo_original_percentual,
             total_grupo_a,total_grupo_b,total_grupo_c,total_grupo_d,encargo_total)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [nome, categoria, regime, uf, None, desc, f'{len(profissionais)} profissionais GOINFRA importados.',
             'Ativo', 'GOINFRA', vigencia, vig_ini, vig_fim, medias['total'],
             medias['A'], medias['B'], medias['C'], medias['D'], medias['total']]).lastrowid
    for letra, titulo in [('A','Encargos Sociais'), ('B','Encargos Trabalhistas'), ('C','Verbas Rescisórias'), ('D','Reincidências')]:
        gid = db.execute("""INSERT INTO grupos_encargos
            (id_perfil,letra,descricao,total_grupo) VALUES (?,?,?,?)""",
            [pid, letra, titulo, medias[letra]]).lastrowid
        db.execute("""INSERT INTO itens_encargo
            (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem)
            VALUES (?,?,?,?,?,?)""",
            [gid, f"Média dos profissionais GOINFRA - Grupo {letra}", 'GOINFRA/GO - Tabela de Encargos Sociais',
             medias[letra], 'Resumo estatístico; use a tabela de profissionais para aplicação por código.', 1])
    for p in profissionais:
        db.execute("""INSERT INTO encargos_goinfra_profissionais
            (id_perfil,codigo_profissional,descricao,unidade,total_grupo_a,total_grupo_b,total_grupo_c,total_grupo_d,encargo_total,parcelas_json)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            [pid, p['codigo'], p['descricao'], p['unidade'], p['A'], p['B'], p['C'], p['D'], p['total'],
             _json.dumps(p['parcelas'], ensure_ascii=False)])
    return pid

def _importar_encargos_goinfra(db):
    ensure_encargos_schema(db)
    files = request.files or {}
    uploads = [
        ('Normal', files.get('arquivo_onerado')),
        ('Desonerado', files.get('arquivo_desonerado')),
    ]
    if not uploads[0][1] or not uploads[1][1]:
        raise ValueError('Envie as duas planilhas GOINFRA/GO: onerada e desonerada. A tabela analítica exige os dois regimes.')
    uf = 'GO'
    payload_ref = uploads[0][1].read()
    uploads[0][1].stream.seek(0)
    ref_ini, ref_fim, ref_vig = _goinfra_detect_referencia(payload_ref)
    vig_ini = (request.form.get('vigencia_inicio') if request.form else None) or ref_ini or '2026-02-01'
    vig_fim = (request.form.get('vigencia_fim') if request.form else None) or ref_fim or '2026-02-28'
    vigencia = (request.form.get('vigencia') if request.form else None) or ref_vig or f"GOINFRA/GO {vig_ini[:7]} a {vig_fim[:7]}"
    criados = []
    for fallback_regime, upload in uploads:
        if not upload or not upload.filename:
            continue
        if not upload.filename.lower().endswith(('.xlsx', '.xlsm')):
            raise ValueError('Use arquivos GOINFRA em formato .xlsx.')
        regime = _goinfra_detect_regime(upload.filename, fallback_regime)
        profissionais = _goinfra_parse_xlsx(upload.read())
        for categoria in ('Horista', 'Mensalista'):
            subset = [p for p in profissionais if _goinfra_categoria_profissional(p) == categoria]
            if subset:
                criados.append(_goinfra_upsert_perfil(db, uf, regime, categoria, vig_ini, vig_fim, vigencia, subset))
    return criados

def _importar_encargos_seinfra(db, vig_ini='2023-10-01', vig_fim='2026-12-31', vigencia=None):
    criados = []
    vigencia = vigencia or f'Tabela 028/028.1 - {vig_ini[:7]} a {vig_fim[:7]}'
    for regime in ('Normal', 'Desonerado'):
        for categoria in ('Horista', 'Mensalista'):
            pid = _upsert_encargo_referencial(
                db, 'SEINFRA', 'CE', regime, categoria,
                vig_ini, vig_fim,
                vigencia,
                'Encargos sociais SEINFRA/CE importados das tabelas 028 e 028.1. Aplicáveis ao estado do Ceará.'
            )
            criados.append(pid)
    return criados

def _importar_encargos_sudecap(db, vig_ini='2026-01-01', vig_fim='2026-12-31', vigencia=None):
    criados = []
    vigencia = vigencia or f'Tabela SUDECAP/BH - {vig_ini[:7]} a {vig_fim[:7]}'
    valores = SUDECAP_ENCARGOS_2025 if str(vig_ini).startswith('2025-') else None
    descricao = (
        'Encargos sociais SUDECAP/BH importados da tabela oficial PBH/SUDECAP. '
        'Aplicáveis ao estado de Minas Gerais.'
    )
    for regime in ('Normal', 'Desonerado'):
        for categoria in ('Horista', 'Mensalista'):
            criados.append(_upsert_encargo_referencial(
                db, 'SUDECAP', 'MG', regime, categoria,
                vig_ini, vig_fim, vigencia, descricao,
                valores=valores
            ))
    return criados

@app.route('/api/encargos/importar-seinfra', methods=['POST'])
def enc_importar_seinfra():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        vig_ini, vig_fim, vigencia = _encargos_import_form('SEINFRA/CE')
        criados = _importar_encargos_seinfra(db, vig_ini, vig_fim, vigencia)
        sincronizar_encargos_sociais_insumos(db, ['SEINFRA'])
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais SEINFRA/CE importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fonte': 'SEINFRA',
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/importar-sudecap', methods=['POST'])
def enc_importar_sudecap():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        vig_ini, vig_fim, vigencia = _encargos_import_form('SUDECAP/BH')
        criados = _importar_encargos_sudecap(db, vig_ini, vig_fim, vigencia)
        sincronizar_encargos_sociais_insumos(db, ['SUDECAP'])
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais SUDECAP/BH importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fonte': 'SUDECAP',
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/importar-sinapi', methods=['POST'])
def enc_importar_sinapi():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        payload, vig_ini, vig_fim, vigencia = _encargos_import_pdf_payload('SINAPI')
        criados = _importar_encargos_sinapi_pdf(db, payload, vig_ini, vig_fim, vigencia)
        sincronizar_encargos_sociais_insumos(db, ['SINAPI'])
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais SINAPI importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fonte': 'SINAPI',
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/importar-sicro', methods=['POST'])
def enc_importar_sicro():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        criados = _importar_encargos_sicro(db)
        sincronizar_encargos_sociais_insumos(db, ['SICRO'])
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais SICRO importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fonte': 'SICRO',
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/importar-goinfra', methods=['POST'])
def enc_importar_goinfra():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        criados = _importar_encargos_goinfra(db)
        sincronizar_encargos_sociais_insumos(db, ['GOINFRA'])
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais GOINFRA/GO importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fonte': 'GOINFRA',
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/importar-referenciais', methods=['POST'])
def enc_importar_referenciais():
    db = get_db()
    ensure_encargos_schema(db)
    try:
        criados = _importar_encargos_seinfra(db) + _importar_encargos_sudecap(db)
        db.commit()
        return jsonify({
            'mensagem': 'Encargos sociais referenciais importados/atualizados com sucesso.',
            'perfis_processados': len(criados),
            'fontes': ['SEINFRA', 'SUDECAP'],
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

@app.route('/api/encargos/perfis/<int:id>', methods=['GET'])
def enc_perfis_get(id):
    db = get_db()
    ensure_encargos_schema(db)
    _calc_encargos(db, id)
    row = db.execute(SEL_PERFIL + " WHERE pe.id_perfil=?", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro': 'Perfil não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/encargos/perfis', methods=['POST'])
def enc_perfis_create():
    d = request.json or {}
    if not d.get('nome_perfil','').strip():
        return jsonify({'erro': 'Nome do perfil é obrigatório.'}), 400
    if d.get('categoria') not in ('Horista','Mensalista'):
        return jsonify({'erro': 'Categoria inválida.'}), 400
    db = get_db()
    ensure_encargos_schema(db)
    cur = db.execute("""INSERT INTO perfis_encargos
        (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
         fonte_referencia,vigencia,vigencia_inicio,vigencia_fim)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        [d['nome_perfil'].strip(),
         d.get('categoria','Horista'),
         d.get('regime','Normal'),
         d.get('uf_referencia') or None,
         d.get('id_data_base') or None,
         d.get('descricao'), d.get('observacoes'), d.get('situacao','Ativo'),
         (d.get('fonte_referencia') or 'SINAPI').strip().upper(),
         d.get('vigencia') or None,
         d.get('vigencia_inicio') or None,
         d.get('vigencia_fim') or None])
    pid = cur.lastrowid
    # Cria grupos vazios A-D
    for letra in ('A','B','C','D'):
        descs = {'A':'Encargos Básicos','B':'Encargos sobre Tempo Trabalhado',
                 'C':'Encargos Rescisórios','D':'Incidência de A sobre B e C'}
        db.execute("INSERT INTO grupos_encargos (id_perfil,letra,descricao) VALUES (?,?,?)",
                   [pid, letra, descs[letra]])
    db.commit()
    row = db.execute(SEL_PERFIL+" WHERE pe.id_perfil=?",[pid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/encargos/perfis/<int:id>', methods=['PUT'])
def enc_perfis_update(id):
    d = request.json or {}
    if not d.get('nome_perfil','').strip():
        return jsonify({'erro': 'Nome do perfil é obrigatório.'}), 400
    db = get_db()
    ensure_encargos_schema(db)
    cur = db.execute("""UPDATE perfis_encargos SET
        nome_perfil=?,categoria=?,regime=?,uf_referencia=?,id_data_base=?,
        descricao=?,observacoes=?,situacao=?,fonte_referencia=?,vigencia=?,
        vigencia_inicio=?,vigencia_fim=?
        WHERE id_perfil=?""",
        [d['nome_perfil'].strip(), d.get('categoria','Horista'),
         d.get('regime','Normal'), d.get('uf_referencia') or None,
         d.get('id_data_base') or None,
         d.get('descricao'), d.get('observacoes'), d.get('situacao','Ativo'),
         (d.get('fonte_referencia') or 'SINAPI').strip().upper(),
         d.get('vigencia') or None,
         d.get('vigencia_inicio') or None,
         d.get('vigencia_fim') or None, id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Perfil não encontrado.'}), 404
    _calc_encargos(db, id)
    row = db.execute(SEL_PERFIL+" WHERE pe.id_perfil=?",[id]).fetchone()
    db.close(); return jsonify(dict(row))

@app.route('/api/encargos/perfis/<int:id>', methods=['DELETE'])
def enc_perfis_delete(id):
    db = get_db()
    ensure_encargos_schema(db)
    cur = db.execute('DELETE FROM perfis_encargos WHERE id_perfil=?',[id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro':'Perfil não encontrado.'}), 404
    return jsonify({'mensagem':'Perfil excluído.'})

@app.route('/api/encargos/perfis/<int:id>/duplicar', methods=['POST'])
def enc_perfis_duplicate(id):
    db = get_db()
    ensure_encargos_schema(db)
    p = db.execute('SELECT * FROM perfis_encargos WHERE id_perfil=?',[id]).fetchone()
    if not p: db.close(); return jsonify({'erro':'Perfil não encontrado.'}), 404
    p = dict(p)
    cur = db.execute("""INSERT INTO perfis_encargos
        (nome_perfil,categoria,regime,uf_referencia,id_data_base,descricao,observacoes,situacao,
         fonte_referencia,vigencia,vigencia_inicio,vigencia_fim)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        ['Cópia de '+p['nome_perfil'], p['categoria'], p['regime'],
         p['uf_referencia'], p['id_data_base'], p['descricao'], p['observacoes'], 'Ativo',
         p.get('fonte_referencia') or 'SINAPI', p.get('vigencia'),
         p.get('vigencia_inicio'), p.get('vigencia_fim')])
    novo_pid = cur.lastrowid
    # Copiar grupos e itens
    grupos = db.execute('SELECT * FROM grupos_encargos WHERE id_perfil=?',[id]).fetchall()
    for g in grupos:
        g = dict(g)
        gc = db.execute("INSERT INTO grupos_encargos (id_perfil,letra,descricao) VALUES (?,?,?)",
                        [novo_pid, g['letra'], g['descricao']])
        novo_gid = gc.lastrowid
        itens = db.execute('SELECT * FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem',[g['id_grupo_enc']]).fetchall()
        for it in itens:
            it = dict(it)
            db.execute("INSERT INTO itens_encargo (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem) VALUES (?,?,?,?,?,?)",
                       [novo_gid, it['descricao'], it['base_legal'], it['percentual'], it['observacoes'], it['ordem']])
    db.commit()
    _calc_encargos(db, novo_pid)
    row = db.execute(SEL_PERFIL+" WHERE pe.id_perfil=?",[novo_pid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

# ── Recalcular D ───────────────────────────────────────────────────────────────
@app.route('/api/encargos/perfis/<int:id>/recalcular-d', methods=['POST'])
def enc_recalcular_d(id):
    """Recalcula D1 e D2 pela fórmula simplificada A×(B+C)."""
    db = get_db()
    ensure_encargos_schema(db)
    p = db.execute('SELECT id_perfil FROM perfis_encargos WHERE id_perfil=?', [id]).fetchone()
    if not p: db.close(); return jsonify({'erro': 'Perfil não encontrado.'}), 404
    totais = _calc_encargos(db, id, recalc_d=True)
    row = db.execute(SEL_PERFIL + " WHERE pe.id_perfil=?", [id]).fetchone()
    db.close()
    return jsonify({'perfil': dict(row), 'totais': totais})

# ── Grupos ─────────────────────────────────────────────────────────────────────
@app.route('/api/encargos/perfis/<int:pid>/grupos', methods=['GET'])
def enc_grupos_list(pid):
    db = get_db()
    ensure_encargos_schema(db)
    grupos = rows_to_list(db.execute(
        "SELECT * FROM grupos_encargos WHERE id_perfil=? ORDER BY letra",[pid]).fetchall())
    for g in grupos:
        g['itens'] = rows_to_list(db.execute(
            "SELECT * FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem",[g['id_grupo_enc']]).fetchall())
    db.close(); return jsonify(grupos)

@app.route('/api/encargos/perfis/<int:pid>/sicro-profissionais', methods=['GET'])
def enc_sicro_profissionais(pid):
    db = get_db()
    ensure_encargos_schema(db)
    p = db.execute("SELECT id_perfil FROM perfis_encargos WHERE id_perfil=?", [pid]).fetchone()
    if not p:
        db.close()
        return jsonify({'erro': 'Perfil não encontrado.'}), 404
    rows = rows_to_list(db.execute("""
        SELECT id_profissional_enc, codigo_profissional, descricao, unidade,
               total_grupo_a, total_grupo_b, total_grupo_c, total_grupo_d, encargo_total
          FROM encargos_sicro_profissionais
         WHERE id_perfil=?
         ORDER BY codigo_profissional
    """, [pid]).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/encargos/sicro-profissionais', methods=['GET'])
def enc_sicro_profissionais_analitico():
    uf = request.args.get('uf','')
    cat = request.args.get('categoria','')
    reg = request.args.get('regime','')
    vig_ini_mes = request.args.get('vigencia_inicio_mes','')
    vig_fim_mes = request.args.get('vigencia_fim_mes','')
    q = request.args.get('q','')

    where = " WHERE UPPER(COALESCE(pe.fonte_referencia,''))='SICRO'"
    params = []
    if uf:
        where += " AND pe.uf_referencia=?"; params.append(uf)
    if cat and not cat.startswith('Profissional'):
        where += " AND pe.categoria=?"; params.append(cat)
    if reg:
        where += " AND pe.regime=?"; params.append(reg)
    if vig_ini_mes:
        where += " AND substr(COALESCE(pe.vigencia_inicio,''),1,7)=?"; params.append(vig_ini_mes)
    if vig_fim_mes:
        where += " AND substr(COALESCE(pe.vigencia_fim,''),1,7)=?"; params.append(vig_fim_mes)
    if q:
        where += " AND (esp.codigo_profissional LIKE ? OR esp.descricao LIKE ?)"
        like = f'%{q}%'; params += [like, like]

    db = get_db()
    ensure_encargos_schema(db)
    rows = rows_to_list(db.execute("""
        SELECT
            pe.uf_referencia,
            pe.categoria,
            pe.vigencia,
            pe.vigencia_inicio,
            pe.vigencia_fim,
            esp.codigo_profissional,
            MAX(esp.descricao) AS descricao,
            MAX(esp.unidade) AS unidade,
            MAX(CASE WHEN pe.regime='Normal' THEN esp.total_grupo_a END) AS normal_a,
            MAX(CASE WHEN pe.regime='Normal' THEN esp.total_grupo_b END) AS normal_b,
            MAX(CASE WHEN pe.regime='Normal' THEN esp.total_grupo_c END) AS normal_c,
            MAX(CASE WHEN pe.regime='Normal' THEN esp.total_grupo_d END) AS normal_d,
            MAX(CASE WHEN pe.regime='Normal' THEN esp.encargo_total END) AS normal_total,
            MAX(CASE WHEN pe.regime='Desonerado' THEN esp.total_grupo_a END) AS desonerado_a,
            MAX(CASE WHEN pe.regime='Desonerado' THEN esp.total_grupo_b END) AS desonerado_b,
            MAX(CASE WHEN pe.regime='Desonerado' THEN esp.total_grupo_c END) AS desonerado_c,
            MAX(CASE WHEN pe.regime='Desonerado' THEN esp.total_grupo_d END) AS desonerado_d,
            MAX(CASE WHEN pe.regime='Desonerado' THEN esp.encargo_total END) AS desonerado_total
          FROM encargos_sicro_profissionais esp
          JOIN perfis_encargos pe ON pe.id_perfil = esp.id_perfil
    """ + where + """
         GROUP BY pe.uf_referencia, pe.categoria, pe.vigencia_inicio, pe.vigencia_fim,
                  pe.vigencia, esp.codigo_profissional
         ORDER BY pe.uf_referencia, pe.categoria, esp.codigo_profissional
    """, params).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/encargos/goinfra-profissionais', methods=['GET'])
def enc_goinfra_profissionais_analitico():
    uf = request.args.get('uf','')
    cat = request.args.get('categoria','')
    reg = request.args.get('regime','')
    vig_ini_mes = request.args.get('vigencia_inicio_mes','')
    vig_fim_mes = request.args.get('vigencia_fim_mes','')
    q = request.args.get('q','')

    where = " WHERE UPPER(COALESCE(pe.fonte_referencia,''))='GOINFRA'"
    params = []
    if uf:
        where += " AND pe.uf_referencia=?"; params.append(uf)
    if cat and not cat.startswith('Profissional'):
        where += " AND pe.categoria=?"; params.append(cat)
    if reg:
        where += " AND pe.regime=?"; params.append(reg)
    if vig_ini_mes:
        where += " AND substr(COALESCE(pe.vigencia_inicio,''),1,7)=?"; params.append(vig_ini_mes)
    if vig_fim_mes:
        where += " AND substr(COALESCE(pe.vigencia_fim,''),1,7)=?"; params.append(vig_fim_mes)
    if q:
        where += " AND (egp.codigo_profissional LIKE ? OR egp.descricao LIKE ?)"
        like = f'%{q}%'; params += [like, like]

    db = get_db()
    ensure_encargos_schema(db)
    rows = rows_to_list(db.execute("""
        SELECT
            pe.uf_referencia,
            pe.categoria,
            pe.vigencia,
            pe.vigencia_inicio,
            pe.vigencia_fim,
            egp.codigo_profissional,
            MAX(egp.descricao) AS descricao,
            MAX(egp.unidade) AS unidade,
            MAX(CASE WHEN pe.regime='Normal' THEN egp.total_grupo_a END) AS normal_a,
            MAX(CASE WHEN pe.regime='Normal' THEN egp.total_grupo_b END) AS normal_b,
            MAX(CASE WHEN pe.regime='Normal' THEN egp.total_grupo_c END) AS normal_c,
            MAX(CASE WHEN pe.regime='Normal' THEN egp.total_grupo_d END) AS normal_d,
            MAX(CASE WHEN pe.regime='Normal' THEN egp.encargo_total END) AS normal_total,
            MAX(CASE WHEN pe.regime='Desonerado' THEN egp.total_grupo_a END) AS desonerado_a,
            MAX(CASE WHEN pe.regime='Desonerado' THEN egp.total_grupo_b END) AS desonerado_b,
            MAX(CASE WHEN pe.regime='Desonerado' THEN egp.total_grupo_c END) AS desonerado_c,
            MAX(CASE WHEN pe.regime='Desonerado' THEN egp.total_grupo_d END) AS desonerado_d,
            MAX(CASE WHEN pe.regime='Desonerado' THEN egp.encargo_total END) AS desonerado_total
          FROM encargos_goinfra_profissionais egp
          JOIN perfis_encargos pe ON pe.id_perfil = egp.id_perfil
    """ + where + """
         GROUP BY pe.uf_referencia, pe.categoria, pe.vigencia_inicio, pe.vigencia_fim,
                  pe.vigencia, egp.codigo_profissional
         ORDER BY pe.uf_referencia, pe.categoria, CAST(egp.codigo_profissional AS INTEGER), egp.codigo_profissional
    """, params).fetchall())
    db.close()
    return jsonify(rows)

# ── Itens ──────────────────────────────────────────────────────────────────────
@app.route('/api/encargos/itens', methods=['POST'])
def enc_item_create():
    d = request.json or {}
    db = get_db()
    ensure_encargos_schema(db)
    cur = db.execute("""INSERT INTO itens_encargo
        (id_grupo_enc,descricao,base_legal,percentual,observacoes,ordem)
        VALUES (?,?,?,?,?,?)""",
        [d['id_grupo_enc'], d.get('descricao',''), d.get('base_legal'),
         float(d.get('percentual') or 0), d.get('observacoes'), d.get('ordem',0)])
    db.commit()
    # Recalcular perfil
    g = db.execute("SELECT id_perfil FROM grupos_encargos WHERE id_grupo_enc=?",
                   [d['id_grupo_enc']]).fetchone()
    if g: _calc_encargos(db, g[0])
    row = dict(db.execute("SELECT * FROM itens_encargo WHERE id_item=?",[cur.lastrowid]).fetchone())
    db.close(); return jsonify(row), 201

@app.route('/api/encargos/itens/<int:id>', methods=['PUT'])
def enc_item_update(id):
    d = request.json or {}
    db = get_db()
    ensure_encargos_schema(db)
    cur = db.execute("""UPDATE itens_encargo SET
        descricao=?,base_legal=?,percentual=?,observacoes=?,ordem=?
        WHERE id_item=?""",
        [d.get('descricao',''), d.get('base_legal'),
         float(d.get('percentual') or 0), d.get('observacoes'), d.get('ordem',0), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Item não encontrado.'}), 404
    g = db.execute("""SELECT ge.id_perfil FROM grupos_encargos ge
                      JOIN itens_encargo ie ON ie.id_grupo_enc=ge.id_grupo_enc
                      WHERE ie.id_item=?""",[id]).fetchone()
    if g: _calc_encargos(db, g[0])
    row = dict(db.execute("SELECT * FROM itens_encargo WHERE id_item=?",[id]).fetchone())
    db.close(); return jsonify(row)

@app.route('/api/encargos/itens/<int:id>', methods=['DELETE'])
def enc_item_delete(id):
    db = get_db()
    ensure_encargos_schema(db)
    g = db.execute("""SELECT ge.id_perfil FROM grupos_encargos ge
                      JOIN itens_encargo ie ON ie.id_grupo_enc=ge.id_grupo_enc
                      WHERE ie.id_item=?""",[id]).fetchone()
    pid = g[0] if g else None
    cur = db.execute('DELETE FROM itens_encargo WHERE id_item=?',[id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro':'Item não encontrado.'}), 404
    if pid: _calc_encargos(db, pid)
    db.close(); return jsonify({'mensagem':'Item excluído.'})

# ── Memória de cálculo (read-only, retorna o detalhamento completo) ────────────
@app.route('/api/encargos/perfis/<int:pid>/memoria', methods=['GET'])
def enc_memoria(pid):
    db = get_db()
    ensure_encargos_schema(db)
    p = db.execute('SELECT * FROM perfis_encargos WHERE id_perfil=?',[pid]).fetchone()
    if not p: db.close(); return jsonify({'erro':'Perfil não encontrado.'}), 404
    p = dict(p)
    totais = _calc_encargos(db, pid, recalc_d=False)
    grupos = rows_to_list(db.execute(
        "SELECT * FROM grupos_encargos WHERE id_perfil=? ORDER BY letra",[pid]).fetchall())
    for g in grupos:
        g['itens'] = rows_to_list(db.execute(
            "SELECT * FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem",
            [g['id_grupo_enc']]).fetchall())
    A = totais['A']; B = totais['B']; C = totais['C']; D = totais['D']
    db.close()
    return jsonify({
        'perfil': p,
        'grupos': grupos,
        'totais': totais,
        'formula': {
            'A': round(A, 4), 'B': round(B, 4), 'C': round(C, 4),
            'D': round(D, 4), 'total': round(totais['total'], 4),
            'fonte_d': 'Valores D1/D2 provenientes do SINAPI (01/2026). '
                       'Use "Recalcular D" para substituir pela fórmula A×(B+C).',
            'formula_texto': f"Total = A + B + C + D = "
                             f"{A:.4f} + {B:.4f} + {C:.4f} + {D:.4f} = {totais['total']:.4f}%"
        }
    })

@app.route('/api/encargos/perfis/<int:pid>/exportar-excel', methods=['GET'])
def enc_memoria_exportar_excel(pid):
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    db = get_db()
    ensure_encargos_schema(db)
    try:
        _calc_encargos(db, pid, recalc_d=False)
        p = db.execute("SELECT * FROM perfis_encargos WHERE id_perfil=?", [pid]).fetchone()
        if not p:
            return jsonify({'erro': 'Perfil não encontrado.'}), 404
        p = dict(p)
        grupos = rows_to_list(db.execute(
            "SELECT * FROM grupos_encargos WHERE id_perfil=? ORDER BY letra", [pid]).fetchall())
        for g in grupos:
            g['itens'] = rows_to_list(db.execute(
                "SELECT * FROM itens_encargo WHERE id_grupo_enc=? ORDER BY ordem",
                [g['id_grupo_enc']]).fetchall())
    finally:
        db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Memória Encargos'

    colors = {
        'A': 'DCEBFF',
        'B': 'DCFCE7',
        'C': 'FFF4D6',
        'D': 'E5E7EB',
    }
    dark = '0F172A'
    blue = '2563EB'
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    ws.merge_cells('A1:E1')
    ws['A1'] = 'Memória de Cálculo - Encargos Sociais'
    ws['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=dark)
    ws['A1'].alignment = Alignment(horizontal='center')

    meta = [
        ('Perfil', p.get('nome_perfil') or ''),
        ('Fonte', p.get('fonte_referencia') or ''),
        ('UF', p.get('uf_referencia') or ''),
        ('Categoria', p.get('categoria') or ''),
        ('Regime', 'Com Desoneração' if p.get('regime') == 'Desonerado' else 'Sem Desoneração'),
        ('Vigência', f"{p.get('vigencia_inicio') or ''} a {p.get('vigencia_fim') or ''}"),
    ]
    row = 3
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color=blue)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    subtotal_cells = {}
    for g in grupos:
        letra = g['letra']
        fill = PatternFill('solid', fgColor=colors.get(letra, 'F8FAFC'))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"Grupo {letra} - {g.get('descricao') or ''}")
        ws.cell(row=row, column=1).font = Font(bold=True, color=dark)
        ws.cell(row=row, column=1).fill = fill
        row += 1

        headers = ['Código', 'Descrição da Parcela', 'Base Legal', 'Percentual (%)', 'Observações']
        for col, header in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=header)
            c.font = Font(bold=True, color='334155')
            c.fill = PatternFill('solid', fgColor='F1F5F9')
            c.border = border
            c.alignment = Alignment(horizontal='center')
        row += 1

        first_item_row = row
        for it in g['itens']:
            desc = it.get('descricao') or ''
            codigo = ''
            if ' - ' in desc:
                codigo, desc = desc.split(' - ', 1)
            ws.cell(row=row, column=1, value=codigo)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value=it.get('base_legal') or '')
            ws.cell(row=row, column=4, value=float(it.get('percentual') or 0))
            ws.cell(row=row, column=5, value=it.get('observacoes') or '')
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
            ws.cell(row=row, column=4).number_format = '0.0000'
            row += 1
        last_item_row = row - 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.cell(row=row, column=1, value=f"Subtotal Grupo {letra}")
        ws.cell(row=row, column=4, value=f"=SUM(D{first_item_row}:D{last_item_row})" if last_item_row >= first_item_row else 0)
        subtotal_cells[letra] = f"D{row}"
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.font = Font(bold=True, color=dark)
            c.fill = fill
            c.border = border
        ws.cell(row=row, column=4).number_format = '0.0000'
        row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value='Total dos Encargos Sociais')
    ws.cell(row=row, column=4, value='=' + '+'.join(subtotal_cells[l] for l in ('A','B','C','D') if l in subtotal_cells))
    for col in range(1, 6):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=13, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=blue)
        c.border = border
    ws.cell(row=row, column=4).number_format = '0.0000'

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value='Fórmula: Total = Grupo A + Grupo B + Grupo C + Grupo D')
    ws.cell(row=row, column=1).font = Font(italic=True, color='475569')

    widths = [14, 48, 34, 16, 32]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A10'
    ws.sheet_view.showGridLines = False

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_safe = re.sub(r'[^A-Za-z0-9_-]+', '_', p.get('nome_perfil') or f'encargos_{pid}').strip('_')[:80]
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'memoria_encargos_{nome_safe}.xlsx'
    )

def _encargo_categoria_from_unidade(unidade, fallback='Horista'):
    u = (unidade or '').strip().lower()
    if u in ('h', 'hr', 'hora', 'horas'):
        return 'Horista'
    if u in ('mes', 'mês', 'meses', 'mensal'):
        return 'Mensalista'
    return fallback or 'Horista'

def _row_get(row, key, default=None):
    try:
        return row[key]
    except Exception:
        return default

def _encargos_norm_fonte(fonte):
    f = (fonte or '').strip().upper()
    aliases = {
        'SEINFRA/CE': 'SEINFRA',
        'SUDECAP/BH': 'SUDECAP',
        'SUDECAP/MG': 'SUDECAP',
        'SUDCAP': 'SUDECAP',
        'GOINFRA/GO': 'GOINFRA',
        'CDHU/SP': 'CDHU',
    }
    return aliases.get(f, f)

def _encargos_mesma_fonte(fonte_item, perfil):
    return _encargos_norm_fonte(fonte_item) == _encargos_norm_fonte(perfil.get('fonte_referencia') or '')

def _encargo_percentual_original(db, perfil_novo, categoria, fonte, uf, id_data_base):
    original_do_perfil = perfil_novo.get('encargo_original_percentual')
    if original_do_perfil not in (None, ''):
        try:
            return float(original_do_perfil or 0)
        except Exception:
            pass
    fonte = (fonte or perfil_novo.get('fonte_referencia') or 'SINAPI').strip().upper()
    uf = (uf or perfil_novo.get('uf_referencia') or '').strip().upper()
    regime = perfil_novo.get('regime') or 'Normal'
    data_ref = None
    if id_data_base:
        row = db.execute("SELECT mes, ano FROM datas_base WHERE id_data_base=?", [id_data_base]).fetchone()
        if row:
            data_ref = f"{int(row['ano']):04d}-{int(row['mes']):02d}-01"
    base_sql = """
        SELECT encargo_total FROM perfis_encargos
        WHERE situacao='Ativo'
          AND categoria=?
          AND regime=?
          AND upper(COALESCE(fonte_referencia,''))=?
          AND (uf_referencia=? OR uf_referencia IS NULL OR uf_referencia='')
    """
    params = [categoria, regime, fonte, uf]
    if data_ref:
        base_sql += " AND (vigencia_inicio IS NULL OR vigencia_inicio<=?) AND (vigencia_fim IS NULL OR vigencia_fim>=?)"
        params.extend([data_ref, data_ref])
    base_sql += " ORDER BY CASE WHEN uf_referencia=? THEN 0 ELSE 1 END, id_perfil LIMIT 1"
    params.append(uf)
    row = db.execute(base_sql, params).fetchone()
    if row:
        return float(row['encargo_total'] or 0)
    row = db.execute("""
        SELECT encargo_total FROM perfis_encargos
        WHERE situacao='Ativo' AND categoria=? AND regime=?
        ORDER BY CASE WHEN upper(COALESCE(fonte_referencia,''))='SINAPI' THEN 0 ELSE 1 END, id_perfil LIMIT 1
    """, [categoria, regime]).fetchone()
    return float(row['encargo_total'] or 0) if row else float(perfil_novo.get('encargo_total') or 0)

def _encargos_comp_secao_totais(db, id_composicao):
    rows = db.execute("""
        SELECT letra_secao, unidade,
               COALESCE(custo_total, COALESCE(quantidade,0) * COALESCE(preco_unitario,0)) AS total
        FROM composicoes_secao_itens
        WHERE id_composicao=?
    """, [id_composicao]).fetchall()
    total = sum(float(r['total'] or 0) for r in rows)
    mo = sum(float(r['total'] or 0) for r in rows if (r['letra_secao'] or '').upper() == 'B')
    unidade_mo = next((r['unidade'] for r in rows if (r['letra_secao'] or '').upper() == 'B'), '')
    return total, mo, unidade_mo

def _encargos_comp_itens_totais(db, id_composicao):
    rows = db.execute("""
        SELECT ic.unidade, ic.coeficiente, ic.preco_unitario, ic.custo_parcial,
               ic.tipo_item, i.tipo_insumo
        FROM itens_composicao ic
        LEFT JOIN insumos i ON i.codigo_insumo = ic.codigo_item
        WHERE ic.id_composicao=?
    """, [id_composicao]).fetchall()
    total = 0.0
    mo = 0.0
    unidade_mo = ''
    for r in rows:
        custo = float(r['custo_parcial'] if r['custo_parcial'] is not None else 0)
        if not custo:
            custo = float(r['coeficiente'] or 0) * float(r['preco_unitario'] or 0)
        total += custo
        tipo = (r['tipo_insumo'] or r['tipo_item'] or '').lower()
        if 'obra' in tipo or 'mao' in tipo or 'mão' in tipo:
            mo += custo
            if not unidade_mo:
                unidade_mo = r['unidade'] or ''
    return total, mo, unidade_mo

def _encargos_custo_composicao_ajustado(db, item_orc, perfil_novo, escopo='todos'):
    id_comp = item_orc['id_composicao']
    comp = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id_comp]).fetchone()
    if not comp:
        return None
    if escopo == 'mesma_fonte' and not _encargos_mesma_fonte(comp['fonte'], perfil_novo):
        return None
    total_sec, mo_sec, unidade_sec = _encargos_comp_secao_totais(db, id_comp)
    total_it, mo_it, unidade_it = _encargos_comp_itens_totais(db, id_comp)
    mo_atual = mo_sec if mo_sec > 0 else mo_it
    total_calc = total_sec if total_sec > 0 else total_it
    if mo_atual <= 0:
        return None
    custo_atual = float(item_orc['custo_unitario'] or 0) or float(comp['custo_unitario'] or 0) or total_calc
    categoria = _encargo_categoria_from_unidade(unidade_sec or unidade_it, perfil_novo.get('categoria'))
    enc_original = _encargo_percentual_original(
        db, perfil_novo, categoria, comp['fonte'], comp['uf_referencia'],
        _row_get(item_orc, 'id_data_base')
    )
    enc_novo = float(perfil_novo['encargo_total'] or 0)
    base_mo = mo_atual / (1 + enc_original / 100.0) if enc_original > -99 else mo_atual
    mo_novo = base_mo * (1 + enc_novo / 100.0)
    custo_novo = max(0.0, custo_atual + (mo_novo - mo_atual))
    return {
        'id_item': item_orc['id_item'],
        'custo_atual': round(custo_atual, 6),
        'custo_novo': round(custo_novo, 6),
        'mo_atual': round(mo_atual, 6),
        'mo_novo': round(mo_novo, 6),
        'encargo_original': round(enc_original, 6),
        'encargo_novo': round(enc_novo, 6),
        'categoria': categoria,
        'fonte': comp['fonte'],
    }

@app.route('/api/encargos/perfis/<int:pid>/aplicar-orcamento', methods=['POST'])
def enc_aplicar_orcamento(pid):
    d = request.json or {}
    id_orcamento = d.get('id_orcamento')
    escopo = d.get('escopo_aplicacao') or 'todos'
    if escopo not in ('todos', 'mesma_fonte'):
        return jsonify({'erro': 'Escopo de aplicação inválido.'}), 400
    if not id_orcamento:
        return jsonify({'erro': 'Selecione um orçamento sintético.'}), 400
    db = get_db()
    ensure_encargos_schema(db)
    try:
        _calc_encargos(db, pid)
        perfil = db.execute("SELECT * FROM perfis_encargos WHERE id_perfil=?", [pid]).fetchone()
        orc = db.execute("SELECT * FROM orcamentos WHERE id_orcamento=?", [id_orcamento]).fetchone()
        if not perfil:
            return jsonify({'erro': 'Perfil de encargos não encontrado.'}), 404
        if not orc:
            return jsonify({'erro': 'Orçamento não encontrado.'}), 404
        perfil = dict(perfil)
        itens = db.execute("""
            SELECT s.*, o.id_data_base, o.uf_referencia AS orc_uf
            FROM orcamento_sintetico s
            JOIN orcamentos o ON o.id_orcamento = s.id_orcamento
            WHERE s.id_orcamento=? AND s.tipo_linha='item'
        """, [id_orcamento]).fetchall()
        detalhes = []
        custo_antes = 0.0
        custo_depois = 0.0
        for it in itens:
            atual = float(it['custo_unitario'] or 0)
            novo = atual
            det = None
            if it['id_composicao']:
                det = _encargos_custo_composicao_ajustado(db, it, perfil, escopo)
            elif it['id_insumo']:
                ins = db.execute("SELECT * FROM insumos WHERE id_insumo=?", [it['id_insumo']]).fetchone()
                if ins and 'obra' in (ins['tipo_insumo'] or '').lower():
                    if escopo == 'mesma_fonte' and not _encargos_mesma_fonte(it['fonte'], perfil):
                        det = None
                    else:
                        categoria = _encargo_categoria_from_unidade(it['unidade'], perfil.get('categoria'))
                        enc_original = _encargo_percentual_original(db, perfil, categoria, it['fonte'], it['orc_uf'], it['id_data_base'])
                        enc_novo = float(perfil['encargo_total'] or 0)
                        base = atual / (1 + enc_original / 100.0) if enc_original > -99 else atual
                        novo = base * (1 + enc_novo / 100.0)
                        det = {'id_item': it['id_item'], 'custo_atual': atual, 'custo_novo': novo,
                               'mo_atual': atual, 'mo_novo': novo, 'encargo_original': enc_original,
                               'encargo_novo': enc_novo, 'categoria': categoria, 'fonte': it['fonte']}
            if det:
                novo = round(float(det['custo_novo'] or 0), 4)
                if abs(novo - atual) > 0.0001:
                    db.execute("UPDATE orcamento_sintetico SET custo_unitario=? WHERE id_item=?", [novo, it['id_item']])
                    detalhes.append(det)
            qtd = float(it['quantidade'] or 0)
            custo_antes += qtd * atual
            custo_depois += qtd * novo
        novo_custo_direto = db.execute("""
            SELECT COALESCE(SUM(quantidade * custo_unitario),0)
            FROM orcamento_sintetico
            WHERE id_orcamento=? AND tipo_linha='item'
        """, [id_orcamento]).fetchone()[0] or 0
        bdi_pct = float(orc['bdi_percentual'] or 0)
        novo_bdi = round(float(novo_custo_direto) * bdi_pct / 100.0, 4)
        db.execute("""
            UPDATE orcamentos
               SET valor_custo_direto=?, valor_bdi=?, valor_total=?
             WHERE id_orcamento=?
        """, [round(float(novo_custo_direto), 4), novo_bdi,
              round(float(novo_custo_direto) + novo_bdi, 4), id_orcamento])
        db.execute("""
            INSERT INTO encargos_orcamento_aplicacoes
                (id_orcamento,id_perfil,encargo_novo_percentual,itens_atualizados,custo_antes,custo_depois,observacoes)
            VALUES (?,?,?,?,?,?,?)
        """, [id_orcamento, pid, float(perfil['encargo_total'] or 0), len(detalhes),
              round(custo_antes, 4), round(custo_depois, 4),
              d.get('observacoes') or f"Aplicado perfil {perfil.get('nome_perfil')} - escopo {escopo}"])
        db.commit()
        return jsonify({
            'mensagem': f"Encargo social aplicado a {len(detalhes)} item(ns) do orçamento.",
            'itens_atualizados': len(detalhes),
            'escopo_aplicacao': escopo,
            'custo_antes': round(custo_antes, 4),
            'custo_depois': round(custo_depois, 4),
            'diferenca': round(custo_depois - custo_antes, 4),
            'detalhes': detalhes[:50],
        })
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 8 — CUSTO HORÁRIO DOS EQUIPAMENTOS
# ═══════════════════════════════════════════════════════════════════════════════

# Constantes SINAPI
TAXA_JUROS_SINAPI = 0.0617   # 6,17% a.a. (Ofício-Circular DNIT 08/2023)
TMA_SINAPI        = 0.025    # Taxa média para impostos/seguros (12/2022)

def _calcular_chp_chi(eq, preco_aq, preco_comb, preco_oper):
    """Calcula CHP e CHI dado os preços de entrada."""
    Va   = float(preco_aq   or 0)
    Pcomb= float(preco_comb or 0)
    Poper= float(preco_oper or 0)

    coef_dep = float(eq['coef_depreciacao'] or 0)
    coef_jur = float(eq['coef_juros']       or 0)
    coef_man = float(eq['coef_manutencao']  or 0)
    consumo  = float(eq['consumo_combustivel_hora'] or 0)
    coef_is  = float(eq['coef_impostos_seguros'] or 0)

    D    = round(coef_dep * Va, 4)
    J    = round(coef_jur * Va, 4)
    M    = round(coef_man * Va, 4)
    CMAT = round(consumo  * Pcomb, 4)
    CMOB = round(Poper, 4)
    IS   = round(coef_is  * Va, 4) if eq['tem_impostos_seguros'] else 0.0

    CHP = round(D + J + M + CMAT + CMOB + IS, 4)
    CHI = round(D + J       + CMOB + IS, 4)

    return {
        'D': D, 'J': J, 'M': M, 'CMAT': CMAT, 'CMOB': CMOB, 'IS': IS,
        'CHP': CHP, 'CHI': CHI,
    }

def _equip_codigo_variantes(*codigos):
    vals = set()
    for codigo in codigos:
        raw = (codigo or '').strip()
        if not raw:
            continue
        bare = raw.split('.', 1)[1] if '.' in raw else raw
        vals.update({raw, bare})
        for prefix in ('SINAPI', 'SICRO', 'SEINFRA', 'SUDECAP', 'GOINFRA', 'CDHU', 'USUARIO'):
            vals.add(f'{prefix}.{bare}')
    return [v for v in vals if v]

def _novo_codigo_composicao_usuario(db, codigo_base):
    base = (codigo_base or 'EQUIP').strip().replace('SINAPI.', '').replace('SICRO.', '').replace('USUARIO.', '')
    cand = f'USUARIO.{base}'
    i = 2
    while db.execute("SELECT 1 FROM composicoes WHERE codigo=? LIMIT 1", [cand]).fetchone():
        cand = f'USUARIO.{base}-{i}'
        i += 1
    return cand

def _criar_composicao_usuario_de_referencia(db, comp_orig, custo_unitario, observacao=''):
    comp_orig = dict(comp_orig)
    cod_novo = _novo_codigo_composicao_usuario(db, comp_orig.get('codigo'))
    cur = db.execute("""
        INSERT INTO composicoes
        (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,
         situacao_ref,custo_unitario,fic,producao_equipe,unidade_producao,situacao,observacoes,
         custo_horario_execucao,custo_unitario_execucao,custo_fic,subtotal_sicro)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [
        cod_novo, 'USUARIO', comp_orig.get('formato'), comp_orig.get('descricao'),
        comp_orig.get('unidade'), comp_orig.get('id_grupo_comp'), comp_orig.get('mes_referencia'),
        comp_orig.get('uf_referencia'), comp_orig.get('situacao_ref'), round(float(custo_unitario or 0), 4),
        comp_orig.get('fic'), comp_orig.get('producao_equipe'), comp_orig.get('unidade_producao'),
        'Ativo', observacao or comp_orig.get('observacoes'), comp_orig.get('custo_horario_execucao'),
        comp_orig.get('custo_unitario_execucao'), comp_orig.get('custo_fic'), comp_orig.get('subtotal_sicro')
    ])
    novo_id = cur.lastrowid

    for it in db.execute("SELECT * FROM itens_composicao WHERE id_composicao=? ORDER BY ordem", [comp_orig['id_composicao']]).fetchall():
        db.execute("""
            INSERT INTO itens_composicao
            (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,preco_unitario,custo_parcial,ordem)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, [
            novo_id, it['tipo_item'], it['codigo_item'], it['descricao'], it['unidade'], it['coeficiente'],
            it['situacao_item'], it['preco_unitario'], it['custo_parcial'], it['ordem']
        ])

    sec_map = {}
    for sec in db.execute("SELECT * FROM composicoes_secoes WHERE id_composicao=? ORDER BY ordem", [comp_orig['id_composicao']]).fetchall():
        cur_sec = db.execute("""
            INSERT INTO composicoes_secoes (id_composicao,letra_secao,nome_secao,custo_total_secao,ordem)
            VALUES (?,?,?,?,?)
        """, [novo_id, sec['letra_secao'], sec['nome_secao'], sec['custo_total_secao'], sec['ordem']])
        sec_map[sec['id_secao']] = cur_sec.lastrowid

    for it in db.execute("SELECT * FROM composicoes_secao_itens WHERE id_composicao=? ORDER BY ordem", [comp_orig['id_composicao']]).fetchall():
        db.execute("""
            INSERT INTO composicoes_secao_itens
            (id_composicao,id_secao,letra_secao,codigo_item,descricao,quantidade,unidade,util_operativa,
             util_improdutiva,custo_hp,custo_hi,preco_unitario,custo_total,cod_transporte,cod_transp_ln,
             cod_transp_rp,cod_transp_p,fit,dmt,ordem)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [
            novo_id, sec_map.get(it['id_secao']), it['letra_secao'], it['codigo_item'], it['descricao'],
            it['quantidade'], it['unidade'], it['util_operativa'], it['util_improdutiva'], it['custo_hp'],
            it['custo_hi'], it['preco_unitario'], it['custo_total'], it['cod_transporte'], it['cod_transp_ln'],
            it['cod_transp_rp'], it['cod_transp_p'], it['fit'], it['dmt'], it['ordem']
        ])

    return novo_id, cod_novo

def _impacto_equipamento_sinapi(db, eq):
    variantes_chp = _equip_codigo_variantes(eq.get('codigo_chp'))
    variantes_chi = _equip_codigo_variantes(eq.get('codigo_chi'))
    todos = list(dict.fromkeys(variantes_chp + variantes_chi))
    comps = []
    if todos:
        ph = ','.join('?' * len(todos))
        comps = rows_to_list(db.execute(f"""
            SELECT id_composicao, codigo, descricao, fonte, formato, unidade, custo_unitario
            FROM composicoes
            WHERE codigo IN ({ph})
            ORDER BY codigo
        """, todos).fetchall())
    comp_ids = [c['id_composicao'] for c in comps]
    orcs = []
    if comp_ids:
        phids = ','.join('?' * len(comp_ids))
        orcs = rows_to_list(db.execute(f"""
            SELECT os.id_item, os.id_orcamento, os.id_composicao, os.codigo, os.descricao,
                   os.quantidade, os.custo_unitario, o.nome_orcamento, ob.nome_obra
            FROM orcamento_sintetico os
            JOIN orcamentos o ON o.id_orcamento=os.id_orcamento
            LEFT JOIN obras ob ON ob.id_obra=o.id_obra
            WHERE os.id_composicao IN ({phids})
            ORDER BY o.nome_orcamento, os.ordem
        """, comp_ids).fetchall())
    return {
        'tipo': 'SINAPI',
        'equipamento': eq,
        'composicoes': comps,
        'orcamentos': orcs,
        'total_composicoes': len(comps),
        'total_orcamentos': len({o['id_item'] for o in orcs}),
        'tem_impacto': bool(comps or orcs),
    }

def _impacto_equipamento_sicro(db, eq):
    variantes = _equip_codigo_variantes(eq.get('codigo_chp'), eq.get('codigo_insumo_equip'))
    comps = []
    if variantes:
        ph = ','.join('?' * len(variantes))
        comps = rows_to_list(db.execute(f"""
            SELECT DISTINCT c.id_composicao, c.codigo, c.descricao, c.fonte, c.formato,
                   c.unidade, c.custo_unitario
            FROM composicoes_secao_itens si
            JOIN composicoes c ON c.id_composicao=si.id_composicao
            WHERE si.letra_secao='A' AND si.codigo_item IN ({ph})
            ORDER BY c.codigo
        """, variantes).fetchall())
    comp_ids = [c['id_composicao'] for c in comps]
    orcs = []
    if comp_ids:
        phids = ','.join('?' * len(comp_ids))
        orcs = rows_to_list(db.execute(f"""
            SELECT os.id_item, os.id_orcamento, os.id_composicao, os.codigo, os.descricao,
                   os.quantidade, os.custo_unitario, o.nome_orcamento, ob.nome_obra
            FROM orcamento_sintetico os
            JOIN orcamentos o ON o.id_orcamento=os.id_orcamento
            LEFT JOIN obras ob ON ob.id_obra=o.id_obra
            WHERE os.id_composicao IN ({phids})
            ORDER BY o.nome_orcamento, os.ordem
        """, comp_ids).fetchall())
    return {
        'tipo': 'SICRO',
        'equipamento': eq,
        'composicoes': comps,
        'orcamentos': orcs,
        'total_composicoes': len(comps),
        'total_orcamentos': len({o['id_item'] for o in orcs}),
        'tem_impacto': bool(comps or orcs),
    }

def _recalcular_composicao_sicro(db, id_composicao):
    secoes = {}
    for r in db.execute("""
        SELECT letra_secao, COALESCE(SUM(COALESCE(custo_total,0)),0) AS total
        FROM composicoes_secao_itens
        WHERE id_composicao=?
        GROUP BY letra_secao
    """, [id_composicao]).fetchall():
        secoes[r['letra_secao']] = float(r['total'] or 0)

    comp = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id_composicao]).fetchone()
    if not comp:
        return 0
    prod = float(comp['producao_equipe'] or 0)
    fic = float(comp['fic'] or 0)
    custo_horario_execucao = secoes.get('A', 0) + secoes.get('B', 0)
    custo_unitario_execucao = (custo_horario_execucao / prod) if prod else 0
    custo_fic = custo_unitario_execucao * fic
    subtotal = custo_unitario_execucao + custo_fic + secoes.get('C', 0) + secoes.get('D', 0)
    total = subtotal + secoes.get('E', 0) + secoes.get('F', 0)

    for letra, total_sec in secoes.items():
        db.execute("""
            UPDATE composicoes_secoes SET custo_total_secao=?
            WHERE id_composicao=? AND letra_secao=?
        """, [round(total_sec, 4), id_composicao, letra])
    db.execute("""
        UPDATE composicoes
        SET custo_horario_execucao=?, custo_unitario_execucao=?, custo_fic=?,
            subtotal_sicro=?, custo_unitario=?
        WHERE id_composicao=?
    """, [
        round(custo_horario_execucao, 4), round(custo_unitario_execucao, 4),
        round(custo_fic, 4), round(subtotal, 4), round(total, 4), id_composicao
    ])
    return round(total, 4)

SEL_EQ = """
    SELECT e.*, f.nome_familia
    FROM equipamentos_sinapi e
    LEFT JOIN familias_equipamentos f ON e.id_familia = f.id_familia
"""

# ── Famílias ───────────────────────────────────────────────────────────────────
@app.route('/api/equipamentos/familias', methods=['GET'])
def eq_familias():
    db = get_db()
    rows = rows_to_list(db.execute("""
        SELECT f.*, COUNT(e.id_equip) AS qtd_equipamentos
        FROM familias_equipamentos f
        LEFT JOIN equipamentos_sinapi e ON e.id_familia = f.id_familia
        GROUP BY f.id_familia ORDER BY f.nome_familia""").fetchall())
    db.close(); return jsonify(rows)

# ── Equipamentos — CRUD ────────────────────────────────────────────────────────
@app.route('/api/equipamentos', methods=['GET'])
def eq_list():
    q          = request.args.get('q','')
    id_familia = request.args.get('id_familia','')
    situacao   = request.args.get('situacao','')
    sistema    = request.args.get('sistema','')
    sql = SEL_EQ + " WHERE 1=1"
    params = []
    if q:          sql += " AND e.descricao LIKE ?"; params.append(f'%{q}%')
    if id_familia: sql += " AND e.id_familia=?";     params.append(id_familia)
    if situacao:   sql += " AND e.situacao=?";        params.append(situacao)
    if sistema:    sql += " AND COALESCE(e.sistema,'SINAPI')=?"; params.append(sistema)
    sql += " ORDER BY f.nome_familia, e.descricao"
    db = get_db()
    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/equipamentos/<int:id>', methods=['GET'])
def eq_get(id):
    db = get_db()
    row = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/equipamentos', methods=['POST'])
def eq_create():
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    db = get_db()
    cur = db.execute("""INSERT INTO equipamentos_sinapi
        (codigo_chp, codigo_chi, codigo_insumo_equip, codigo_insumo_comb, codigo_operador,
         descricao, id_familia, coef_depreciacao, coef_juros, coef_manutencao,
         consumo_combustivel_hora, unidade_combustivel, tem_impostos_seguros,
         coef_impostos_seguros, situacao)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [d.get('codigo_chp'), d.get('codigo_chi'), d.get('codigo_insumo_equip'),
         d.get('codigo_insumo_comb'), d.get('codigo_operador'), d['descricao'].strip(),
         d.get('id_familia') or None, d.get('coef_depreciacao'), d.get('coef_juros'),
         d.get('coef_manutencao'), d.get('consumo_combustivel_hora'),
         d.get('unidade_combustivel','L'), 1 if d.get('tem_impostos_seguros') else 0,
         d.get('coef_impostos_seguros'), d.get('situacao','Ativo')])
    db.commit()
    row = db.execute(SEL_EQ + " WHERE e.id_equip=?", [cur.lastrowid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/equipamentos/<int:id>', methods=['PUT'])
def eq_update(id):
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    db = get_db()
    cur = db.execute("""UPDATE equipamentos_sinapi SET
        codigo_chp=?,codigo_chi=?,codigo_insumo_equip=?,codigo_insumo_comb=?,
        codigo_operador=?,descricao=?,id_familia=?,coef_depreciacao=?,coef_juros=?,
        coef_manutencao=?,consumo_combustivel_hora=?,unidade_combustivel=?,
        tem_impostos_seguros=?,coef_impostos_seguros=?,situacao=?
        WHERE id_equip=?""",
        [d.get('codigo_chp'), d.get('codigo_chi'), d.get('codigo_insumo_equip'),
         d.get('codigo_insumo_comb'), d.get('codigo_operador'), d['descricao'].strip(),
         d.get('id_familia') or None, d.get('coef_depreciacao'), d.get('coef_juros'),
         d.get('coef_manutencao'), d.get('consumo_combustivel_hora'),
         d.get('unidade_combustivel','L'), 1 if d.get('tem_impostos_seguros') else 0,
         d.get('coef_impostos_seguros'), d.get('situacao','Ativo'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    row = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id]).fetchone()
    db.close(); return jsonify(dict(row))

@app.route('/api/equipamentos/<int:id>', methods=['DELETE'])
def eq_delete(id):
    db = get_db()
    total = db.execute("SELECT COUNT(*) FROM precos_equipamentos WHERE id_equip=?", [id]).fetchone()[0]
    if total > 0:
        db.close()
        return jsonify({'erro': f'Equipamento possui {total} registro(s) de preço. Exclua-os primeiro.'}), 409
    cur = db.execute("DELETE FROM equipamentos_sinapi WHERE id_equip=?", [id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    return jsonify({'mensagem': 'Equipamento excluído.'})

# ── Cálculo CHP/CHI (simulação sem salvar) ─────────────────────────────────────
@app.route('/api/equipamentos/<int:id>/calcular', methods=['POST'])
def eq_calcular(id):
    d = request.json or {}
    db = get_db()
    eq = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id]).fetchone()
    db.close()
    if not eq: return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    eq = dict(eq)
    resultado = _calcular_chp_chi(eq,
        d.get('preco_aquisicao', 0),
        d.get('preco_combustivel', 0),
        d.get('preco_operador_hora', 0))
    return jsonify({**resultado, 'equipamento': eq})

@app.route('/api/equipamentos/<int:id>/impacto', methods=['GET'])
def eq_impacto(id):
    db = get_db()
    row = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id]).fetchone()
    if not row:
        db.close()
        return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    eq = dict(row)
    sistema = (eq.get('sistema') or 'SINAPI').upper()
    impacto = _impacto_equipamento_sicro(db, eq) if sistema == 'SICRO' else _impacto_equipamento_sinapi(db, eq)
    db.close()
    return jsonify(impacto)

@app.route('/api/equipamentos/<int:id>/aplicar-custo', methods=['POST'])
def eq_aplicar_custo(id):
    d = request.json or {}
    try:
        chp = float(d.get('chp') or 0)
        chi = float(d.get('chi') or 0)
    except (TypeError, ValueError):
        return jsonify({'erro': 'Valores de CHP/CHI inválidos.'}), 400
    if chp <= 0 and chi <= 0:
        return jsonify({'erro': 'Informe ao menos um valor válido de CHP ou CHI.'}), 400
    modo = d.get('modo') or 'preservar'

    db = get_db()
    try:
        row = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id]).fetchone()
        if not row:
            return jsonify({'erro': 'Equipamento não encontrado.'}), 404
        eq = dict(row)
        sistema = (eq.get('sistema') or 'SINAPI').upper()
        db.execute("""
            UPDATE equipamentos_sinapi
            SET custo_produtivo=?, custo_improdutivo=?
            WHERE id_equip=?
        """, [chp if chp > 0 else None, chi if chi > 0 else None, id])

        if sistema == 'SICRO':
            impacto = _impacto_equipamento_sicro(db, eq)
            variantes = _equip_codigo_variantes(eq.get('codigo_chp'), eq.get('codigo_insumo_equip'))
            novo_insumo = None
            if modo == 'preservar':
                ins = None
                if variantes:
                    ph = ','.join('?' * len(variantes))
                    ins = db.execute(f"SELECT * FROM insumos WHERE codigo_insumo IN ({ph}) LIMIT 1", variantes).fetchone()
                codigo_base = eq.get('codigo_insumo_equip') or eq.get('codigo_chp') or f'EQ{id}'
                codigo_novo = _novo_codigo_preservado(db, codigo_base)
                if ins:
                    cur = db.execute("""
                        INSERT INTO insumos
                        (codigo_insumo,descricao,tipo_insumo,id_unidade,id_grupo,origem,encargos_aplicaveis,
                         situacao,observacoes,encargos_sociais_percentual)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, [
                        codigo_novo, ins['descricao'], ins['tipo_insumo'], ins['id_unidade'], ins['id_grupo'],
                        'USUARIO', ins['encargos_aplicaveis'], 'Ativo',
                        d.get('observacoes') or 'Cópia criada a partir de custo horário editado.',
                        ins['encargos_sociais_percentual']
                    ])
                else:
                    cur = db.execute("""
                        INSERT INTO insumos
                        (codigo_insumo,descricao,tipo_insumo,origem,situacao,observacoes)
                        VALUES (?,?,?,?,?,?)
                    """, [
                        codigo_novo, eq.get('descricao'), 'Equipamento', 'USUARIO', 'Ativo',
                        d.get('observacoes') or 'Criado a partir de custo horário SICRO editado.'
                    ])
                novo_insumo_id = cur.lastrowid
                db.execute("""
                    INSERT INTO precos_insumos
                    (id_insumo,uf_referencia,preco_desonerado,preco_nao_desonerado,preco_referencia,observacoes)
                    VALUES (?,?,?,?,?,?)
                """, [novo_insumo_id, d.get('uf_referencia') or eq.get('uf_referencia'), chp, chp, chp, d.get('observacoes')])
                novo_insumo = dict(db.execute("SELECT * FROM insumos WHERE id_insumo=?", [novo_insumo_id]).fetchone())
                db.commit()
                return jsonify({
                    'mensagem': 'Novo insumo de usuário criado. Composições e orçamentos existentes foram preservados.',
                    'tipo': 'SICRO',
                    'novo_insumo': novo_insumo,
                    'impacto': impacto,
                })

            if not variantes:
                return jsonify({'erro': 'Equipamento SICRO sem código para localizar composições.'}), 400
            ph = ','.join('?' * len(variantes))
            itens = db.execute(f"""
                SELECT id_item_secao, id_composicao, quantidade, util_operativa, util_improdutiva
                FROM composicoes_secao_itens
                WHERE letra_secao='A' AND codigo_item IN ({ph})
            """, variantes).fetchall()
            comp_ids = set()
            for it in itens:
                comp_ids.add(it['id_composicao'])
                qtd = float(it['quantidade'] or 0)
                uo = float(it['util_operativa'] or 0)
                ui = float(it['util_improdutiva'] or 0)
                hp = chp if chp > 0 else 0
                hi = chi if chi > 0 else 0
                custo_total = qtd * (uo * hp + ui * hi)
                db.execute("""
                    UPDATE composicoes_secao_itens
                    SET custo_hp=?, custo_hi=?, custo_total=?
                    WHERE id_item_secao=?
                """, [hp, hi, round(custo_total, 4), it['id_item_secao']])
            recalculadas = []
            for cid in comp_ids:
                total = _recalcular_composicao_sicro(db, cid)
                recalculadas.append({'id_composicao': cid, 'custo_unitario': total})
            orc_atualizados = 0
            if modo == 'atualizar_orcamentos' and comp_ids:
                phids = ','.join('?' * len(comp_ids))
                cur = db.execute(f"""
                    UPDATE orcamento_sintetico
                    SET custo_unitario = COALESCE((
                        SELECT c.custo_unitario FROM composicoes c
                        WHERE c.id_composicao = orcamento_sintetico.id_composicao
                    ), custo_unitario)
                    WHERE id_composicao IN ({phids})
                """, list(comp_ids))
                orc_atualizados = cur.rowcount
            db.commit()
            return jsonify({
                'mensagem': 'Custo horário SICRO aplicado às composições analíticas.',
                'tipo': 'SICRO',
                'composicoes_recalculadas': recalculadas,
                'orcamentos_atualizados': orc_atualizados,
                'impacto': impacto,
            })

        impacto = _impacto_equipamento_sinapi(db, eq)
        comp_map = {}
        for comp in impacto['composicoes']:
            codigo = str(comp.get('codigo') or '')
            valor = chp if (eq.get('codigo_chp') and codigo.split('.', 1)[-1] == str(eq.get('codigo_chp')).split('.', 1)[-1]) else chi
            if valor <= 0:
                continue
            original = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [comp['id_composicao']]).fetchone()
            novo_id, cod_novo = _criar_composicao_usuario_de_referencia(
                db, original, valor,
                d.get('observacoes') or 'Composição criada a partir de custo horário de equipamento editado.'
            )
            comp_map[comp['id_composicao']] = {'novo_id': novo_id, 'codigo': cod_novo, 'custo_unitario': round(valor, 4)}

        orc_atualizados = 0
        if modo == 'atualizar_orcamentos' and comp_map:
            for old_id, novo in comp_map.items():
                novo_comp = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [novo['novo_id']]).fetchone()
                cur = db.execute("""
                    UPDATE orcamento_sintetico
                    SET id_composicao=?, codigo=?, fonte='USUARIO', descricao=?, custo_unitario=?
                    WHERE id_composicao=?
                """, [novo['novo_id'], novo_comp['codigo'], novo_comp['descricao'], novo_comp['custo_unitario'], old_id])
                orc_atualizados += cur.rowcount
        db.commit()
        return jsonify({
            'mensagem': 'Composição(ões) de usuário criada(s) a partir do custo horário editado.',
            'tipo': 'SINAPI',
            'composicoes_criadas': list(comp_map.values()),
            'orcamentos_atualizados': orc_atualizados,
            'impacto': impacto,
        })
    except Exception as e:
        db.rollback()
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-800:]}), 500
    finally:
        db.close()

# ── Preços cadastrados (histórico) ─────────────────────────────────────────────
SEL_PRECO_EQ = """
    SELECT p.*, db2.mes, db2.ano, db2.descricao AS desc_data_base,
           fr.nome_fonte, e.descricao AS desc_equip
    FROM precos_equipamentos p
    LEFT JOIN datas_base db2 ON p.id_data_base = db2.id_data_base
    LEFT JOIN fontes_referencia fr ON p.id_fonte = fr.id_fonte
    LEFT JOIN equipamentos_sinapi e ON p.id_equip = e.id_equip
"""

@app.route('/api/equipamentos/<int:id_eq>/precos', methods=['GET'])
def eq_precos_list(id_eq):
    db = get_db()
    rows = rows_to_list(db.execute(
        SEL_PRECO_EQ + " WHERE p.id_equip=? ORDER BY p.id_preco_eq DESC", [id_eq]).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/equipamentos/<int:id_eq>/precos', methods=['POST'])
def eq_precos_create(id_eq):
    d = request.json or {}
    db = get_db()
    eq = db.execute(SEL_EQ + " WHERE e.id_equip=?", [id_eq]).fetchone()
    if not eq: db.close(); return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    eq = dict(eq)

    Va    = float(d.get('preco_aquisicao')    or 0)
    Pcomb = float(d.get('preco_combustivel')  or 0)
    Poper = float(d.get('preco_operador_hora')or 0)
    res   = _calcular_chp_chi(eq, Va, Pcomb, Poper)

    cur = db.execute("""INSERT INTO precos_equipamentos
        (id_equip, id_data_base, id_fonte, uf_referencia,
         preco_aquisicao, preco_combustivel, preco_operador_hora,
         custo_depreciacao, custo_juros, custo_manutencao,
         custo_materiais, custo_mao_obra, custo_imp_seguros,
         chp_calculado, chi_calculado, observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [id_eq, d.get('id_data_base') or None, d.get('id_fonte') or None,
         d.get('uf_referencia') or None,
         Va, Pcomb, Poper,
         res['D'], res['J'], res['M'], res['CMAT'], res['CMOB'], res['IS'],
         res['CHP'], res['CHI'], d.get('observacoes')])
    db.commit()
    row = db.execute(SEL_PRECO_EQ + " WHERE p.id_preco_eq=?", [cur.lastrowid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/precos-equipamentos/<int:id>', methods=['DELETE'])
def eq_preco_delete(id):
    db = get_db()
    cur = db.execute("DELETE FROM precos_equipamentos WHERE id_preco_eq=?", [id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro': 'Registro não encontrado.'}), 404
    return jsonify({'mensagem': 'Preço excluído.'})

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 4 — BDI
# ═══════════════════════════════════════════════════════════════════════════════

def _ano_bdi_perfil(p):
    try:
        if p.get('ano_orcamento'):
            return int(p.get('ano_orcamento'))
    except (TypeError, ValueError):
        pass
    vig = str(p.get('vigencia') or '')
    m = re.search(r'(20\d{2}|19\d{2})', vig)
    return int(m.group(1)) if m else 2026

def _cprb_por_regime_ano(regime, ano):
    if regime != 'Desonerado':
        return 0.0
    if ano <= 2024:
        return 4.5
    if ano == 2025:
        return 3.6
    if ano == 2026:
        return 2.7
    if ano == 2027:
        return 1.8
    return 0.0

def _componente_eh(desc, termo):
    return termo in (desc or '').lower()

def _bdi_parametros_reforma(db, p):
    ano = _ano_bdi_perfil(p)
    cbs = _float_or_zero(p.get('cbs_percentual')) / 100.0
    ibs = _float_or_zero(p.get('ibs_percentual')) / 100.0
    fator_efetivo = _float_or_zero(p.get('fator_efetivo_ivaeq'))
    mat = _float_or_zero(p.get('percentual_mat_ivaeq'))
    credito_bdi = _float_or_zero(p.get('credito_bdi_ivaeq'))
    if ano >= 2027 and (cbs or ibs):
        ivaeq = max(0.0, (cbs + ibs) * (fator_efetivo - mat - credito_bdi))
    else:
        ivaeq = _float_or_zero(p.get('ivaeq_percentual')) / 100.0
    iss = None
    origem_ivaeq = 'CBS/IBS e parâmetros da Reforma Tributária' if ano >= 2027 and (cbs or ibs) else 'manual'
    origem_iss = 'manual'

    id_orc = p.get('id_orcamento_ivaeq')
    if id_orc:
        orc = db.execute(SEL_ORC + " WHERE o.id_orcamento=?", [id_orc]).fetchone()
        if orc:
            orc = _enriquecer_orcamento_reforma(db, dict(orc))
            ivaeq = _float_or_zero(orc.get('ivaeq_percentual')) / 100.0
            ano = int(orc.get('ano_tributario') or ano)
            origem_ivaeq = f"orçamento {id_orc}"
            id_mun = orc.get('id_municipio')
            if id_mun:
                row = db.execute("""
                    SELECT aliquota_iss FROM municipio_aliquotas_anuais
                    WHERE id_municipio=? AND ano=?
                """, [id_mun, ano]).fetchone()
                if row:
                    iss = _float_or_zero(row['aliquota_iss'])
                    origem_iss = f"município do orçamento {id_orc}"

    if iss is None:
        iss_manual = p.get('iss_percentual_manual')
        if iss_manual is not None:
            iss = _float_or_zero(iss_manual)
        else:
            row = db.execute("""
                SELECT percentual FROM componentes_bdi
                WHERE id_perfil_bdi=? AND ativo=1 AND grupo='T'
                  AND lower(descricao) LIKE '%iss%'
                ORDER BY ordem LIMIT 1
            """, [p['id_perfil_bdi']]).fetchone()
            iss = _float_or_zero(row['percentual']) if row else 0.0
            origem_iss = 'componente ISS do perfil'

    return {
        'ano': ano,
        'ivaeq': ivaeq,
        'iss': iss / 100.0,
        'cbs': cbs,
        'ibs': ibs,
        'fator_efetivo': fator_efetivo,
        'percentual_mat': mat,
        'credito_bdi': credito_bdi,
        'origem_ivaeq': origem_ivaeq,
        'origem_iss': origem_iss,
    }

def _calc_bdi(db, pid):
    """Regras OrçaSmart:
       até 2026: [(1+AC+S+R)*(1+DF)*(1+L)/(1-T)-1]*100
       2027-2032: [(1+AC+S+R)*(1+DF)*(1+L)*(1+IVAeq)/(1-T)-1]*100
       2033+: [(1+AC+S+R)*(1+DF)*(1+L)*(1+IVAeq)-1]*100
       T considera ISS e CPRB aplicável; PIS/COFINS entram apenas até 2026.
    """
    ensure_bdi_reforma_fields(db)
    ensure_municipio_aliquotas_table(db)
    p = db.execute("SELECT * FROM perfis_bdi WHERE id_perfil_bdi=?", [pid]).fetchone()
    if not p:
        return {'AC':0, 'S':0, 'R':0, 'DF':0, 'L':0, 'T':0, 'bdi':0}
    p = dict(p)
    ano = _ano_bdi_perfil(p)

    rows = db.execute("""
        SELECT grupo, descricao, percentual
        FROM componentes_bdi
        WHERE id_perfil_bdi=? AND ativo=1
    """, [pid]).fetchall()
    g = {}
    iss_comp = 0.0
    pis_cofins = 0.0
    cprb_comp = 0.0
    simples_tributos = 0.0
    for r in rows:
        grp = r['grupo']
        pct = _float_or_zero(r['percentual']) / 100.0
        desc = r['descricao'] or ''
        if grp == 'T':
            if p.get('regime_tributario') == 'Simples Nacional':
                if not (_componente_eh(desc, 'irpj') or _componente_eh(desc, 'csll')):
                    simples_tributos += pct
                    if _componente_eh(desc, 'iss'):
                        iss_comp += pct
                    elif _componente_eh(desc, 'cprb') or _componente_eh(desc, 'previdenci'):
                        cprb_comp += pct
                    elif _componente_eh(desc, 'pis') or _componente_eh(desc, 'cofins') or _componente_eh(desc, 'cbs'):
                        pis_cofins += pct
                continue
            if _componente_eh(desc, 'iss'):
                iss_comp += pct
            elif ano <= 2026 and (_componente_eh(desc, 'pis') or _componente_eh(desc, 'cofins')):
                pis_cofins += pct
            continue
        g[grp] = g.get(grp, 0.0) + pct

    params = _bdi_parametros_reforma(db, p)
    ano = params['ano']
    iss = params['iss'] if p.get('id_orcamento_ivaeq') or p.get('iss_percentual_manual') is not None else iss_comp
    regime_prev = p.get('regime_previdenciario') or ('Desonerado' if p.get('regime_tributario') == 'Desonerado' else 'Onerado')
    cprb = _cprb_por_regime_ano('Desonerado' if regime_prev == 'Desonerado' else p.get('regime_tributario'), ano) / 100.0

    AC = g.get('AC',0); S = g.get('S',0); R = g.get('R',0)
    DF = g.get('DF',0); L = g.get('L',0); Outros = g.get('Outros',0)
    if p.get('regime_tributario') == 'Simples Nacional':
        T = simples_tributos
        iss = iss_comp
        cprb = cprb_comp
    else:
        T = iss + cprb + pis_cofins
    K = (1 + AC + S + R) * (1 + DF) * (1 + L)
    ivaeq = params['ivaeq'] if ano >= 2027 else 0.0
    if p.get('regime_tributario') == 'Simples Nacional':
        denom = max(0.001, 1 - T)
        bdi = ((K / denom) - 1) * 100
    else:
        denom = 1.0 if ano >= 2033 else max(0.001, 1 - T)
        bdi = ((K * (1 + ivaeq) / denom) - 1) * 100
    bdi = round(bdi + Outros * 100, 4)

    db.execute("UPDATE perfis_bdi SET bdi_percentual=?, ano_orcamento=?, ivaeq_percentual=?, iss_percentual_manual=? WHERE id_perfil_bdi=?",
               [bdi, ano, round(params['ivaeq'] * 100, 4), round(iss * 100, 4), pid])
    db.commit()
    return {
        'AC':round(AC*100,4), 'S':round(S*100,4), 'R':round(R*100,4),
        'DF':round(DF*100,4), 'L':round(L*100,4), 'T':round(T*100,4),
        'ISS':round(iss*100,4), 'CPRB':round(cprb*100,4), 'PIS_COFINS':round(pis_cofins*100,4),
        'CBS':round(params.get('cbs',0)*100,4), 'IBS':round(params.get('ibs',0)*100,4),
        'FATOR_EFETIVO':round(params.get('fator_efetivo',0)*100,4),
        'PERCENTUAL_MAT':round(params.get('percentual_mat',0)*100,4),
        'IVAeq':round(params['ivaeq']*100,4), 'ano': ano,
        'origem_ivaeq': params['origem_ivaeq'], 'origem_iss': params['origem_iss'],
        'bdi': bdi
    }

SEL_BDI = """
    SELECT b.*, COUNT(c.id_componente) AS qtd_componentes
    FROM perfis_bdi b
    LEFT JOIN componentes_bdi c ON c.id_perfil_bdi = b.id_perfil_bdi AND c.ativo=1
    WHERE 1=1
"""

@app.route('/api/bdi/perfis', methods=['GET'])
def bdi_list():
    tipo   = request.args.get('tipo','')
    regime = request.args.get('regime','')
    ano    = request.args.get('ano','')
    quartil = request.args.get('quartil','')
    q      = request.args.get('q','')
    sql    = SEL_BDI
    params = []
    if tipo:   sql += " AND b.tipo_obra=?";          params.append(tipo)
    if regime: sql += " AND b.regime_tributario=?";   params.append(regime)
    if ano:    sql += " AND b.ano_orcamento=?";        params.append(ano)
    if quartil: sql += " AND b.quartil=?";             params.append(quartil)
    faixa = request.args.get('faixa_simples','')
    if faixa: sql += " AND b.simples_faixa=?";          params.append(faixa)
    if q:      sql += " AND b.nome_perfil LIKE ?";    params.append(f'%{q}%')
    sql += " GROUP BY b.id_perfil_bdi ORDER BY b.tipo_obra, b.nome_perfil"
    db = get_db()
    ensure_bdi_reforma_fields(db)
    # Recalcular todos
    pids = [r[0] for r in db.execute("SELECT id_perfil_bdi FROM perfis_bdi").fetchall()]
    for pid in pids:
        _calc_bdi(db, pid)
    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/bdi/perfis/<int:id>', methods=['GET'])
def bdi_get(id):
    db = get_db()
    ensure_bdi_reforma_fields(db)
    row = db.execute(SEL_BDI + " AND b.id_perfil_bdi=? GROUP BY b.id_perfil_bdi", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro': 'Perfil não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/bdi/perfis', methods=['POST'])
def bdi_create():
    d = request.json or {}
    if not d.get('nome_perfil','').strip():
        return jsonify({'erro': 'Nome do perfil é obrigatório.'}), 400
    db = get_db()
    ensure_bdi_reforma_fields(db)
    cur = db.execute("""INSERT INTO perfis_bdi
        (nome_perfil,tipo_obra,regime_tributario,descricao,usa_reforma_tributaria,vigencia,observacoes,
         ano_orcamento,quartil,cbs_percentual,ibs_percentual,fator_efetivo_ivaeq,percentual_mat_ivaeq,
         credito_bdi_ivaeq,ivaeq_percentual,iss_percentual_manual,id_orcamento_ivaeq,
         regime_previdenciario,simples_faixa,simples_faixa_label,simples_receita_limite,
         simples_aliquota_efetiva,simples_irpj_percentual,simples_csll_percentual)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [d['nome_perfil'].strip(), d.get('tipo_obra'), d.get('regime_tributario','Normal'),
         d.get('descricao'), 1 if d.get('usa_reforma_tributaria') else 0,
         d.get('vigencia'), d.get('observacoes'),
         d.get('ano_orcamento'), d.get('quartil'),
         _float_or_zero(d.get('cbs_percentual')), _float_or_zero(d.get('ibs_percentual')),
         _float_or_zero(d.get('fator_efetivo_ivaeq') if d.get('fator_efetivo_ivaeq') is not None else 0.5),
         _float_or_zero(d.get('percentual_mat_ivaeq') if d.get('percentual_mat_ivaeq') is not None else 0.4),
         _float_or_zero(d.get('credito_bdi_ivaeq')),
         _float_or_zero(d.get('ivaeq_percentual')),
         d.get('iss_percentual_manual'), d.get('id_orcamento_ivaeq'),
         d.get('regime_previdenciario') or 'Onerado',
         d.get('simples_faixa'), d.get('simples_faixa_label'),
         d.get('simples_receita_limite'),
         _float_or_zero(d.get('simples_aliquota_efetiva')),
         _float_or_zero(d.get('simples_irpj_percentual')),
         _float_or_zero(d.get('simples_csll_percentual'))])
    pid = cur.lastrowid
    # Criar grupos vazios padrão
    ordem_grupos = {'AC': 1, 'S': 2, 'R': 3, 'DF': 4, 'L': 5, 'T': 6}
    for grp in ('AC','S','R','DF','L','T'):
        descs = {'AC':'Administração Central','S':'Seguros e Garantias',
                 'R':'Riscos','DF':'Despesas Financeiras','L':'Lucro','T':'Tributos'}
        db.execute("""INSERT INTO componentes_bdi
            (id_perfil_bdi,grupo,codigo,descricao,percentual,ordem)
            VALUES (?,?,?,?,0,?)""",
            [pid, grp, grp+'1', descs[grp], ordem_grupos[grp]])
    db.commit()
    _calc_bdi(db, pid)
    row = db.execute(SEL_BDI + " AND b.id_perfil_bdi=? GROUP BY b.id_perfil_bdi", [pid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/bdi/perfis/<int:id>', methods=['PUT'])
def bdi_update(id):
    d = request.json or {}
    if not d.get('nome_perfil','').strip():
        return jsonify({'erro': 'Nome do perfil é obrigatório.'}), 400
    db = get_db()
    ensure_bdi_reforma_fields(db)
    cur = db.execute("""UPDATE perfis_bdi SET
        nome_perfil=?,tipo_obra=?,regime_tributario=?,descricao=?,
        usa_reforma_tributaria=?,vigencia=?,observacoes=?,situacao=?,
        ano_orcamento=?,quartil=?,cbs_percentual=?,ibs_percentual=?,fator_efetivo_ivaeq=?,
        percentual_mat_ivaeq=?,credito_bdi_ivaeq=?,ivaeq_percentual=?,iss_percentual_manual=?,id_orcamento_ivaeq=?,
        regime_previdenciario=?,simples_faixa=?,simples_faixa_label=?,simples_receita_limite=?,
        simples_aliquota_efetiva=?,simples_irpj_percentual=?,simples_csll_percentual=?
        WHERE id_perfil_bdi=?""",
        [d['nome_perfil'].strip(), d.get('tipo_obra'), d.get('regime_tributario','Normal'),
         d.get('descricao'), 1 if d.get('usa_reforma_tributaria') else 0,
         d.get('vigencia'), d.get('observacoes'), d.get('situacao','Ativo'),
         d.get('ano_orcamento'), d.get('quartil'),
         _float_or_zero(d.get('cbs_percentual')), _float_or_zero(d.get('ibs_percentual')),
         _float_or_zero(d.get('fator_efetivo_ivaeq') if d.get('fator_efetivo_ivaeq') is not None else 0.5),
         _float_or_zero(d.get('percentual_mat_ivaeq') if d.get('percentual_mat_ivaeq') is not None else 0.4),
         _float_or_zero(d.get('credito_bdi_ivaeq')),
         _float_or_zero(d.get('ivaeq_percentual')),
         d.get('iss_percentual_manual'), d.get('id_orcamento_ivaeq'),
         d.get('regime_previdenciario') or 'Onerado',
         d.get('simples_faixa'), d.get('simples_faixa_label'),
         d.get('simples_receita_limite'),
         _float_or_zero(d.get('simples_aliquota_efetiva')),
         _float_or_zero(d.get('simples_irpj_percentual')),
         _float_or_zero(d.get('simples_csll_percentual')), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro': 'Perfil não encontrado.'}), 404
    _calc_bdi(db, id)
    row = db.execute(SEL_BDI + " AND b.id_perfil_bdi=? GROUP BY b.id_perfil_bdi", [id]).fetchone()
    db.close(); return jsonify(dict(row))

@app.route('/api/bdi/perfis/<int:id>', methods=['DELETE'])
def bdi_delete(id):
    db = get_db()
    cur = db.execute("DELETE FROM perfis_bdi WHERE id_perfil_bdi=?", [id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro': 'Perfil não encontrado.'}), 404
    return jsonify({'mensagem': 'Perfil BDI excluído.'})

@app.route('/api/bdi/perfis/<int:id>/duplicar', methods=['POST'])
def bdi_duplicate(id):
    db = get_db()
    ensure_bdi_reforma_fields(db)
    p = db.execute("SELECT * FROM perfis_bdi WHERE id_perfil_bdi=?", [id]).fetchone()
    if not p: db.close(); return jsonify({'erro': 'Perfil não encontrado.'}), 404
    p = dict(p)
    cur = db.execute("""INSERT INTO perfis_bdi
        (nome_perfil,tipo_obra,regime_tributario,descricao,usa_reforma_tributaria,vigencia,
         ano_orcamento,quartil,cbs_percentual,ibs_percentual,fator_efetivo_ivaeq,percentual_mat_ivaeq,
         credito_bdi_ivaeq,ivaeq_percentual,iss_percentual_manual,id_orcamento_ivaeq,
         regime_previdenciario,simples_faixa,simples_faixa_label,simples_receita_limite,
         simples_aliquota_efetiva,simples_irpj_percentual,simples_csll_percentual)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        ['Cópia de '+p['nome_perfil'], p['tipo_obra'], p['regime_tributario'],
         p['descricao'], p['usa_reforma_tributaria'], p['vigencia'],
         p.get('ano_orcamento'), p.get('quartil'), p.get('cbs_percentual'), p.get('ibs_percentual'),
         p.get('fator_efetivo_ivaeq'), p.get('percentual_mat_ivaeq'), p.get('credito_bdi_ivaeq'),
         p.get('ivaeq_percentual'), p.get('iss_percentual_manual'),
         p.get('id_orcamento_ivaeq'), p.get('regime_previdenciario'), p.get('simples_faixa'),
         p.get('simples_faixa_label'), p.get('simples_receita_limite'),
         p.get('simples_aliquota_efetiva'), p.get('simples_irpj_percentual'), p.get('simples_csll_percentual')])
    novo = cur.lastrowid
    comps = db.execute("SELECT * FROM componentes_bdi WHERE id_perfil_bdi=?", [id]).fetchall()
    for c in comps:
        c = dict(c)
        db.execute("""INSERT INTO componentes_bdi
            (id_perfil_bdi,grupo,codigo,descricao,base_legal,percentual,incide_sobre,ativo,ordem,observacoes)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            [novo, c['grupo'], c['codigo'], c['descricao'], c['base_legal'],
             c['percentual'], c['incide_sobre'], c['ativo'], c['ordem'], c['observacoes']])
    db.commit()
    _calc_bdi(db, novo)
    row = db.execute(SEL_BDI + " AND b.id_perfil_bdi=? GROUP BY b.id_perfil_bdi", [novo]).fetchone()
    db.close(); return jsonify(dict(row)), 201

# ── Componentes ────────────────────────────────────────────────────────────────
@app.route('/api/bdi/perfis/<int:pid>/componentes', methods=['GET'])
def bdi_comp_list(pid):
    db = get_db()
    ensure_bdi_reforma_fields(db)
    rows = rows_to_list(db.execute("""SELECT * FROM componentes_bdi
        WHERE id_perfil_bdi=? ORDER BY grupo, ordem""", [pid]).fetchall())
    db.close(); return jsonify(rows)

@app.route('/api/bdi/componentes', methods=['POST'])
def bdi_comp_create():
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    db = get_db()
    cur = db.execute("""INSERT INTO componentes_bdi
        (id_perfil_bdi,grupo,codigo,descricao,base_legal,percentual,incide_sobre,ativo,ordem,observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        [d['id_perfil_bdi'], d.get('grupo','Outros'), d.get('codigo'),
         d['descricao'].strip(), d.get('base_legal'),
         float(d.get('percentual') or 0), d.get('incide_sobre','CD'),
         1 if d.get('ativo', 1) else 0, d.get('ordem', 99), d.get('observacoes')])
    pid = d['id_perfil_bdi']
    db.commit()
    _calc_bdi(db, pid)
    row = dict(db.execute("SELECT * FROM componentes_bdi WHERE id_componente=?", [cur.lastrowid]).fetchone())
    db.close(); return jsonify(row), 201

@app.route('/api/bdi/componentes/<int:id>', methods=['PUT'])
def bdi_comp_update(id):
    d = request.json or {}
    db = get_db()
    cur = db.execute("""UPDATE componentes_bdi SET
        grupo=?,codigo=?,descricao=?,base_legal=?,percentual=?,
        incide_sobre=?,ativo=?,ordem=?,observacoes=?
        WHERE id_componente=?""",
        [d.get('grupo','Outros'), d.get('codigo'), d.get('descricao','').strip(),
         d.get('base_legal'), float(d.get('percentual') or 0),
         d.get('incide_sobre','CD'), 1 if d.get('ativo',True) else 0,
         d.get('ordem',0), d.get('observacoes'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro': 'Componente não encontrado.'}), 404
    pid = db.execute("SELECT id_perfil_bdi FROM componentes_bdi WHERE id_componente=?", [id]).fetchone()
    if pid: _calc_bdi(db, pid[0])
    row = dict(db.execute("SELECT * FROM componentes_bdi WHERE id_componente=?", [id]).fetchone())
    db.close(); return jsonify(row)

@app.route('/api/bdi/componentes/<int:id>', methods=['DELETE'])
def bdi_comp_delete(id):
    db = get_db()
    pid = db.execute("SELECT id_perfil_bdi FROM componentes_bdi WHERE id_componente=?", [id]).fetchone()
    cur = db.execute("DELETE FROM componentes_bdi WHERE id_componente=?", [id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro': 'Componente não encontrado.'}), 404
    if pid: _calc_bdi(db, pid[0])
    db.close(); return jsonify({'mensagem': 'Componente excluído.'})

# ── Memória de cálculo BDI ─────────────────────────────────────────────────────
@app.route('/api/bdi/perfis/<int:pid>/memoria', methods=['GET'])
def bdi_memoria(pid):
    db = get_db()
    ensure_bdi_reforma_fields(db)
    p = db.execute("SELECT * FROM perfis_bdi WHERE id_perfil_bdi=?", [pid]).fetchone()
    if not p: db.close(); return jsonify({'erro': 'Perfil não encontrado.'}), 404
    p = dict(p)
    totais = _calc_bdi(db, pid)
    comps  = rows_to_list(db.execute("""SELECT * FROM componentes_bdi
        WHERE id_perfil_bdi=? AND ativo=1 ORDER BY grupo, ordem""", [pid]).fetchall())
    AC=totais['AC']; S=totais['S']; R=totais['R']
    DF=totais['DF']; L=totais['L']; T=totais['T']
    ISS=totais.get('ISS', 0); CPRB=totais.get('CPRB', 0); IVAeq=totais.get('IVAeq', 0)
    ano=totais.get('ano', _ano_bdi_perfil(p))
    bdi = totais['bdi']
    db.close()
    if p.get('regime_tributario') == 'Simples Nacional':
        expressao = 'BDI Simples = {[(1+AC+S+R)×(1+DF)×(1+L)/(1-T Simples)] - 1} × 100'
        texto = (f"BDI = {{[(1+{(AC+S+R)/100:.4f})×(1+{DF/100:.4f})×"
                 f"(1+{L/100:.4f}) / (1-{T/100:.4f})] - 1}} × 100 = {bdi:.4f}%")
    elif ano <= 2026:
        expressao = 'BDI = {[(1+AC+S+R)×(1+DF)×(1+L)/(1-T)] - 1} × 100'
        texto = (f"BDI = {{[(1+{(AC+S+R)/100:.4f})×(1+{DF/100:.4f})×"
                 f"(1+{L/100:.4f}) / (1-{T/100:.4f})] - 1}} × 100 = {bdi:.4f}%")
    elif ano < 2033:
        expressao = 'BDI = {[(1+AC+S+R)×(1+DF)×(1+L)×(1+IVAeq)/(1-T)] - 1} × 100'
        texto = (f"BDI = {{[(1+{(AC+S+R)/100:.4f})×(1+{DF/100:.4f})×"
                 f"(1+{L/100:.4f})×(1+{IVAeq/100:.4f}) / (1-{T/100:.4f})] - 1}} × 100 = {bdi:.4f}%")
    else:
        expressao = 'BDI = {[(1+AC+S+R)×(1+DF)×(1+L)×(1+IVAeq)] - 1} × 100'
        texto = (f"BDI = {{[(1+{(AC+S+R)/100:.4f})×(1+{DF/100:.4f})×"
                 f"(1+{L/100:.4f})×(1+{IVAeq/100:.4f})] - 1}} × 100 = {bdi:.4f}%")
    return jsonify({
        'perfil': p,
        'componentes': comps,
        'totais_grupo': totais,
        'formula': {
            'expressao': expressao,
            'AC': AC, 'S': S, 'R': R, 'DF': DF, 'L': L, 'T': T,
            'ISS': ISS, 'CPRB': CPRB, 'IVAeq': IVAeq, 'ano': ano,
            'bdi': bdi,
            'texto': texto,
            'fonte': 'OrçaSmart: TCU Acórdão 2622/2013-Plenário adaptado à transição da reforma tributária',
        }
    })

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 5 — COMPOSIÇÕES DE CUSTO
# ═══════════════════════════════════════════════════════════════════════════════

SEL_COMP = """
    SELECT c.*, g.nome_grupo AS nome_grupo_comp
    FROM composicoes c
    LEFT JOIN grupos_composicoes g ON c.id_grupo_comp = g.id_grupo_comp
"""

def _comp_codigo_variantes(codigo):
    cod = (codigo or '').strip()
    if not cod:
        return []
    variantes = {cod}
    if '.' in cod:
        variantes.add(cod.split('.')[-1])
    for prefixo in ('SINAPI.', 'SICRO.', 'SEINFRA.', 'SUDECAP.', 'GOINFRA.', 'CDHU.', 'USUARIO.'):
        if cod.startswith(prefixo):
            variantes.add(cod[len(prefixo):])
        else:
            variantes.add(prefixo + cod)
    return [v for v in variantes if v]

def _impacto_composicao(db, id_composicao):
    comp = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id_composicao]).fetchone()
    if not comp:
        return None
    comp = dict(comp)

    parents = {}
    fila = [comp]
    vistos = {id_composicao}
    while fila:
        atual = fila.pop(0)
        variantes = _comp_codigo_variantes(atual.get('codigo'))
        if not variantes:
            continue
        qs = ','.join(['?'] * len(variantes))
        rows = rows_to_list(db.execute(f"""
            SELECT DISTINCT c.*
            FROM itens_composicao ic
            JOIN composicoes c ON c.id_composicao = ic.id_composicao
            WHERE UPPER(COALESCE(ic.tipo_item,'')) = 'COMPOSICAO'
              AND ic.codigo_item IN ({qs})
              AND c.id_composicao <> ?
        """, variantes + [atual['id_composicao']]).fetchall())
        for row in rows:
            cid = row['id_composicao']
            if cid not in parents:
                parents[cid] = row
            if cid not in vistos:
                vistos.add(cid)
                fila.append(row)

    variantes_origem = _comp_codigo_variantes(comp.get('codigo'))
    params_dir = [id_composicao]
    where_dir = "os.id_composicao = ?"
    if variantes_origem:
        where_dir += " OR os.codigo IN (" + ",".join(["?"] * len(variantes_origem)) + ")"
        params_dir += variantes_origem

    diretos = rows_to_list(db.execute(f"""
        SELECT os.id_item, os.id_orcamento, os.descricao, os.codigo, os.quantidade,
               os.custo_unitario, os.id_composicao,
               o.nome_orcamento, o.versao, o.status,
               ob.nome_obra
        FROM orcamento_sintetico os
        JOIN orcamentos o  ON o.id_orcamento = os.id_orcamento
        LEFT JOIN obras ob ON ob.id_obra = o.id_obra
        WHERE {where_dir}
        ORDER BY o.nome_orcamento, os.ordem
    """, params_dir).fetchall())
    for row in diretos:
        row['impacto_tipo'] = 'direto'

    indiretos = []
    parent_ids = list(parents.keys())
    if parent_ids:
        qs = ','.join(['?'] * len(parent_ids))
        indiretos = rows_to_list(db.execute(f"""
            SELECT os.id_item, os.id_orcamento, os.descricao, os.codigo, os.quantidade,
                   os.custo_unitario, os.id_composicao,
                   o.nome_orcamento, o.versao, o.status,
                   ob.nome_obra
            FROM orcamento_sintetico os
            JOIN orcamentos o  ON o.id_orcamento = os.id_orcamento
            LEFT JOIN obras ob ON ob.id_obra = o.id_obra
            WHERE os.id_composicao IN ({qs})
            ORDER BY o.nome_orcamento, os.ordem
        """, parent_ids).fetchall())
        for row in indiretos:
            row['impacto_tipo'] = 'indireto'

    combinados = {}
    for row in diretos + indiretos:
        item_id = row.get('id_item')
        if item_id not in combinados:
            combinados[item_id] = row
        elif combinados[item_id].get('impacto_tipo') != 'direto':
            combinados[item_id] = row

    return {
        'composicao': comp,
        'composicoes_auxiliares': list(parents.values()),
        'orcamentos_diretos': diretos,
        'orcamentos_indiretos': indiretos,
        'orcamentos': list(combinados.values()),
        'qtd_orcamentos': len(combinados),
        'qtd_composicoes_auxiliares': len(parents),
    }

def _recalcular_composicao_unitaria(db, id_composicao):
    itens = rows_to_list(db.execute(
        "SELECT * FROM itens_composicao WHERE id_composicao=? ORDER BY ordem, id_item",
        [id_composicao]
    ).fetchall())
    total = 0.0
    for it in itens:
        coef = float(it.get('coeficiente') or 0)
        preco = it.get('preco_unitario')
        if (it.get('tipo_item') or '').upper() == 'COMPOSICAO':
            variantes = _comp_codigo_variantes(it.get('codigo_item'))
            if variantes:
                qs = ','.join(['?'] * len(variantes))
                ref = db.execute(
                    f"SELECT custo_unitario FROM composicoes WHERE codigo IN ({qs}) ORDER BY id_composicao DESC LIMIT 1",
                    variantes
                ).fetchone()
                if ref:
                    preco = ref['custo_unitario']
        preco = float(preco or 0)
        parcial = round(coef * preco, 4)
        db.execute(
            "UPDATE itens_composicao SET preco_unitario=?, custo_parcial=? WHERE id_item=?",
            [preco, parcial, it['id_item']]
        )
        total += parcial
    total = round(total, 4)
    db.execute("UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?", [total, id_composicao])
    return total

def _propagar_composicoes_auxiliares(db, parent_ids):
    ids = list(dict.fromkeys(parent_ids or []))
    if not ids:
        return {}
    custos = {}
    for _ in range(max(2, len(ids) + 1)):
        for cid in reversed(ids):
            custos[cid] = _recalcular_composicao_unitaria(db, cid)
    return custos

def _atualizar_orcamentos_por_composicoes(db, comp_ids):
    ids = list(dict.fromkeys(comp_ids or []))
    for cid in ids:
        comp = db.execute("SELECT descricao, custo_unitario FROM composicoes WHERE id_composicao=?", [cid]).fetchone()
        if comp:
            db.execute(
                "UPDATE orcamento_sintetico SET descricao=?, custo_unitario=? WHERE id_composicao=?",
                [comp['descricao'], comp['custo_unitario'], cid]
            )

# ── Grupos de composições ──────────────────────────────────────────────────────
@app.route('/api/composicoes/grupos', methods=['GET'])
def comp_grupos():
    fonte = request.args.get('fonte','')
    db = get_db()
    sql = """SELECT g.*, COUNT(c.id_composicao) AS qtd_composicoes
             FROM grupos_composicoes g
             LEFT JOIN composicoes c ON c.id_grupo_comp=g.id_grupo_comp
             WHERE 1=1"""
    params = []
    if fonte:
        sql += " AND g.fonte=?"; params.append(fonte)
    sql += " GROUP BY g.id_grupo_comp ORDER BY g.nome_grupo"
    rows = rows_to_list(db.execute(sql, params).fetchall())
    db.close(); return jsonify(rows)

# ── Listar composições ─────────────────────────────────────────────────────────
@app.route('/api/composicoes', methods=['GET'])
def comp_list():
    fonte   = request.args.get('fonte','')
    formato = request.args.get('formato','')
    grupo   = request.args.get('id_grupo_comp','')
    q       = request.args.get('q','')
    uf      = request.args.get('uf','')
    mes_ref = request.args.get('mes_ref','')
    regime  = request.args.get('regime','')
    limit   = int(request.args.get('limit', 100))
    offset  = int(request.args.get('offset', 0))

    sql = SEL_COMP + " WHERE 1=1"
    params = []
    if fonte:   sql += " AND c.fonte=?";            params.append(fonte)
    if formato: sql += " AND c.formato=?";           params.append(formato)
    if grupo:   sql += " AND c.id_grupo_comp=?";     params.append(grupo)
    if uf:      sql += " AND c.uf_referencia=?";     params.append(uf)
    if mes_ref: sql += " AND c.mes_referencia=?";    params.append(mes_ref)
    if regime == 'Desonerado':
        sql += " AND (LOWER(COALESCE(c.situacao_ref,'')) LIKE '%desonerado%' OR LOWER(COALESCE(c.situacao_ref,'')) LIKE '%com desoner%')"
    elif regime == 'Onerado':
        sql += """ AND (
            LOWER(COALESCE(c.situacao_ref,'')) = 'onerado'
            OR LOWER(COALESCE(c.situacao_ref,'')) LIKE '%sem desoner%'
            OR (LOWER(COALESCE(c.situacao_ref,'')) LIKE '%onerado%'
                AND LOWER(COALESCE(c.situacao_ref,'')) NOT LIKE '%desonerado%')
        )"""
    if q:       sql += " AND (c.descricao LIKE ? OR c.codigo LIKE ?)"; params += [f'%{q}%',f'%{q}%']
    sql += " ORDER BY c.fonte, c.codigo LIMIT ? OFFSET ?"
    params += [limit, offset]

    db = get_db()
    rows  = rows_to_list(db.execute(sql, params).fetchall())
    count_params = params[:-2]  # limit e offset são sempre os 2 últimos
    count_sql = ("SELECT COUNT(*) FROM composicoes c WHERE 1=1" +
                 (" AND c.fonte=?"                               if fonte   else "") +
                 (" AND c.formato=?"                             if formato else "") +
                 (" AND c.id_grupo_comp=?"                       if grupo   else "") +
                 (" AND c.uf_referencia=?"                       if uf      else "") +
                 (" AND c.mes_referencia=?"                      if mes_ref else "") +
                 (" AND (LOWER(COALESCE(c.situacao_ref,'')) LIKE '%desonerado%' OR LOWER(COALESCE(c.situacao_ref,'')) LIKE '%com desoner%')" if regime == 'Desonerado' else "") +
                 (""" AND (
                     LOWER(COALESCE(c.situacao_ref,'')) = 'onerado'
                     OR LOWER(COALESCE(c.situacao_ref,'')) LIKE '%sem desoner%'
                     OR (LOWER(COALESCE(c.situacao_ref,'')) LIKE '%onerado%'
                         AND LOWER(COALESCE(c.situacao_ref,'')) NOT LIKE '%desonerado%')
                 )""" if regime == 'Onerado' else "") +
                 (" AND (c.descricao LIKE ? OR c.codigo LIKE ?)" if q       else ""))
    total = db.execute(count_sql, count_params).fetchone()[0]
    db.close()
    return jsonify({'items': rows, 'total': total, 'limit': limit, 'offset': offset})

# ── Recalcular custos de composições SINAPI/SICRO em lote ────────────────────
@app.route('/api/composicoes/recalcular-custos', methods=['POST'])
def comp_recalcular_custos():
    """Recalcula custo_unitario das composições via itens_composicao × precos_insumos.
    Aceita parâmetros no body JSON para filtrar por UF, data-base, regime e modo."""
    d      = request.json or {}
    uf     = d.get('uf', '')           # '' = todas as UFs
    mes_ref= d.get('mes_ref', '')      # '' = todas as datas-base  (ex: '04/2026')
    regime = d.get('regime', 'ambos')  # 'desonerado' | 'nao_desonerado' | 'ambos'
    modo   = d.get('modo', 'sem_custo')# 'sem_custo' | 'todos'

    db = get_db()

    # Monta a expressão de preço conforme regime
    if regime == 'desonerado':
        preco_expr = "COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia)"
    elif regime == 'nao_desonerado':
        preco_expr = "COALESCE(p.preco_nao_desonerado, p.preco_desonerado, p.preco_referencia)"
    else:  # ambos — prioridade desonerado
        preco_expr = "COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia)"

    # Cláusulas extras de filtro na UPDATE
    extra_where = ""
    if uf:      extra_where += " AND uf_referencia = " + repr(uf)
    if mes_ref: extra_where += " AND mes_referencia = " + repr(mes_ref)
    if modo == 'sem_custo':
        extra_where += " AND (custo_unitario IS NULL OR custo_unitario = 0)"

    SQL_UPDATE = f"""
        UPDATE composicoes
        SET custo_unitario = (
            SELECT COALESCE(SUM(
                ic.coeficiente * COALESCE(
                    CASE WHEN ic.tipo_item = 'COMPOSICAO' THEN (
                        SELECT NULLIF(COALESCE(c2.custo_unitario, 0), 0)
                        FROM composicoes c2
                        WHERE c2.codigo  = ic.codigo_item
                           OR c2.codigo  = 'SINAPI.' || ic.codigo_item
                           OR c2.codigo  = 'SICRO.'  || ic.codigo_item
                           OR c2.codigo  = 'SEINFRA.' || ic.codigo_item
                           OR c2.codigo  = 'SUDECAP.' || ic.codigo_item
                           OR c2.codigo  = 'GOINFRA.' || ic.codigo_item
                           OR c2.codigo  = 'CDHU.' || ic.codigo_item
                        LIMIT 1
                    ) END,
                    (SELECT {preco_expr}
                     FROM precos_insumos p
                     JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
                     WHERE p.id_insumo = (
                         SELECT id_insumo FROM insumos
                         WHERE codigo_insumo = ic.codigo_item LIMIT 1
                     )
                     ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1),
                    ic.preco_unitario,
                    0
                )
            ), 0)
            FROM itens_composicao ic
            WHERE ic.id_composicao = composicoes.id_composicao
        )
        WHERE formato = 'UNITARIO'
          AND EXISTS (
              SELECT 1 FROM itens_composicao ic2
              WHERE ic2.id_composicao = composicoes.id_composicao
          )
          {extra_where}
    """

    total = 0
    passagens = 0
    for _ in range(4):
        cur = db.execute(SQL_UPDATE)
        n   = cur.rowcount
        db.commit()
        passagens += 1
        total += n
        if n == 0:
            break

    com_custo = db.execute(
        "SELECT COUNT(*) FROM composicoes WHERE custo_unitario > 0 AND formato='UNITARIO'"
    ).fetchone()[0]
    sem_custo = db.execute(
        "SELECT COUNT(*) FROM composicoes WHERE (custo_unitario IS NULL OR custo_unitario=0) AND formato='UNITARIO'"
    ).fetchone()[0]
    db.close()

    return jsonify({
        'atualizados': total,
        'passagens':   passagens,
        'com_custo':   com_custo,
        'sem_custo':   sem_custo,
        'mensagem':    (f'{total} composição(ões) recalculada(s) em {passagens} passagem(ns). '
                        f'{sem_custo} ainda sem custo (insumos sem preço cadastrado).'),
    })


# ── Excluir composições em lote ───────────────────────────────────────────────
@app.route('/api/composicoes/excluir-lote', methods=['POST'])
def comp_excluir_lote():
    """Exclui composições em lote com base nos filtros fornecidos.
    Se dry_run=True, retorna apenas a contagem sem excluir."""
    d       = request.json or {}
    fonte   = d.get('fonte', '')
    formato = d.get('formato', '')
    uf      = d.get('uf', '')
    mes_ref = d.get('mes_ref', '')
    grupo   = d.get('id_grupo_comp', '')
    dry_run = bool(d.get('dry_run', False))

    # Exige ao menos um filtro para evitar exclusão total acidental
    if not any([fonte, formato, uf, mes_ref, grupo]):
        return jsonify({'erro': 'Informe pelo menos um critério de seleção para excluir.'}), 400

    where = " WHERE 1=1"
    params = []
    if fonte:   where += " AND fonte=?";          params.append(fonte)
    if formato: where += " AND formato=?";         params.append(formato)
    if uf:      where += " AND uf_referencia=?";   params.append(uf)
    if mes_ref: where += " AND mes_referencia=?";  params.append(mes_ref)
    if grupo:   where += " AND id_grupo_comp=?";   params.append(int(grupo))

    db = get_db()
    ids = [
        row['id_composicao']
        for row in db.execute("SELECT id_composicao FROM composicoes" + where, params).fetchall()
    ]
    total = len(ids)

    if dry_run:
        db.close()
        return jsonify({'total': total})

    excluidos = 0
    for i in range(0, len(ids), 500):
        lote = ids[i:i + 500]
        placeholders = ','.join(['?'] * len(lote))
        db.execute(f"DELETE FROM composicoes_secao_itens WHERE id_composicao IN ({placeholders})", lote)
        db.execute(f"DELETE FROM composicoes_secoes WHERE id_composicao IN ({placeholders})", lote)
        db.execute(f"DELETE FROM itens_composicao WHERE id_composicao IN ({placeholders})", lote)
        cur = db.execute(f"DELETE FROM composicoes WHERE id_composicao IN ({placeholders})", lote)
        excluidos += cur.rowcount
    db.commit()
    db.close()
    return jsonify({'excluidos': excluidos, 'mensagem': f'{excluidos} composição(ões) excluída(s) com sucesso.'})

@app.route('/api/composicoes/<int:id>', methods=['GET'])
def comp_get(id):
    db = get_db()
    row = db.execute(SEL_COMP + " WHERE c.id_composicao=?", [id]).fetchone()
    if not row: db.close(); return jsonify({'erro': 'Composição não encontrada.'}), 404
    c = dict(row)

    # ── Itens (formato UNITÁRIO) com preços resolvidos dinamicamente ──────────
    # preco_unitario é buscado em cascata:
    #   1. Para tipo=COMPOSICAO → custo_unitario da composição referenciada
    #   2. Para tipos de insumo → último preço em precos_insumos (mais recente por data-base)
    #   3. Fallback → preco_unitario armazenado no próprio item (ex: composições do usuário)
    raw_itens = db.execute("""
        SELECT ic.*,
            COALESCE(
                CASE WHEN ic.tipo_item = 'COMPOSICAO' THEN (
                    SELECT custo_unitario FROM composicoes
                    WHERE codigo = ic.codigo_item
                       OR codigo = 'SINAPI.' || ic.codigo_item
                       OR codigo = 'SICRO.'  || ic.codigo_item
                       OR codigo = 'SEINFRA.' || ic.codigo_item
                       OR codigo = 'SUDECAP.' || ic.codigo_item
                       OR codigo = 'GOINFRA.' || ic.codigo_item
                       OR codigo = 'CDHU.' || ic.codigo_item
                    LIMIT 1
                ) END,
                (SELECT COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia)
                 FROM precos_insumos p
                 JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
                 WHERE p.id_insumo = (
                     SELECT id_insumo FROM insumos
                     WHERE codigo_insumo = ic.codigo_item LIMIT 1
                 )
                 ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1),
                ic.preco_unitario
            ) AS preco_resolvido
        FROM itens_composicao ic
        WHERE ic.id_composicao = ?
        ORDER BY ic.ordem
    """, [id]).fetchall()

    itens = []
    for r in raw_itens:
        item = dict(r)
        preco = item.pop('preco_resolvido') or 0
        item['preco_unitario'] = preco
        item['custo_parcial']  = round((item.get('coeficiente') or 0) * preco, 6)
        itens.append(item)
    c['itens'] = itens

    # Custo total calculado (soma dos parciais dos itens com preço)
    custo_calc = round(sum(it['custo_parcial'] for it in itens), 4)
    c['custo_calculado'] = custo_calc

    # Lazy-store: se custo_unitario não estava na tabela mas calculamos agora, gravamos
    # para que a listagem mostre o valor sem precisar recalcular sempre.
    if custo_calc > 0 and not (c.get('custo_unitario') or 0):
        try:
            db.execute(
                "UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?",
                [custo_calc, id]
            )
            db.commit()
            c['custo_unitario'] = custo_calc
        except Exception:
            pass  # não crítico — não bloqueia a resposta

    # ── Seções + itens (formato PRODUÇÃO HORÁRIA) ─────────────────────────────
    secoes = rows_to_list(db.execute(
        "SELECT * FROM composicoes_secoes WHERE id_composicao=? ORDER BY ordem", [id]).fetchall())
    for s in secoes:
        s['itens'] = rows_to_list(db.execute(
            "SELECT * FROM composicoes_secao_itens WHERE id_secao=? ORDER BY ordem",
            [s['id_secao']]).fetchall())
    c['secoes'] = secoes
    db.close(); return jsonify(c)

# ── CRUD Composições (usuário) ─────────────────────────────────────────────────
@app.route('/api/composicoes', methods=['POST'])
def comp_create():
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    db = get_db()
    cur = db.execute("""INSERT INTO composicoes
        (codigo,fonte,formato,descricao,unidade,id_grupo_comp,
         mes_referencia,uf_referencia,fic,producao_equipe,unidade_producao,
         situacao_ref,situacao,observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [d.get('codigo') or None, d.get('fonte','USUARIO'),
         d.get('formato','UNITARIO'), d['descricao'].strip(),
         d.get('unidade'), d.get('id_grupo_comp') or None,
         d.get('mes_referencia'), d.get('uf_referencia'),
         d.get('fic'), d.get('producao_equipe'), d.get('unidade_producao'),
         d.get('situacao_ref'), d.get('situacao','Ativo'), d.get('observacoes')])
    db.commit()
    row = db.execute(SEL_COMP + " WHERE c.id_composicao=?", [cur.lastrowid]).fetchone()
    db.close(); return jsonify(dict(row)), 201

@app.route('/api/composicoes/<int:id>', methods=['PUT'])
def comp_update(id):
    d = request.json or {}
    if not d.get('descricao','').strip():
        return jsonify({'erro': 'Descrição é obrigatória.'}), 400
    db = get_db()
    impacto = _impacto_composicao(db, id)
    if impacto and (impacto.get('composicoes_auxiliares') or impacto.get('orcamentos')):
        db.close()
        return jsonify({
            'erro': 'Composicao utilizada no sistema. Use a edicao com tratamento de impacto para preservar ou atualizar composicoes e orcamentos vinculados.'
        }), 409
    cur = db.execute("""UPDATE composicoes SET
        codigo=?,descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,
        uf_referencia=?,fic=?,producao_equipe=?,unidade_producao=?,
        situacao_ref=?,situacao=?,observacoes=?
        WHERE id_composicao=?""",
        [d.get('codigo'), d['descricao'].strip(), d.get('unidade'),
         d.get('id_grupo_comp') or None, d.get('mes_referencia'), d.get('uf_referencia'),
         d.get('fic'), d.get('producao_equipe'), d.get('unidade_producao'),
         d.get('situacao_ref'), d.get('situacao','Ativo'), d.get('observacoes'), id])
    db.commit()
    if cur.rowcount == 0: db.close(); return jsonify({'erro': 'Composição não encontrada.'}), 404
    row = db.execute(SEL_COMP + " WHERE c.id_composicao=?", [id]).fetchone()
    db.close(); return jsonify(dict(row))


# ── Verificar uso nos orçamentos sintéticos ───────────────────────────────────
@app.route('/api/composicoes/<int:id>/uso-orcamentos', methods=['GET'])
def comp_uso_orcamentos(id):
    db = get_db()
    impacto = _impacto_composicao(db, id)
    db.close()
    if not impacto:
        return jsonify([])
    return jsonify(impacto['orcamentos'])

@app.route('/api/composicoes/<int:id>/impacto', methods=['GET'])
def comp_impacto(id):
    db = get_db()
    impacto = _impacto_composicao(db, id)
    db.close()
    if not impacto:
        return jsonify({'erro': 'Composicao nao encontrada.'}), 404
    return jsonify(impacto)


@app.route('/api/composicoes/<int:id>/excluir-com-vinculo', methods=['POST'])
def comp_excluir_com_vinculo(id):
    d    = request.json or {}
    acao = d.get('acao', 'desvincular')
    db   = get_db()
    try:
        impacto = _impacto_composicao(db, id)
        if not impacto:
            return jsonify({'erro': 'Composicao nao encontrada.'}), 404
        comp = impacto['composicao']
        parent_ids = [c['id_composicao'] for c in impacto.get('composicoes_auxiliares', [])]
        if acao == 'remover':
            db.execute("DELETE FROM orcamento_sintetico WHERE id_composicao=?", [id])
            variantes = _comp_codigo_variantes(comp.get('codigo'))
            if parent_ids and variantes:
                qs_parent = ','.join(['?'] * len(parent_ids))
                qs_var = ','.join(['?'] * len(variantes))
                db.execute(f"""
                    DELETE FROM itens_composicao
                    WHERE id_composicao IN ({qs_parent})
                      AND UPPER(COALESCE(tipo_item,'')) = 'COMPOSICAO'
                      AND codigo_item IN ({qs_var})
                """, parent_ids + variantes)
                _propagar_composicoes_auxiliares(db, parent_ids)
                _atualizar_orcamentos_por_composicoes(db, parent_ids)
        else:
            db.execute("UPDATE orcamento_sintetico SET id_composicao=NULL WHERE id_composicao=?", [id])
        db.execute("DELETE FROM composicoes_secao_itens WHERE id_composicao=?", [id])
        db.execute("DELETE FROM composicoes_secoes WHERE id_composicao=?", [id])
        db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id])
        cur = db.execute("DELETE FROM composicoes WHERE id_composicao=?", [id])
        db.commit()
        if cur.rowcount == 0:
            return jsonify({'erro': 'Composicao nao encontrada.'}), 404
        return jsonify({'mensagem': 'Composicao excluida com sucesso.'})
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-500:]}), 500
    finally:
        db.close()


@app.route('/api/composicoes/<int:id>/editar-com-vinculo', methods=['POST'])
def comp_editar_com_vinculo(id):
    d           = request.json or {}
    dados       = d.get('dados', {})
    itens_novos = d.get('itens', [])
    acao_orc    = d.get('acao_orcamentos', 'manter')

    if not dados.get('descricao','').strip():
        return jsonify({'erro': 'Descricao e obrigatoria.'}), 400

    db = get_db()
    try:
        comp_orig = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id]).fetchone()
        if not comp_orig:
            return jsonify({'erro': 'Composicao nao encontrada.'}), 404
        comp_orig  = dict(comp_orig)
        fonte_orig = comp_orig.get('fonte','USUARIO')
        impacto    = _impacto_composicao(db, id) or {}
        parent_ids = [c['id_composicao'] for c in impacto.get('composicoes_auxiliares', [])]
        tem_impacto = bool(parent_ids or impacto.get('orcamentos'))
        fontes_referenciais = ('SINAPI', 'SICRO', 'SEINFRA', 'SUDECAP', 'GOINFRA', 'CDHU')
        criar_nova = fonte_orig in fontes_referenciais or (acao_orc == 'manter' and tem_impacto)
        cod_novo   = None

        if criar_nova:
            cod_base = (dados.get('codigo') or comp_orig.get('codigo',''))
            for prefixo in ('SINAPI.', 'SICRO.', 'SEINFRA.', 'SUDECAP.', 'GOINFRA.', 'CDHU.', 'USUARIO.'):
                cod_base = cod_base.replace(prefixo, '')
            cod_novo = 'USUARIO.' + cod_base
            sufixo = 2
            while db.execute("SELECT 1 FROM composicoes WHERE codigo=?", [cod_novo]).fetchone():
                cod_novo = 'USUARIO.' + cod_base + '-' + str(sufixo); sufixo += 1

            cur = db.execute(
                "INSERT INTO composicoes (codigo,fonte,formato,descricao,unidade,id_grupo_comp,"
                "mes_referencia,uf_referencia,fic,producao_equipe,unidade_producao,"
                "situacao_ref,situacao,observacoes,custo_unitario) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                [cod_novo, 'USUARIO',
                 dados.get('formato', comp_orig['formato']),
                 dados['descricao'].strip(),
                 dados.get('unidade', comp_orig.get('unidade')),
                 dados.get('id_grupo_comp') or None,
                 dados.get('mes_referencia', comp_orig.get('mes_referencia')),
                 dados.get('uf_referencia') or None,
                 dados.get('fic'), dados.get('producao_equipe'), dados.get('unidade_producao'),
                 dados.get('situacao_ref'), 'Ativo', dados.get('observacoes'), 0])
            db.commit()
            id_resultado = cur.lastrowid
        else:
            db.execute(
                "UPDATE composicoes SET codigo=?,descricao=?,unidade=?,id_grupo_comp=?,"
                "mes_referencia=?,uf_referencia=?,fic=?,producao_equipe=?,unidade_producao=?,"
                "situacao_ref=?,situacao=?,observacoes=? WHERE id_composicao=?",
                [dados.get('codigo'), dados['descricao'].strip(),
                 dados.get('unidade'), dados.get('id_grupo_comp') or None,
                 dados.get('mes_referencia'), dados.get('uf_referencia') or None,
                 dados.get('fic'), dados.get('producao_equipe'), dados.get('unidade_producao'),
                 dados.get('situacao_ref'), 'Ativo', dados.get('observacoes'), id])
            db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id])
            db.commit()
            id_resultado = id

        for ordem, it in enumerate(itens_novos):
            db.execute(
                "INSERT INTO itens_composicao (id_composicao,tipo_item,codigo_item,descricao,"
                "unidade,coeficiente,preco_unitario,custo_parcial,situacao_item,ordem) VALUES (?,?,?,?,?,?,?,?,?,?)",
                [id_resultado, it.get('tipo_item','INSUMO'), it.get('codigo_item'),
                 it.get('descricao',''), it.get('unidade'),
                 float(it.get('coeficiente') or 0),
                 it.get('preco_unitario'), it.get('custo_parcial'),
                 it.get('situacao_item'), ordem])
        db.commit()

        # Se veio de SICRO, também salvar em composicoes_secoes/itens (mantém estrutura original)
        if criar_nova and itens_novos and itens_novos[0].get('_secao'):
            from itertools import groupby
            SEC_NOMES = {'A':'Equipamentos','B':'Mão de Obra','C':'Material',
                         'D':'Atividades Auxiliares','E':'Tempo Fixo','F':'Momento de Transporte'}
            # Agrupar por seção preservando ordem
            secao_grupos = {}
            for it in itens_novos:
                s = it.get('_secao','A')
                if s not in secao_grupos: secao_grupos[s] = []
                secao_grupos[s].append(it)
            for ordem_sec, (letra, its) in enumerate(secao_grupos.items()):
                id_sec = db.execute(
                    "INSERT INTO composicoes_secoes (id_composicao,letra_secao,nome_secao,ordem) VALUES (?,?,?,?)",
                    [id_resultado, letra, SEC_NOMES.get(letra, letra), ordem_sec]).lastrowid
                db.commit()
                for ordem_it, it in enumerate(its):
                    db.execute(
                        "INSERT INTO composicoes_secao_itens "
                        "(id_composicao,id_secao,letra_secao,codigo_item,descricao,quantidade,unidade,"
                        "util_operativa,util_improdutiva,custo_hp,custo_hi,preco_unitario,custo_total,ordem) "
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [id_resultado, id_sec, letra,
                         it.get('codigo_item'), it.get('descricao'),
                         float(it.get('coeficiente') or 0), it.get('unidade'),
                         it.get('util_operativa'), it.get('util_improdutiva'),
                         it.get('preco_unitario') if letra=='A' else None,
                         it.get('custo_hi'), it.get('preco_unitario'),
                         it.get('custo_parcial'), ordem_it])
                db.commit()

        custo = db.execute(
            "SELECT COALESCE(SUM(coeficiente * COALESCE(preco_unitario,0)),0) FROM itens_composicao WHERE id_composicao=?",
            [id_resultado]).fetchone()[0]
        db.execute("UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?", [round(custo,4), id_resultado])
        db.commit()

        if acao_orc == 'atualizar':
            db.execute(
                "UPDATE orcamento_sintetico SET id_composicao=?,descricao=?,custo_unitario=? WHERE id_composicao=?",
                [id_resultado, dados['descricao'].strip(), round(custo,4), id])
        if acao_orc in ('atualizar', 'alterar_composicoes'):
            if parent_ids and id_resultado != id:
                comp_nova = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id_resultado]).fetchone()
                variantes = _comp_codigo_variantes(comp_orig.get('codigo'))
                if comp_nova and variantes:
                    comp_nova = dict(comp_nova)
                    qs_parent = ','.join(['?'] * len(parent_ids))
                    qs_var = ','.join(['?'] * len(variantes))
                    db.execute(f"""
                        UPDATE itens_composicao
                        SET codigo_item=?, descricao=?, unidade=?, preco_unitario=?,
                            custo_parcial=ROUND(COALESCE(coeficiente,0) * ?, 4)
                        WHERE id_composicao IN ({qs_parent})
                          AND UPPER(COALESCE(tipo_item,'')) = 'COMPOSICAO'
                          AND codigo_item IN ({qs_var})
                    """, [comp_nova.get('codigo'), comp_nova.get('descricao'), comp_nova.get('unidade'),
                          round(custo,4), round(custo,4)] + parent_ids + variantes)
            if parent_ids:
                _propagar_composicoes_auxiliares(db, parent_ids)
                if acao_orc == 'atualizar':
                    _atualizar_orcamentos_por_composicoes(db, parent_ids)
            db.commit()

        row = db.execute(SEL_COMP + " WHERE c.id_composicao=?", [id_resultado]).fetchone()
        return jsonify({
            'composicao':  dict(row),
            'id_resultado': id_resultado,
            'criou_nova':   criar_nova,
            'cod_novo':     cod_novo,
            'mensagem': ('Nova composicao USUARIO criada (codigo: ' + (cod_novo or '') + ').' if criar_nova else 'Composicao atualizada.'),
        })
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-800:]}), 500
    finally:
        db.close()



@app.route('/api/composicoes/<int:id>', methods=['DELETE'])
def comp_delete(id):
    db = get_db()
    impacto = _impacto_composicao(db, id)
    if impacto and (impacto.get('composicoes_auxiliares') or impacto.get('orcamentos')):
        db.close()
        return jsonify({
            'erro': 'Composicao utilizada no sistema. Use a exclusao com tratamento de impacto para preservar historico ou recalcular os vinculos.'
        }), 409
    db.execute("DELETE FROM composicoes_secao_itens WHERE id_composicao=?", [id])
    db.execute("DELETE FROM composicoes_secoes WHERE id_composicao=?", [id])
    db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id])
    cur = db.execute("DELETE FROM composicoes WHERE id_composicao=?", [id])
    db.commit(); db.close()
    if cur.rowcount == 0: return jsonify({'erro': 'Composição não encontrada.'}), 404
    return jsonify({'mensagem': 'Composição excluída.'})

# ── Itens (unitário) ────────────────────────────────────────────────────────────
@app.route('/api/composicoes/<int:cid>/itens', methods=['POST'])
def comp_item_create(cid):
    d = request.json or {}
    db = get_db()
    max_ord = db.execute("SELECT COALESCE(MAX(ordem),0) FROM itens_composicao WHERE id_composicao=?", [cid]).fetchone()[0]
    cur = db.execute("""INSERT INTO itens_composicao
        (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,ordem)
        VALUES (?,?,?,?,?,?,?,?)""",
        [cid, d.get('tipo_item','INSUMO'), d.get('codigo_item'),
         d.get('descricao',''), d.get('unidade'), float(d.get('coeficiente') or 0),
         d.get('situacao_item'), max_ord+1])
    db.commit()
    row = dict(db.execute("SELECT * FROM itens_composicao WHERE id_item=?", [cur.lastrowid]).fetchone())
    db.close(); return jsonify(row), 201

@app.route('/api/composicoes/itens/<int:id>', methods=['PUT'])
def comp_item_update(id):
    d = request.json or {}
    db = get_db()
    db.execute("""UPDATE itens_composicao SET
        tipo_item=?,codigo_item=?,descricao=?,unidade=?,coeficiente=?,situacao_item=?
        WHERE id_item=?""",
        [d.get('tipo_item','INSUMO'), d.get('codigo_item'), d.get('descricao',''),
         d.get('unidade'), float(d.get('coeficiente') or 0), d.get('situacao_item'), id])
    db.commit()
    row = dict(db.execute("SELECT * FROM itens_composicao WHERE id_item=?", [id]).fetchone())
    db.close(); return jsonify(row)

@app.route('/api/composicoes/itens/<int:id>', methods=['DELETE'])
def comp_item_delete(id):
    db = get_db()
    db.execute("DELETE FROM itens_composicao WHERE id_item=?", [id])
    db.commit(); db.close()
    return jsonify({'mensagem': 'Item excluído.'})

# ── Stats para dashboard ────────────────────────────────────────────────────────
@app.route('/api/composicoes/stats', methods=['GET'])
def comp_stats():
    db = get_db()
    por_fonte = rows_to_list(db.execute("""
        SELECT fonte, COUNT(*) AS total FROM composicoes GROUP BY fonte""").fetchall())
    por_formato = rows_to_list(db.execute("""
        SELECT formato, COUNT(*) AS total FROM composicoes GROUP BY formato""").fetchall())
    db.close()
    return jsonify({'por_fonte': por_fonte, 'por_formato': por_formato})

# ─── START ────────────────────────────────────────────────────────────────────
# ─── ORÇAMENTO SINTÉTICO (Módulo 6) ──────────────────────────────────────────

SEL_ORC_FULL = """
    SELECT o.*, ob.nome_obra, ob.uf AS obra_uf,
           db.mes AS data_base_mes, db.ano AS data_base_ano,
           b.bdi_percentual AS bdi_perf_percentual, b.nome_perfil AS bdi_nome_perfil
    FROM orcamentos o
    LEFT JOIN obras ob       ON o.id_obra       = ob.id_obra
    LEFT JOIN datas_base db  ON o.id_data_base  = db.id_data_base
    LEFT JOIN perfis_bdi b   ON o.id_bdi_perfil = b.id_perfil_bdi
"""

@app.route('/api/orcamentos/<int:id>/sintetico', methods=['GET'])
def os_list(id):
    db = get_db()
    ensure_orcamento_sintetico_bdi_linha(db)
    rows = rows_to_list(db.execute("""
        SELECT s.*
        FROM orcamento_sintetico s
        WHERE s.id_orcamento = ?
        ORDER BY s.ordem, s.id_item
    """, [id]).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/orcamentos/<int:id>/sintetico', methods=['POST'])
def os_create(id):
    d = request.json or {}
    if not d.get('descricao', '').strip() and d.get('tipo_linha') == 'item':
        d['descricao'] = 'Novo item'
    db = get_db()
    ensure_orcamento_sintetico_bdi_linha(db)
    max_ord = db.execute(
        "SELECT COALESCE(MAX(ordem),0) FROM orcamento_sintetico WHERE id_orcamento=?", [id]
    ).fetchone()[0]
    cur = db.execute("""
        INSERT INTO orcamento_sintetico
          (id_orcamento, item_num, tipo_linha, profundidade, ordem, tipo_item,
           id_composicao, id_insumo, codigo, fonte, descricao, unidade, quantidade, custo_unitario,
           bdi_percentual_linha)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, [
        id,
        d.get('item_num', ''),
        d.get('tipo_linha', 'item'),
        d.get('profundidade', 1),
        d.get('ordem', max_ord + 1),
        d.get('tipo_item'),
        d.get('id_composicao'),
        d.get('id_insumo'),
        d.get('codigo', ''),
        d.get('fonte', ''),
        d.get('descricao', ''),
        d.get('unidade', ''),
        d.get('quantidade', 0),
        d.get('custo_unitario', 0),
        d.get('bdi_percentual_linha'),
    ])
    db.commit()
    row = dict(db.execute("SELECT * FROM orcamento_sintetico WHERE id_item=?", [cur.lastrowid]).fetchone())
    db.close()
    return jsonify(row), 201

@app.route('/api/orcamentos/sintetico/<int:id>', methods=['PUT'])
def os_update(id):
    d = request.json or {}
    CAMPOS = ['item_num','tipo_linha','profundidade','ordem','tipo_item',
              'id_composicao','id_insumo','codigo','fonte','descricao',
              'unidade','quantidade','custo_unitario','bdi_percentual_linha']
    sets, vals = [], []
    for c in CAMPOS:
        if c in d:
            sets.append(f"{c}=?"); vals.append(d[c])
    if not sets:
        return jsonify({'erro': 'Nenhum campo para atualizar.'}), 400
    vals.append(id)
    db = get_db()
    ensure_orcamento_sintetico_bdi_linha(db)
    db.execute(f"UPDATE orcamento_sintetico SET {','.join(sets)} WHERE id_item=?", vals)
    db.commit()
    row = db.execute("SELECT * FROM orcamento_sintetico WHERE id_item=?", [id]).fetchone()
    db.close()
    if not row:
        return jsonify({'erro': 'Item não encontrado.'}), 404
    return jsonify(dict(row))

@app.route('/api/orcamentos/sintetico/<int:id>', methods=['DELETE'])
def os_delete(id):
    db = get_db()
    row = db.execute("SELECT * FROM orcamento_sintetico WHERE id_item=?", [id]).fetchone()
    if not row:
        db.close(); return jsonify({'erro': 'Item não encontrado.'}), 404
    r = dict(row)
    if r['tipo_linha'] == 'section' and r.get('item_num'):
        prefix = r['item_num'] + '.'
        db.execute("""DELETE FROM orcamento_sintetico
                      WHERE id_orcamento=? AND (id_item=? OR item_num LIKE ?)""",
                   [r['id_orcamento'], id, prefix + '%'])
    else:
        db.execute("DELETE FROM orcamento_sintetico WHERE id_item=?", [id])
    db.commit()
    db.close()
    return jsonify({'mensagem': 'Item excluído.'})

@app.route('/api/orcamentos/<int:id>/sintetico/reordenar', methods=['POST'])
def os_reorder(id):
    items = request.json or []
    db = get_db()
    for it in items:
        db.execute("""UPDATE orcamento_sintetico
                      SET ordem=?, item_num=?, profundidade=?
                      WHERE id_item=? AND id_orcamento=?""",
                   [it.get('ordem'), it.get('item_num'), it.get('profundidade'),
                    it.get('id_item'), id])
    db.commit()
    db.close()
    return jsonify({'mensagem': 'Reordenado.'})

@app.route('/api/orcamentos/<int:id>/sintetico/restaurar', methods=['PUT'])
def os_restore(id):
    d = request.json or {}
    items = d.get('itens', [])
    if isinstance(items, dict) and isinstance(items.get('value'), list):
        items = items['value']
    if (isinstance(items, list) and len(items) == 1 and
            isinstance(items[0], dict) and isinstance(items[0].get('value'), list)):
        items = items[0]['value']
    if not isinstance(items, list):
        return jsonify({'erro': 'Lista de itens invalida.'}), 400
    if any(not isinstance(it, dict) or 'tipo_linha' not in it for it in items):
        return jsonify({'erro': 'Itens do snapshot invalidos.'}), 400

    campos = [
        'id_item', 'id_orcamento', 'item_num', 'tipo_linha', 'profundidade',
        'ordem', 'tipo_item', 'id_composicao', 'id_insumo', 'codigo', 'fonte',
        'descricao', 'unidade', 'quantidade', 'custo_unitario', 'bdi_percentual_linha'
    ]
    db = get_db()
    ensure_orcamento_sintetico_bdi_linha(db)
    try:
        db.execute("DELETE FROM orcamento_sintetico WHERE id_orcamento=?", [id])
        for ordem, it in enumerate(items, start=1):
            vals = {
                'id_item': it.get('id_item'),
                'id_orcamento': id,
                'item_num': it.get('item_num', ''),
                'tipo_linha': it.get('tipo_linha', 'item'),
                'profundidade': it.get('profundidade', 1),
                'ordem': it.get('ordem') or ordem,
                'tipo_item': it.get('tipo_item'),
                'id_composicao': it.get('id_composicao'),
                'id_insumo': it.get('id_insumo'),
                'codigo': it.get('codigo', ''),
                'fonte': it.get('fonte', ''),
                'descricao': it.get('descricao', ''),
                'unidade': it.get('unidade', ''),
                'quantidade': it.get('quantidade', 0),
                'custo_unitario': it.get('custo_unitario', 0),
                'bdi_percentual_linha': it.get('bdi_percentual_linha'),
            }
            if vals['id_item']:
                db.execute(f"""
                    INSERT INTO orcamento_sintetico ({','.join(campos)})
                    VALUES ({','.join(['?'] * len(campos))})
                """, [vals[c] for c in campos])
            else:
                sem_id = campos[1:]
                db.execute(f"""
                    INSERT INTO orcamento_sintetico ({','.join(sem_id)})
                    VALUES ({','.join(['?'] * len(sem_id))})
                """, [vals[c] for c in sem_id])

        if 'bdi_percentual' in d or 'id_bdi_perfil' in d:
            db.execute("""UPDATE orcamentos
                          SET bdi_percentual=?, id_bdi_perfil=?
                          WHERE id_orcamento=?""",
                       [d.get('bdi_percentual', 0), d.get('id_bdi_perfil'), id])

        db.commit()
        rows = rows_to_list(db.execute("""
            SELECT * FROM orcamento_sintetico
            WHERE id_orcamento=?
            ORDER BY ordem, id_item
        """, [id]).fetchall())
        db.close()
        return jsonify({'mensagem': 'Orcamento restaurado.', 'itens': rows})
    except Exception as e:
        db.rollback()
        db.close()
        return jsonify({'erro': str(e)}), 500

@app.route('/api/orcamentos/<int:id>/bdi', methods=['PUT'])
def os_update_bdi(id):
    d = request.json or {}
    db = get_db()
    try:
        db.execute("""UPDATE orcamentos
                      SET bdi_percentual=?, id_bdi_perfil=?
                      WHERE id_orcamento=?""",
                   [d.get('bdi_percentual', 0), d.get('id_bdi_perfil'), id])
        db.commit()
    except Exception as e:
        db.close(); return jsonify({'erro': str(e)}), 500
    db.close()
    return jsonify({'mensagem': 'BDI atualizado.'})

@app.route('/api/orcamentos/<int:id>/sintetico/totais', methods=['PUT'])
def os_update_totais(id):
    d = request.json or {}
    db = get_db()
    try:
        db.execute("""UPDATE orcamentos
                      SET valor_custo_direto=?, valor_bdi=?, valor_total=?
                      WHERE id_orcamento=?""",
                   [d.get('custo_direto', 0), d.get('valor_bdi', 0), d.get('total', 0), id])
        db.commit()
    except Exception:
        pass  # Colunas podem não existir em bancos antigos — não é crítico
    db.close()
    return jsonify({'mensagem': 'Totais atualizados.'})

def _pav_norm(txt):
    txt = unicodedata.normalize('NFD', str(txt or '').lower())
    txt = ''.join(ch for ch in txt if unicodedata.category(ch) != 'Mn')
    return re.sub(r'\s+', ' ', txt).strip()

def _pav_comp_custo(db, id_composicao):
    row = db.execute("SELECT custo_unitario FROM composicoes WHERE id_composicao=?", [id_composicao]).fetchone()
    custo = float(row['custo_unitario'] or 0) if row else 0
    if custo > 0:
        return custo
    itens = db.execute("""
        SELECT ic.coeficiente,
            COALESCE(
                CASE WHEN ic.tipo_item = 'COMPOSICAO' THEN (
                    SELECT custo_unitario FROM composicoes
                    WHERE codigo = ic.codigo_item
                       OR codigo = 'SINAPI.' || ic.codigo_item
                       OR codigo = 'SICRO.' || ic.codigo_item
                    LIMIT 1
                ) END,
                (SELECT COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia)
                 FROM precos_insumos p
                 JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
                 WHERE p.id_insumo = (
                     SELECT id_insumo FROM insumos WHERE codigo_insumo = ic.codigo_item LIMIT 1
                 )
                 ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1),
                ic.preco_unitario
            ) AS preco_resolvido
        FROM itens_composicao ic
        WHERE ic.id_composicao=?
    """, [id_composicao]).fetchall()
    custo = round(sum(float(r['coeficiente'] or 0) * float(r['preco_resolvido'] or 0) for r in itens), 4)
    if custo > 0:
        try:
            db.execute("UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?", [custo, id_composicao])
        except Exception:
            pass
    return custo

def _pav_servicos_from_camadas(ctx, camadas):
    area = float(ctx.get('area') or 0)
    servicos = []
    for camada in camadas:
        nome = camada.get('nome', '')
        material = camada.get('material', '')
        esp = float(camada.get('esp') or 0)
        if esp <= 0:
            continue
        vol = area * esp / 100.0
        nmat = _pav_norm(nome + ' ' + material)
        if 'subleito' in nmat:
            continue
        if 'placa de concreto' in nmat:
            tipo = 'placa_concreto'
        elif 'revestimento' in nmat and 'intertravado' in nmat:
            tipo = 'intertravado'
        elif 'revestimento' in nmat:
            tipo = 'revestimento_asfaltico'
        elif 'base' == _pav_norm(nome):
            tipo = 'base'
        elif 'sub-base' in nmat or 'sub base' in nmat:
            tipo = 'subbase'
        elif 'reforco' in nmat:
            tipo = 'reforco'
        else:
            tipo = 'camada_granular'
        servicos.append({
            'tipo': tipo,
            'camada': nome,
            'material': material,
            'esp_cm': esp,
            'area_m2': area,
            'volume_m3': vol,
            'descricao': f"{nome} - {material} ({esp:.1f} cm)",
        })
    if any(s['tipo'] in ('base', 'subbase', 'reforco') for s in servicos) and any(s['tipo'] == 'revestimento_asfaltico' for s in servicos):
        servicos.insert(0, {
            'tipo': 'imprimacao',
            'camada': 'Imprimacao',
            'material': 'Imprimacao betuminosa sobre base granular',
            'esp_cm': 0,
            'area_m2': area,
            'volume_m3': 0,
            'descricao': 'Imprimacao betuminosa ligante sobre base',
        })
    return servicos

def _pav_perfil_servico(tipo, material=''):
    material = _pav_norm(material)
    perfis = {
        'imprimacao': {
            'termos': ['imprimacao', 'imprimante', 'ligante', 'betuminos'],
            'bonus': ['execucao', 'aplicacao', 'pavimentacao'],
            'penaliza': ['transporte', 'drenagem', 'sinalizacao', 'meio fio', 'sarjeta'],
            'unidades': ['m2'],
        },
        'revestimento_asfaltico': {
            'termos': ['cbuq', 'concreto asfaltico', 'concreto betuminoso', 'camada de rolamento', 'massa asfaltica', 'asf'],
            'bonus': ['usinad', 'quente', 'aplicacao', 'execucao', 'pavimentacao'],
            'penaliza': ['transporte', 'carga', 'descarga', 'fresagem', 'recapeamento', 'tapa buraco', 'sinalizacao', 'canal', 'drenagem', 'usina de asfalto', 'instalacao', 'montagem'],
            'unidades': ['t', 'ton', 'm3'],
        },
        'intertravado': {
            'termos': ['intertravado', 'bloco de concreto', 'piso intertravado', 'paver'],
            'bonus': ['assentamento', 'execucao', 'pavimento'],
            'penaliza': ['meio fio', 'guia', 'drenagem'],
            'unidades': ['m2'],
        },
        'placa_concreto': {
            'termos': ['pavimento de concreto', 'placa de concreto', 'concreto de cimento portland', 'concreto simples'],
            'bonus': ['execucao', 'junta', 'pavimento rigido'],
            'penaliza': ['estrutura', 'forma', 'edificacao', 'drenagem'],
            'unidades': ['m3', 'm2'],
        },
        'base': {
            'termos': ['base', 'brita graduada', 'bgs', 'brita graduada simples', 'solo cimento', 'bgtc'],
            'bonus': ['execucao', 'compactacao', 'pavimentacao', 'estabilizada' if ('cimento' in material or 'bgtc' in material) else 'granular'],
            'penaliza': ['transporte', 'drenagem', 'sinalizacao', 'subleito', 'regularizacao', 'gesso', 'argamassa', 'reboco', 'emboço', 'emoco', 'alvenaria', 'parede', 'revestimento ceramico'],
            'unidades': ['m3', 'm2'],
        },
        'subbase': {
            'termos': ['sub-base', 'sub base', 'solo brita', 'brita graduada', 'material granular'],
            'bonus': ['execucao', 'compactacao', 'pavimentacao'],
            'penaliza': ['transporte', 'drenagem', 'sinalizacao', 'gesso', 'argamassa', 'reboco', 'emboço', 'emoco', 'alvenaria', 'parede', 'revestimento ceramico'],
            'unidades': ['m3', 'm2'],
        },
        'reforco': {
            'termos': ['reforco do subleito', 'regularizacao do subleito', 'solo selecionado', 'estabilizacao'],
            'bonus': ['compactacao', 'execucao', 'pavimentacao'],
            'penaliza': ['transporte', 'drenagem', 'sinalizacao'],
            'unidades': ['m3', 'm2'],
        },
    }
    return perfis.get(tipo, perfis['base'])

def _pav_sicro_relevante(tipo, codigo, desc):
    cod = str(codigo or '').upper().replace('SICRO.', '')
    d = _pav_norm(desc)
    bloqueios = [
        'chapisco', 'argamassa', 'central de britagem', 'central de concreto',
        'montagem e desmontagem', 'instalacao da usina', 'rampa para acesso',
        'canal ', 'drenagem', 'grelha', 'meio fio', 'sinalizacao',
        'transporte', 'carga, manobra', 'demolicao', 'remendo profundo'
    ]
    if any(b in d for b in bloqueios):
        return False
    if tipo == 'imprimacao':
        return cod in ('4011351', '4011352') or ('imprima' in d and 'remendo' not in d)
    if tipo == 'revestimento_asfaltico':
        faixa_401 = cod.startswith('40114') and ('concreto asf' in d or 'pre-misturado a quente' in d or 'macadame betuminoso' in d)
        faixa_641 = (cod.startswith('641607') or cod.startswith('641608')) and 'concreto asf' in d
        return faixa_401 or faixa_641
    if tipo == 'base':
        familia = cod.startswith(('401121', '401122', '401127', '401128', '401129', '401130', '401134', '401154', '401156'))
        texto = ('base' in d or 'brita graduada' in d or 'macadame' in d or 'solo melhorado' in d or 'solo cimento' in d)
        return familia and texto and 'subleito' not in d and 'concreto para sub-base' not in d
    if tipo == 'subbase':
        familia = cod.startswith(('401121', '401122', '401127', '401128', '401130', '401154', '401156'))
        texto = ('sub-base' in d or 'sub base' in d or 'base ou sub-base' in d or 'brita graduada' in d or 'solo estabilizado' in d or 'solo melhorado' in d)
        return familia and texto and 'concreto para sub-base' not in d
    if tipo == 'reforco':
        return cod.startswith(('40112', '40113')) and ('subleito' in d or 'solo' in d)
    return True

def _pav_buscar_composicao(db, fonte, uf, mes_ref, servico):
    perfil = _pav_perfil_servico(servico['tipo'], servico.get('material', ''))
    fonte = (fonte or '').upper()
    params = [fonte]
    sql = """
        SELECT id_composicao, codigo, descricao, unidade, fonte, uf_referencia,
               mes_referencia, custo_unitario
        FROM composicoes
        WHERE fonte=?
    """
    if uf:
        sql += " AND (uf_referencia=? OR COALESCE(uf_referencia,'')='')"
        params.append(uf)
    rows = db.execute(sql + " ORDER BY codigo LIMIT 12000", params).fetchall()
    if not rows and uf:
        rows = db.execute(sql.replace(" AND (uf_referencia=? OR COALESCE(uf_referencia,'')='')", "") + " ORDER BY codigo LIMIT 12000", [fonte]).fetchall()

    best = None
    for r in rows:
        desc = _pav_norm((r['codigo'] or '') + ' ' + (r['descricao'] or '') + ' ' + (r['unidade'] or ''))
        if fonte == 'SICRO' and not _pav_sicro_relevante(servico['tipo'], r['codigo'], r['descricao']):
            continue
        score = 0
        for termo in perfil['termos']:
            if _pav_norm(termo) in desc:
                score += 24
        for termo in perfil['bonus']:
            if _pav_norm(termo) in desc:
                score += 7
        for termo in perfil['penaliza']:
            if _pav_norm(termo) in desc:
                score -= 30
        if servico['tipo'] in ('base', 'subbase'):
            contexto_granular = any(t in desc for t in [
                'pavimentacao', 'brita', 'solo', 'granular', 'bgs', 'bgtc',
                'sub-base', 'sub base', 'base estabilizada', 'base para pavimento'
            ])
            if not contexto_granular:
                score -= 45
            if servico['tipo'] == 'base' and ('base e sub-base' in desc or 'base e sub base' in desc):
                score += 8
        un = _pav_norm(r['unidade'])
        if un in perfil['unidades']:
            score += 12
        if uf and (r['uf_referencia'] or '').upper() == uf:
            score += 10
        if mes_ref and (r['mes_referencia'] or '') == mes_ref:
            score += 10
        if r['custo_unitario']:
            score += 4
        if fonte == 'SICRO':
            cod = str(r['codigo'] or '').upper().replace('SICRO.', '')
            if servico['tipo'] == 'imprimacao' and cod in ('4011351', '4011352'):
                score += 50
            elif servico['tipo'] == 'revestimento_asfaltico' and cod.startswith('40114'):
                score += 45
            elif servico['tipo'] in ('base', 'subbase') and cod.startswith(('401127', '401154')):
                score += 45
            elif servico['tipo'] in ('base', 'subbase') and cod.startswith(('401121', '401122', '401128', '401129', '401130')):
                score += 30
        if best is None or score > best['score']:
            best = {'row': dict(r), 'score': score}
    min_score = 45 if fonte == 'SICRO' else 18
    if not best or best['score'] < min_score:
        return None
    comp = best['row']
    comp['custo_unitario'] = _pav_comp_custo(db, comp['id_composicao']) or float(comp.get('custo_unitario') or 0)
    comp['_score_ia'] = best['score']
    return comp

def _pav_quantidade_por_unidade(servico, unidade):
    un = _pav_norm(unidade)
    area = float(servico.get('area_m2') or 0)
    vol = float(servico.get('volume_m3') or 0)
    if un in ('m2', 'm²'):
        return area
    if un in ('m3', 'm³'):
        return vol
    if un in ('t', 'ton', 'tonelada', 'toneladas'):
        dens = 2.4 if servico['tipo'] == 'revestimento_asfaltico' else 1.9
        return vol * dens
    return vol or area

@app.route('/api/pavimentos/gerar-orcamento', methods=['POST'])
def pav_gerar_orcamento():
    d = request.json or {}
    id_obra = d.get('id_obra')
    fonte = (d.get('fonte') or 'SICRO').upper()
    uf = (d.get('uf_referencia') or '').upper()
    id_data_base = d.get('id_data_base')
    ctx = d.get('ctx') or {}
    camadas = d.get('camadas') or []
    if fonte not in ('SINAPI', 'SICRO'):
        return jsonify({'erro': 'Fonte deve ser SINAPI ou SICRO.'}), 400
    if not id_obra:
        return jsonify({'erro': 'Selecione uma obra de destino.'}), 400
    if not camadas:
        return jsonify({'erro': 'Dimensione o pavimento antes de gerar o orçamento.'}), 400

    db = get_db()
    try:
        obra = db.execute('SELECT * FROM obras WHERE id_obra=?', [id_obra]).fetchone()
        if not obra:
            return jsonify({'erro': 'Obra não encontrada.'}), 404
        data_base = db.execute('SELECT * FROM datas_base WHERE id_data_base=?', [id_data_base]).fetchone() if id_data_base else None
        mes_ref = f"{int(data_base['mes']):02d}/{int(data_base['ano'])}" if data_base else ''
        servicos = _pav_servicos_from_camadas(ctx, camadas)
        if not servicos:
            return jsonify({'erro': 'Nenhuma camada orçável foi encontrada no perfil calculado.'}), 400

        nome_orc = f"Pavimentacao - {ctx.get('trecho') or ctx.get('obra') or 'dimensionamento'}"
        cur = db.execute("""INSERT INTO orcamentos
            (id_obra,nome_orcamento,descricao,id_data_base,uf_referencia,versao,status,observacoes,regime_previdenciario)
            VALUES (?,?,?,?,?,?,?,?,?)""", [
                id_obra, nome_orc[:180],
                'Orçamento detalhado gerado automaticamente pelo módulo de dimensionamento de pavimentos.',
                id_data_base, uf, '1.0', 'Em elaboração',
                'Escopo restrito à etapa de pavimentação: camadas do pavimento, base, sub-base, reforço e imprimação quando aplicável.',
                d.get('regime_previdenciario') or 'Onerado'
            ])
        id_orc = cur.lastrowid
        try:
            db.execute("UPDATE orcamentos SET bdi_percentual=? WHERE id_orcamento=?", [float(d.get('bdi_percentual') or 0), id_orc])
        except Exception:
            pass

        ordem = 1
        db.execute("""INSERT INTO orcamento_sintetico
            (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
            VALUES (?,?,?,?,?,?)""", [id_orc, '1', 'section', 0, ordem, 'PAVIMENTAÇÃO'])

        itens = []
        avisos = []
        for idx, serv in enumerate(servicos, start=1):
            comp = _pav_buscar_composicao(db, fonte, uf, mes_ref, serv)
            if not comp:
                avisos.append(f"Sem composição {fonte} confiável para: {serv['descricao']}")
                comp = {
                    'id_composicao': None, 'codigo': '', 'fonte': fonte,
                    'descricao': serv['descricao'], 'unidade': 'm3' if serv['volume_m3'] else 'm2',
                    'custo_unitario': 0, '_score_ia': 0
                }
            qtd = _pav_quantidade_por_unidade(serv, comp.get('unidade'))
            ordem += 1
            db.execute("""INSERT INTO orcamento_sintetico
                (id_orcamento,item_num,tipo_linha,profundidade,ordem,tipo_item,
                 id_composicao,codigo,fonte,descricao,unidade,quantidade,custo_unitario)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""", [
                    id_orc, f'1.{idx}', 'item', 1, ordem, 'composicao',
                    comp.get('id_composicao'), comp.get('codigo') or '', fonte,
                    comp.get('descricao') or serv['descricao'], comp.get('unidade') or '',
                    round(qtd, 4), float(comp.get('custo_unitario') or 0)
                ])
            itens.append({
                'servico': serv['descricao'], 'codigo': comp.get('codigo') or '',
                'descricao': comp.get('descricao') or serv['descricao'],
                'unidade': comp.get('unidade') or '', 'quantidade': round(qtd, 4),
                'custo_unitario': float(comp.get('custo_unitario') or 0),
                'score_ia': comp.get('_score_ia', 0),
            })

        total = db.execute("""SELECT COALESCE(SUM(quantidade*custo_unitario),0)
                              FROM orcamento_sintetico
                              WHERE id_orcamento=? AND tipo_linha='item'""", [id_orc]).fetchone()[0]
        bdi_pct = float(d.get('bdi_percentual') or 0)
        try:
            db.execute("""UPDATE orcamentos SET valor_custo_direto=?, valor_bdi=?, valor_total=?
                          WHERE id_orcamento=?""", [total, total * bdi_pct / 100, total * (1 + bdi_pct / 100), id_orc])
        except Exception:
            pass
        db.commit()
        return jsonify({
            'mensagem': 'Orçamento detalhado de pavimentação gerado.',
            'id_orcamento': id_orc,
            'total_itens': len(itens),
            'total_custo_direto': total,
            'fonte': fonte,
            'uf': uf,
            'mes_referencia': mes_ref,
            'itens': itens,
            'avisos': avisos,
        }), 201
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500
    finally:
        db.close()

# ── Exportar Orçamento Sintético (Excel / PDF) ─────────────────────────────────

def _get_os_dados(db, id_orcamento):
    """Retorna dados completos do orçamento + itens com valores calculados."""
    ensure_orcamento_sintetico_bdi_linha(db)
    try:
        orc = dict(db.execute(SEL_ORC_FULL + " WHERE o.id_orcamento=?", [id_orcamento]).fetchone())
    except Exception:
        orc = dict(db.execute("""SELECT o.*, ob.nome_obra FROM orcamentos o
                                  LEFT JOIN obras ob ON o.id_obra=ob.id_obra
                                  WHERE o.id_orcamento=?""", [id_orcamento]).fetchone())
    itens = rows_to_list(db.execute("""
        SELECT * FROM orcamento_sintetico WHERE id_orcamento=?
        ORDER BY ordem, id_item""", [id_orcamento]).fetchall())
    bdi = float(orc.get('bdi_percentual') or 0)
    # Calcular valores e totais por seção
    for it in itens:
        if it['tipo_linha'] == 'section':
            it['valor'] = 0
        else:
            try:
                bdi_item = bdi if it.get('bdi_percentual_linha') in (None, '') else float(it.get('bdi_percentual_linha') or 0)
            except (TypeError, ValueError):
                bdi_item = bdi
            custo = float(it.get('custo_unitario') or 0)
            qtd = float(it.get('quantidade') or 0)
            it['bdi_efetivo_percentual'] = bdi_item
            it['preco_unitario_com_bdi'] = round(custo * (1 + bdi_item / 100), 4)
            it['valor'] = round(qtd * it['preco_unitario_com_bdi'], 2)
    # Preencher valor das seções somando filhos
    _fill_section_values(itens)
    return orc, itens, bdi

def _fill_section_values(itens):
    stack = []  # (profundidade, acumulador_idx)
    for i, it in enumerate(itens):
        if it['tipo_linha'] == 'section':
            depth = it.get('profundidade', 0)
            # Fechar seções de mesma ou maior profundidade
            while stack and stack[-1][0] >= depth:
                _, idx = stack.pop()
                # sum all itens after this section until next same-level section
                pass
            stack.append((depth, i))
    # Simpler approach: for each section sum all subsequent items until another section of same/lower depth
    for i, it in enumerate(itens):
        if it['tipo_linha'] != 'section':
            continue
        depth = it.get('profundidade', 0)
        total = 0
        for j in range(i+1, len(itens)):
            nxt = itens[j]
            if nxt['tipo_linha'] == 'section' and nxt.get('profundidade', 0) <= depth:
                break
            if nxt['tipo_linha'] == 'item':
                total += nxt.get('valor', 0)
        it['valor'] = round(total, 2)

@app.route('/api/orcamentos/<int:id>/exportar/excel', methods=['GET'])
def os_exportar_excel(id):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from flask import send_file
    db = get_db()
    orc_row = db.execute("SELECT 1 FROM orcamentos WHERE id_orcamento=?", [id]).fetchone()
    if not orc_row: db.close(); return jsonify({'erro': 'Orçamento não encontrado.'}), 404
    orc, itens, bdi = _get_os_dados(db, id)
    db.close()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Orçamento Sintético'

    # ── Estilos ──────────────────────────────────────────────────────────────
    DARK   = '0F172A'
    MID    = '1E293B'
    SEC0   = '0F172A'   # seção nível 0 (fundo escuro)
    SEC1   = '1E293B'   # seção nível 1
    ALT    = 'F8FAFF'   # linha par
    HEADER = '334155'   # cabeçalho da tabela

    def mk_font(bold=False, color='000000', size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic,
                    name='Calibri')

    def mk_fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def mk_border(bottom_color='CBD5E1', all_thin=False):
        s = Side(style='thin', color=bottom_color)
        if all_thin:
            return Border(left=s, right=s, top=s, bottom=s)
        return Border(bottom=s)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(vertical='center', wrap_text=True, indent=1)
    right  = Alignment(horizontal='right', vertical='center')

    FMT_MOEDA = 'R$ #,##0.00'
    FMT_PCT   = '0.00"%"'

    # ── Cabeçalho do documento ────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    ws['A1'] = orc.get('nome_obra', '—')
    ws['A1'].font = mk_font(bold=True, color='FFFFFF', size=13)
    ws['A1'].fill = mk_fill(DARK)
    ws['A1'].alignment = Alignment(vertical='center', indent=2)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Orçamento: {orc.get('nome_orcamento','—')}   |   Versão: {orc.get('versao','—')}   |   Status: {orc.get('status','—')}   |   Base: {orc.get('data_base_mes','?')}/{orc.get('data_base_ano','?')}"
    ws['A2'].font = mk_font(color='FFFFFF', size=9)
    ws['A2'].fill = mk_fill(MID)
    ws['A2'].alignment = Alignment(vertical='center', indent=2)
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A3:J3')
    ws['A3'] = f"BDI Aplicado: {bdi:.4f}%   |   Custo Direto: R$ {float(orc.get('valor_custo_direto') or 0):,.2f}   |   VALOR TOTAL DO ORÇAMENTO: R$ {float(orc.get('valor_total') or 0):,.2f}"
    ws['A3'].font = mk_font(bold=True, color='FFFFFF', size=10)
    ws['A3'].fill = mk_fill('6366F1')
    ws['A3'].alignment = Alignment(vertical='center', indent=2)
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 6  # espaço

    # ── Cabeçalho da tabela ───────────────────────────────────────────────────
    headers = ['Item', 'Código', 'Fonte', 'Descrição dos Serviços', 'Un.', 'Quantidade', 'Custo Unit. (R$)', 'Preço Unit. (R$)', 'Valor (R$)', '% s/ Total']
    ws.append([''] * 10)  # row 4 spacing
    ws.append(headers)   # row 5
    HDR_ROW = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(HDR_ROW, c)
        cell.value = h
        cell.font  = mk_font(bold=True, color='FFFFFF', size=9)
        cell.fill  = mk_fill(HEADER)
        cell.alignment = center
        cell.border = mk_border(all_thin=True, bottom_color='475569')
    ws.row_dimensions[HDR_ROW].height = 28

    # ── Itens ──────────────────────────────────────────────────────────────────
    vt = float(orc.get('valor_total') or 0)
    row_idx = HDR_ROW + 1

    for n, it in enumerate(itens):
        tipo  = it.get('tipo_linha', 'item')
        depth = it.get('profundidade', 0)
        v     = float(it.get('valor') or 0)
        cu    = float(it.get('custo_unitario') or 0)
        pu    = float(it.get('preco_unitario_com_bdi') or (cu * (1 + bdi / 100)))
        qt    = float(it.get('quantidade') or 0)
        pct   = v / vt * 100 if vt > 0 else 0

        indent = depth * 3

        if tipo == 'section':
            is_sec0 = depth == 0
            fill_c  = SEC0 if is_sec0 else SEC1
            desc    = f"{'   ' * depth}{it.get('descricao','')}"
            ws.append([
                it.get('item_num',''), '', '', desc,
                '', '', '', '',
                v, pct
            ])
            for c in range(1, 11):
                cell = ws.cell(row_idx, c)
                cell.font      = mk_font(bold=True, color='FFFFFF', size=9 if not is_sec0 else 10)
                cell.fill      = mk_fill(fill_c)
                cell.alignment = right if c in (9, 10) else (center if c == 1 else left)
                cell.border    = mk_border(bottom_color='334155')
                if c == 9:  cell.number_format = FMT_MOEDA
                if c == 10: cell.number_format = FMT_PCT
            ws.row_dimensions[row_idx].height = 22 if is_sec0 else 18
        else:
            fill_c = ALT if row_idx % 2 == 0 else 'FFFFFF'
            desc   = f"{'  ' * depth}{it.get('descricao','')}"
            ws.append([
                it.get('item_num',''),
                it.get('codigo',''),
                it.get('fonte',''),
                desc,
                it.get('unidade',''),
                qt if qt else None,
                cu if cu else None,
                pu if pu else None,
                v if v else None,
                pct if pct else None,
            ])
            for c in range(1, 11):
                cell = ws.cell(row_idx, c)
                cell.font      = mk_font(size=8)
                cell.fill      = mk_fill(fill_c)
                cell.border    = mk_border()
                cell.alignment = right if c in (6,7,8,9,10) else (center if c in (1,2,3,5) else left)
                if c in (6,):     cell.number_format = '#,##0.000'
                if c in (7,8,9):  cell.number_format = FMT_MOEDA
                if c == 10:       cell.number_format = FMT_PCT
            ws.row_dimensions[row_idx].height = 16

        row_idx += 1

    # ── Rodapé totais ─────────────────────────────────────────────────────────
    ws.append([''] * 10)
    row_idx += 1
    ws.append(['', '', '', 'CUSTO DIRETO (sem BDI)', '', '', '', '', float(orc.get('valor_custo_direto') or 0), ''])
    for c in range(1, 11):
        cell = ws.cell(row_idx + 1, c)
        cell.font = mk_font(bold=True, color='FFFFFF', size=9)
        cell.fill = mk_fill('475569')
        cell.alignment = right if c in (9,10) else left
        if c == 9: cell.number_format = FMT_MOEDA
    ws.append(['', '', '', f'BDI ({bdi:.4f}%)', '', '', '', '', float(orc.get('valor_bdi') or 0), ''])
    for c in range(1, 11):
        cell = ws.cell(row_idx + 2, c)
        cell.font = mk_font(bold=True, color='FFFFFF', size=9)
        cell.fill = mk_fill('334155')
        cell.alignment = right if c in (9,10) else left
        if c == 9: cell.number_format = FMT_MOEDA
    ws.append(['', '', '', 'VALOR TOTAL DO ORÇAMENTO', '', '', '', '', float(orc.get('valor_total') or 0), 100.0])
    for c in range(1, 11):
        cell = ws.cell(row_idx + 3, c)
        cell.font = mk_font(bold=True, color='FFFFFF', size=11)
        cell.fill = mk_fill('6366F1')
        cell.alignment = right if c in (9,10) else left
        if c == 9:  cell.number_format = FMT_MOEDA
        if c == 10: cell.number_format = FMT_PCT

    # ── Larguras das colunas ──────────────────────────────────────────────────
    widths = [7, 10, 12, 55, 6, 11, 16, 16, 16, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{HDR_ROW + 1}'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nome_safe = (orc.get('nome_orcamento') or 'orcamento').replace(' ','_')[:40]
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'orcamento_{nome_safe}.xlsx')


@app.route('/api/orcamentos/<int:id>/exportar/pdf', methods=['GET'])
def os_exportar_pdf(id):
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from flask import send_file

    db = get_db()
    orc_row = db.execute("SELECT 1 FROM orcamentos WHERE id_orcamento=?", [id]).fetchone()
    if not orc_row: db.close(); return jsonify({'erro': 'Orçamento não encontrado.'}), 404
    orc, itens, bdi = _get_os_dados(db, id)
    db.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=12*mm, rightMargin=12*mm,
                             topMargin=14*mm, bottomMargin=14*mm)

    styles = getSampleStyleSheet()
    C_DARK     = colors.HexColor('#0F172A')
    C_MID      = colors.HexColor('#1E293B')
    C_SEC0     = colors.HexColor('#0F172A')
    C_SEC1     = colors.HexColor('#1E293B')
    C_HEADER   = colors.HexColor('#334155')
    C_ACCENT   = colors.HexColor('#6366F1')
    C_ALT      = colors.HexColor('#F8FAFF')
    C_WHITE    = colors.white
    C_TEXT     = colors.HexColor('#1E293B')
    C_LIGHT    = colors.HexColor('#F1F5F9')
    C_BORDER   = colors.HexColor('#CBD5E1')

    def sty_para(text, size=8, bold=False, color=C_TEXT, align=TA_LEFT, indent=0):
        s = ParagraphStyle('x', fontSize=size, leading=size*1.3,
                           textColor=color, alignment=align,
                           fontName='Helvetica-Bold' if bold else 'Helvetica',
                           leftIndent=indent)
        return Paragraph(text or '—', s)

    vt     = float(orc.get('valor_total') or 0)
    cd     = float(orc.get('valor_custo_direto') or 0)
    vbdi   = float(orc.get('valor_bdi') or 0)
    nome_obra = orc.get('nome_obra','—')
    nome_orc  = orc.get('nome_orcamento','—')

    story = []

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    story.append(Paragraph(f'<font size="14"><b>{nome_obra}</b></font>', ParagraphStyle('h', alignment=TA_LEFT, textColor=C_WHITE, backColor=C_DARK, fontSize=14, leading=20, leftIndent=6, spaceAfter=2)))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f'Orçamento: <b>{nome_orc}</b> &nbsp;|&nbsp; Versão: {orc.get("versao","—")} &nbsp;|&nbsp; Status: {orc.get("status","—")} &nbsp;|&nbsp; Base: {orc.get("data_base_mes","?")}/{orc.get("data_base_ano","?")} &nbsp;|&nbsp; BDI: <b>{bdi:.4f}%</b>',
        ParagraphStyle('sub', fontSize=8, textColor=C_WHITE, backColor=C_MID, leftIndent=6, rightIndent=6, leading=14, spaceAfter=2)
    ))
    story.append(Spacer(1, 2*mm))

    # KPI bar
    vt_fmt  = f'R$ {vt:,.2f}'
    cd_fmt  = f'R$ {cd:,.2f}'
    kpi_data = [[
        sty_para('CUSTO DIRETO', size=7, color=C_WHITE, align=TA_CENTER),
        sty_para(cd_fmt.replace(',','.').replace('.',',',1), size=10, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para(f'BDI ({bdi:.2f}%)', size=7, color=C_WHITE, align=TA_CENTER),
        sty_para(f'R$ {vbdi:,.2f}'.replace(',','.').replace('.',',',1), size=10, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para('VALOR TOTAL DO ORÇAMENTO', size=7, color=C_WHITE, align=TA_CENTER),
        sty_para(f'R$ {vt:,.2f}'.replace(',','.').replace('.',',',1), size=12, bold=True, color=C_WHITE, align=TA_CENTER),
    ]]
    kpi_table = Table(kpi_data, colWidths=[35*mm, 45*mm, 35*mm, 45*mm, 55*mm, 60*mm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (3,-1), C_MID),
        ('BACKGROUND', (4,0), (-1,-1), C_ACCENT),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [C_MID, C_MID]),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#334155')),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 4*mm))

    # ── Tabela de itens ────────────────────────────────────────────────────────
    PAGE_W = landscape(A4)[0] - 24*mm
    col_widths = [12*mm, 16*mm, 18*mm, None, 10*mm, 18*mm, 26*mm, 26*mm, 26*mm, 18*mm]
    # Desc = remaining width
    fixed = sum(w for w in col_widths if w)
    col_widths[3] = PAGE_W - fixed

    hdr_cells = [
        sty_para('Item',   7, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para('Código', 7, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para('Fonte',  7, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para('Descrição dos Serviços', 7, bold=True, color=C_WHITE),
        sty_para('Un.',    7, bold=True, color=C_WHITE, align=TA_CENTER),
        sty_para('Quant.', 7, bold=True, color=C_WHITE, align=TA_RIGHT),
        sty_para('Custo Unit.', 7, bold=True, color=C_WHITE, align=TA_RIGHT),
        sty_para('Preço Unit.', 7, bold=True, color=C_WHITE, align=TA_RIGHT),
        sty_para('Valor (R$)',  7, bold=True, color=C_WHITE, align=TA_RIGHT),
        sty_para('%',      7, bold=True, color=C_WHITE, align=TA_RIGHT),
    ]
    data = [hdr_cells]
    row_styles = []  # list of (row_idx, bg_color, text_color, bold, height)

    for n, it in enumerate(itens):
        row_num = len(data)
        tipo  = it.get('tipo_linha','item')
        depth = it.get('profundidade', 0)
        v     = float(it.get('valor') or 0)
        cu    = float(it.get('custo_unitario') or 0)
        pu    = float(it.get('preco_unitario_com_bdi') or (cu * (1 + bdi / 100)))
        qt    = float(it.get('quantidade') or 0)
        pct   = v / vt * 100 if vt > 0 else 0
        indent = depth * 8

        def br(s): return (f'R$ {s:,.2f}').replace(',','X').replace('.',',').replace('X','.')

        if tipo == 'section':
            bg  = C_SEC0 if depth == 0 else C_SEC1
            sz  = 9 if depth == 0 else 8
            desc_txt = ('   ' * depth) + (it.get('descricao','') or '')
            row = [
                sty_para(it.get('item_num',''), sz, bold=True, color=C_WHITE, align=TA_CENTER),
                sty_para('', sz, color=C_WHITE),
                sty_para('', sz, color=C_WHITE),
                sty_para(f'<b>{desc_txt}</b>', sz, color=C_WHITE),
                sty_para('', sz, color=C_WHITE),
                sty_para('', sz, color=C_WHITE),
                sty_para('', sz, color=C_WHITE),
                sty_para('', sz, color=C_WHITE),
                sty_para(br(v), sz, bold=True, color=C_WHITE, align=TA_RIGHT),
                sty_para(f'{pct:.2f}%', sz, bold=True, color=C_WHITE, align=TA_RIGHT),
            ]
            row_styles.append((row_num, bg, True))
        else:
            bg = C_ALT if row_num % 2 == 0 else C_WHITE
            desc_txt = it.get('descricao','') or ''
            row = [
                sty_para(it.get('item_num',''), 7, align=TA_CENTER),
                sty_para(it.get('codigo','') or '', 7, align=TA_CENTER),
                sty_para(it.get('fonte','') or '', 7, align=TA_CENTER),
                sty_para(desc_txt, 7, indent=indent),
                sty_para(it.get('unidade','') or '', 7, align=TA_CENTER),
                sty_para(f'{qt:,.3f}' if qt else '', 7, align=TA_RIGHT),
                sty_para(br(cu) if cu else '', 7, align=TA_RIGHT),
                sty_para(br(pu) if pu else '', 7, align=TA_RIGHT),
                sty_para(br(v) if v else '', 7, align=TA_RIGHT),
                sty_para(f'{pct:.2f}%' if pct else '', 7, align=TA_RIGHT),
            ]
            row_styles.append((row_num, bg, False))
        data.append(row)

    # Rodapé
    def add_total_row(label, value, bg, sz=9):
        nonlocal data
        row_num = len(data)
        row = [sty_para('','',color=C_WHITE)]*3 + \
              [sty_para(f'<b>{label}</b>', sz, bold=True, color=C_WHITE)] + \
              [sty_para('','',color=C_WHITE)]*4 + \
              [sty_para(f'<b>{br(value)}</b>', sz, bold=True, color=C_WHITE, align=TA_RIGHT),
               sty_para('','',color=C_WHITE)]
        data.append(row)
        row_styles.append((row_num, bg, True))

    add_total_row('CUSTO DIRETO (sem BDI)', cd, C_HEADER)
    add_total_row(f'BDI ({bdi:.4f}%)', vbdi, C_MID)
    add_total_row('VALOR TOTAL DO ORÇAMENTO', vt, C_ACCENT, sz=10)

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    # Estilos da tabela
    base_style = [
        ('BACKGROUND', (0,0), (-1,0), C_HEADER),
        ('TEXTCOLOR', (0,0), (-1,0), C_WHITE),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.3, C_BORDER),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [C_WHITE, C_ALT]),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
    ]
    for row_num, bg, is_sec in row_styles:
        base_style.append(('BACKGROUND', (0, row_num), (-1, row_num), bg))

    tbl.setStyle(TableStyle(base_style))
    story.append(tbl)

    # ── Build ──────────────────────────────────────────────────────────────────
    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_DARK)
        canvas.setFont('Helvetica', 7)
        canvas.drawString(12*mm, 8*mm, f'Emitido em {__import__("datetime").date.today().strftime("%d/%m/%Y")}   |   {nome_obra} — {nome_orc}')
        canvas.drawRightString(landscape(A4)[0] - 12*mm, 8*mm, f'Pág. {doc.page}')
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    buf.seek(0)
    nome_safe = (orc.get('nome_orcamento') or 'orcamento').replace(' ','_')[:40]
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'orcamento_{nome_safe}.pdf')


# ── Recalcular custos unitários zerados de composições vinculadas ─────────────
@app.route('/api/orcamentos/<int:id>/recalcular-custos', methods=['POST'])
def os_recalcular_custos(id):
    """Percorre os itens do orçamento que têm id_composicao e custo_unitario=0,
       recalcula o custo via itens_composicao × precos_insumos e grava."""
    db = get_db()

    SQL_CUSTO_COMP = """
        SELECT COALESCE(SUM(
            ic.coeficiente * COALESCE(
                CASE WHEN ic.tipo_item = 'COMPOSICAO' THEN (
                    SELECT custo_unitario FROM composicoes
                    WHERE codigo = ic.codigo_item
                       OR codigo = 'SINAPI.' || ic.codigo_item
                       OR codigo = 'SICRO.'  || ic.codigo_item
                    LIMIT 1
                ) END,
                (SELECT COALESCE(p.preco_desonerado,
                                 p.preco_nao_desonerado,
                                 p.preco_referencia)
                 FROM precos_insumos p
                 JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
                 WHERE p.id_insumo = (
                     SELECT id_insumo FROM insumos
                     WHERE codigo_insumo = ic.codigo_item LIMIT 1
                 )
                 ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1),
                ic.preco_unitario,
                0
            )
        ), 0) AS custo_calc
        FROM itens_composicao ic
        WHERE ic.id_composicao = ?
    """

    itens = db.execute("""
        SELECT id_item, id_composicao, custo_unitario
        FROM orcamento_sintetico
        WHERE id_orcamento = ? AND tipo_linha = 'item'
          AND id_composicao IS NOT NULL
          AND (custo_unitario IS NULL OR custo_unitario = 0)
    """, [id]).fetchall()

    atualizados = 0
    for row in itens:
        custo = db.execute(SQL_CUSTO_COMP, [row['id_composicao']]).fetchone()
        custo_calc = round(float(custo['custo_calc'] or 0), 4)
        if custo_calc > 0:
            db.execute(
                "UPDATE orcamento_sintetico SET custo_unitario=? WHERE id_item=?",
                [custo_calc, row['id_item']]
            )
            atualizados += 1

    db.commit()
    db.close()
    return jsonify({'atualizados': atualizados,
                    'mensagem': f'{atualizados} item(ns) recalculado(s).'})

# ── Endpoint estendido de orçamento (inclui BDI) ─────────────────────────────
@app.route('/api/orcamentos/<int:id>/completo', methods=['GET'])
def orc_get_completo(id):
    db = get_db()
    try:
        row = db.execute(SEL_ORC_FULL + " WHERE o.id_orcamento=?", [id]).fetchone()
    except Exception:
        row = db.execute("""SELECT o.*, ob.nome_obra FROM orcamentos o
                            LEFT JOIN obras ob ON o.id_obra=ob.id_obra
                            WHERE o.id_orcamento=?""", [id]).fetchone()
    db.close()
    if not row:
        return jsonify({'erro': 'Orçamento não encontrado.'}), 404
    return jsonify(dict(row))


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTAÇÃO DE ORÇAMENTO SINTÉTICO (PDF / Excel via IA)
# ═══════════════════════════════════════════════════════════════════════════════

_PROMPT_IMPORTAR_SINTETICO = """Você é engenheiro orçamentista sênior especializado em obras de construção civil brasileiras.

Receberá o conteúdo textual de um orçamento sintético (extraído de PDF ou planilha Excel).
Sua tarefa é interpretar esse conteúdo e estruturá-lo em um JSON padronizado.

REGRAS OBRIGATÓRIAS:
1. Responda SOMENTE com o JSON puro, sem nenhum texto antes ou depois, sem markdown, sem ```json
2. Preserve ao máximo as descrições originais dos serviços
3. Identifique seções/grupos de serviços (linhas sem valor unitário que funcionam como cabeçalhos)
4. Para itens com código reconhecível (SINAPI, SICRO, etc.), preserve o código
5. Se um valor de custo_unitario não estiver disponível, use 0
6. Normalize unidades: M2→m2, M3→m3, KG→kg, UN→un, M→m, etc.
7. Para linhas que são claramente seções/grupos (sem quantidade/preço unitário), use tipo_linha="section"
8. Para linhas com código, quantidade e preço, use tipo_linha="item"
9. Infere a hierarquia: seções de nível 0, subseções de nível 1, itens de nível 1 ou 2

FORMATO JSON DE SAÍDA:
{
  "titulo": "Nome identificado do orçamento ou null",
  "observacoes": "Notas sobre o arquivo interpretado",
  "secoes": [
    {
      "descricao": "NOME DA SEÇÃO EM MAIÚSCULAS",
      "subsecoes": [
        {
          "descricao": "Nome da subseção (se existir, senão null)",
          "itens": [
            {
              "codigo": "103689",
              "fonte": "SINAPI",
              "descricao": "Descrição completa do serviço",
              "unidade": "m2",
              "quantidade": 2.88,
              "custo_unitario": 462.36
            }
          ]
        }
      ]
    }
  ]
}

IDENTIFICAÇÃO DE FONTE:
- Códigos numéricos de 5-6 dígitos sem ponto → SINAPI
- Códigos como "68.01.04", "67.02.07" → pode ser SINAPI ou outra tabela
- Prefixo "C" seguido de número → USUARIO ou CP
- Códigos SICRO geralmente têm formato diferente
- Sem código → USUARIO

CONTEÚDO DO ARQUIVO A INTERPRETAR:
{conteudo}
"""

_PROMPT_MATCH_IMPORTACAO = """Você é engenheiro orçamentista sênior com acesso ao banco de dados SINAPI/SICRO do sistema.

Receberá itens de um orçamento importado e uma lista de composições disponíveis no banco de dados.
Sua tarefa é associar cada item importado à melhor composição disponível no banco, quando possível.

ITENS IMPORTADOS:
{itens}

COMPOSIÇÕES DISPONÍVEIS (id | código | descrição | unidade | custo_unitario | fonte):
{composicoes}

REGRAS:
1. Para cada item, procure a composição mais semelhante por código ou por descrição
2. Se encontrar correspondência por código exato: use o id_composicao dessa composição
3. Se não encontrar por código, tente por similaridade de descrição e unidade
4. Se não houver correspondência adequada: id_composicao = null (mantém o custo_unitario original)
5. Responda SOMENTE com JSON puro, sem texto adicional

FORMATO DE SAÍDA:
{
  "matches": [
    {
      "idx": 0,
      "id_composicao": 123,
      "codigo_matched": "103689",
      "custo_unitario_matched": 462.36,
      "confianca": "alta"
    }
  ]
}
"""


@app.route('/api/orcamentos/<int:id>/importar-sintetico', methods=['POST'])
def importar_sintetico(id):
    """
    Importa um orçamento sintético a partir de PDF ou Excel via IA.
    Recebe multipart/form-data com o arquivo e opcionalmente modo_merge.
    """
    import base64, io

    # Verificar orçamento existe
    db = get_db()
    orc = db.execute('SELECT * FROM orcamentos WHERE id_orcamento=?', [id]).fetchone()
    if not orc:
        db.close()
        return jsonify({'erro': 'Orçamento não encontrado.'}), 404

    arquivo = request.files.get('arquivo')
    if not arquivo:
        db.close()
        return jsonify({'erro': 'Nenhum arquivo enviado. Envie um PDF ou Excel.'}), 400

    modo_merge = request.form.get('modo_merge', 'substituir')  # 'substituir' ou 'adicionar'

    nome_arquivo = arquivo.filename or ''
    ext = nome_arquivo.rsplit('.', 1)[-1].lower() if '.' in nome_arquivo else ''

    conteudo_texto = ''
    mensagem_extracao = ''

    # ── Extrair conteúdo do arquivo ──────────────────────────────────────────
    if ext in ('xlsx', 'xls', 'xlsm', 'ods'):
        try:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()), read_only=True, data_only=True)
                linhas = []
                for ws in wb.worksheets:
                    linhas.append(f'=== Planilha: {ws.title} ===')
                    for row in ws.iter_rows(values_only=True):
                        # Remover colunas None no final para evitar linhas
                        # de dezenas de KB em planilhas com range largo
                        row_values = list(row)
                        while row_values and row_values[-1] is None:
                            row_values.pop()
                        linha = '\t'.join(str(c) if c is not None else '' for c in row_values)
                        if linha.strip():
                            linhas.append(linha)
                conteudo_texto = '\n'.join(linhas)
                mensagem_extracao = f'Excel extraído: {len(linhas)} linhas de {len(wb.worksheets)} planilha(s).'
            except Exception:
                arquivo.stream.seek(0)
                import pandas as pd
                engine = 'xlrd' if ext == 'xls' else ('odf' if ext == 'ods' else 'openpyxl')
                dfs = pd.read_excel(io.BytesIO(arquivo.read()), sheet_name=None, engine=engine)
                linhas = []
                for sheet_name, df in dfs.items():
                    linhas.append(f'=== Planilha: {sheet_name} ===')
                    linhas.append(df.to_csv(sep='\t', index=False))
                conteudo_texto = '\n'.join(linhas)
                mensagem_extracao = f'Excel extraído via pandas: {len(dfs)} planilha(s).'
        except Exception as e:
            db.close()
            return jsonify({'erro': f'Falha ao ler arquivo Excel: {str(e)}'}), 400

    elif ext == 'pdf':
        dados_b64 = base64.b64encode(arquivo.read()).decode()
        # Usar Claude Vision para interpretar o PDF diretamente
        try:
            texto_pdf = _call_claude_ia([{
                'role': 'user',
                'content': [
                    {
                        'type': 'document',
                        'source': {
                            'type': 'base64',
                            'media_type': 'application/pdf',
                            'data': dados_b64,
                        }
                    },
                    {
                        'type': 'text',
                        'text': (
                            'Extraia TODO o conteúdo textual deste orçamento sintético de obra, '
                            'preservando a estrutura de tabela com colunas: código, descrição, unidade, quantidade, custo unitário, valor total. '
                            'Separe colunas com \\t e linhas com \\n. '
                            'Preserve exatamente os números e textos. '
                            'Responda apenas com o conteúdo extraído, sem explicações.'
                        )
                    }
                ]
            }], max_tokens=8000)
            conteudo_texto = texto_pdf
            mensagem_extracao = 'PDF extraído via IA.'
        except Exception as e:
            db.close()
            return jsonify({'erro': f'Falha ao processar PDF: {str(e)}'}), 500
    else:
        # Tentar ler como texto simples (CSV, TXT, etc.)
        try:
            conteudo_texto = arquivo.read().decode('utf-8', errors='replace')
            mensagem_extracao = 'Arquivo de texto lido diretamente.'
        except Exception as e:
            db.close()
            return jsonify({'erro': f'Formato não suportado: {ext}. Use PDF, Excel (.xlsx/.xls) ou CSV.'}), 400

    if not conteudo_texto.strip():
        db.close()
        return jsonify({'erro': 'O arquivo parece estar vazio ou não foi possível extrair conteúdo.'}), 400

    # Limitar tamanho do conteúdo (evitar ultrapassar limite de tokens)
    conteudo_truncado = conteudo_texto[:18000]
    if len(conteudo_texto) > 18000:
        mensagem_extracao += f' (conteúdo truncado em 18000 chars de {len(conteudo_texto)} totais)'

    # ── Chamar IA para estruturar o orçamento ────────────────────────────────
    try:
        resposta_ia = _call_claude_ia([{
            'role': 'user',
            'content': _PROMPT_IMPORTAR_SINTETICO.replace('{conteudo}', conteudo_truncado)
        }], max_tokens=8000)
        estrutura = _clean_json(resposta_ia)
    except Exception as e:
        db.close()
        return jsonify({'erro': f'Falha na interpretação por IA: {str(e)}'}), 500

    secoes = estrutura.get('secoes', [])
    if not secoes:
        db.close()
        return jsonify({'erro': 'A IA não conseguiu identificar seções no arquivo. Verifique o formato.'}), 422

    # ── Coletar todos os itens com código para tentar match no banco ─────────
    todos_itens_flat = []
    for sec in secoes:
        for subsec in sec.get('subsecoes', []):
            for item in subsec.get('itens', []):
                if item.get('codigo'):
                    todos_itens_flat.append(item)

    # Buscar composições que possam ter match (pelos códigos encontrados)
    matches_por_codigo = {}
    if todos_itens_flat:
        codigos = list({it['codigo'] for it in todos_itens_flat if it.get('codigo')})
        placeholders = ','.join('?' * len(codigos))
        comps_match = db.execute(f"""
            SELECT id_composicao, codigo, descricao, unidade, custo_unitario, fonte
            FROM composicoes
            WHERE codigo IN ({placeholders})
        """, codigos).fetchall() if codigos else []
        for c in comps_match:
            matches_por_codigo[c['codigo']] = dict(c)

        # Se houver itens sem match direto, tentar via IA (limitado a 40 itens)
        sem_match = [it for it in todos_itens_flat if it.get('codigo') not in matches_por_codigo]
        if sem_match and len(sem_match) <= 40:
            try:
                # Buscar composições similares para tentar match
                termos_busca = ' '.join(
                    it.get('descricao', '')[:30] for it in sem_match[:5]
                )
                comp_amostra = db.execute("""
                    SELECT id_composicao, codigo, descricao, unidade, custo_unitario, fonte
                    FROM composicoes
                    ORDER BY RANDOM() LIMIT 200
                """).fetchall()
                lista_comp_txt = '\n'.join(
                    f"{c['id_composicao']} | {c['codigo']} | {c['descricao'][:60]} | {c['unidade']} | {c['custo_unitario']} | {c['fonte']}"
                    for c in comp_amostra
                )
                itens_txt = '\n'.join(
                    f"{i} | {it.get('codigo','')} | {it.get('descricao','')[:60]} | {it.get('unidade','')} | {it.get('custo_unitario',0)}"
                    for i, it in enumerate(sem_match)
                )
                resp_match = _call_claude_ia([{
                    'role': 'user',
                    'content': _PROMPT_MATCH_IMPORTACAO.replace('{itens}', itens_txt).replace('{composicoes}', lista_comp_txt)
                }], max_tokens=3000)
                match_data = _clean_json(resp_match)
                for m in match_data.get('matches', []):
                    if m.get('id_composicao') and m.get('confianca') in ('alta', 'média', 'media'):
                        sem_match_item = sem_match[m['idx']] if m['idx'] < len(sem_match) else None
                        if sem_match_item:
                            sem_match_item['_id_composicao'] = m['id_composicao']
                            sem_match_item['_custo_matched'] = m.get('custo_unitario_matched')
            except Exception:
                pass  # Match por IA é opcional; seguimos sem ele

    # ── Gravar no banco ───────────────────────────────────────────────────────
    if modo_merge == 'substituir':
        db.execute('DELETE FROM orcamento_sintetico WHERE id_orcamento=?', [id])

    ordem = db.execute(
        'SELECT COALESCE(MAX(ordem),0) FROM orcamento_sintetico WHERE id_orcamento=?', [id]
    ).fetchone()[0]

    total_itens = 0
    total_secoes = 0

    for sec_i, sec in enumerate(secoes):
        ordem += 1
        total_secoes += 1
        sec_num_base = sec_i + 1
        sec_num = str(sec_num_base)

        db.execute("""INSERT INTO orcamento_sintetico
            (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
            VALUES (?,?,?,?,?,?)""",
            [id, sec_num, 'section', 0, ordem,
             (sec.get('descricao') or 'SEÇÃO').upper()])

        for sub_j, subsec in enumerate(sec.get('subsecoes', [])):
            sub_desc = subsec.get('descricao')
            itens_desta = subsec.get('itens', [])

            if sub_desc:
                # Tem subseção real → inserir como section de profundidade 1
                ordem += 1
                sub_num = f'{sec_num}.{sub_j + 1}'
                db.execute("""INSERT INTO orcamento_sintetico
                    (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
                    VALUES (?,?,?,?,?,?)""",
                    [id, sub_num, 'section', 1, ordem, sub_desc.upper()])
                prof_item = 2
                num_prefixo = sub_num
            else:
                prof_item = 1
                num_prefixo = sec_num

            for it_k, item in enumerate(itens_desta):
                ordem += 1
                total_itens += 1
                it_num = f'{num_prefixo}.{it_k + 1}' if sub_desc else f'{sec_num}.{it_k + 1}'

                codigo = item.get('codigo', '')
                comp_match = matches_por_codigo.get(codigo)
                id_composicao = None
                custo_unit = float(item.get('custo_unitario') or 0)
                fonte = item.get('fonte', '')

                if comp_match:
                    id_composicao = comp_match['id_composicao']
                    if custo_unit == 0:
                        custo_unit = float(comp_match.get('custo_unitario') or 0)
                    if not fonte:
                        fonte = comp_match.get('fonte', '')
                elif item.get('_id_composicao'):
                    id_composicao = item['_id_composicao']
                    if custo_unit == 0 and item.get('_custo_matched'):
                        custo_unit = float(item['_custo_matched'])

                if not fonte:
                    if codigo and codigo.isdigit() and len(codigo) >= 5:
                        fonte = 'SINAPI'
                    elif codigo:
                        fonte = 'USUARIO'

                db.execute("""INSERT INTO orcamento_sintetico
                    (id_orcamento,item_num,tipo_linha,profundidade,ordem,tipo_item,
                     id_composicao,codigo,fonte,descricao,unidade,quantidade,custo_unitario)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    [id, it_num, 'item', prof_item, ordem,
                     'composicao' if id_composicao else 'composicao',
                     id_composicao,
                     codigo,
                     fonte,
                     item.get('descricao', ''),
                     item.get('unidade', ''),
                     float(item.get('quantidade') or 0),
                     custo_unit])

    db.commit()

    # Atualizar totais do orçamento
    total = db.execute("""
        SELECT COALESCE(SUM(quantidade * custo_unitario), 0)
        FROM orcamento_sintetico
        WHERE id_orcamento=? AND tipo_linha='item'
    """, [id]).fetchone()[0]
    db.execute("""UPDATE orcamentos SET valor_custo_direto=?, valor_total=?
                  WHERE id_orcamento=?""", [total, total, id])
    db.commit()
    db.close()

    return jsonify({
        'mensagem': f'Importação concluída: {total_itens} item(ns) em {total_secoes} seção(ões).',
        'total_itens': total_itens,
        'total_secoes': total_secoes,
        'extracao': mensagem_extracao,
        'titulo_detectado': estrutura.get('titulo'),
        'observacoes_ia': estrutura.get('observacoes'),
    }), 200


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTAR ORÇAMENTO SINTÉTICO — EXCEL DIRETO (SEM IA)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/orcamentos/<int:id>/importar-sintetico-excel', methods=['POST'])
def importar_sintetico_excel_direto(id):
    """
    Importa orçamento sintético a partir de Excel de forma direta, sem uso de IA.
    Mantém exatamente o conteúdo da planilha: código, descrição, unidade, quantidade
    e custo unitário. Suporta .xlsx, .xls, .xlsm e .ods.
    Estrutura esperada: colunas Código | Descrição | Unidade | Quantidade | Custo Unit. | Valor
    Linhas sem unidade são tratadas como cabeçalhos de seção.
    """
    import io

    db = get_db()
    orc = db.execute('SELECT * FROM orcamentos WHERE id_orcamento=?', [id]).fetchone()
    if not orc:
        db.close()
        return jsonify({'erro': 'Orçamento não encontrado.'}), 404

    arquivo = request.files.get('arquivo')
    if not arquivo:
        db.close()
        return jsonify({'erro': 'Nenhum arquivo enviado.'}), 400

    nome_arquivo = arquivo.filename or ''
    ext = nome_arquivo.rsplit('.', 1)[-1].lower() if '.' in nome_arquivo else ''

    if ext not in ('xlsx', 'xls', 'xlsm', 'ods'):
        db.close()
        return jsonify({'erro': 'Apenas arquivos Excel (.xlsx, .xls, .xlsm, .ods) são aceitos nesta opção.'}), 400

    modo_merge = request.form.get('modo_merge', 'substituir')

    try:
        import pandas as pd
        engine = 'xlrd' if ext == 'xls' else ('odf' if ext == 'ods' else 'openpyxl')
        df = pd.read_excel(io.BytesIO(arquivo.read()), header=None, engine=engine, dtype=object)
    except Exception as e:
        db.close()
        return jsonify({'erro': f'Falha ao ler arquivo Excel: {str(e)}'}), 400

    # ── Helpers ──────────────────────────────────────────────────────────────
    import math

    def is_empty(v):
        if v is None:
            return True
        if isinstance(v, float) and math.isnan(v):
            return True
        return str(v).strip() in ('', 'nan', 'None', 'NaN')

    def clean_str(v):
        return '' if is_empty(v) else str(v).strip()

    def clean_num(v):
        if is_empty(v):
            return 0.0
        try:
            s = str(v).replace(',', '.').strip()
            return float(s)
        except Exception:
            return 0.0

    # ── Verificar se há pelo menos 6 colunas significativas ──────────────────
    max_cols = df.shape[1]
    if max_cols < 2:
        db.close()
        return jsonify({'erro': 'Arquivo sem colunas suficientes. Verifique o formato.'}), 422

    # ── Detectar linha de cabeçalho e coluna-base ─────────────────────────────
    # Procurar a linha que tem "descrição" ou "código" para pular
    HEADER_MARKERS = {'código', 'codigo', 'descrição', 'descricao', 'descrição dos serviços',
                      'unid', 'unid.', 'unidade', 'quantidade', 'custo', 'valor'}

    def is_header_row(row):
        for v in row:
            if not is_empty(v) and clean_str(v).lower() in HEADER_MARKERS:
                return True
        return False

    # ── Parse rows ────────────────────────────────────────────────────────────
    rows_parsed = []

    for idx, row in df.iterrows():
        # Get at most 6 columns (idx 0..5)
        def col(n):
            return row.iloc[n] if n < len(row) else None

        c0 = col(0)  # Código
        c1 = col(1)  # Descrição
        c2 = col(2)  # Unidade
        c3 = col(3)  # Quantidade
        c4 = col(4)  # Custo Unit.
        c5 = col(5)  # Valor

        desc = clean_str(c1)
        if not desc:
            continue  # Linha vazia

        # Pular linha de cabeçalho
        if is_header_row(row):
            continue

        unit = clean_str(c2)
        code = clean_str(c0)
        has_total = not is_empty(c5)

        if unit:
            # Linha de item (tem unidade)
            rows_parsed.append({
                'type': 'item',
                'code': code,
                'desc': desc,
                'unit': unit,
                'qty': clean_num(c3),
                'cost': clean_num(c4),
                'has_total': has_total,
            })
        else:
            # Linha de seção / subseção
            rows_parsed.append({
                'type': 'section',
                'desc': desc,
                'has_total': has_total,
            })

    if not rows_parsed:
        db.close()
        return jsonify({'erro': 'Nenhum item ou seção encontrado no arquivo. Verifique o formato da planilha.'}), 422

    # ── Buscar composições pelo código (match direto, sem IA) ─────────────────
    all_codes = list({r['code'] for r in rows_parsed if r['type'] == 'item' and r['code']})
    matches_por_codigo = {}
    if all_codes:
        placeholders = ','.join('?' * len(all_codes))
        comps = db.execute(
            f'SELECT id_composicao, codigo, descricao, unidade, custo_unitario, fonte '
            f'FROM composicoes WHERE codigo IN ({placeholders})',
            all_codes
        ).fetchall()
        for c in comps:
            matches_por_codigo[c['codigo']] = dict(c)

    # ── Limpar e gravar no banco ──────────────────────────────────────────────
    if modo_merge == 'substituir':
        db.execute('DELETE FROM orcamento_sintetico WHERE id_orcamento=?', [id])

    ordem = db.execute(
        'SELECT COALESCE(MAX(ordem), 0) FROM orcamento_sintetico WHERE id_orcamento=?', [id]
    ).fetchone()[0]

    total_itens  = 0
    total_secoes = 0

    # Estado de seção corrente
    sec_num              = 0
    current_section_num  = None
    current_sub_num      = None
    sec_item_count       = 0
    sub_item_count       = 0
    sub_count            = 0  # contador de subseções dentro da seção atual

    def ensure_section():
        """Garante que há pelo menos uma seção pai criada."""
        nonlocal sec_num, current_section_num, current_sub_num
        nonlocal sec_item_count, sub_count, total_secoes, ordem
        if current_section_num is None:
            sec_num += 1
            current_section_num = str(sec_num)
            current_sub_num = None
            sec_item_count = 0
            sub_count = 0
            total_secoes += 1
            ordem += 1
            db.execute("""INSERT INTO orcamento_sintetico
                (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
                VALUES (?,?,?,?,?,?)""",
                [id, current_section_num, 'section', 0, ordem, 'GERAL'])

    for r in rows_parsed:
        ordem += 1

        if r['type'] == 'section':
            if r['has_total']:
                # Seção principal (profundidade 0)
                sec_num += 1
                current_section_num = str(sec_num)
                current_sub_num = None
                sec_item_count = 0
                sub_count = 0
                total_secoes += 1
                db.execute("""INSERT INTO orcamento_sintetico
                    (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
                    VALUES (?,?,?,?,?,?)""",
                    [id, current_section_num, 'section', 0, ordem, r['desc'].upper()])
            else:
                # Subseção (profundidade 1)
                ensure_section()
                sub_count += 1
                current_sub_num = f'{current_section_num}.{sub_count}'
                sub_item_count = 0
                total_secoes += 1
                db.execute("""INSERT INTO orcamento_sintetico
                    (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
                    VALUES (?,?,?,?,?,?)""",
                    [id, current_sub_num, 'section', 1, ordem, r['desc'].upper()])

        else:  # item
            ensure_section()

            if current_sub_num:
                sub_item_count += 1
                it_num   = f'{current_sub_num}.{sub_item_count}'
                prof_item = 2
            else:
                sec_item_count += 1
                it_num   = f'{current_section_num}.{sec_item_count}'
                prof_item = 1

            codigo    = r['code']
            custo_unit = r['cost']
            unidade   = r['unit']
            fonte     = ''
            id_composicao = None

            # Tentativa de match por código
            if codigo and codigo in matches_por_codigo:
                comp = matches_por_codigo[codigo]
                id_composicao = comp['id_composicao']
                if custo_unit == 0:
                    custo_unit = float(comp.get('custo_unitario') or 0)
                if not unidade:
                    unidade = comp.get('unidade', '')
                fonte = comp.get('fonte', '')

            if not fonte:
                if codigo and str(codigo).isdigit() and len(str(codigo)) >= 5:
                    fonte = 'SINAPI'
                elif codigo:
                    fonte = 'USUARIO'

            db.execute("""INSERT INTO orcamento_sintetico
                (id_orcamento,item_num,tipo_linha,profundidade,ordem,tipo_item,
                 id_composicao,codigo,fonte,descricao,unidade,quantidade,custo_unitario)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [id, it_num, 'item', prof_item, ordem,
                 'composicao',
                 id_composicao,
                 codigo,
                 fonte,
                 r['desc'],
                 unidade,
                 r['qty'],
                 custo_unit])
            total_itens += 1

    db.commit()

    # Atualizar totais do orçamento
    total_custo = db.execute("""
        SELECT COALESCE(SUM(quantidade * custo_unitario), 0)
        FROM orcamento_sintetico
        WHERE id_orcamento=? AND tipo_linha='item'
    """, [id]).fetchone()[0]
    db.execute("""UPDATE orcamentos SET valor_custo_direto=?, valor_total=?
                  WHERE id_orcamento=?""", [total_custo, total_custo, id])
    db.commit()
    db.close()

    return jsonify({
        'mensagem': f'Importação direta concluída: {total_itens} item(ns) em {total_secoes} seção(ões).',
        'total_itens': total_itens,
        'total_secoes': total_secoes,
    }), 200


# ═══════════════════════════════════════════════════════════════════════════════
# CURVA ABC — Serviços e Insumos
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/orcamentos/<int:id>/curva-abc-servicos', methods=['GET'])
def curva_abc_servicos(id):
    """
    Curva ABC dos serviços do orçamento sintético.
    Serviços com mesmo código (ou mesma descrição quando sem código) são consolidados
    em uma única linha, somando quantidades e valores. O custo unitário exibido é o
    custo médio ponderado (Σ custo_total / Σ quantidade).
    """
    db = get_db()
    ensure_orcamento_sintetico_bdi_linha(db)
    orc = db.execute("""
        SELECT o.bdi_percentual, o.nome_orcamento, o.versao, o.status,
               ob.nome_obra
        FROM orcamentos o
        LEFT JOIN obras ob ON o.id_obra = ob.id_obra
        WHERE o.id_orcamento = ?
    """, [id]).fetchone()
    if not orc:
        db.close()
        return jsonify({'erro': 'Orçamento não encontrado.'}), 404

    bdi_pct = float(orc['bdi_percentual'] or 0)

    itens = db.execute("""
        SELECT id_item, item_num, descricao, unidade, quantidade,
               custo_unitario, bdi_percentual_linha, codigo, fonte, tipo_item, id_composicao
        FROM orcamento_sintetico
        WHERE id_orcamento = ? AND tipo_linha = 'item'
        ORDER BY ordem, id_item
    """, [id]).fetchall()
    db.close()

    # ── Consolidar serviços com mesmo código/descrição ──────────────────────
    # Chave: código não-vazio → usa código; caso contrário usa descrição em maiúsculas
    def _chave_serv(row):
        cod = (row.get('codigo') or '').strip().upper()
        return cod if cod else (row.get('descricao') or '').strip().upper()

    consolidado = {}   # chave → dados agregados

    for row in itens:
        it       = dict(row)
        qtd      = float(it.get('quantidade')    or 0)
        custo_su = float(it.get('custo_unitario') or 0)
        try:
            bdi_item = bdi_pct if it.get('bdi_percentual_linha') in (None, '') else float(it.get('bdi_percentual_linha') or 0)
        except (TypeError, ValueError):
            bdi_item = bdi_pct
        preco_bdi = custo_su * (1 + bdi_item / 100)
        valor     = preco_bdi * qtd

        key = _chave_serv(it)
        if key not in consolidado:
            consolidado[key] = {
                'codigo':          (it.get('codigo') or '').strip(),
                'descricao':       it.get('descricao') or '',
                'unidade':         it.get('unidade') or '',
                'fonte':           it.get('fonte') or '',
                'tipo_item':       it.get('tipo_item') or '',
                'id_composicao':   it.get('id_composicao'),
                'bdi_percentual':  bdi_item,
                'soma_qtd':        0.0,
                'soma_custo_direto': 0.0,  # Σ (custo_su × qtd)
                'valor_total':     0.0,
                'soma_bdi_ponderado': 0.0,
                'ocorrencias':     [],
            }
        consolidado[key]['soma_qtd']          += qtd
        consolidado[key]['soma_custo_direto'] += custo_su * qtd
        consolidado[key]['valor_total']       += valor
        consolidado[key]['soma_bdi_ponderado'] += bdi_item * (custo_su * qtd)
        # Registrar cada ocorrência para detalhe opcional
        consolidado[key]['ocorrencias'].append({
            'item_num':  it.get('item_num', ''),
            'quantidade': qtd,
            'custo_unitario': custo_su,
            'bdi_percentual': bdi_item,
            'preco_bdi': round(preco_bdi, 4),
            'valor':     round(valor, 2),
        })

    # ── Montar result com custo médio ponderado ─────────────────────────────
    result = []
    total_geral = 0.0
    for dados in consolidado.values():
        qtd_total  = dados['soma_qtd']
        valor_total= dados['valor_total']
        # Custo médio ponderado s/ BDI: Σ(cu×qtd) / Σqtd
        custo_medio = (dados['soma_custo_direto'] / qtd_total) if qtd_total > 0 else 0.0
        bdi_medio = (dados['soma_bdi_ponderado'] / dados['soma_custo_direto']) if dados['soma_custo_direto'] > 0 else bdi_pct
        preco_medio_bdi = (valor_total / qtd_total) if qtd_total > 0 else 0.0

        result.append({
            'codigo':                dados['codigo'],
            'descricao':             dados['descricao'],
            'unidade':               dados['unidade'],
            'fonte':                 dados['fonte'],
            'tipo_item':             dados['tipo_item'],
            'id_composicao':         dados['id_composicao'],
            'bdi_percentual':        round(bdi_medio, 4),
            'quantidade':            round(qtd_total, 4),
            'custo_unitario':        round(custo_medio, 4),
            'preco_unitario_com_bdi': round(preco_medio_bdi, 4),
            'valor_total':           round(valor_total, 2),
            'ocorrencias':           dados['ocorrencias'],
            'consolidado':           len(dados['ocorrencias']) > 1,
        })
        total_geral += valor_total

    # ── Ordenar, calcular % e classificar ───────────────────────────────────
    result.sort(key=lambda x: x['valor_total'], reverse=True)

    acum = 0.0
    for i, it in enumerate(result):
        it['rank'] = i + 1
        pct   = (it['valor_total'] / total_geral * 100) if total_geral else 0
        acum += pct
        it['percentual']           = round(pct,  4)
        it['percentual_acumulado'] = round(acum, 4)
        it['classe'] = 'A' if acum <= 50 else ('B' if acum <= 80 else 'C')

    # ── Resumo por classe ────────────────────────────────────────────────────
    resumo = {}
    for cls in ('A', 'B', 'C'):
        cls_itens = [x for x in result if x['classe'] == cls]
        resumo[cls] = {
            'qtd':   len(cls_itens),
            'valor': round(sum(x['valor_total'] for x in cls_itens), 2),
            'pct':   round(sum(x['percentual']  for x in cls_itens), 2),
        }

    return jsonify({
        'orcamento':      dict(orc),
        'itens':          result,
        'total_geral':    round(total_geral, 2),
        'bdi_percentual': bdi_pct,
        'resumo':         resumo,
    })


@app.route('/api/orcamentos/<int:id>/curva-abc-insumos', methods=['GET'])
def curva_abc_insumos(id):
    """
    Curva ABC de Insumos — Metodologia com decomposição recursiva:

    Para cada serviço do orçamento sintético:
      1. Localiza a composição analítica vinculada.
      2. Para cada item da composição:
         - Se for INSUMO → acumula diretamente.
         - Se for COMPOSICAO → busca RECURSIVAMENTE seus insumos,
           propagando o coeficiente pai × coeficiente filho.
           Repete até atingir apenas insumos (sem limite de níveis).
      3. Consolida insumos pelo código (ou descrição quando sem código).
      4. Calcula preço médio ponderado e ordena por custo_total decrescente.
      5. Classifica A / B / C.
    """
    db = get_db()
    orc = db.execute("""
        SELECT o.bdi_percentual, o.nome_orcamento, o.versao, o.status,
               ob.nome_obra
        FROM orcamentos o
        LEFT JOIN obras ob ON o.id_obra = ob.id_obra
        WHERE o.id_orcamento = ?
    """, [id]).fetchone()
    if not orc:
        db.close()
        return jsonify({'erro': 'Orçamento não encontrado.'}), 404

    # ── 1. Serviços do orçamento sintético ──────────────────────────────────
    servicos = db.execute("""
        SELECT id_item, item_num, descricao, unidade, quantidade,
               custo_unitario, codigo, fonte, tipo_item, id_composicao
        FROM orcamento_sintetico
        WHERE id_orcamento = ? AND tipo_linha = 'item'
        ORDER BY ordem, id_item
    """, [id]).fetchall()

    # ── 2. Cache de composições (para busca eficiente por código) ────────────
    # comp_cache: codigo_normalizado → id_composicao
    comp_rows = db.execute(
        "SELECT id_composicao, codigo FROM composicoes"
    ).fetchall()
    comp_cache = {}
    for r in comp_rows:
        raw = (r['codigo'] or '').strip().upper()
        if not raw:
            continue
        # Strip any existing prefix to get the bare code
        bare = raw.replace('SINAPI.', '').replace('SICRO.', '').replace('SEINFRA.', '').replace('SUDECAP.', '').replace('GOINFRA.', '').replace('CDHU.', '').replace('USUARIO.', '').strip()
        # Register all variants used by reference importers.
        for k in [bare, 'SINAPI.' + bare, 'SICRO.' + bare, 'SEINFRA.' + bare, 'SUDECAP.' + bare, 'GOINFRA.' + bare, 'CDHU.' + bare, raw]:
            if k and k not in comp_cache:
                comp_cache[k] = r['id_composicao']

    # SQL: todos os itens de uma composição (incluindo sub-composições)
    SQL_ITENS = """
        SELECT ic.codigo_item, ic.descricao, ic.unidade,
               ic.coeficiente, ic.tipo_item, ic.preco_unitario
        FROM itens_composicao ic
        WHERE ic.id_composicao = ?
        ORDER BY ic.ordem
    """

    # SQL: resolver preço de um insumo pelo código
    SQL_PRECO = """
        SELECT COALESCE(p.preco_desonerado,
                        p.preco_nao_desonerado,
                        p.preco_referencia) AS preco,
               COALESCE(p.ibs_percentual, 0) AS ibs_percentual,
               COALESCE(p.cbs_percentual, 0) AS cbs_percentual
        FROM precos_insumos p
        JOIN datas_base db2 ON db2.id_data_base = p.id_data_base
        WHERE p.id_insumo = (
            SELECT id_insumo FROM insumos
            WHERE codigo_insumo = ? LIMIT 1
        )
        ORDER BY db2.ano DESC, db2.mes DESC LIMIT 1
    """

    # ── 3. Consolidado: chave → somas ────────────────────────────────────────
    consolidado: dict = {}

    def _chave(codigo, descricao):
        c = (codigo or '').strip().upper()
        return c if c else (descricao or '').strip().upper()

    def _add(codigo, descricao, unidade, tipo, qtd_insumo, preco,
             ibs_percentual, cbs_percentual, serv_desc, serv_num):
        custo_item = qtd_insumo * preco
        valor_ibs = custo_item * (float(ibs_percentual or 0) / 100.0)
        valor_cbs = custo_item * (float(cbs_percentual or 0) / 100.0)
        key = _chave(codigo, descricao)
        if key not in consolidado:
            consolidado[key] = {
                'codigo':      (codigo or '').strip(),
                'descricao':   descricao or '',
                'unidade':     unidade or '',
                'tipo_item':   tipo or 'INSUMO',
                'qtd_total':   0.0,
                'soma_custo':  0.0,
                'valor_ibs':   0.0,
                'valor_cbs':   0.0,
                'ocorrencias': [],
            }
        consolidado[key]['qtd_total']  += qtd_insumo
        consolidado[key]['soma_custo'] += custo_item
        consolidado[key]['valor_ibs']  += valor_ibs
        consolidado[key]['valor_cbs']  += valor_cbs
        consolidado[key]['ocorrencias'].append({
            'servico':    serv_desc,
            'item_num':   serv_num,
            'qtd_insumo': round(qtd_insumo, 6),
            'preco':      round(preco, 4),
            'custo':      round(custo_item, 2),
            # campos herdados para compatibilidade com a UI
            'qtd_servico':  round(qtd_insumo, 6),
            'coeficiente':  1.0,
            'ibs_percentual': round(float(ibs_percentual or 0), 4),
            'cbs_percentual': round(float(cbs_percentual or 0), 4),
            'valor_ibs': round(valor_ibs, 2),
            'valor_cbs': round(valor_cbs, 2),
        })

    def _resolver_preco(codigo_item, preco_armazenado):
        """Retorna o melhor preço disponível para um insumo."""
        if codigo_item:
            r = db.execute(SQL_PRECO, [codigo_item.strip()]).fetchone()
            if r and r['preco']:
                return {
                    'preco': float(r['preco']),
                    'ibs_percentual': float(r['ibs_percentual'] or 0),
                    'cbs_percentual': float(r['cbs_percentual'] or 0),
                }
        return {
            'preco': float(preco_armazenado or 0),
            'ibs_percentual': 0.0,
            'cbs_percentual': 0.0,
        }

    # ── 4. Expansão recursiva de uma composição ──────────────────────────────
    def _expandir(id_comp, fator, serv_desc, serv_num, visitados=None):
        """
        Percorre todos os itens de id_comp.
        - Se o item for INSUMO:  acumula com quantidade = coef × fator.
        - Se o item for COMPOSICAO: localiza a sub-composição e chama
          recursivamente com fator = coef × fator (propagação de coeficientes).
        visitados evita loops infinitos em referências circulares.
        """
        if visitados is None:
            visitados = set()
        if id_comp in visitados:
            return
        visitados = visitados | {id_comp}   # imutável para não contaminar outros ramos

        itens = db.execute(SQL_ITENS, [id_comp]).fetchall()
        for row in itens:
            ci    = dict(row)
            coef  = float(ci.get('coeficiente') or 0)
            if coef <= 0:
                continue
            tipo  = (ci.get('tipo_item') or 'INSUMO').upper()
            cod   = (ci.get('codigo_item') or '').strip()

            if tipo == 'COMPOSICAO':
                # Tentar resolver a sub-composição pelo código (todas variantes)
                cod_up = cod.upper()
                bare   = cod_up.replace('SINAPI.','').replace('SICRO.','').replace('SEINFRA.','').replace('SUDECAP.','').replace('GOINFRA.','').replace('CDHU.','').replace('USUARIO.','').strip()
                sub_id = (comp_cache.get(bare) or
                          comp_cache.get('SINAPI.' + bare) or
                          comp_cache.get('SICRO.'  + bare) or
                          comp_cache.get('SEINFRA.' + bare) or
                          comp_cache.get('SUDECAP.' + bare) or
                          comp_cache.get('GOINFRA.' + bare) or
                          comp_cache.get('CDHU.' + bare) or
                          comp_cache.get(cod_up))
                if sub_id:
                    # Propaga: fator_filho = coeficiente_aqui × fator_pai
                    _expandir(sub_id, coef * fator, serv_desc, serv_num, visitados)
                else:
                    # Sub-composição não encontrada no banco — trata como insumo direto
                    preco = _resolver_preco(cod, ci.get('preco_unitario'))
                    _add(cod, ci.get('descricao',''), ci.get('unidade',''),
                         'COMPOSICAO', coef * fator, preco['preco'],
                         preco['ibs_percentual'], preco['cbs_percentual'], serv_desc, serv_num)
            else:
                # INSUMO (ou MO, MATERIAL, EQUIPAMENTO, etc.)
                preco = _resolver_preco(cod, ci.get('preco_unitario'))
                _add(cod, ci.get('descricao',''), ci.get('unidade',''),
                     tipo, coef * fator, preco['preco'],
                     preco['ibs_percentual'], preco['cbs_percentual'], serv_desc, serv_num)

    # ── 5. Processar cada serviço do orçamento sintético ────────────────────
    for row in servicos:
        s       = dict(row)
        qtd_s   = float(s.get('quantidade') or 0)
        id_comp = s.get('id_composicao')
        serv_d  = s.get('descricao', '')
        serv_n  = s.get('item_num', '')

        if not qtd_s:
            continue

        if not id_comp:
            # Sem id_composicao → tentar localizar a composição pelo código do item
            cod_item = (s.get('codigo') or '').strip().upper()
            if cod_item:
                id_comp = (comp_cache.get(cod_item) or
                           comp_cache.get('SINAPI.' + cod_item) or
                           comp_cache.get('SICRO.'  + cod_item) or
                           comp_cache.get('SEINFRA.' + cod_item) or
                           comp_cache.get('SUDECAP.' + cod_item) or
                           comp_cache.get('GOINFRA.' + cod_item) or
                           comp_cache.get('CDHU.' + cod_item))

        if id_comp:
            # Composição vinculada → expansão recursiva com fator = qtd_servico
            _expandir(id_comp, qtd_s, serv_d, serv_n)
        else:
            # Sem composição em nenhum nível — tratar como insumo/serviço direto
            preco = float(s.get('custo_unitario') or 0)
            if preco > 0:
                _add(s.get('codigo',''), serv_d, s.get('unidade',''),
                     s.get('tipo_item') or 'INSUMO',
                     qtd_s, preco, 0.0, 0.0, serv_d, serv_n)

    db.close()

    # ── 6. Montar resultado final ────────────────────────────────────────────
    result = []
    for dados in consolidado.values():
        qtd   = dados['qtd_total']
        custo = dados['soma_custo']
        preco_medio = (custo / qtd) if qtd > 0 else 0.0
        result.append({
            'codigo':           dados['codigo'],
            'descricao':        dados['descricao'],
            'unidade':          dados['unidade'],
            'tipo_item':        dados['tipo_item'],
            'quantidade_total': round(qtd,   6),
            'custo_unitario':   round(preco_medio, 4),
            'custo_total':      round(custo,  2),
            'valor_ibs':        round(dados['valor_ibs'], 2),
            'valor_cbs':        round(dados['valor_cbs'], 2),
            'ocorrencias':      dados['ocorrencias'],
        })

    # ── 7. Ordenar, calcular % e classificar ────────────────────────────────
    result.sort(key=lambda x: x['custo_total'], reverse=True)
    total_geral = sum(x['custo_total'] for x in result)

    acum = 0.0
    for i, it in enumerate(result):
        it['rank'] = i + 1
        pct   = (it['custo_total'] / total_geral * 100) if total_geral else 0
        acum += pct
        it['percentual']           = round(pct,  4)
        it['percentual_acumulado'] = round(acum, 4)
        it['classe'] = 'A' if acum <= 50 else ('B' if acum <= 80 else 'C')

    # ── 8. Resumo por classe ─────────────────────────────────────────────────
    resumo = {}
    for cls in ('A', 'B', 'C'):
        cls_itens = [x for x in result if x['classe'] == cls]
        resumo[cls] = {
            'qtd':   len(cls_itens),
            'valor': round(sum(x['custo_total'] for x in cls_itens), 2),
            'pct':   round(sum(x['percentual']  for x in cls_itens), 2),
        }

    total_ibs = sum(x.get('valor_ibs', 0) for x in result)
    total_cbs = sum(x.get('valor_cbs', 0) for x in result)

    return jsonify({
        'orcamento':     dict(orc),
        'itens':         result,
        'total_geral':   round(total_geral, 2),
        'total_ibs':     round(total_ibs, 2),
        'total_cbs':     round(total_cbs, 2),
        'resumo':        resumo,
        'total_insumos': len(result),
        'metodologia':   'Insumos obtidos por decomposição recursiva das composições analíticas vinculadas a cada serviço do orçamento sintético.',
    })

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO IA — Análise de Projetos e Geração de Orçamento Sintético
# ═══════════════════════════════════════════════════════════════════════════════
import base64, threading, uuid, json as _json

# Instalar json-repair automaticamente se não estiver presente
try:
    import json_repair as _json_repair
except ImportError:
    import subprocess, sys as _sys
    print("📦  Instalando dependência json-repair...")
    subprocess.check_call([_sys.executable, '-m', 'pip', 'install', 'json-repair', '-q'])
    import json_repair as _json_repair
    print("✅  json-repair instalado com sucesso.")

_ANALISE_JOBS: dict = {}          # job_id → {status, progresso, etapa, resultado, erro}
_FORMATOS_OK  = {'png','jpg','jpeg','pdf','dxf','ifc'}
_MAX_FILES    = 20

# ── Prompts ────────────────────────────────────────────────────────────────────
_PROMPT_EXTRAIR = """Você é um engenheiro civil especialista em orçamentação de obras.

TAREFA: Analisar este projeto de construção civil e extrair TODOS os quantitativos possíveis.

REGRAS OBRIGATÓRIAS:
1. Sua resposta deve ser SOMENTE o objeto JSON abaixo, sem NENHUM texto antes ou depois
2. Não use markdown, não use ```json, não escreva explicações — apenas o JSON puro
3. Mesmo que a imagem seja de baixa qualidade, extraia o que for visível
4. Use as cotas dimensionais visíveis no projeto para calcular áreas e volumes
5. Se ver uma planta baixa com cotas, calcule área total e por ambiente
6. Se ver projeto hidráulico/elétrico/sanitário, conte os pontos e meça tubulações

FORMATO OBRIGATÓRIO (responda exatamente assim, sem modificar as chaves):
{"tipo_documento": "planta_arquitetonica", "escala": "1:50", "quantidades": [{"servico": "Area total construida", "quantidade": 41.88, "unidade": "m2", "observacao": "calculado pelas cotas 5.99x6.99"}, {"servico": "Alvenaria de blocos", "quantidade": 45.0, "unidade": "m2", "observacao": "estimado pelo perimetro das paredes"}], "confianca": "alta", "observacoes_gerais": "Planta baixa residencial unifamiliar"}

ELEMENTOS A BUSCAR CONFORME O TIPO DE PROJETO:
- Planta arquitetônica: área total, área por cômodo, perímetro de paredes, pé-direito, portas (un), janelas (un/m²)
- Planta estrutural: pilares (un), vigas (m), lajes (m²), escadas, fundações
- Projeto hidráulico: pontos de água fria/quente (un), tubulações (m), reservatórios (L), conexões (un)
- Projeto sanitário: ramais de esgoto (m), caixas de inspeção (un), fossas, tubos (m), diâmetros
- Instalações elétricas: pontos de tomada (un), pontos de iluminação (un), eletrodutos (m), circuitos (un), carga total (W)
- Cobertura: área de telhado (m²), tipo de telha, calhas (m), rufos (m)
- Fundações: blocos (un/m³), lastro (m²), baldrame (m), escavação (m³)

IMPORTANTE: quantidade deve ser um número (não string). Unidade deve ser m2, m3, m, un, L, W, etc."""

_PROMPT_MATCH = """Você é engenheiro orçamentista sênior especializado em obras de construção civil.

QUANTITATIVOS EXTRAÍDOS DOS PROJETOS:
{quantidades}

COMPOSIÇÕES DISPONÍVEIS (formato: id | código | descrição | unidade | custo_unit | fonte):
{composicoes}

Tarefa: gerar o orçamento sintético associando os quantitativos às composições mais adequadas.

Regras IMPORTANTES:
1. Use APENAS ids da lista acima — nunca invente ids
2. Um serviço pode precisar de múltiplas composições (ex: alvenaria → assentamento + vergas + encunhamento)
3. Se nenhuma composição serve, use null para id_composicao
4. Organize em seções lógicas (SERV. PRELIMINARES, FUNDAÇÕES, SUPERESTRUTURA, ALVENARIA, etc.)
5. A quantidade deve estar na mesma unidade da composição escolhida
6. Preencha custo_unitario com o valor da composição selecionada (da lista acima)

Responda APENAS em JSON sem texto adicional:
{{
  "secoes": [
    {{
      "descricao": "NOME DA SEÇÃO EM MAIÚSCULAS",
      "itens": [
        {{
          "id_composicao": 123,
          "codigo": "103328",
          "fonte": "SINAPI",
          "descricao": "Descrição do serviço",
          "unidade": "m2",
          "quantidade": 450.0,
          "custo_unitario": 106.48,
          "justificativa": "resumo de 1 linha explicando a escolha"
        }}
      ]
    }}
  ],
  "cobertura_pct": 75,
  "observacoes": "notas gerais sobre o orçamento e o que ficou faltando"
}}"""


def _call_claude_ia(messages, max_tokens=6000):
    """Chama a API Claude. Requer ANTHROPIC_API_KEY."""
    import urllib.request, urllib.error
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        raise EnvironmentError(
            "ANTHROPIC_API_KEY não configurada.\n"
            "Defina a variável de ambiente antes de usar a análise por IA:\n"
            "  Windows: set ANTHROPIC_API_KEY=sua_chave\n"
            "  Linux/Mac: export ANTHROPIC_API_KEY=sua_chave"
        )
    body = _json.dumps({
        'model': 'claude-opus-4-5',
        'max_tokens': max_tokens,
        'messages': messages,
    }).encode()
    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers={
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01',
            'content-type': 'application/json',
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            data = _json.loads(resp.read())

        stop_reason = data.get('stop_reason', '')
        content     = data.get('content', [])

        # Extrair texto dos content blocks
        text = ''
        for block in content:
            if block.get('type') == 'text':
                text = block.get('text', '')
                break

        if not text:
            tipos = [b.get('type') for b in content]
            raise RuntimeError(
                f"API retornou sem texto. stop_reason={stop_reason!r}, "
                f"content_types={tipos}. "
                f"Verifique sua cota em console.anthropic.com"
            )

        # Se resposta foi cortada por limite de tokens, avisar mas continuar
        # (json-repair lida com JSON truncado)
        if stop_reason == 'max_tokens':
            text += '\n/* RESPOSTA TRUNCADA POR LIMITE DE TOKENS */'

        return text

    except urllib.error.HTTPError as e:
        corpo = e.read().decode()[:400]
        raise RuntimeError(f"Erro HTTP {e.code} na API Claude: {corpo}")
    except urllib.error.URLError as e:
        reason = getattr(e, "reason", e)
        reason_text = str(reason)
        if "WinError 10013" in reason_text:
            raise RuntimeError(
                "A chamada externa para a API Claude foi bloqueada pelo Windows ou pelo ambiente "
                "em que o servidor foi iniciado (WinError 10013). Reinicie o servidor em um terminal "
                "com permissao de rede e verifique firewall/antivirus/proxy."
            )
        raise RuntimeError(f"Erro de rede ao chamar a API Claude: {reason_text}")


def _clean_json(text: str) -> dict:
    """Extrai e repara JSON da resposta da IA, tolerando todos os formatos comuns."""
    import re

    original = text or ''

    # Verificação precoce: resposta vazia
    if not original.strip():
        raise _json.JSONDecodeError(
            "A IA retornou resposta vazia. Verifique sua cota em console.anthropic.com",
            '', 0)

    t = original.strip()

    # 1. Extrair conteúdo entre ```...```
    m = re.search(r'```(?:json)?\s*([\s\S]*?)```', t)
    if m:
        t = m.group(1).strip()

    # 2. Se não começa com {, achar o primeiro objeto JSON
    if not t.startswith('{'):
        m = re.search(r'\{[\s\S]*\}', t)
        if m:
            t = m.group(0)

    # 3. Se ainda vazio após limpeza, mostrar o que a IA retornou
    if not t:
        preview = original[:400].replace('\n', ' ')
        raise _json.JSONDecodeError(
            f"A IA não retornou JSON válido. Resposta recebida: {preview!r}",
            original, 0)

    # 4. Tentar parse direto
    try:
        return _json.loads(t)
    except _json.JSONDecodeError:
        pass

    # 5. Corrigir vírgulas extras
    t_fix = re.sub(r',\s*([}\]])', r'\1', t)
    try:
        return _json.loads(t_fix)
    except _json.JSONDecodeError:
        pass

    # 6. Usar json-repair (lida com truncamentos, aspas erradas, etc.)
    try:
        repaired = _json_repair.repair_json(t_fix)
        result = _json.loads(repaired)
        if isinstance(result, dict):
            return result
    except Exception:
        pass

    # 7. Último recurso: truncar no último item completo
    try:
        ultimo = t.rfind('}, {')
        if ultimo > 0:
            t_trunc  = t[:ultimo + 1]
            t_trunc += ']' * max(0, t_trunc.count('[') - t_trunc.count(']'))
            t_trunc += '}' * max(0, t_trunc.count('{') - t_trunc.count('}'))
            return _json.loads(t_trunc)
    except Exception:
        pass

    # Tudo falhou — mostrar prévia para diagnóstico
    preview = original[:300].replace('\n', ' ')
    raise _json.JSONDecodeError(
        f"JSON inválido após todas as tentativas. Resposta da IA: {preview!r}",
        t, 0)


# ── Parsers por formato ────────────────────────────────────────────────────────

def _parse_ifc(data: bytes) -> dict:
    try:
        import ifcopenshell, tempfile as _tf
        with _tf.NamedTemporaryFile(suffix='.ifc', delete=False) as f:
            f.write(data); fname = f.name
        ifc = ifcopenshell.open(fname)
        os.unlink(fname)

        schema = ifc.schema   # 'IFC2X3' ou 'IFC4'
        qtd    = []

        def _cnt(tipo):
            """Conta elementos de um tipo — tolerante a tipos inexistentes no schema."""
            try:
                return ifc.by_type(tipo)
            except RuntimeError:
                return []

        def _rect_dims(item):
            """Retorna (xdim, ydim, depth) de um IfcExtrudedAreaSolid com perfil retangular."""
            if item.is_a('IfcExtrudedAreaSolid'):
                p = item.SweptArea
                if p.is_a('IfcRectangleProfileDef'):
                    return p.XDim, p.YDim, item.Depth
            return None

        def _geom_items(element):
            """Gera todos os IfcExtrudedAreaSolid de um elemento."""
            try:
                for rep in element.Representation.Representations:
                    for item in rep.Items:
                        yield item
            except Exception:
                return

        # ── Pavimentos ─────────────────────────────────────────────────────
        pavimentos = _cnt('IfcBuildingStorey')
        if pavimentos:
            qtd.append({'servico': 'Pavimentos (andares)',
                        'quantidade': len(pavimentos), 'unidade': 'un',
                        'observacao': ', '.join(p.Name or f'Pav.{i+1}'
                                                for i, p in enumerate(pavimentos))})

        # ── Paredes — área real ─────────────────────────────────────────────
        # IFC4: IfcWallStandardCase é o tipo concreto; IfcWall é o supertipo
        paredes = _cnt('IfcWallStandardCase') or _cnt('IfcWall')
        areas_p, esp_p, vols_p = [], set(), []
        for w in paredes:
            for item in _geom_items(w):
                dims = _rect_dims(item)
                if dims:
                    xd, yd, depth = dims
                    # XDim = comprimento, YDim = espessura, Depth = altura
                    # (na extrusão vertical a direção é Z)
                    comprimento = xd
                    espessura   = yd
                    altura      = depth
                    areas_p.append(comprimento * altura)   # área visível da face
                    vols_p.append(comprimento * espessura * altura)
                    esp_p.add(round(espessura, 3))
                    break  # um sólido por elemento é suficiente

        if paredes:
            qtd.append({'servico': 'Paredes — contagem',
                        'quantidade': len(paredes), 'unidade': 'un',
                        'observacao': 'Contagem direta do modelo BIM'})
        if areas_p:
            qtd.append({'servico': 'Paredes — área de formas/revestimento',
                        'quantidade': round(sum(areas_p), 2), 'unidade': 'm2',
                        'observacao': (f'Soma das áreas das faces ({len(areas_p)} paredes medidas). '
                                       f'Espessuras: {sorted(esp_p)} m')})
            qtd.append({'servico': 'Paredes — volume de concreto/alvenaria',
                        'quantidade': round(sum(vols_p), 4), 'unidade': 'm3',
                        'observacao': 'Volume total das paredes (comprimento × espessura × altura)'})

        # ── Pilares ─────────────────────────────────────────────────────────
        pilares = _cnt('IfcColumn')
        vols_col, alts_col, secs_col = [], [], []
        for c in pilares:
            for item in _geom_items(c):
                dims = _rect_dims(item)
                if dims:
                    xd, yd, depth = dims
                    vols_col.append(xd * yd * depth)
                    alts_col.append(depth)
                    secs_col.append(f'{xd*100:.0f}×{yd*100:.0f}cm')
                    break

        if pilares:
            qtd.append({'servico': 'Pilares — contagem',
                        'quantidade': len(pilares), 'unidade': 'un',
                        'observacao': 'Contagem direta do modelo BIM'})
        if vols_col:
            sec_comum = max(set(secs_col), key=secs_col.count) if secs_col else '?'
            qtd.append({'servico': 'Pilares — volume de concreto armado',
                        'quantidade': round(sum(vols_col), 4), 'unidade': 'm3',
                        'observacao': (f'Volume total. Seção mais comum: {sec_comum}. '
                                       f'Altura média: {sum(alts_col)/len(alts_col):.2f}m')})

        # ── Vigas ─────────────────────────────────────────────────────────
        vigas = _cnt('IfcBeam')
        vols_v, comps_v = [], []
        for v in vigas:
            for item in _geom_items(v):
                dims = _rect_dims(item)
                if dims:
                    xd, yd, depth = dims
                    vols_v.append(xd * yd * depth)
                    comps_v.append(depth)
                    break

        if vigas:
            qtd.append({'servico': 'Vigas — contagem',
                        'quantidade': len(vigas), 'unidade': 'un',
                        'observacao': 'Contagem direta do modelo BIM'})
        if vols_v:
            qtd.append({'servico': 'Vigas — volume de concreto armado',
                        'quantidade': round(sum(vols_v), 4), 'unidade': 'm3',
                        'observacao': f'Volume total. Comprimento total: {sum(comps_v):.2f}m'})
            qtd.append({'servico': 'Vigas — comprimento total',
                        'quantidade': round(sum(comps_v), 2), 'unidade': 'm',
                        'observacao': 'Soma dos comprimentos de todas as vigas'})

        # ── Lajes ─────────────────────────────────────────────────────────
        lajes = _cnt('IfcSlab')
        areas_l, vols_l, esps_l = [], [], []
        for s in lajes:
            for item in _geom_items(s):
                dims = _rect_dims(item)
                if dims:
                    xd, yd, depth = dims
                    areas_l.append(xd * yd)
                    vols_l.append(xd * yd * depth)
                    esps_l.append(depth)
                    break

        if lajes:
            qtd.append({'servico': 'Lajes — contagem',
                        'quantidade': len(lajes), 'unidade': 'un',
                        'observacao': 'Contagem direta do modelo BIM'})
        if areas_l:
            esp_med = sum(esps_l)/len(esps_l) if esps_l else 0
            qtd.append({'servico': 'Lajes — área total',
                        'quantidade': round(sum(areas_l), 2), 'unidade': 'm2',
                        'observacao': f'Área total das lajes. Espessura média: {esp_med:.3f}m'})
            qtd.append({'servico': 'Lajes — volume de concreto',
                        'quantidade': round(sum(vols_l), 4), 'unidade': 'm3',
                        'observacao': 'Volume total das lajes'})

        # ── Esquadrias ────────────────────────────────────────────────────
        for tipo, nome in [('IfcDoor', 'Portas'), ('IfcWindow', 'Janelas')]:
            elems = _cnt(tipo)
            if elems:
                qtd.append({'servico': nome, 'quantidade': len(elems),
                            'unidade': 'un', 'observacao': 'Contagem direta do modelo BIM'})

        # ── Escadas ───────────────────────────────────────────────────────
        escadas = _cnt('IfcStair')
        if escadas:
            qtd.append({'servico': 'Escadas', 'quantidade': len(escadas),
                        'unidade': 'un', 'observacao': 'Contagem direta do modelo BIM'})

        # ── Fundações (nomes variam por schema) ───────────────────────────
        for tipo in ('IfcFooting', 'IfcPile', 'IfcDeepFoundation', 'IfcFoundation'):
            try:
                elems = ifc.by_type(tipo)
                if elems:
                    qtd.append({'servico': f'Elementos de fundação ({tipo})',
                                'quantidade': len(elems), 'unidade': 'un',
                                'observacao': 'Contagem direta do modelo BIM'})
            except RuntimeError:
                pass   # tipo não existe neste schema — ignorar silenciosamente

        # ── Coberturas ────────────────────────────────────────────────────
        for tipo in ('IfcRoof', 'IfcCovering'):
            try:
                elems = ifc.by_type(tipo)
                if elems:
                    qtd.append({'servico': f'Cobertura ({tipo})',
                                'quantidade': len(elems), 'unidade': 'un',
                                'observacao': 'Contagem direta do modelo BIM'})
                    break
            except RuntimeError:
                pass

        obs = (f'Modelo BIM IFC ({schema}) — {len(pavimentos)} pavimento(s). '
               f'{len(qtd)} quantitativos extraídos da geometria.')

        return {'tipo_documento': 'modelo_bim_ifc', 'quantidades': qtd,
                'confianca': 'alta', 'observacoes_gerais': obs,
                'escala': None}

    except ImportError:
        return {'tipo_documento': 'modelo_bim_ifc', 'quantidades': [],
                'confianca': 'baixa',
                'observacoes_gerais': 'ifcopenshell não instalado — execute: pip install ifcopenshell'}
    except Exception as e:
        import traceback
        return {'tipo_documento': 'modelo_bim_ifc', 'quantidades': [],
                'confianca': 'baixa',
                'observacoes_gerais': f'Erro ao processar IFC: {type(e).__name__}: {e}'}


def _parse_dxf(data: bytes) -> dict:
    try:
        import ezdxf, io, math
        doc = ezdxf.read(io.StringIO(data.decode('utf-8', errors='replace')))
        msp = doc.modelspace()

        layers, total_len, hatch_count = {}, 0.0, 0
        blocks_inserts = {}

        for ent in msp:
            lyr = getattr(getattr(ent, 'dxf', None), 'layer', '0')
            layers[lyr] = layers.get(lyr, 0) + 1
            t = ent.dxftype()
            if t == 'LINE':
                s, e = ent.dxf.start, ent.dxf.end
                total_len += math.sqrt(sum((e[i]-s[i])**2 for i in range(3)))
            elif t in ('LWPOLYLINE', 'POLYLINE', 'SPLINE'):
                try: total_len += ent.length()
                except: pass
            elif t == 'HATCH':
                hatch_count += 1
            elif t == 'INSERT':
                bn = getattr(ent.dxf, 'name', 'BLOCO')
                blocks_inserts[bn] = blocks_inserts.get(bn, 0) + 1

        qtd = []
        if total_len > 0:
            qtd.append({'servico': 'Comprimento total de traços (verificar escala)',
                       'quantidade': round(total_len, 2), 'unidade': 'un.DXF',
                       'observacao': 'Soma de linhas/polilínhas. Multiplicar pelo fator de escala do desenho.'})
        if hatch_count > 0:
            qtd.append({'servico': f'Hatches (áreas hachuradas)', 'quantidade': hatch_count,
                       'unidade': 'un', 'observacao': 'Contagem de regiões hachuradas'})
        for blk, cnt_blk in sorted(blocks_inserts.items(), key=lambda x: -x[1])[:15]:
            qtd.append({'servico': f'Bloco "{blk}"', 'quantidade': cnt_blk,
                       'unidade': 'un', 'observacao': 'Contagem de inserções deste bloco no desenho'})

        layer_info = ', '.join(f'{k}({v})' for k,v in list(layers.items())[:8])
        return {'tipo_documento': 'desenho_cad_dxf', 'quantidades': qtd,
                'confianca': 'media',
                'observacoes_gerais': (f'DXF com {len(layers)} camadas e {len(doc.modelspace())} entidades. '
                                      f'Layers: {layer_info}. '
                                      '⚠️ Valores em unidades do CAD — verificar escala do projeto.')}
    except ImportError:
        return {'tipo_documento': 'desenho_cad_dxf', 'quantidades': [],
                'confianca': 'baixa', 'observacoes_gerais': 'ezdxf não instalado — pip install ezdxf'}
    except Exception as e:
        return {'tipo_documento': 'desenho_cad_dxf', 'quantidades': [],
                'confianca': 'baixa', 'observacoes_gerais': f'Erro ao ler DXF: {e}'}


def _analyze_via_vision(data: bytes, mime: str, fname: str) -> dict:
    """Envia imagem ou PDF ao Claude Vision e extrai quantitativos."""
    b64 = base64.b64encode(data).decode()
    source = {'type': 'base64', 'media_type': mime, 'data': b64}
    content_block = ({'type': 'document', 'source': source}
                     if mime == 'application/pdf'
                     else {'type': 'image', 'source': source})
    raw_text = ''
    try:
        raw_text = _call_claude_ia([{
            'role': 'user',
            'content': [content_block,
                        {'type': 'text', 'text': f'Arquivo: {fname}\n\n{_PROMPT_EXTRAIR}'}]
        }], max_tokens=3000)
        result = _clean_json(raw_text)
        # Se retornou "invalido" mesmo com quantidades potenciais, guardar o raw
        if result.get('tipo_documento') == 'invalido':
            result['resposta_bruta'] = raw_text[:500]
        return result
    except _json.JSONDecodeError as e:
        # Salvar a resposta bruta para debug visível no UI
        return {
            'tipo_documento': 'erro_parse',
            'quantidades': [],
            'confianca': 'baixa',
            'observacoes_gerais': (
                f'A IA respondeu mas o JSON estava malformado. '
                f'Erro: {e}. '
                f'Resposta (primeiros 300 chars): {raw_text[:300]}'
            )
        }
    except Exception as e:
        return {
            'tipo_documento': 'erro_api',
            'quantidades': [],
            'confianca': 'baixa',
            'observacoes_gerais': f'Erro na chamada à API: {type(e).__name__}: {e}'
        }


# ── Worker de análise ──────────────────────────────────────────────────────────

def _worker_analise(job_id: str, id_obra: int, files_info: list, db_path: str):
    def upd(etapa, prog):
        _ANALISE_JOBS[job_id].update({'etapa': etapa, 'progresso': prog})

    try:
        _ANALISE_JOBS[job_id] = {'status':'processando','progresso':0,
                                  'etapa':'Iniciando análise...','resultado':None,'erro':None}

        # ── 1. Extrair quantitativos de cada arquivo ───────────────────────
        brutos, n = [], max(len(files_info), 1)
        for i, (fname, fdata, ftype) in enumerate(files_info):
            upd(f'Analisando: {fname}', int(i / n * 50))
            if   ftype == 'ifc':               r = _parse_ifc(fdata)
            elif ftype == 'dxf':               r = _parse_dxf(fdata)
            elif ftype in ('png','jpg','jpeg'): r = _analyze_via_vision(fdata, f'image/{"jpeg" if ftype != "png" else "png"}', fname)
            elif ftype == 'pdf':               r = _analyze_via_vision(fdata, 'application/pdf', fname)
            else: continue
            r['arquivo'] = fname
            brutos.append(r)

        # ── 2. Consolidar texto de quantitativos ───────────────────────────
        upd('Consolidando quantitativos...', 55)
        linhas = []
        for r in brutos:
            linhas.append(f'\n=== {r.get("arquivo","?")} [{r.get("tipo_documento","?")}] confiança={r.get("confianca","?")} ===')
            for q in r.get('quantidades', []):
                linhas.append(f'  • {q["servico"]}: {q["quantidade"]} {q["unidade"]}  ({q.get("observacao","")})')
            if r.get('observacoes_gerais'):
                linhas.append(f'  Obs: {r["observacoes_gerais"]}')

        sem_qtd = not any(r.get('quantidades') for r in brutos)
        if sem_qtd:
            _ANALISE_JOBS[job_id].update({
                'status': 'concluido', 'progresso': 100, 'etapa': 'Concluído',
                'resultado': {'secoes': [], 'cobertura_pct': 0,
                              'observacoes': 'Não foi possível extrair quantitativos. '
                              'Verifique se os arquivos são projetos de engenharia/arquitetura legíveis.',
                              'quantitativos_brutos': brutos}})
            return

        # ── 3. Buscar composições do banco (limitado às mais relevantes) ─────
        upd('Buscando composições no banco de dados...', 62)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row

        # Extrair palavras-chave dos quantitativos para filtrar composições relevantes
        palavras_qtd = ' '.join(linhas).lower()
        keywords = []
        mapa = {
            'parede|alvenaria|bloco':           'alvenaria',
            'laje|concreto|concretagem':        'concretagem',
            'fundaç|sapata|estaca|baldrame':    'fundação',
            'cobertura|telha|telhamento':       'cobertura',
            'hidráulico|agua|tubulação|esgoto': 'hidráulico',
            'elétrico|eletroduto|tomada':       'elétrico',
            'revestimento|cerâmica|pintura':    'revestimento',
            'piso|contrapiso':                  'piso',
            'esquadria|porta|janela':           'esquadria',
            'escavaç|terraplenagem':            'terraplenagem',
        }
        for pat, kw in mapa.items():
            if any(p in palavras_qtd for p in pat.split('|')):
                keywords.append(kw)

        # Buscar composições por relevância (filtro por palavras-chave se houver)
        if keywords:
            like_clause = ' OR '.join(f"LOWER(descricao) LIKE '%{k}%'" for k in keywords)
            comps = [dict(r) for r in conn.execute(f"""
                SELECT id_composicao, codigo, fonte,
                       SUBSTR(descricao, 1, 80) AS descricao,
                       unidade, COALESCE(custo_unitario,0) AS custo_unitario
                FROM composicoes WHERE situacao='Ativo'
                  AND ({like_clause})
                ORDER BY fonte, id_composicao LIMIT 150
            """).fetchall()]
            # Completar com genéricas se poucos resultados
            if len(comps) < 50:
                ids_ja = tuple(c['id_composicao'] for c in comps) or (0,)
                extra = [dict(r) for r in conn.execute(f"""
                    SELECT id_composicao, codigo, fonte,
                           SUBSTR(descricao, 1, 80) AS descricao,
                           unidade, COALESCE(custo_unitario,0) AS custo_unitario
                    FROM composicoes WHERE situacao='Ativo'
                      AND id_composicao NOT IN ({','.join('?'*len(ids_ja))})
                    ORDER BY fonte, id_composicao LIMIT {150 - len(comps)}
                """, list(ids_ja)).fetchall()]
                comps += extra
        else:
            comps = [dict(r) for r in conn.execute("""
                SELECT id_composicao, codigo, fonte,
                       SUBSTR(descricao, 1, 80) AS descricao,
                       unidade, COALESCE(custo_unitario,0) AS custo_unitario
                FROM composicoes WHERE situacao='Ativo'
                ORDER BY fonte, id_composicao LIMIT 150
            """).fetchall()]
        conn.close()

        if not comps:
            _ANALISE_JOBS[job_id].update({
                'status': 'concluido', 'progresso': 100, 'etapa': 'Concluído',
                'resultado': {'secoes': [], 'cobertura_pct': 0,
                              'observacoes': 'Nenhuma composição ativa encontrada no banco. '
                              'Importe composições do SINAPI/SICRO antes de usar a análise por IA.',
                              'quantitativos_brutos': brutos}})
            return

        # ── 4. Claude faz o matching ──────────────────────────────────────
        upd(f'IA gerando orçamento ({len(comps)} composições disponíveis)...', 70)
        comps_txt = '\n'.join(
            f'{c["id_composicao"]}|{c["codigo"]}|{c["descricao"]}|{c["unidade"]}|{c["custo_unitario"]:.2f}|{c["fonte"]}'
            for c in comps
        )
        prompt = _PROMPT_MATCH.format(
            quantidades='\n'.join(linhas),
            composicoes=comps_txt
        )
        resp = _call_claude_ia([{'role': 'user', 'content': prompt}], max_tokens=6000)
        resultado = _clean_json(resp)
        resultado['quantitativos_brutos'] = brutos

        upd('Concluído!', 100)
        _ANALISE_JOBS[job_id].update({'status':'concluido','progresso':100,
                                       'etapa':'Concluído','resultado':resultado})

    except EnvironmentError as e:
        _ANALISE_JOBS[job_id].update({'status':'erro_config','progresso':0,
                                       'etapa':'Configuração necessária','erro':str(e)})
    except Exception as e:
        import traceback
        _ANALISE_JOBS[job_id].update({'status':'erro','progresso':0,
                                       'etapa':'Erro na análise',
                                       'erro': str(e),
                                       'detalhe': traceback.format_exc()})


# ── Rotas ──────────────────────────────────────────────────────────────────────

@app.route('/api/obras/<int:id_obra>/analisar-projetos', methods=['POST'])
def analisar_projetos(id_obra):
    arquivos = request.files.getlist('arquivo')
    if not arquivos:
        return jsonify({'erro': 'Nenhum arquivo enviado.'}), 400
    if len(arquivos) > _MAX_FILES:
        return jsonify({'erro': f'Máximo de {_MAX_FILES} arquivos por análise.'}), 400

    files_info = []
    for f in arquivos:
        if not f.filename: continue
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in _FORMATOS_OK:
            return jsonify({'erro': f'Formato não suportado: "{f.filename}". '
                                    f'Formatos aceitos: {", ".join(sorted(_FORMATOS_OK)).upper()}'}), 400
        files_info.append((f.filename, f.read(), ext))

    if not files_info:
        return jsonify({'erro': 'Nenhum arquivo válido.'}), 400

    job_id = str(uuid.uuid4())
    _ANALISE_JOBS[job_id] = {'status':'aguardando','progresso':0,'etapa':'Na fila...','resultado':None,'erro':None}
    threading.Thread(target=_worker_analise,
                     args=(job_id, id_obra, files_info, DB_PATH),
                     daemon=True).start()
    return jsonify({'job_id': job_id})


@app.route('/api/analise/<job_id>', methods=['GET'])
def analise_get(job_id):
    job = _ANALISE_JOBS.get(job_id)
    if not job: return jsonify({'erro': 'Análise não encontrada ou expirada.'}), 404
    return jsonify(job)


@app.route('/api/obras/<int:id_obra>/orcamento-ia', methods=['POST'])
def criar_orcamento_ia(id_obra):
    d = request.json or {}
    secoes = d.get('secoes', [])
    if not secoes:
        return jsonify({'erro': 'Nenhuma seção para criar.'}), 400

    nome = d.get('nome_orcamento', 'Orçamento — Gerado por IA').strip()
    db   = get_db()

    # Criar o orçamento pai
    cur = db.execute("""INSERT INTO orcamentos
        (id_obra, nome_orcamento, descricao, status, versao)
        VALUES (?,?,?,?,?)""",
        [id_obra, nome,
         'Rascunho gerado automaticamente por IA a partir de análise de projetos. '
         'Revisar e validar todos os itens antes de aprovar.',
         'Em elaboração', '1.0-IA'])
    db.commit()
    id_orc = cur.lastrowid

    # Criar itens do orçamento sintético
    ordem = 0
    total_itens = 0
    for sec_i, sec in enumerate(secoes):
        ordem += 1
        sec_num = str(sec_i + 1)
        db.execute("""INSERT INTO orcamento_sintetico
            (id_orcamento,item_num,tipo_linha,profundidade,ordem,descricao)
            VALUES (?,?,?,?,?,?)""",
            [id_orc, sec_num, 'section', 0, ordem, sec['descricao'].upper()])

        for it_j, it in enumerate(sec.get('itens', [])):
            ordem += 1; total_itens += 1
            db.execute("""INSERT INTO orcamento_sintetico
                (id_orcamento,item_num,tipo_linha,profundidade,ordem,
                 tipo_item,id_composicao,codigo,fonte,descricao,unidade,quantidade,custo_unitario)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [id_orc, f'{sec_num}.{it_j+1}', 'item', 1, ordem,
                 'composicao', it.get('id_composicao') or None,
                 it.get('codigo',''), it.get('fonte',''),
                 it.get('descricao',''), it.get('unidade',''),
                 float(it.get('quantidade') or 0),
                 float(it.get('custo_unitario') or 0)])

    db.commit()
    db.close()
    return jsonify({'id_orcamento': id_orc, 'total_itens': total_itens,
                    'mensagem': f'Orçamento criado com {total_itens} item(ns) em {len(secoes)} seção(ões).'})


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 9 — Demonstrativos de Produções Horárias (SICRO)
# ═══════════════════════════════════════════════════════════════════════════════

SEL_PEM = """
    SELECT s.*, COUNT(e.id_pem_equip) AS qtd_equipamentos
    FROM pem_servicos s
    LEFT JOIN pem_equipamentos e ON e.id_pem = s.id_pem
"""

@app.route('/api/pem/stats', methods=['GET'])
def pem_stats():
    db = get_db()
    result = {
        'total_servicos':    db.execute("SELECT COUNT(*) FROM pem_servicos").fetchone()[0],
        'total_equipamentos':db.execute("SELECT COUNT(*) FROM pem_equipamentos").fetchone()[0],
        'total_variaveis':   db.execute("SELECT COUNT(*) FROM pem_variaveis").fetchone()[0],
        'com_formula':       db.execute("SELECT COUNT(*) FROM pem_equipamentos WHERE formula != '' AND formula IS NOT NULL").fetchone()[0],
        'com_ligacao_sicro': db.execute("""
            SELECT COUNT(DISTINCT s.id_pem) FROM pem_servicos s
            INNER JOIN composicoes c ON c.codigo = s.codigo AND c.fonte = 'SICRO'
        """).fetchone()[0],
    }
    db.close()
    return jsonify(result)


@app.route('/api/pem', methods=['GET'])
def pem_list():
    q      = request.args.get('q', '')
    limit  = int(request.args.get('limit', 50))
    offset = int(request.args.get('offset', 0))

    sql    = SEL_PEM + " WHERE 1=1"
    params = []
    if q:
        sql += " AND (s.codigo LIKE ? OR s.servico LIKE ?)"
        like = f'%{q}%'; params += [like, like]
    sql += " GROUP BY s.id_pem ORDER BY s.codigo"
    count_sql = f"SELECT COUNT(*) FROM pem_servicos s WHERE 1=1"
    count_params = []
    if q:
        count_sql += " AND (s.codigo LIKE ? OR s.servico LIKE ?)"; count_params += [like, like]

    db    = get_db()
    total = db.execute(count_sql, count_params).fetchone()[0]
    rows  = rows_to_list(db.execute(sql + f" LIMIT {limit} OFFSET {offset}", params).fetchall())
    db.close()
    return jsonify({'total': total, 'items': rows})


@app.route('/api/pem/<int:id>', methods=['GET'])
def pem_get(id):
    db  = get_db()
    s   = db.execute(SEL_PEM + " WHERE s.id_pem=? GROUP BY s.id_pem", [id]).fetchone()
    if not s:
        db.close(); return jsonify({'erro': 'Demonstrativo não encontrado.'}), 404
    result = dict(s)

    equips = rows_to_list(db.execute(
        "SELECT * FROM pem_equipamentos WHERE id_pem=? ORDER BY ordem", [id]).fetchall())
    for eq in equips:
        eq['variaveis'] = rows_to_list(db.execute(
            "SELECT * FROM pem_variaveis WHERE id_pem_equip=? ORDER BY letra",
            [eq['id_pem_equip']]).fetchall())
    result['equipamentos'] = equips

    # Verificar se há composição SICRO vinculada
    comp = db.execute("""
        SELECT id_composicao, codigo, descricao
        FROM composicoes
        WHERE fonte='SICRO' AND (codigo=? OR codigo=?)
        LIMIT 1
    """, [result['codigo'], 'SICRO.' + str(result['codigo'])]).fetchone()
    result['composicao_vinculada'] = dict(comp) if comp else None

    db.close()
    return jsonify(result)


def _pem_producao_equipe(servico, equips):
    candidatos = []
    for eq in equips:
        ph = float(eq.get('producao_horaria') or 0)
        nu = float(eq.get('num_unidades') or 0) or 1
        if ph > 0:
            candidatos.append(ph * nu)
    if candidatos:
        return min(candidatos)
    return float(servico.get('producao_equipe') or 0)


def _preco_insumo_por_uf_data(db, codigo, uf, id_data_base):
    if not codigo or not uf or not id_data_base:
        return None
    codigos = [codigo]
    if codigo.startswith('SICRO.'):
        codigos.append(codigo.replace('SICRO.', '', 1))
    else:
        codigos.append('SICRO.' + codigo)
    row = db.execute(f"""
        SELECT COALESCE(p.preco_desonerado, p.preco_nao_desonerado, p.preco_referencia) AS preco
        FROM precos_insumos p
        JOIN insumos i ON i.id_insumo = p.id_insumo
        WHERE i.codigo_insumo IN ({','.join(['?'] * len(codigos))})
          AND p.uf_referencia=?
          AND p.id_data_base=?
        ORDER BY CASE WHEN i.origem='SICRO' THEN 0 ELSE 1 END
        LIMIT 1
    """, codigos + [uf, id_data_base]).fetchone()
    return float(row['preco']) if row and row['preco'] is not None else None


@app.route('/api/pem/<int:id>/criar-composicao-usuario', methods=['POST'])
def pem_criar_composicao_usuario(id):
    d = request.json or {}
    uf = (d.get('uf') or '').strip().upper()
    id_data_base = d.get('id_data_base')
    if not uf:
        return jsonify({'erro': 'UF é obrigatória.'}), 400
    if not id_data_base:
        return jsonify({'erro': 'Data-base é obrigatória.'}), 400

    db = get_db()
    try:
        serv = db.execute("SELECT * FROM pem_servicos WHERE id_pem=?", [id]).fetchone()
        if not serv:
            return jsonify({'erro': 'Demonstrativo não encontrado.'}), 404
        serv = dict(serv)

        data_base = db.execute("SELECT * FROM datas_base WHERE id_data_base=?", [id_data_base]).fetchone()
        if not data_base:
            return jsonify({'erro': 'Data-base não encontrada.'}), 404
        mes_ref = f"{int(data_base['mes']):02d}/{int(data_base['ano'])}"

        base = db.execute("""
            SELECT * FROM composicoes
            WHERE fonte='SICRO' AND (codigo=? OR codigo=?)
            LIMIT 1
        """, [serv['codigo'], 'SICRO.' + str(serv['codigo'])]).fetchone()
        if not base:
            return jsonify({'erro': 'Composição SICRO vinculada não encontrada.'}), 404
        base = dict(base)

        equips = rows_to_list(db.execute(
            "SELECT * FROM pem_equipamentos WHERE id_pem=? ORDER BY ordem", [id]).fetchall())
        overrides = d.get('equipamentos') or []
        if isinstance(overrides, list):
            by_id = {int(eq.get('id_pem_equip')): eq for eq in equips if eq.get('id_pem_equip') is not None}
            for ov in overrides:
                if not isinstance(ov, dict):
                    continue
                try:
                    eq = by_id.get(int(ov.get('id_pem_equip')))
                except Exception:
                    eq = None
                if not eq:
                    continue
                for campo in ['producao_horaria', 'num_unidades',
                              'utilizacao_operativa', 'utilizacao_improdutiva']:
                    if campo in ov and ov[campo] is not None:
                        eq[campo] = float(ov[campo] or 0)
        eq_by_codigo = {str(eq.get('codigo_equip') or '').strip(): eq for eq in equips}
        producao_equipe = _pem_producao_equipe(serv, equips)
        if producao_equipe <= 0:
            return jsonify({'erro': 'Produção da equipe inválida para gerar a composição.'}), 400

        codigo_base = str(serv['codigo'])
        codigo_novo = 'USUARIO.' + codigo_base
        sufixo = 2
        while db.execute("SELECT 1 FROM composicoes WHERE codigo=?", [codigo_novo]).fetchone():
            codigo_novo = f'USUARIO.{codigo_base}-{sufixo}'
            sufixo += 1

        cur = db.execute("""
            INSERT INTO composicoes
              (codigo, fonte, formato, descricao, unidade, id_grupo_comp,
               mes_referencia, uf_referencia, situacao_ref, custo_unitario,
               fic, producao_equipe, unidade_producao, situacao, observacoes)
            VALUES (?, 'USUARIO', 'PRODUCAO_HORARIA', ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, 'Ativo', ?)
        """, [
            codigo_novo,
            base.get('descricao') or serv.get('servico'),
            base.get('unidade') or serv.get('unidade'),
            base.get('id_grupo_comp'),
            mes_ref,
            uf,
            base.get('situacao_ref'),
            base.get('fic'),
            producao_equipe,
            base.get('unidade_producao') or serv.get('unidade'),
            f'Criada a partir do demonstrativo de produção horária {serv["codigo"]}. '
            f'UF {uf}, data-base {mes_ref}.',
        ])
        id_novo = cur.lastrowid

        total_geral = 0.0
        secoes_base = rows_to_list(db.execute(
            "SELECT * FROM composicoes_secoes WHERE id_composicao=? ORDER BY ordem",
            [base['id_composicao']]).fetchall())
        for sec in secoes_base:
            itens_base = rows_to_list(db.execute("""
                SELECT * FROM composicoes_secao_itens
                WHERE id_secao=?
                ORDER BY ordem
            """, [sec['id_secao']]).fetchall())

            sec_total = 0.0
            id_sec_nova = db.execute("""
                INSERT INTO composicoes_secoes
                  (id_composicao, letra_secao, nome_secao, ordem, custo_total_secao)
                VALUES (?, ?, ?, ?, 0)
            """, [id_novo, sec['letra_secao'], sec.get('nome_secao'), sec.get('ordem')]).lastrowid

            for it in itens_base:
                novo = dict(it)
                novo.update({'id_composicao': id_novo, 'id_secao': id_sec_nova})

                if sec['letra_secao'] == 'A':
                    eq = eq_by_codigo.get(str(it.get('codigo_item') or '').strip())
                    if eq:
                        novo['quantidade'] = float(eq.get('num_unidades') or it.get('quantidade') or 0)
                        novo['util_operativa'] = float(eq.get('utilizacao_operativa') or 0)
                        novo['util_improdutiva'] = float(eq.get('utilizacao_improdutiva') or 0)

                    custo_hp = it.get('custo_hp')
                    custo_hi = it.get('custo_hi')
                    eq_custo = db.execute("""
                        SELECT custo_produtivo, custo_improdutivo
                        FROM equipamentos_sinapi
                        WHERE sistema='SICRO' AND codigo_chp=?
                        LIMIT 1
                    """, [it.get('codigo_item')]).fetchone()
                    if eq_custo:
                        custo_hp = eq_custo['custo_produtivo'] if eq_custo['custo_produtivo'] is not None else custo_hp
                        custo_hi = eq_custo['custo_improdutivo'] if eq_custo['custo_improdutivo'] is not None else custo_hi
                    novo['custo_hp'] = custo_hp
                    novo['custo_hi'] = custo_hi
                    novo['custo_total'] = (
                        float(novo.get('quantidade') or 0) *
                        ((float(novo.get('util_operativa') or 0) * float(custo_hp or 0)) +
                         (float(novo.get('util_improdutiva') or 0) * float(custo_hi or 0)))
                    )
                else:
                    preco = _preco_insumo_por_uf_data(db, it.get('codigo_item'), uf, id_data_base)
                    if preco is not None and preco > 0:
                        novo['preco_unitario'] = preco
                    novo['custo_total'] = float(novo.get('quantidade') or 0) * float(novo.get('preco_unitario') or 0)

                sec_total += float(novo.get('custo_total') or 0)
                db.execute("""
                    INSERT INTO composicoes_secao_itens
                      (id_composicao, id_secao, letra_secao, codigo_item, descricao,
                       quantidade, unidade, util_operativa, util_improdutiva,
                       custo_hp, custo_hi, preco_unitario, custo_total,
                       cod_transporte, cod_transp_ln, cod_transp_rp, cod_transp_p,
                       fit, dmt, ordem)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, [
                    id_novo, id_sec_nova, novo.get('letra_secao'), novo.get('codigo_item'),
                    novo.get('descricao'), novo.get('quantidade'), novo.get('unidade'),
                    novo.get('util_operativa'), novo.get('util_improdutiva'),
                    novo.get('custo_hp'), novo.get('custo_hi'), novo.get('preco_unitario'),
                    novo.get('custo_total'), novo.get('cod_transporte'), novo.get('cod_transp_ln'),
                    novo.get('cod_transp_rp'), novo.get('cod_transp_p'), novo.get('fit'),
                    novo.get('dmt'), novo.get('ordem'),
                ])

            total_geral += sec_total
            db.execute("UPDATE composicoes_secoes SET custo_total_secao=? WHERE id_secao=?",
                       [sec_total, id_sec_nova])

        custo_unitario = total_geral / producao_equipe if producao_equipe else 0
        db.execute("UPDATE composicoes SET custo_unitario=? WHERE id_composicao=?",
                   [custo_unitario, id_novo])
        db.commit()

        row = db.execute("SELECT * FROM composicoes WHERE id_composicao=?", [id_novo]).fetchone()
        return jsonify({
            'mensagem': f'Composição {codigo_novo} criada com sucesso.',
            'composicao': dict(row),
        }), 201
    except Exception as e:
        db.rollback()
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1000:]}), 500
    finally:
        db.close()


@app.route('/api/pem/equipamentos/<int:id>', methods=['PUT'])
def pem_equip_update(id):
    d  = request.json or {}
    db = get_db()
    CAMPOS = ['producao_horaria','num_unidades','utilizacao_operativa','utilizacao_improdutiva','formula']
    sets, vals = [], []
    for c in CAMPOS:
        if c in d: sets.append(f"{c}=?"); vals.append(d[c])
    if not sets:
        db.close(); return jsonify({'erro': 'Nenhum campo para atualizar.'}), 400
    vals.append(id)
    db.execute(f"UPDATE pem_equipamentos SET {','.join(sets)} WHERE id_pem_equip=?", vals)
    db.commit()
    row = db.execute("SELECT * FROM pem_equipamentos WHERE id_pem_equip=?", [id]).fetchone()
    db.close()
    if not row: return jsonify({'erro': 'Equipamento não encontrado.'}), 404
    return jsonify(dict(row))


@app.route('/api/pem/equipamentos/<int:id>/variaveis', methods=['PUT'])
def pem_vars_update(id):
    """Atualiza em lote todas as variáveis de um equipamento."""
    variaveis = request.json or []   # [{letra, valor}, ...]
    db = get_db()
    for var in variaveis:
        letra = var.get('letra', '').lower()
        valor = var.get('valor')
        if not letra: continue
        if valor is None:
            db.execute("DELETE FROM pem_variaveis WHERE id_pem_equip=? AND letra=?", [id, letra])
        else:
            nome  = var.get('nome_variavel', '')
            unid  = var.get('unidade', '')
            db.execute("""INSERT OR REPLACE INTO pem_variaveis
                (id_pem_equip, letra, nome_variavel, unidade, valor)
                VALUES (?,?,?,?,?)""", [id, letra, nome, unid, float(valor)])
    db.commit()
    rows = rows_to_list(db.execute(
        "SELECT * FROM pem_variaveis WHERE id_pem_equip=? ORDER BY letra", [id]).fetchall())
    db.close()
    return jsonify(rows)



# ═══════════════════════════════════════════════════════════════════════════════
# EVENTOGRAMA — Tabela de Eventos Geradores de Pagamento
# ═══════════════════════════════════════════════════════════════════════════════

_GRUPOS_EVT = [
    ('Serviços Preliminares',        ['preliminar','mobilizaç','desmobiliz','canteiro','sondagem','levantamento topograf','topograf','licença','placa de obra','tapume','demoliç','limpeza de terreno','cercament']),
    ('Fundações',                    ['fundaç','estaca','sapata','radier','baldrame','tubulão','microestaca','cortina de estacas','estacamento']),
    ('Estrutura',                    ['estrutura','concreto','pilar','viga','laje','armação','armadura','escoramento','steel frame','wood frame','pré-moldado','pré moldado','estrutural']),
    ('Alvenaria e Vedação',          ['alvenaria','bloco','tijolo','vedaç','parede de','divisória','drywall','parede seca']),
    ('Cobertura e Impermeabilização',['cobertura','telhado','telha','cumeeira','calha','rufo','estrutura de madeira','tesoura','impermeabiliz','manta','membrana']),
    ('Instalações Hidrossanitárias', ['hidráulic','sanitári','esgoto','água fria','água quente','caixa d\'água','cisterna','fossa','poço','vaso sanit','lavatório','tanque','torneira','chuveiro','sifão']),
    ('Instalações Elétricas',        ['elétric','eletric','luminária','tomada','interruptor','quadro de distribuição','quadro elétric','spda','para-raio','cabeamento','fiação','conduíte','eletroduto']),
    ('Instalações Especiais',        ['ar condicionado','climatizaç','incêndio','alarme','automação','cabeamento estruturado','cftv','interfone','elevador','escada rolante','gás','solar','fotovoltaic']),
    ('Revestimentos Internos',       ['revestimento int','reboco','chapisco','emboço','gesso','contrapiso','regulariz','azulejo','cerâmica','porcelanato','forro']),
    ('Revestimentos Externos',       ['revestimento ext','fachada','pastilha','granito','mármore','pedra','ACM','textura','granitina']),
    ('Esquadrias',                   ['esquadria','porta','janela','vidro','vidraça','caixilho','veneziana','portão','guarda-corpo','corrimão','gradil']),
    ('Pintura',                      ['pintura','tinta','verniz','massa corrida','selador','textura de parede','lixamento','primer','esmalte','látex']),
    ('Pavimentação e Urbanização',   ['paviment','calçada','piso externo','passeio','paisagism','jardinagem','cerca','muro de arrimo','urbanizaç','meio-fio','sarjeta','drenagem']),
    ('Limpeza e Entrega',            ['limpeza final','entrega','desmontagem','remoção de entulho','limpeza geral','as built','as-built']),
]

def _classificar_grupo_evg(descricao):
    d = (descricao or '').lower()
    for grupo, kws in _GRUPOS_EVT:
        if any(kw in d for kw in kws):
            return grupo
    return 'Outros Serviços'

def _calc_valor_item_evg(item, bdi_pct):
    qt  = float(item.get('quantidade') or 0)
    cu  = float(item.get('custo_unitario') or 0)
    bdi = 1 + float(bdi_pct or 0) / 100
    return round(qt * cu * bdi, 2)

def _recalc_evento_evg(db, id_evento, bdi_pct):
    itens = db.execute("""
        SELECT s.quantidade, s.custo_unitario
        FROM ev_evento_itens ei
        JOIN orcamento_sintetico s ON s.id_item=ei.id_item
        WHERE ei.id_evento=?""", [id_evento]).fetchall()
    valor = sum(_calc_valor_item_evg(dict(i), bdi_pct) for i in itens)
    subs  = db.execute("SELECT valor_calculado FROM ev_eventos WHERE id_evento_pai=?", [id_evento]).fetchall()
    valor += sum(float(r[0] or 0) for r in subs)
    db.execute("UPDATE ev_eventos SET valor_calculado=? WHERE id_evento=?", [valor, id_evento])
    return valor

def _get_bdi_evg(db, id_orcamento):
    r = db.execute("SELECT bdi_percentual FROM orcamentos WHERE id_orcamento=?", [id_orcamento]).fetchone()
    return float(r[0] or 0) if r else 0.0

def _get_all_eventos_evg(db, id_eventograma):
    eventos = rows_to_list(db.execute("""
        SELECT e.*, COUNT(ei.id) AS qtd_itens
        FROM ev_eventos e
        LEFT JOIN ev_evento_itens ei ON ei.id_evento=e.id_evento
        WHERE e.id_eventograma=?
        GROUP BY e.id_evento
        ORDER BY e.ordem, e.numero_evento
    """, [id_eventograma]).fetchall())
    for ev in eventos:
        ev['itens'] = rows_to_list(db.execute("""
            SELECT ei.id AS id_vinculo, s.*
            FROM ev_evento_itens ei
            JOIN orcamento_sintetico s ON s.id_item=ei.id_item
            WHERE ei.id_evento=?
            ORDER BY s.ordem
        """, [ev['id_evento']]).fetchall())
        ev['subeventos'] = []
    idx = {e['id_evento']: e for e in eventos}
    raiz = []
    for e in eventos:
        pai = e.get('id_evento_pai')
        if pai and pai in idx:
            idx[pai]['subeventos'].append(e)
        else:
            raiz.append(e)
    return raiz

@app.route('/api/eventogramas', methods=['GET'])
def evg_list():
    id_orc = request.args.get('id_orcamento','')
    db = get_db()
    q = """SELECT eg.*, o.nome_orcamento, o.valor_total, ob.nome_obra,
                  COUNT(DISTINCT ev.id_evento) AS qtd_eventos
           FROM eventogramas eg
           JOIN orcamentos o ON o.id_orcamento=eg.id_orcamento
           JOIN obras ob ON ob.id_obra=o.id_obra
           LEFT JOIN ev_eventos ev ON ev.id_eventograma=eg.id_eventograma AND ev.id_evento_pai IS NULL
           WHERE 1=1"""
    params = []
    if id_orc:
        q += " AND eg.id_orcamento=?"; params.append(id_orc)
    q += " GROUP BY eg.id_eventograma ORDER BY eg.data_criacao DESC"
    rows = rows_to_list(db.execute(q, params).fetchall())
    db.close()
    return jsonify(rows)

@app.route('/api/eventogramas', methods=['POST'])
def evg_create():
    d = request.json or {}
    id_orc = d.get('id_orcamento')
    if not id_orc: return jsonify({'erro':'id_orcamento obrigatório.'}), 400
    db = get_db()
    orc = db.execute("SELECT * FROM orcamentos WHERE id_orcamento=?", [id_orc]).fetchone()
    if not orc: db.close(); return jsonify({'erro':'Orçamento não encontrado.'}), 404
    cur = db.execute("""
        INSERT INTO eventogramas (id_orcamento,nome,descricao,modo_geracao,status,valor_total_ref,observacoes)
        VALUES (?,?,?,?,?,?,?)""",
        [id_orc, d.get('nome','Eventograma'), d.get('descricao'),
         d.get('modo_geracao','manual'), 'Rascunho',
         float(orc['valor_total'] or 0), d.get('observacoes')])
    db.commit()
    row = dict(db.execute("SELECT * FROM eventogramas WHERE id_eventograma=?", [cur.lastrowid]).fetchone())
    db.close()
    return jsonify(row), 201

@app.route('/api/eventogramas/<int:id>', methods=['GET'])
def evg_get(id):
    db = get_db()
    row = db.execute("""
        SELECT eg.*, o.nome_orcamento, o.valor_total, o.bdi_percentual, ob.nome_obra, ob.id_obra
        FROM eventogramas eg
        JOIN orcamentos o ON o.id_orcamento=eg.id_orcamento
        JOIN obras ob ON ob.id_obra=o.id_obra
        WHERE eg.id_eventograma=?""", [id]).fetchone()
    if not row: db.close(); return jsonify({'erro':'Eventograma não encontrado.'}), 404
    result = dict(row)
    result['eventos'] = _get_all_eventos_evg(db, id)
    bdi = float(row['bdi_percentual'] or 0)
    itens = rows_to_list(db.execute("""
        SELECT s.* FROM orcamento_sintetico s WHERE s.id_orcamento=?
        ORDER BY s.ordem, s.id_item""", [result['id_orcamento']]).fetchall())
    alocados = {r[0] for r in db.execute("""
        SELECT DISTINCT ei.id_item FROM ev_evento_itens ei
        JOIN ev_eventos ev ON ev.id_evento=ei.id_evento
        WHERE ev.id_eventograma=?""", [id]).fetchall()}
    for it in itens:
        it['alocado'] = it['id_item'] in alocados
        it['valor']   = _calc_valor_item_evg(it, bdi)
    result['itens_orcamento'] = itens
    result['bdi_percentual']  = bdi
    db.close()
    return jsonify(result)

@app.route('/api/eventogramas/<int:id>', methods=['PUT'])
def evg_update(id):
    d = request.json or {}
    db = get_db()
    db.execute("""UPDATE eventogramas SET nome=?,descricao=?,status=?,observacoes=?,
                  data_atualizacao=datetime('now') WHERE id_eventograma=?""",
               [d.get('nome'), d.get('descricao'), d.get('status'), d.get('observacoes'), id])
    db.commit()
    row = dict(db.execute("SELECT * FROM eventogramas WHERE id_eventograma=?", [id]).fetchone())
    db.close()
    return jsonify(row)

@app.route('/api/eventogramas/<int:id>', methods=['DELETE'])
def evg_delete(id):
    db = get_db()
    db.execute("DELETE FROM eventogramas WHERE id_eventograma=?", [id])
    db.commit(); db.close()
    return jsonify({'status':'ok'})

@app.route('/api/eventogramas/<int:id>/gerar', methods=['POST'])
def evg_gerar(id):
    d = request.json or {}
    modo   = d.get('modo','automatico')
    limpar = d.get('limpar_existentes', True)
    db = get_db()
    evg = db.execute("SELECT * FROM eventogramas WHERE id_eventograma=?", [id]).fetchone()
    if not evg: db.close(); return jsonify({'erro':'Não encontrado.'}), 404
    id_orc = evg['id_orcamento']
    bdi    = _get_bdi_evg(db, id_orc)
    itens  = rows_to_list(db.execute(
        "SELECT * FROM orcamento_sintetico WHERE id_orcamento=? ORDER BY ordem, id_item", [id_orc]
    ).fetchall())
    if limpar:
        db.execute("DELETE FROM ev_eventos WHERE id_eventograma=?", [id])
    grupos = {}
    secao_atual = None
    for it in itens:
        if it['tipo_linha'] == 'section' and it.get('profundidade',0) == 0:
            secao_atual = it['descricao']
        if it['tipo_linha'] != 'item':
            continue
        grupo = _classificar_grupo_evg(secao_atual or '')
        if grupo == 'Outros Serviços':
            grupo = _classificar_grupo_evg(it['descricao'])
        grupos.setdefault(grupo, []).append(it)
    ordem_ref = [g[0] for g in _GRUPOS_EVT] + ['Outros Serviços']
    grupos_ord = sorted(grupos.items(), key=lambda x: ordem_ref.index(x[0]) if x[0] in ordem_ref else 999)
    criados = []
    num = 1
    for grupo, items_g in grupos_ord:
        valor_g = sum(_calc_valor_item_evg(it, bdi) for it in items_g)
        cur = db.execute("""
            INSERT INTO ev_eventos (id_eventograma,numero_evento,descricao,grupo,
                criterio_medicao,valor_calculado,ordem)
            VALUES (?,?,?,?,?,?,?)""",
            [id, f'{num:02d}', grupo, grupo,
             'Medição física com base nas quantidades executadas e atestadas pelo fiscal',
             valor_g, num])
        id_ev = cur.lastrowid
        for it in items_g:
            try:
                db.execute("INSERT INTO ev_evento_itens (id_evento,id_item) VALUES (?,?)", [id_ev, it['id_item']])
            except: pass
        criados.append({'id_evento':id_ev,'grupo':grupo,'qtd_itens':len(items_g),'valor':valor_g})
        num += 1
    db.execute("UPDATE eventogramas SET modo_geracao=?,data_atualizacao=datetime('now') WHERE id_eventograma=?", [modo, id])
    db.commit(); db.close()
    return jsonify({'status':'ok','eventos_criados':len(criados),'detalhes':criados})

@app.route('/api/eventogramas/<int:id>/eventos', methods=['GET'])
def evg_eventos_list(id):
    db = get_db()
    r = _get_all_eventos_evg(db, id)
    db.close()
    return jsonify(r)

@app.route('/api/eventogramas/<int:id>/eventos', methods=['POST'])
def evg_evento_create(id):
    d = request.json or {}
    db = get_db()
    pai = d.get('id_evento_pai')
    if pai:
        max_ord = db.execute("SELECT COALESCE(MAX(ordem),0) FROM ev_eventos WHERE id_evento_pai=?", [pai]).fetchone()[0]
        num = d.get('numero_evento') or f'{max_ord+1:02d}'
    else:
        max_ord = db.execute("SELECT COALESCE(MAX(ordem),0) FROM ev_eventos WHERE id_eventograma=? AND id_evento_pai IS NULL", [id]).fetchone()[0]
        num = d.get('numero_evento') or f'{max_ord+1:02d}'
    cur = db.execute("""
        INSERT INTO ev_eventos (id_eventograma,id_evento_pai,numero_evento,descricao,grupo,
            criterio_medicao,condicao_pagamento,prazo_marco,docs_comprobatorios,
            observacoes,valor_calculado,ordem)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        [id, pai, num, d.get('descricao','Novo Evento'), d.get('grupo'),
         d.get('criterio_medicao'), d.get('condicao_pagamento'),
         d.get('prazo_marco'), d.get('docs_comprobatorios'),
         d.get('observacoes'), 0, d.get('ordem', max_ord+1)])
    db.commit()
    row = dict(db.execute("SELECT * FROM ev_eventos WHERE id_evento=?", [cur.lastrowid]).fetchone())
    db.close()
    return jsonify(row), 201

@app.route('/api/eventogramas/<int:eid>/eventos/<int:id>', methods=['PUT'])
def evg_evento_update(eid, id):
    d = request.json or {}
    db = get_db()
    db.execute("""UPDATE ev_eventos SET numero_evento=?,descricao=?,grupo=?,
        criterio_medicao=?,condicao_pagamento=?,prazo_marco=?,
        docs_comprobatorios=?,observacoes=?,ordem=?
        WHERE id_evento=? AND id_eventograma=?""",
        [d.get('numero_evento'), d.get('descricao'), d.get('grupo'),
         d.get('criterio_medicao'), d.get('condicao_pagamento'), d.get('prazo_marco'),
         d.get('docs_comprobatorios'), d.get('observacoes'), d.get('ordem',0), id, eid])
    db.commit()
    row = dict(db.execute("SELECT * FROM ev_eventos WHERE id_evento=?", [id]).fetchone())
    db.close()
    return jsonify(row)

@app.route('/api/eventogramas/<int:eid>/eventos/<int:id>', methods=['DELETE'])
def evg_evento_delete(eid, id):
    db = get_db()
    db.execute("DELETE FROM ev_eventos WHERE id_evento=? AND id_eventograma=?", [id, eid])
    db.commit(); db.close()
    return jsonify({'status':'ok'})

@app.route('/api/eventogramas/<int:eid>/eventos/<int:id>/itens', methods=['POST'])
def evg_item_add(eid, id):
    d = request.json or {}
    ids = d.get('ids', [d.get('id_item')] if d.get('id_item') else [])
    db = get_db()
    evg = db.execute("SELECT id_orcamento FROM eventogramas WHERE id_eventograma=?", [eid]).fetchone()
    bdi = _get_bdi_evg(db, evg['id_orcamento']) if evg else 0
    ok = 0
    for item_id in ids:
        try:
            db.execute("INSERT INTO ev_evento_itens (id_evento,id_item) VALUES (?,?)", [id, item_id])
            ok += 1
        except: pass
    _recalc_evento_evg(db, id, bdi)
    ev = db.execute("SELECT id_evento_pai FROM ev_eventos WHERE id_evento=?", [id]).fetchone()
    if ev and ev[0]: _recalc_evento_evg(db, ev[0], bdi)
    db.commit(); db.close()
    return jsonify({'status':'ok','inseridos':ok})

@app.route('/api/eventogramas/<int:eid>/eventos/<int:id>/itens/<int:item_id>', methods=['DELETE'])
def evg_item_remove(eid, id, item_id):
    db = get_db()
    evg = db.execute("SELECT id_orcamento FROM eventogramas WHERE id_eventograma=?", [eid]).fetchone()
    bdi = _get_bdi_evg(db, evg['id_orcamento']) if evg else 0
    db.execute("DELETE FROM ev_evento_itens WHERE id_evento=? AND id_item=?", [id, item_id])
    _recalc_evento_evg(db, id, bdi)
    ev = db.execute("SELECT id_evento_pai FROM ev_eventos WHERE id_evento=?", [id]).fetchone()
    if ev and ev[0]: _recalc_evento_evg(db, ev[0], bdi)
    db.commit(); db.close()
    return jsonify({'status':'ok'})

@app.route('/api/eventogramas/<int:eid>/eventos/<int:id>/itens/mover', methods=['POST'])
def evg_item_mover(eid, id):
    d = request.json or {}
    destino = d.get('id_evento_destino')
    ids     = d.get('ids', [])
    if not destino or not ids: return jsonify({'erro':'Parâmetros inválidos.'}), 400
    db = get_db()
    evg = db.execute("SELECT id_orcamento FROM eventogramas WHERE id_eventograma=?", [eid]).fetchone()
    bdi = _get_bdi_evg(db, evg['id_orcamento']) if evg else 0
    for item_id in ids:
        db.execute("DELETE FROM ev_evento_itens WHERE id_evento=? AND id_item=?", [id, item_id])
        try:
            db.execute("INSERT INTO ev_evento_itens (id_evento,id_item) VALUES (?,?)", [destino, item_id])
        except: pass
    _recalc_evento_evg(db, id, bdi)
    _recalc_evento_evg(db, destino, bdi)
    db.commit(); db.close()
    return jsonify({'status':'ok'})

@app.route('/api/eventogramas/<int:id>/reordenar', methods=['POST'])
def evg_reordenar(id):
    ordens = request.json or []
    db = get_db()
    for item in ordens:
        db.execute("UPDATE ev_eventos SET ordem=?,numero_evento=? WHERE id_evento=? AND id_eventograma=?",
                   [item.get('ordem'), item.get('numero_evento'), item.get('id_evento'), id])
    db.commit(); db.close()
    return jsonify({'status':'ok'})

@app.route('/api/eventogramas/<int:id>/validar', methods=['GET'])
def evg_validar(id):
    db = get_db()
    evg = db.execute("""
        SELECT eg.*, o.valor_total, o.bdi_percentual, o.id_orcamento
        FROM eventogramas eg JOIN orcamentos o ON o.id_orcamento=eg.id_orcamento
        WHERE eg.id_eventograma=?""", [id]).fetchone()
    if not evg: db.close(); return jsonify({'erro':'Não encontrado.'}), 404
    vt_orc = float(evg['valor_total'] or 0)
    alertas = []
    todos_itens = {r[0] for r in db.execute("""
        SELECT id_item FROM orcamento_sintetico WHERE id_orcamento=? AND tipo_linha='item'
    """, [evg['id_orcamento']]).fetchall()}
    alocados_raw = db.execute("""
        SELECT ei.id_item, COUNT(*) FROM ev_evento_itens ei
        JOIN ev_eventos ev ON ev.id_evento=ei.id_evento
        WHERE ev.id_eventograma=? GROUP BY ei.id_item""", [id]).fetchall()
    alocados_cnt = {r[0]: r[1] for r in alocados_raw}
    ids_aloc  = set(alocados_cnt.keys())
    nao_aloc  = todos_itens - ids_aloc
    duplicados = {k for k,v in alocados_cnt.items() if v > 1}
    if nao_aloc:
        alertas.append({'tipo':'warning','msg':f'{len(nao_aloc)} item(ns) não alocado(s) em nenhum evento.'})
    if duplicados:
        alertas.append({'tipo':'error','msg':f'{len(duplicados)} item(ns) com alocação duplicada.'})
    eventos = db.execute("""
        SELECT ev.id_evento, ev.numero_evento, ev.descricao, ev.valor_calculado, COUNT(ei.id) AS qtd
        FROM ev_eventos ev
        LEFT JOIN ev_evento_itens ei ON ei.id_evento=ev.id_evento
        WHERE ev.id_eventograma=?
        GROUP BY ev.id_evento""", [id]).fetchall()
    sem_itens = [dict(e) for e in eventos if e['qtd'] == 0]
    if sem_itens:
        alertas.append({'tipo':'warning','msg':f'{len(sem_itens)} evento(s) sem itens vinculados.'})
    soma = sum(float(e['valor_calculado'] or 0) for e in db.execute(
        "SELECT valor_calculado FROM ev_eventos WHERE id_eventograma=? AND id_evento_pai IS NULL", [id]).fetchall())
    if vt_orc > 0 and abs(soma-vt_orc)/vt_orc > 0.001:
        alertas.append({'tipo':'error','msg':f'Soma dos eventos (R$ {soma:,.2f}) difere do total do orçamento (R$ {vt_orc:,.2f}).'})
    for ev in eventos:
        vev = float(ev['valor_calculado'] or 0)
        if vt_orc > 0 and vev/vt_orc > 0.40:
            alertas.append({'tipo':'warning','msg':f'Evento {ev["numero_evento"]} com valor elevado ({vev/vt_orc*100:.1f}% do contrato).'})
    if not list(eventos):
        alertas.append({'tipo':'error','msg':'Nenhum evento cadastrado.'})
    db.close()
    return jsonify({
        'alertas': alertas, 'total_alertas': len(alertas),
        'qtd_itens_total': len(todos_itens), 'qtd_itens_alocados': len(ids_aloc),
        'qtd_itens_nao_alocados': len(nao_aloc),
        'qtd_eventos': len(list(eventos)), 'soma_eventos': round(soma,2),
        'valor_orcamento': round(vt_orc,2),
        'percentual_alocado': round(len(ids_aloc)/len(todos_itens)*100,2) if todos_itens else 0,
    })

@app.route('/api/eventogramas/<int:id>/exportar/json', methods=['GET'])
def evg_exportar_json(id):
    import json as _json
    from flask import Response
    db = get_db()
    evg = db.execute("""
        SELECT eg.*, o.nome_orcamento, o.valor_total, o.bdi_percentual, ob.nome_obra
        FROM eventogramas eg JOIN orcamentos o ON o.id_orcamento=eg.id_orcamento
        JOIN obras ob ON ob.id_obra=o.id_obra
        WHERE eg.id_eventograma=?""", [id]).fetchone()
    if not evg: db.close(); return jsonify({'erro':'Não encontrado.'}), 404
    result = dict(evg)
    result['eventos'] = _get_all_eventos_evg(db, id)
    db.close()
    return Response(_json.dumps(result, ensure_ascii=False, indent=2),
                    mimetype='application/json',
                    headers={'Content-Disposition': f'attachment; filename=eventograma_{id}.json'})

@app.route('/api/eventogramas/<int:id>/exportar/excel', methods=['GET'])
def evg_exportar_excel(id):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    db = get_db()
    evg = db.execute("""
        SELECT eg.*, o.nome_orcamento, o.valor_total, o.bdi_percentual, ob.nome_obra
        FROM eventogramas eg JOIN orcamentos o ON o.id_orcamento=eg.id_orcamento
        JOIN obras ob ON ob.id_obra=o.id_obra
        WHERE eg.id_eventograma=?""", [id]).fetchone()
    if not evg: db.close(); return jsonify({'erro':'Não encontrado.'}), 404
    bdi    = float(evg['bdi_percentual'] or 0)
    vt_orc = float(evg['valor_total'] or 0)
    eventos_raiz = _get_all_eventos_evg(db, id)
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Eventograma'
    hdr_fill = PatternFill('solid', fgColor='0F172A')
    alt_fill = PatternFill('solid', fgColor='F8FAFF')
    sub_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    sub_font = Font(bold=True, color='FFFFFF', size=9)
    thin  = Border(bottom=Side(style='thin', color='CBD5E1'))
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(vertical='center', wrap_text=True)
    right  = Alignment(horizontal='right', vertical='center')
    cols = ['Nº','Descrição do Evento','Grupo','Itens (qtd)','Valor (R$)','% Contrato','% Acumulado','Saldo (R$)','Critério de Aceite','Documentação Exigida','Prazo/Marco','Observações']
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(1, c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    acum = 0.0; rn = 2
    def _wr(ev, sub=False):
        nonlocal acum, rn
        v = float(ev.get('valor_calculado') or 0)
        pct = v/vt_orc*100 if vt_orc else 0
        if not sub: acum += v
        ws.append([ev.get('numero_evento',''), ev.get('descricao',''), ev.get('grupo',''),
                   len(ev.get('itens',[])), v, pct,
                   acum/vt_orc*100 if vt_orc else 0, vt_orc-acum,
                   ev.get('criterio_medicao',''), ev.get('docs_comprobatorios',''),
                   ev.get('prazo_marco',''), ev.get('observacoes','')])
        fl = sub_fill if sub else (alt_fill if rn%2==0 else PatternFill())
        fn = sub_font if sub else Font(size=9)
        for c in range(1, len(cols)+1):
            cell = ws.cell(rn, c)
            cell.fill = fl; cell.font = fn; cell.border = thin
            cell.alignment = right if c in (5,6,7,8) else (center if c in (1,4) else left)
            if c == 5: cell.number_format = 'R$ #,##0.00'
            if c in (6,7): cell.number_format = '0.00'
        rn += 1
        for s in ev.get('subeventos',[]): _wr(s, True)
    for ev in eventos_raiz: _wr(ev)
    widths = [8,38,22,10,16,12,12,18,32,32,20,28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32; ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('Memória de Vinculação')
    ws2.append(['Nº Evento','Evento','Cód.','Descrição Item','Un','Qtd','Custo Unit.','Valor','Fonte'])
    for c in range(1,10):
        cell = ws2.cell(1, c); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center
    r2 = 2
    def _wm(ev):
        nonlocal r2
        for it in ev.get('itens',[]):
            v = _calc_valor_item_evg(it, bdi)
            ws2.append([ev.get('numero_evento',''), ev.get('descricao',''),
                        it.get('codigo',''), it.get('descricao',''),
                        it.get('unidade',''), it.get('quantidade',0),
                        it.get('custo_unitario',0), v, it.get('fonte','')])
            fl = alt_fill if r2%2==0 else PatternFill()
            for c in range(1,10):
                ws2.cell(r2,c).fill = fl; ws2.cell(r2,c).font = Font(size=9)
                ws2.cell(r2,c).alignment = right if c in (6,7,8) else left
                ws2.cell(r2,c).border = thin
            r2 += 1
        for s in ev.get('subeventos',[]): _wm(s)
    for ev in eventos_raiz: _wm(ev)
    for i,w in enumerate([8,32,10,42,6,10,14,14,10],1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'eventograma_{id}.xlsx')



# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTAÇÃO SINAPI — Insumos e Composições
# ═══════════════════════════════════════════════════════════════════════════════

UFS_SINAPI = ['AC','AL','AM','AP','BA','CE','DF','ES','GO','MA',
              'MG','MS','MT','PA','PB','PE','PI','PR','RJ','RN',
              'RO','RR','RS','SC','SE','SP','TO']

TIPO_SINAPI_MAP = {
    'MATERIAL':    'Material',
    'EQUIPAMENTO': 'Equipamento',
    'MÃO DE OBRA': 'Mão de Obra',
    'MAO DE OBRA': 'Mão de Obra',
    'SERVIÇOS':    'Serviço Auxiliar',
    'SERVICOS':    'Serviço Auxiliar',
    'SERVICO':     'Serviço Auxiliar',
    'SERVICO AUXILIAR': 'Serviço Auxiliar',
}

def _parse_mes_ref(xl_file):
    """Lê o mês/ano de referência do cabeçalho de qualquer aba."""
    import re
    for aba in ['ISD', 'ICD', 'Analítico', 'CSD', 'CCD']:
        try:
            df = xl_file.parse(aba, header=None, nrows=5)
            for i in range(5):
                row = df.iloc[i]
                for j, val in enumerate(row):
                    if str(val).strip().lower().startswith('mês de referência') or \
                       str(val).strip().lower().startswith('mes de referencia'):
                        # Next cell has the value
                        for k in range(j+1, min(j+5, len(row))):
                            v = str(row.iloc[k]).strip()
                            m = re.match(r'(\d{1,2})/(\d{4})', v)
                            if m:
                                return int(m.group(1)), int(m.group(2))
        except Exception:
            continue
    return None, None

def _parse_insumos_aba(xl_file, aba, desonerado: bool):
    """Lê insumos de uma aba ISD ou ICD.
    Retorna lista de dicts: {codigo, descricao, tipo, unidade, origem_preco, precos: {UF: valor}}
    """
    import re
    df = xl_file.parse(aba, header=None)

    # Linha 3 tem as UFs (col 3+)
    ufs_row = df.iloc[3]
    uf_cols = {}
    for col_idx in range(3, len(ufs_row)):
        val = str(ufs_row.iloc[col_idx]).strip()
        if val in UFS_SINAPI:
            uf_cols[val] = col_idx

    # Dados a partir da linha 10 (0-indexed)
    insumos = []
    for row_idx in range(10, len(df)):
        row = df.iloc[row_idx]
        try:
            tipo_raw = str(row.iloc[0]).strip().upper()
            codigo   = str(row.iloc[1]).strip()
            descricao= str(row.iloc[2]).strip()
            unidade  = str(row.iloc[3]).strip()
            origem   = str(row.iloc[4]).strip() if len(row) > 4 else ''

            if not codigo or codigo == 'nan' or descricao == 'nan':
                continue
            # Normalizar código: remover ".0" de floats (ex: 45333.0 → 45333)
            if codigo.endswith('.0') and codigo[:-2].isdigit():
                codigo = codigo[:-2]
            # Deve ser um código numérico
            if not re.match(r'^\d+$', codigo):
                continue

            tipo = TIPO_SINAPI_MAP.get(tipo_raw, 'Material')

            precos = {}
            for uf, col_idx in uf_cols.items():
                if col_idx < len(row):
                    v = row.iloc[col_idx]
                    try:
                        precos[uf] = float(str(v).replace(',','.')) if str(v) not in ('nan','') else None
                    except Exception:
                        precos[uf] = None

            insumos.append({
                'codigo':      codigo,
                'descricao':   descricao,
                'tipo':        tipo,
                'unidade':     unidade,
                'origem_preco': origem,
                'precos':      precos,
                'desonerado':  desonerado,
            })
        except Exception:
            continue
    return insumos

def _parse_analitico(xl_file):
    """Lê composições da aba Analítico.
    Retorna lista de dicts: {codigo, descricao, unidade, grupo, situacao, itens: [...]}
    """
    import re
    df = xl_file.parse('Analítico', header=None)

    composicoes = {}
    comp_order  = []
    current_comp = None

    for row_idx in range(10, len(df)):
        row = df.iloc[row_idx]
        try:
            grupo      = str(row.iloc[0]).strip() if str(row.iloc[0]) != 'nan' else ''
            col1       = str(row.iloc[1]).strip()  # código composição (sempre presente)
            tipo_item  = str(row.iloc[2]).strip() if str(row.iloc[2]) != 'nan' else ''
            col3       = str(row.iloc[3]).strip() if str(row.iloc[3]) != 'nan' else ''
            col4       = str(row.iloc[4]).strip() if str(row.iloc[4]) != 'nan' else ''
            col5       = str(row.iloc[5]).strip() if str(row.iloc[5]) != 'nan' else ''
            col6       = str(row.iloc[6]).strip() if str(row.iloc[6]) != 'nan' else ''
            col7       = str(row.iloc[7]).strip() if str(row.iloc[7]) != 'nan' else ''

            # Normalizar códigos (int64 vira "104658.0" ou "104658")
            def _norm_cod(s):
                s = s.strip()
                if s.endswith('.0') and s[:-2].isdigit(): return s[:-2]
                return s
            col1 = _norm_cod(col1)
            col3 = _norm_cod(col3)

            if not col1 or not re.match(r'^\d+', col1):
                continue

            # Linha de cabeçalho de composição: col[2] (tipo_item) está vazio
            if not tipo_item:
                # col1=código, col3 está vazio (é o código do item), col4=descrição, col5=unidade, col7=situação
                codigo_comp = col1
                descricao   = col4 if col4 else col3
                unidade_c   = col5
                situacao    = col7

                current_comp = {
                    'codigo':    codigo_comp,
                    'descricao': descricao,
                    'unidade':   unidade_c,
                    'grupo':     grupo,
                    'situacao':  situacao,
                    'itens':     [],
                }
                composicoes[codigo_comp] = current_comp
                comp_order.append(codigo_comp)
            else:
                # Linha de item: tipo_item = 'INSUMO' | 'COMPOSICAO'
                if current_comp is None:
                    continue
                tipo_norm = tipo_item.upper()
                if tipo_norm not in ('INSUMO', 'COMPOSICAO', 'EQUIPAMENTO'):
                    continue

                codigo_item = col3
                descricao_item = col4
                unidade_item   = col5
                try:
                    coeficiente = float(str(col6).replace(',','.')) if col6 else 0
                except Exception:
                    coeficiente = 0
                situacao_item = col7

                current_comp['itens'].append({
                    'tipo_item':   tipo_norm,
                    'codigo_item': codigo_item,
                    'descricao':   descricao_item,
                    'unidade':     unidade_item,
                    'coeficiente': coeficiente,
                    'situacao':    situacao_item,
                })
        except Exception:
            continue

    return [composicoes[c] for c in comp_order]


@app.route('/api/sinapi/analisar', methods=['POST'])
def sinapi_analisar():
    """Analisa o arquivo SINAPI e retorna metadados antes de importar."""
    import tempfile
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Arquivo não enviado.'}), 400
    f = request.files['arquivo']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.xlsm'):
        return jsonify({'erro': 'Use arquivo .xlsx ou .xls.'}), 400

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        f.save(tmp.name); tmp_path = tmp.name

    try:
        import pandas as pd
        xl = pd.ExcelFile(tmp_path)
        abas = xl.sheet_names

        # Mês de referência
        mes, ano = _parse_mes_ref(xl)

        # Quais abas existem
        tem_isd = 'ISD' in abas
        tem_icd = 'ICD' in abas
        tem_anal = 'Analítico' in abas

        # Contagem prévia
        qtd_isd  = 0
        qtd_icd  = 0
        qtd_comp = 0
        if tem_isd:
            df_isd = xl.parse('ISD', header=None)
            qtd_isd = max(0, len(df_isd) - 10)
        if tem_icd:
            df_icd = xl.parse('ICD', header=None)
            qtd_icd = max(0, len(df_icd) - 10)
        if tem_anal:
            df_anal = xl.parse('Analítico', header=None)
            qtd_comp = max(0, (df_anal.iloc[10:, 2].isna() | (df_anal.iloc[10:, 2] == '')).sum())

        # Verificar sobreposição no banco se mês/ano conhecido
        db = get_db()
        sobreposicao = {}
        if mes and ano:
            db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
            if db_row:
                id_db = db_row[0]
                cnt_ins = db.execute("""
                    SELECT COUNT(*) FROM precos_insumos pi
                    JOIN insumos i ON i.id_insumo=pi.id_insumo
                    WHERE pi.id_data_base=? AND i.origem='SINAPI'""", [id_db]).fetchone()[0]
                cnt_comp = db.execute("""
                    SELECT COUNT(*) FROM composicoes
                    WHERE mes_referencia=? AND fonte='SINAPI'""",
                    [f'{mes:02d}/{ano}']).fetchone()[0]
                if cnt_ins > 0 or cnt_comp > 0:
                    sobreposicao = {'insumos': int(cnt_ins), 'composicoes': int(cnt_comp),
                                    'id_data_base': int(id_db)}
        db.close()

        return jsonify({
            'mes': int(mes) if mes else None,
            'ano': int(ano) if ano else None,
            'abas': list(abas),
            'tem_isd': bool(tem_isd), 'tem_icd': bool(tem_icd), 'tem_analitico': bool(tem_anal),
            'qtd_insumos_isd': int(qtd_isd),
            'qtd_insumos_icd': int(qtd_icd),
            'qtd_composicoes':  int(qtd_comp),
            'sobreposicao':    sobreposicao,
        })
    except Exception as e:
        return jsonify({'erro': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


@app.route('/api/sinapi/importar', methods=['POST'])
def sinapi_importar():
    import tempfile, pandas as pd

    mes_param  = request.form.get('mes', '').strip()
    ano_param  = request.form.get('ano', '').strip()
    uf_param   = request.form.get('uf', '').strip().upper()
    imp_isd    = request.form.get('importar_isd', 'true').lower() == 'true'
    imp_icd    = request.form.get('importar_icd', 'true').lower() == 'true'
    imp_anal   = request.form.get('importar_analitico', 'true').lower() == 'true'
    sobrepor   = request.form.get('sobrepor', 'false').lower() == 'true'

    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Arquivo não enviado.'}), 400

    f   = request.files['arquivo']
    ext = os.path.splitext(f.filename)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        f.save(tmp.name); tmp_path = tmp.name

    db = None
    try:
        xl = pd.ExcelFile(tmp_path)
        mes_arq, ano_arq = _parse_mes_ref(xl)
        mes = int(mes_param) if mes_param else mes_arq
        ano = int(ano_param) if ano_param else ano_arq
        if not mes or not ano:
            return jsonify({'erro': 'Mês/ano de referência não identificado. Informe manualmente no passo anterior.'}), 400

        mes_ref_str = f'{mes:02d}/{ano}'
        db = get_db()

        # Data-base
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
        id_data_base = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)", [mes, ano, f'SINAPI {mes:02d}/{ano}']).lastrowid
        db.commit()

        # Fonte
        fn = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='SINAPI'").fetchone()
        id_fonte = fn[0] if fn else db.execute(
            "INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia) VALUES (?,?,?,?)",
            ['SINAPI','Oficial','Caixa Econômica Federal / IBGE','Nacional']).lastrowid
        db.commit()

        # Unidades cache
        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        def get_unidade(sigla):
            if not sigla or str(sigla) == 'nan': return None
            sigla = str(sigla).upper()[:20]
            if sigla not in unid_cache:
                uid = db.execute("INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)",
                                 [sigla, sigla, 'Outro']).lastrowid
                db.commit(); unid_cache[sigla] = uid
            return unid_cache[sigla]

        ufs = [uf_param] if (uf_param and uf_param != 'TODAS' and uf_param in UFS_SINAPI) else UFS_SINAPI
        resultado = {'data_base': mes_ref_str,
                     'insumos_inseridos': 0, 'insumos_atualizados': 0,
                     'precos_inseridos': 0,  'precos_atualizados': 0,
                     'composicoes_inseridas': 0, 'composicoes_atualizadas': 0,
                     'itens_inseridos': 0, 'composicoes_recalculadas': 0, 'alertas': []}

        # ── Processar aba de insumos (ISD ou ICD) ────────────────────────────
        def processar_aba_insumos(aba, desonerado):
            insumos = _parse_insumos_aba(xl, aba, desonerado=desonerado)
            if not insumos: return

            # Cache de insumos existentes
            cod_map = {r[0]: r[1] for r in db.execute(
                "SELECT codigo_insumo, id_insumo FROM insumos WHERE origem='SINAPI'")}

            novos, updates_ins, novos_precos, updates_preco = [], [], [], []

            for ins in insumos:
                cod = ins['codigo']
                if not cod: continue
                id_un = get_unidade(ins['unidade'])

                if cod in cod_map:
                    id_insumo = cod_map[cod]
                    if sobrepor:
                        updates_ins.append((ins['descricao'], ins['tipo'], id_un, id_insumo))
                else:
                    db.execute("INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao) VALUES (?,?,?,?,'SINAPI','Ativo')",
                               [cod, ins['descricao'], ins['tipo'], id_un])
                    db.commit()
                    id_insumo = db.execute("SELECT id_insumo FROM insumos WHERE codigo_insumo=? AND origem='SINAPI'", [cod]).fetchone()[0]
                    cod_map[cod] = id_insumo
                    resultado['insumos_inseridos'] += 1

                for uf in ufs:
                    preco = ins['precos'].get(uf)
                    if preco is None: continue
                    p_row = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia=?",
                                       [id_insumo, id_data_base, uf]).fetchone()
                    if p_row:
                        if sobrepor:
                            col = 'preco_desonerado' if desonerado else 'preco_nao_desonerado'
                            db.execute(f"UPDATE precos_insumos SET {col}=?,preco_referencia=? WHERE id_preco=?",
                                       [preco, preco, p_row[0]])
                            resultado['precos_atualizados'] += 1
                    else:
                        col = 'preco_desonerado' if desonerado else 'preco_nao_desonerado'
                        db.execute(f"INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,{col},preco_referencia) VALUES (?,?,?,?,?,?)",
                                   [id_insumo, id_data_base, id_fonte, uf, preco, preco])
                        resultado['precos_inseridos'] += 1

            if updates_ins:
                db.executemany("UPDATE insumos SET descricao=?,tipo_insumo=?,id_unidade=? WHERE id_insumo=?", updates_ins)
                resultado['insumos_atualizados'] += len(updates_ins)
            db.commit()

        if imp_isd and 'ISD' in xl.sheet_names:
            processar_aba_insumos('ISD', desonerado=False)
        if imp_icd and 'ICD' in xl.sheet_names:
            processar_aba_insumos('ICD', desonerado=True)

        # ── Analítico ─────────────────────────────────────────────────────────
        aba_anal = next((a for a in ['Analítico','Analitico'] if a in xl.sheet_names), None)
        if imp_anal and aba_anal:
            grp_cache = {r[0]: r[1] for r in db.execute("SELECT nome_grupo,id_grupo_comp FROM grupos_composicoes")}

            def get_grupo(nome):
                if nome not in grp_cache:
                    gid = db.execute("INSERT INTO grupos_composicoes (nome_grupo,fonte) VALUES (?,'SINAPI')", [nome]).lastrowid
                    db.commit(); grp_cache[nome] = gid
                return grp_cache[nome]

            comp_cache = {r[0]: r[1] for r in db.execute(
                "SELECT codigo,id_composicao FROM composicoes WHERE fonte='SINAPI'")}

            itens_batch = []
            for comp in _parse_analitico(xl):
                cod = comp.get('codigo','')
                if not cod: continue
                id_un  = get_unidade(comp['unidade'])
                id_grp = get_grupo(comp.get('grupo','') or 'SINAPI')
                mes_r  = f'{mes:02d}/{ano}'

                if cod in comp_cache:
                    id_comp = comp_cache[cod]
                    if sobrepor:
                        db.execute("UPDATE composicoes SET descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,uf_referencia=?,situacao_ref=? WHERE id_composicao=?",
                                   [comp['descricao'], comp['unidade'], id_grp, mes_r, comp.get('uf',''), comp.get('situacao',''), id_comp])
                        db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id_comp])
                        resultado['composicoes_atualizadas'] += 1
                    else:
                        continue
                else:
                    id_comp = db.execute(
                        "INSERT INTO composicoes (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,situacao_ref,situacao) VALUES (?,'SINAPI','UNITARIO',?,?,?,?,?,?,'Ativo')",
                        [cod, comp['descricao'], comp['unidade'], id_grp, mes_r, comp.get('uf',''), comp.get('situacao','')]).lastrowid
                    comp_cache[cod] = id_comp
                    resultado['composicoes_inseridas'] += 1

                for ordem, it in enumerate(comp['itens']):
                    tipo_db = 'INSUMO' if it['tipo_item'] == 'INSUMO' else 'COMPOSICAO'
                    itens_batch.append((id_comp, tipo_db, it['codigo_item'], it['descricao'],
                                        it['unidade'], it['coeficiente'], it.get('situacao',''), ordem))
                    resultado['itens_inseridos'] += 1
                    if len(itens_batch) >= 500:
                        db.executemany("INSERT INTO itens_composicao (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,ordem) VALUES (?,?,?,?,?,?,?,?)",
                                       itens_batch)
                        db.commit(); itens_batch = []

            if itens_batch:
                db.executemany("INSERT INTO itens_composicao (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,ordem) VALUES (?,?,?,?,?,?,?,?)",
                               itens_batch)
                db.commit()

        resultado['mensagem'] = (
            f"Insumos: {resultado['insumos_inseridos']} inseridos, {resultado['insumos_atualizados']} atualizados. "
            f"Preços: {resultado['precos_inseridos']} inseridos, {resultado['precos_atualizados']} atualizados. "
            f"Composições: {resultado['composicoes_inseridas']} inseridas, {resultado['composicoes_atualizadas']} atualizadas. "
            f"Itens: {resultado['itens_inseridos']} inseridos. "
            f"Use 'Recalcular Custos SINAPI' em Composições para calcular os custos."
        )
        return jsonify(resultado)

    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1000:]}), 500
    finally:
        if db:
            try: db.close()
            except: pass
        try: os.unlink(tmp_path)
        except: pass


# ═══════════════════════════════════════════════════════════════════════════════
# SEINFRA/CE — IMPORTAÇÃO DE INSUMOS E COMPOSIÇÕES
# ═══════════════════════════════════════════════════════════════════════════════

def _seinfra_float(v):
    import pandas as pd
    try:
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace('.', '').replace(',', '.')
            if not s or s.lower() == 'nan':
                return None
            return float(s)
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None

def _seinfra_s(v):
    import pandas as pd
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except Exception:
        pass
    return str(v).strip()

def _seinfra_tipo_por_grupo(grupo):
    g = (grupo or '').upper()
    if 'MAO DE OBRA' in g or 'MÃO DE OBRA' in g:
        return 'Mão de Obra'
    if 'EQUIP' in g or 'EQ.' in g or 'CHORARIO' in g or 'CUSTO HORARIO' in g:
        return 'Equipamento'
    if 'SERVI' in g and 'EMPREIT' in g:
        return 'Serviço Auxiliar'
    return 'Material'

def _seinfra_parse_ref_from_insumos(path):
    import pandas as pd, re
    df = pd.read_excel(path, sheet_name='insumos', header=None, nrows=5)
    texto = ' '.join(_seinfra_s(v) for v in df.values.flatten())
    # Preferir data de emissão do arquivo SEINFRA quando não houver campo explícito de referência.
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})', texto)
    if m:
        return int(m.group(2)), int(m.group(3))
    m = re.search(r'(0?[1-9]|1[0-2])[/\-](20\d{2}|19\d{2})', texto)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def _seinfra_parse_insumos(path):
    import pandas as pd, re
    df = pd.read_excel(path, sheet_name='insumos', header=None)
    insumos, grupo = [], ''
    for _, row in df.iterrows():
        c0, c1, c2, c3 = (_seinfra_s(row.iloc[i]) if i < len(row) else '' for i in range(4))
        if c0 and not c1 and not re.match(r'^[A-Za-z]\d+', c0) and c0.lower() != 'insumo':
            grupo = c0
            continue
        if not re.match(r'^[A-Za-z]\d+', c0):
            continue
        preco = _seinfra_float(row.iloc[3] if len(row) > 3 else None)
        if preco is None:
            continue
        insumos.append({
            'codigo': c0.upper(),
            'descricao': c1,
            'unidade': c2.upper()[:20],
            'preco': preco,
            'grupo': grupo,
            'tipo': _seinfra_tipo_por_grupo(grupo),
        })
    return insumos

def _seinfra_parse_comp_header(txt):
    import re
    s = _seinfra_s(txt)
    m = re.match(r'^(C\d+)\s*-\s*(.+?)\s*-\s*([A-Za-z0-9²³/]+)\s*$', s)
    if not m:
        return None
    return {'codigo': m.group(1).upper(), 'descricao': m.group(2).strip(), 'unidade': m.group(3).strip().upper()}

def _seinfra_parse_composicoes(path):
    import pandas as pd, re
    xl = pd.ExcelFile(path)
    composicoes, atual, secao = [], None, ''

    def finalizar():
        nonlocal atual
        if atual and atual.get('codigo'):
            atual['custo_unitario'] = atual.get('custo_unitario') or sum((it.get('custo_parcial') or 0) for it in atual['itens'])
            composicoes.append(atual)
        atual = None

    for sh in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sh, header=None)
        for _, row in df.iterrows():
            v0 = _seinfra_s(row.iloc[0] if len(row) > 0 else '')
            cab = _seinfra_parse_comp_header(v0)
            if cab:
                finalizar()
                atual = {**cab, 'itens': [], 'custo_unitario': 0}
                secao = ''
                continue
            if not atual:
                continue
            v1 = _seinfra_s(row.iloc[1] if len(row) > 1 else '')
            v2 = _seinfra_s(row.iloc[2] if len(row) > 2 else '')
            v3 = _seinfra_s(row.iloc[3] if len(row) > 3 else '')
            v4 = _seinfra_s(row.iloc[4] if len(row) > 4 else '')
            if v0 and not re.match(r'^[A-Za-z]\d+', v0) and v0.upper() not in ('TOTAL:', 'TOTAL'):
                if v0.upper() not in ('NAN',) and v0.lower() != 'insumo':
                    secao = v0
                continue
            if v3.upper().startswith('TOTAL SIMPLES') or v3.upper().startswith('VALOR GERAL'):
                val = _seinfra_float(row.iloc[5] if len(row) > 5 else None)
                if val is not None:
                    atual['custo_unitario'] = val
                continue
            if not re.match(r'^[A-Za-z]\d+', v0):
                continue
            coef = _seinfra_float(row.iloc[3] if len(row) > 3 else None)
            preco = _seinfra_float(row.iloc[4] if len(row) > 4 else None)
            total = _seinfra_float(row.iloc[5] if len(row) > 5 else None)
            if coef is None:
                continue
            tipo = 'COMPOSICAO' if v0.upper().startswith('C') else 'INSUMO'
            atual['itens'].append({
                'tipo_item': tipo,
                'codigo_item': v0.upper(),
                'descricao': v1,
                'unidade': v2.upper()[:20],
                'coeficiente': coef,
                'preco_unitario': preco,
                'custo_parcial': total,
                'secao': secao,
            })
    finalizar()
    return composicoes

def _seinfra_get_unidade(db, cache, sigla):
    sigla = (_seinfra_s(sigla) or 'UN').upper()[:20]
    if sigla not in cache:
        cache[sigla] = db.execute(
            "INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)",
            [sigla, sigla, 'Outro']).lastrowid
        db.commit()
    return cache[sigla]

@app.route('/api/seinfra/importar', methods=['POST'])
def seinfra_importar():
    import tempfile, os, traceback
    import pandas as pd

    obrig = ['insumos_onerado', 'composicoes_onerado', 'insumos_desonerado', 'composicoes_desonerado']
    faltando = [k for k in obrig if k not in request.files]
    if faltando:
        return jsonify({'erro': 'Envie os quatro arquivos da SEINFRA/CE.'}), 400

    tmp_paths = {}
    db = None
    try:
        for k in obrig:
            f = request.files[k]
            ext = os.path.splitext(f.filename or '')[1].lower()
            if ext not in ('.xls', '.xlsx', '.xlsm'):
                return jsonify({'erro': f'Arquivo inválido em {k}. Use .xls/.xlsx.'}), 400
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            f.save(tmp.name); tmp.close()
            tmp_paths[k] = tmp.name

        mes_arq, ano_arq = _seinfra_parse_ref_from_insumos(tmp_paths['insumos_onerado'])
        mes = int(request.form.get('mes') or mes_arq or 10)
        ano = int(request.form.get('ano') or ano_arq or 2023)
        mes_ref = f'{mes:02d}/{ano}'
        sobrepor = request.form.get('sobrepor', 'true').lower() == 'true'

        ins_on = _seinfra_parse_insumos(tmp_paths['insumos_onerado'])
        ins_des = _seinfra_parse_insumos(tmp_paths['insumos_desonerado'])
        comp_on = _seinfra_parse_composicoes(tmp_paths['composicoes_onerado'])
        comp_des = _seinfra_parse_composicoes(tmp_paths['composicoes_desonerado'])
        if not ins_on and not ins_des:
            return jsonify({'erro': 'Nenhum insumo SEINFRA encontrado nos arquivos.'}), 400
        if not comp_on and not comp_des:
            return jsonify({'erro': 'Nenhuma composição SEINFRA encontrada nos arquivos.'}), 400

        db = get_db()
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
        id_data_base = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)",
            [mes, ano, f'SEINFRA/CE {mes_ref}']).lastrowid
        fonte = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='Seinfra/CE'").fetchone()
        id_fonte = fonte[0] if fonte else db.execute(
            "INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes) VALUES (?,?,?,?,?)",
            ['Seinfra/CE', 'Oficial', 'Secretaria da Infraestrutura do Ceará', 'CE',
             'Fonte estadual SEINFRA/CE. UF fixa: CE.']).lastrowid
        db.commit()

        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        cod_map = {r[0]: r[1] for r in db.execute("SELECT codigo_insumo,id_insumo FROM insumos WHERE origem='SEINFRA'")}
        res = {'data_base': mes_ref, 'uf': 'CE', 'insumos_inseridos': 0, 'insumos_atualizados': 0,
               'precos_inseridos': 0, 'precos_atualizados': 0, 'composicoes_inseridas': 0,
               'composicoes_atualizadas': 0, 'itens_inseridos': 0}

        def importar_insumos(lista, desonerado):
            col = 'preco_desonerado' if desonerado else 'preco_nao_desonerado'
            for ins in lista:
                cod = ins['codigo']
                uid = _seinfra_get_unidade(db, unid_cache, ins['unidade'])
                if cod in cod_map:
                    id_ins = cod_map[cod]
                    if sobrepor:
                        db.execute("UPDATE insumos SET descricao=?,tipo_insumo=?,id_unidade=?,observacoes=? WHERE id_insumo=?",
                                   [ins['descricao'], ins['tipo'], uid, ins.get('grupo'), id_ins])
                        res['insumos_atualizados'] += 1
                else:
                    id_ins = db.execute(
                        "INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao,observacoes) VALUES (?,?,?,?,?,?,?)",
                        [cod, ins['descricao'], ins['tipo'], uid, 'SEINFRA', 'Ativo', ins.get('grupo')]).lastrowid
                    cod_map[cod] = id_ins
                    res['insumos_inseridos'] += 1
                row = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia='CE'",
                                 [id_ins, id_data_base]).fetchone()
                preco = ins['preco']
                if row:
                    if sobrepor:
                        db.execute(f"UPDATE precos_insumos SET {col}=?,preco_referencia=? WHERE id_preco=?",
                                   [preco, preco, row[0]])
                        res['precos_atualizados'] += 1
                else:
                    db.execute(f"INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,{col},preco_referencia) VALUES (?,?,?,?,?,?)",
                               [id_ins, id_data_base, id_fonte, 'CE', preco, preco])
                    res['precos_inseridos'] += 1
            db.commit()

        importar_insumos(ins_on, desonerado=False)
        importar_insumos(ins_des, desonerado=True)

        grp_cache = {r[0]: r[1] for r in db.execute("SELECT nome_grupo,id_grupo_comp FROM grupos_composicoes")}
        def get_grupo(nome):
            if nome not in grp_cache:
                grp_cache[nome] = db.execute(
                    "INSERT INTO grupos_composicoes (nome_grupo,fonte) VALUES (?,?)",
                    [nome, 'SEINFRA']).lastrowid
                db.commit()
            return grp_cache[nome]

        comp_cache = {r[0]: r[1] for r in db.execute("SELECT codigo,id_composicao FROM composicoes WHERE fonte='SEINFRA'")}
        def importar_comps(lista, regime):
            suffix = 'ON' if regime == 'Onerado' else 'DES'
            id_grp = get_grupo(f'SEINFRA/CE {mes_ref} - {regime}')
            for comp in lista:
                cod = f"SEINFRA.{comp['codigo']}.{suffix}"
                if cod in comp_cache:
                    id_comp = comp_cache[cod]
                    if sobrepor:
                        db.execute("""UPDATE composicoes SET descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,
                                      uf_referencia='CE',situacao_ref=?,custo_unitario=?,situacao='Ativo'
                                      WHERE id_composicao=?""",
                                   [comp['descricao'], comp['unidade'], id_grp, mes_ref, regime, comp['custo_unitario'], id_comp])
                        db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id_comp])
                        res['composicoes_atualizadas'] += 1
                    else:
                        continue
                else:
                    id_comp = db.execute("""INSERT INTO composicoes
                        (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,situacao_ref,custo_unitario,situacao)
                        VALUES (?,'SEINFRA','UNITARIO',?,?,?,?,?,?,?,'Ativo')""",
                        [cod, comp['descricao'], comp['unidade'], id_grp, mes_ref, 'CE', regime, comp['custo_unitario']]).lastrowid
                    comp_cache[cod] = id_comp
                    res['composicoes_inseridas'] += 1
                batch = []
                for ordem, it in enumerate(comp['itens']):
                    batch.append((id_comp, it['tipo_item'], it['codigo_item'], it['descricao'], it['unidade'],
                                  it['coeficiente'], it.get('secao'), it.get('preco_unitario'), it.get('custo_parcial'), ordem))
                if batch:
                    db.executemany("""INSERT INTO itens_composicao
                        (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,preco_unitario,custo_parcial,ordem)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""", batch)
                    res['itens_inseridos'] += len(batch)
            db.commit()

        importar_comps(comp_on, 'Onerado')
        importar_comps(comp_des, 'Desonerado')
        res['mensagem'] = (
            f"SEINFRA/CE {mes_ref}: {res['insumos_inseridos']} insumos novos, "
            f"{res['precos_inseridos']} preços novos, {res['composicoes_inseridas']} composições novas."
        )
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1200:]}), 500
    finally:
        if db:
            try: db.close()
            except Exception: pass
        for p in tmp_paths.values():
            try: os.unlink(p)
            except Exception: pass


# ═══════════════════════════════════════════════════════════════════════════════
# SUDECAP/BH — IMPORTAÇÃO DE INSUMOS E COMPOSIÇÕES
# ═══════════════════════════════════════════════════════════════════════════════

def _sudecap_codigo(v):
    import re
    return bool(re.match(r'^\d{2}(?:\.\d{2})+$', _seinfra_s(v)))

def _sudecap_tipo(codigo):
    c = _seinfra_s(codigo)
    if c.startswith('55.'):
        return 'Mão de Obra'
    if c.startswith('54.') or c.startswith('50.'):
        return 'Equipamento'
    return 'Material'

def _sudecap_parse_ref(path):
    import os, re, pandas as pd
    df = pd.read_excel(path, sheet_name=0, header=None, nrows=6)
    texto = ' '.join(_seinfra_s(v) for v in df.values.flatten())
    m = re.search(r'REFER[ÊE]NCIA:\s*(\d{1,2})/(\d{2,4})', texto, re.I)
    if m:
        ano = int(m.group(2))
        return int(m.group(1)), (2000 + ano if ano < 100 else ano)
    m = re.search(r'(20\d{2})[._-](\d{1,2})', os.path.basename(path))
    if m:
        return int(m.group(2)), int(m.group(1))
    return None, None

def _sudecap_parse_insumos(path):
    import pandas as pd
    df = pd.read_excel(path, sheet_name=0, header=None)
    out = []
    for _, row in df.iterrows():
        cod = _seinfra_s(row.iloc[0] if len(row) > 0 else '').upper()
        if not _sudecap_codigo(cod):
            continue
        preco = _seinfra_float(row.iloc[4] if len(row) > 4 else None)
        if preco is None:
            continue
        out.append({
            'codigo': cod,
            'descricao': _seinfra_s(row.iloc[2] if len(row) > 2 else ''),
            'unidade': _seinfra_s(row.iloc[3] if len(row) > 3 else 'UN').upper()[:20],
            'preco': preco,
            'tipo': _sudecap_tipo(cod),
            'origem_planilha': _seinfra_s(row.iloc[1] if len(row) > 1 else 'SUDECAP'),
        })
    return out

def _sudecap_parse_composicoes(paths):
    import pandas as pd
    comps = []
    atual = None

    def finalizar():
        nonlocal atual
        if atual and atual.get('codigo') and atual.get('itens'):
            comps.append(atual)
        atual = None

    for path in paths:
        df = pd.read_excel(path, sheet_name=0, header=None)
        for _, row in df.iterrows():
            c0 = _seinfra_s(row.iloc[0] if len(row) > 0 else '').upper()
            c1 = _seinfra_s(row.iloc[1] if len(row) > 1 else '')
            c2 = _seinfra_s(row.iloc[2] if len(row) > 2 else '')
            und = _seinfra_s(row.iloc[7] if len(row) > 7 else '').upper()[:20]
            consumo = _seinfra_float(row.iloc[9] if len(row) > 9 else None)

            if _sudecap_codigo(c0) and c1 and und and consumo is None:
                finalizar()
                atual = {'codigo': c0, 'descricao': c1, 'unidade': und, 'itens': []}
                continue
            if not atual:
                continue
            if _sudecap_codigo(c1) and c2 and consumo is not None:
                atual['itens'].append({
                    'codigo_item': c1.upper(),
                    'descricao': c2,
                    'unidade': und or 'UN',
                    'coeficiente': consumo,
                })
        finalizar()
    return comps

def _sudecap_calcular_custos(comps, ins_precos, regime):
    comp_codes = {c['codigo'] for c in comps}
    custos = {}
    item_type = {}
    for comp in comps:
        item_type[comp['codigo']] = 'COMPOSICAO'
    for _ in range(12):
        mudou = False
        for comp in comps:
            total = 0.0
            ok = True
            for it in comp['itens']:
                cod = it['codigo_item']
                preco = None
                if cod in ins_precos:
                    preco = ins_precos[cod].get(regime)
                    item_type[cod] = 'INSUMO'
                elif cod in custos:
                    preco = custos[cod]
                    item_type[cod] = 'COMPOSICAO'
                elif cod in comp_codes:
                    ok = False
                    break
                else:
                    preco = 0.0
                    item_type[cod] = 'INSUMO'
                total += (it.get('coeficiente') or 0) * (preco or 0)
            if ok and round(custos.get(comp['codigo'], -1), 6) != round(total, 6):
                custos[comp['codigo']] = total
                mudou = True
        if not mudou:
            break
    return custos, item_type

@app.route('/api/sudecap/importar', methods=['POST'])
def sudecap_importar():
    import tempfile, os, traceback

    obrig = ['insumos_onerado', 'insumos_desonerado', 'composicoes_construcao', 'composicoes_custo_horario']
    faltando = [k for k in obrig if k not in request.files]
    if faltando:
        return jsonify({'erro': 'Envie os quatro arquivos da SUDECAP/BH.'}), 400

    tmp_paths = {}
    db = None
    try:
        for k in obrig:
            f = request.files[k]
            ext = os.path.splitext(f.filename or '')[1].lower()
            if ext not in ('.xls', '.xlsx', '.xlsm'):
                return jsonify({'erro': f'Arquivo inválido em {k}. Use .xls/.xlsx.'}), 400
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            f.save(tmp.name); tmp.close()
            tmp_paths[k] = tmp.name

        mes_arq, ano_arq = _sudecap_parse_ref(tmp_paths['insumos_onerado'])
        mes = int(request.form.get('mes') or mes_arq or 1)
        ano = int(request.form.get('ano') or ano_arq or 2026)
        mes_ref = f'{mes:02d}/{ano}'
        sobrepor = request.form.get('sobrepor', 'true').lower() == 'true'

        ins_on = _sudecap_parse_insumos(tmp_paths['insumos_onerado'])
        ins_des = _sudecap_parse_insumos(tmp_paths['insumos_desonerado'])
        comps = _sudecap_parse_composicoes([tmp_paths['composicoes_construcao'], tmp_paths['composicoes_custo_horario']])
        if not ins_on and not ins_des:
            return jsonify({'erro': 'Nenhum insumo SUDECAP encontrado nos arquivos.'}), 400
        if not comps:
            return jsonify({'erro': 'Nenhuma composição SUDECAP encontrada nos arquivos.'}), 400

        db = get_db()
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
        id_data_base = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)",
            [mes, ano, f'SUDECAP/BH {mes_ref}']).lastrowid
        fonte = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='Sudecap/BH'").fetchone()
        id_fonte = fonte[0] if fonte else db.execute(
            "INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes) VALUES (?,?,?,?,?)",
            ['Sudecap/BH', 'Oficial', 'Superintendência de Desenvolvimento da Capital', 'Belo Horizonte/MG',
             'Fonte municipal SUDECAP/BH. UF fixa: MG.']).lastrowid
        db.commit()

        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        cod_map = {r[0]: r[1] for r in db.execute("SELECT codigo_insumo,id_insumo FROM insumos WHERE origem='SUDECAP'")}
        res = {'data_base': mes_ref, 'uf': 'MG', 'insumos_inseridos': 0, 'insumos_atualizados': 0,
               'precos_inseridos': 0, 'precos_atualizados': 0, 'composicoes_inseridas': 0,
               'composicoes_atualizadas': 0, 'itens_inseridos': 0, 'composicoes_sem_custo': 0}

        ins_precos = {}
        def importar_insumos(lista, regime):
            col = 'preco_desonerado' if regime == 'desonerado' else 'preco_nao_desonerado'
            for ins in lista:
                cod = ins['codigo']
                ins_precos.setdefault(cod, {})[regime] = ins['preco']
                uid = _seinfra_get_unidade(db, unid_cache, ins['unidade'])
                if cod in cod_map:
                    id_ins = cod_map[cod]
                    if sobrepor:
                        db.execute("UPDATE insumos SET descricao=?,tipo_insumo=?,id_unidade=?,observacoes=? WHERE id_insumo=?",
                                   [ins['descricao'], ins['tipo'], uid, ins.get('origem_planilha'), id_ins])
                        res['insumos_atualizados'] += 1
                else:
                    id_ins = db.execute(
                        "INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao,observacoes) VALUES (?,?,?,?,?,?,?)",
                        [cod, ins['descricao'], ins['tipo'], uid, 'SUDECAP', 'Ativo', ins.get('origem_planilha')]).lastrowid
                    cod_map[cod] = id_ins
                    res['insumos_inseridos'] += 1
                row = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia='MG'",
                                 [id_ins, id_data_base]).fetchone()
                preco = ins['preco']
                if row:
                    if sobrepor:
                        db.execute(f"UPDATE precos_insumos SET {col}=?,preco_referencia=? WHERE id_preco=?",
                                   [preco, preco, row[0]])
                        res['precos_atualizados'] += 1
                else:
                    db.execute(f"INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,{col},preco_referencia) VALUES (?,?,?,?,?,?)",
                               [id_ins, id_data_base, id_fonte, 'MG', preco, preco])
                    res['precos_inseridos'] += 1
            db.commit()

        importar_insumos(ins_on, 'onerado')
        importar_insumos(ins_des, 'desonerado')

        custos_on, tipos_on = _sudecap_calcular_custos(comps, ins_precos, 'onerado')
        custos_des, tipos_des = _sudecap_calcular_custos(comps, ins_precos, 'desonerado')

        grp_cache = {r[0]: r[1] for r in db.execute("SELECT nome_grupo,id_grupo_comp FROM grupos_composicoes")}
        def get_grupo(nome):
            if nome not in grp_cache:
                grp_cache[nome] = db.execute(
                    "INSERT INTO grupos_composicoes (nome_grupo,fonte) VALUES (?,?)",
                    [nome, 'SUDECAP']).lastrowid
                db.commit()
            return grp_cache[nome]

        comp_cache = {r[0]: r[1] for r in db.execute("SELECT codigo,id_composicao FROM composicoes WHERE fonte='SUDECAP'")}
        def importar_comps(regime_label, suffix, custos, tipos):
            id_grp = get_grupo(f'SUDECAP/BH {mes_ref} - {regime_label}')
            for comp in comps:
                cod = f"SUDECAP.{comp['codigo']}.{suffix}"
                custo = custos.get(comp['codigo'], 0.0)
                if not custo:
                    res['composicoes_sem_custo'] += 1
                if cod in comp_cache:
                    id_comp = comp_cache[cod]
                    if sobrepor:
                        db.execute("""UPDATE composicoes SET descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,
                                      uf_referencia='MG',situacao_ref=?,custo_unitario=?,situacao='Ativo'
                                      WHERE id_composicao=?""",
                                   [comp['descricao'], comp['unidade'], id_grp, mes_ref, regime_label, custo, id_comp])
                        db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id_comp])
                        res['composicoes_atualizadas'] += 1
                    else:
                        continue
                else:
                    id_comp = db.execute("""INSERT INTO composicoes
                        (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,situacao_ref,custo_unitario,situacao)
                        VALUES (?,'SUDECAP','UNITARIO',?,?,?,?,?,?,?,'Ativo')""",
                        [cod, comp['descricao'], comp['unidade'], id_grp, mes_ref, 'MG', regime_label, custo]).lastrowid
                    comp_cache[cod] = id_comp
                    res['composicoes_inseridas'] += 1
                batch = []
                for ordem, it in enumerate(comp['itens']):
                    cod_item = it['codigo_item']
                    tipo_item = tipos.get(cod_item, 'INSUMO')
                    preco = custos.get(cod_item) if tipo_item == 'COMPOSICAO' else (ins_precos.get(cod_item, {}).get('desonerado' if suffix == 'DES' else 'onerado') or 0)
                    batch.append((id_comp, tipo_item, cod_item, it['descricao'], it['unidade'], it['coeficiente'],
                                  preco, (it['coeficiente'] or 0) * (preco or 0), ordem))
                if batch:
                    db.executemany("""INSERT INTO itens_composicao
                        (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,preco_unitario,custo_parcial,ordem)
                        VALUES (?,?,?,?,?,?,?,?,?)""", batch)
                    res['itens_inseridos'] += len(batch)
            db.commit()

        importar_comps('Onerado', 'ON', custos_on, tipos_on)
        importar_comps('Desonerado', 'DES', custos_des, tipos_des)
        res['mensagem'] = (
            f"SUDECAP/BH {mes_ref}: {res['insumos_inseridos']} insumos novos, "
            f"{res['precos_inseridos']} preços novos, {res['composicoes_inseridas']} composições novas."
        )
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1200:]}), 500
    finally:
        if db:
            try: db.close()
            except Exception: pass
        for p in tmp_paths.values():
            try: os.unlink(p)
            except Exception: pass


# ═══════════════════════════════════════════════════════════════════════════════
# GOINFRA/GO — IMPORTAÇÃO DE INSUMOS E COMPOSIÇÕES EM PDF
# ═══════════════════════════════════════════════════════════════════════════════

def _goinfra_pdf_reader(path):
    import sys
    from pathlib import Path
    try:
        from pypdf import PdfReader
    except Exception:
        dep = Path.home() / '.cache' / 'codex-runtimes' / 'codex-primary-runtime' / 'dependencies' / 'python' / 'Lib' / 'site-packages'
        if dep.exists():
            sys.path.append(str(dep))
        from pypdf import PdfReader
    return PdfReader(path)

def _goinfra_pdf_text(path, max_pages=None):
    reader = _goinfra_pdf_reader(path)
    pages = reader.pages[:max_pages] if max_pages else reader.pages
    return '\n'.join((p.extract_text() or '') for p in pages)

def _goinfra_num(v):
    s = _seinfra_s(v).replace('.', '').replace(',', '.')
    try:
        return float(s)
    except Exception:
        return None

def _goinfra_parse_ref(path):
    import re
    text = _goinfra_pdf_text(path, max_pages=1)
    m = re.search(r'Data base:\s*(\d{2})/(\d{2})/(\d{4})', text, re.I)
    if m:
        return int(m.group(2)), int(m.group(3))
    m = re.search(r'([A-ZÇ]+)/(20\d{2})', text.upper())
    meses = {'JANEIRO':1,'FEVEREIRO':2,'MARCO':3,'MARÇO':3,'ABRIL':4,'MAIO':5,'JUNHO':6,
             'JULHO':7,'AGOSTO':8,'SETEMBRO':9,'OUTUBRO':10,'NOVEMBRO':11,'DEZEMBRO':12}
    if m:
        return meses.get(m.group(1), 1), int(m.group(2))
    return None, None

def _goinfra_parse_labor(path):
    import re
    text = _goinfra_pdf_text(path)
    out = []
    for line in text.splitlines():
        line = _seinfra_s(line)
        m = re.match(r'^(\d{4})\s+(.+?)\s+(h|M\S*S)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$', line, re.I)
        if not m:
            continue
        out.append({
            'codigo': m.group(1),
            'descricao': m.group(2).strip(),
            'unidade': 'H' if m.group(3).lower().startswith('h') else 'MES',
            'preco': _goinfra_num(m.group(6)) or 0,
            'tipo': 'Mão de Obra',
        })
    return out

def _goinfra_parse_material(path):
    import re
    text = _goinfra_pdf_text(path)
    records, atual = [], None
    skip_prefixes = ('GOINFRA', 'Custo Referencial', 'Página:', 'Tabela:', 'Data base:', 'Código',
                     'Auxiliar', 'Descrição', 'SCO -', '01 - ', '2026 - ')
    for raw in text.splitlines():
        line = _seinfra_s(raw)
        if not line or any(line.startswith(p) for p in skip_prefixes):
            continue
        if re.match(r'^[A-Z]?\d{3,4}\s+', line):
            if atual:
                records.append(atual)
            atual = line
        elif atual:
            atual += ' ' + line
    if atual:
        records.append(atual)

    out = []
    for rec in records:
        m = re.match(r'^([A-Z]?\d{3,4})\s+(.+)\s+([A-Za-z0-9²³/]+)\s+([\d.]+,\d{2})$', rec)
        if not m:
            continue
        out.append({
            'codigo': m.group(1),
            'descricao': m.group(2).strip(),
            'unidade': m.group(3).upper()[:20],
            'preco': _goinfra_num(m.group(4)) or 0,
            'tipo': 'Material',
        })
    return out

def _goinfra_parse_composicoes(path):
    import re
    reader = _goinfra_pdf_reader(path)
    comps = []
    comp = None
    current_section = ''
    collecting_desc = False

    def finalizar():
        nonlocal comp
        if comp and comp.get('codigo'):
            if not comp.get('custo_unitario'):
                comp['custo_unitario'] = sum(it['custo_parcial'] for it in comp['itens'])
            comps.append(comp)
        comp = None

    for page in reader.pages:
        text = page.extract_text() or ''
        lines = [_seinfra_s(x) for x in text.splitlines() if _seinfra_s(x)]
        for line in lines:
            if line.startswith('GOINFRA') or line.startswith('Relat') or line.startswith('P�gina:') or line.startswith('Página:') or line.startswith('Tabela de pre') or line.startswith('Data base:') or line.startswith('SCO -'):
                continue
            mserv = re.match(r'^Servi\S*o:\s*(\d{6})\s*-\s*(.*)$', line, re.I)
            if mserv:
                finalizar()
                comp = {'codigo': '', 'descricao': '', 'unidade': '', 'itens': [], 'custo_unitario': 0}
                comp['codigo'] = mserv.group(1)
                rest = mserv.group(2).strip()
                if 'Unidade:' in rest:
                    desc, und = rest.split('Unidade:', 1)
                    comp['descricao'] = desc.strip()
                    comp['unidade'] = und.strip().split()[0].upper()[:20]
                    collecting_desc = False
                else:
                    comp['descricao'] = rest
                    collecting_desc = True
                continue
            if not comp:
                continue
            if collecting_desc:
                if line.startswith('Unidade:'):
                    comp['unidade'] = line.replace('Unidade:', '').strip().split()[0].upper()[:20]
                    collecting_desc = False
                    continue
                if not line.startswith('Código') and not line.startswith('C�digo'):
                    comp['descricao'] = (comp['descricao'] + ' ' + line).strip()
                    continue
            if '(B)' in line or 'Mãos-de-obra' in line or 'M�os-de-obra' in line:
                current_section = 'B - Mão de Obra'
                continue
            if '(C)' in line or 'Materiais' in line:
                current_section = 'C - Materiais'
                continue
            if '(D)' in line or 'Equipamentos' in line:
                current_section = 'D - Equipamentos'
                continue
            mdirect = re.search(r'Custo direto total.*?([\d.]+,\d{2})\s*$', line)
            if mdirect:
                comp['custo_unitario'] = _goinfra_num(mdirect.group(1)) or 0
                continue
            mitem = re.match(r'^([A-Z]?\d{3,6})\s+(.+?)\s+((?:[\d.]+,\d{2,7}\s+){1,6}[\d.]+,\d{2})$', line)
            if mitem:
                nums = re.findall(r'[\d.]+,\d+', mitem.group(3))
                if len(nums) < 2:
                    continue
                coef = _goinfra_num(nums[-2]) or 0
                total = _goinfra_num(nums[-1]) or 0
                preco = (total / coef) if coef else (_goinfra_num(nums[-3]) if len(nums) >= 3 else 0)
                comp['itens'].append({
                    'codigo_item': mitem.group(1),
                    'descricao': mitem.group(2).strip(),
                    'unidade': '',
                    'coeficiente': coef,
                    'preco_unitario': preco,
                    'custo_parcial': total,
                    'secao': current_section,
                })
    finalizar()
    return comps

@app.route('/api/goinfra/importar', methods=['POST'])
def goinfra_importar():
    import tempfile, os, traceback

    obrig = ['mao_obra_onerado', 'mao_obra_desonerado', 'material', 'composicoes_onerado', 'composicoes_desonerado']
    faltando = [k for k in obrig if k not in request.files]
    if faltando:
        return jsonify({'erro': 'Envie os cinco arquivos da GOINFRA/GO.'}), 400

    tmp_paths = {}
    db = None
    try:
        for k in obrig:
            f = request.files[k]
            ext = os.path.splitext(f.filename or '')[1].lower()
            if ext != '.pdf':
                return jsonify({'erro': f'Arquivo inválido em {k}. Use PDF.'}), 400
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            f.save(tmp.name); tmp.close()
            tmp_paths[k] = tmp.name

        mes_arq, ano_arq = _goinfra_parse_ref(tmp_paths['mao_obra_onerado'])
        mes = int(request.form.get('mes') or mes_arq or 2)
        ano = int(request.form.get('ano') or ano_arq or 2026)
        mes_ref = f'{mes:02d}/{ano}'
        sobrepor = request.form.get('sobrepor', 'true').lower() == 'true'

        mo_on = _goinfra_parse_labor(tmp_paths['mao_obra_onerado'])
        mo_des = _goinfra_parse_labor(tmp_paths['mao_obra_desonerado'])
        mat = _goinfra_parse_material(tmp_paths['material'])
        comp_on = _goinfra_parse_composicoes(tmp_paths['composicoes_onerado'])
        comp_des = _goinfra_parse_composicoes(tmp_paths['composicoes_desonerado'])
        if not (mo_on or mo_des or mat):
            return jsonify({'erro': 'Nenhum insumo GOINFRA encontrado nos PDFs.'}), 400
        if not (comp_on or comp_des):
            return jsonify({'erro': 'Nenhuma composição GOINFRA encontrada nos PDFs.'}), 400

        db = get_db()
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
        id_data_base = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)",
            [mes, ano, f'GOINFRA/GO {mes_ref}']).lastrowid
        fonte = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='Goinfra/GO'").fetchone()
        id_fonte = fonte[0] if fonte else db.execute(
            "INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes) VALUES (?,?,?,?,?)",
            ['Goinfra/GO', 'Oficial', 'Agência Goiana de Infraestrutura e Transportes', 'GO',
             'Fonte estadual GOINFRA/GO. UF fixa: GO.']).lastrowid
        db.commit()

        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        cod_map = {r[0]: r[1] for r in db.execute("SELECT codigo_insumo,id_insumo FROM insumos WHERE origem='GOINFRA'")}
        res = {'data_base': mes_ref, 'uf': 'GO', 'insumos_inseridos': 0, 'insumos_atualizados': 0,
               'precos_inseridos': 0, 'precos_atualizados': 0, 'composicoes_inseridas': 0,
               'composicoes_atualizadas': 0, 'itens_inseridos': 0}

        def importar_insumos(lista, regime=None):
            for ins in lista:
                col = 'preco_desonerado' if regime == 'desonerado' else 'preco_nao_desonerado'
                cod = ins['codigo']
                uid = _seinfra_get_unidade(db, unid_cache, ins['unidade'])
                if cod in cod_map:
                    id_ins = cod_map[cod]
                    if sobrepor:
                        db.execute("UPDATE insumos SET descricao=?,tipo_insumo=?,id_unidade=?,observacoes=? WHERE id_insumo=?",
                                   [ins['descricao'], ins['tipo'], uid, 'GOINFRA', id_ins])
                        res['insumos_atualizados'] += 1
                else:
                    id_ins = db.execute(
                        "INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao,observacoes) VALUES (?,?,?,?,?,?,?)",
                        [cod, ins['descricao'], ins['tipo'], uid, 'GOINFRA', 'Ativo', 'GOINFRA']).lastrowid
                    cod_map[cod] = id_ins
                    res['insumos_inseridos'] += 1
                row = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia='GO'",
                                 [id_ins, id_data_base]).fetchone()
                preco = ins['preco']
                if row:
                    if sobrepor:
                        if regime:
                            db.execute(f"UPDATE precos_insumos SET {col}=?,preco_referencia=? WHERE id_preco=?",
                                       [preco, preco, row[0]])
                        else:
                            db.execute("UPDATE precos_insumos SET preco_desonerado=?,preco_nao_desonerado=?,preco_referencia=? WHERE id_preco=?",
                                       [preco, preco, preco, row[0]])
                        res['precos_atualizados'] += 1
                else:
                    if regime:
                        db.execute(f"INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,{col},preco_referencia) VALUES (?,?,?,?,?,?)",
                                   [id_ins, id_data_base, id_fonte, 'GO', preco, preco])
                    else:
                        db.execute("INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,preco_desonerado,preco_nao_desonerado,preco_referencia) VALUES (?,?,?,?,?,?,?)",
                                   [id_ins, id_data_base, id_fonte, 'GO', preco, preco, preco])
                    res['precos_inseridos'] += 1
            db.commit()

        importar_insumos(mo_on, 'onerado')
        importar_insumos(mo_des, 'desonerado')
        importar_insumos(mat, None)

        grp_cache = {r[0]: r[1] for r in db.execute("SELECT nome_grupo,id_grupo_comp FROM grupos_composicoes")}
        def get_grupo(nome):
            if nome not in grp_cache:
                grp_cache[nome] = db.execute(
                    "INSERT INTO grupos_composicoes (nome_grupo,fonte) VALUES (?,?)",
                    [nome, 'GOINFRA']).lastrowid
                db.commit()
            return grp_cache[nome]

        comp_cache = {r[0]: r[1] for r in db.execute("SELECT codigo,id_composicao FROM composicoes WHERE fonte='GOINFRA'")}
        def importar_comps(lista, regime_label, suffix):
            id_grp = get_grupo(f'GOINFRA/GO {mes_ref} - {regime_label}')
            for comp in lista:
                cod = f"GOINFRA.{comp['codigo']}.{suffix}"
                if cod in comp_cache:
                    id_comp = comp_cache[cod]
                    if sobrepor:
                        db.execute("""UPDATE composicoes SET descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,
                                      uf_referencia='GO',situacao_ref=?,custo_unitario=?,situacao='Ativo'
                                      WHERE id_composicao=?""",
                                   [comp['descricao'], comp['unidade'], id_grp, mes_ref, regime_label, comp['custo_unitario'], id_comp])
                        db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id_comp])
                        res['composicoes_atualizadas'] += 1
                    else:
                        continue
                else:
                    id_comp = db.execute("""INSERT INTO composicoes
                        (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,situacao_ref,custo_unitario,situacao)
                        VALUES (?,'GOINFRA','UNITARIO',?,?,?,?,?,?,?,'Ativo')""",
                        [cod, comp['descricao'], comp['unidade'], id_grp, mes_ref, 'GO', regime_label, comp['custo_unitario']]).lastrowid
                    comp_cache[cod] = id_comp
                    res['composicoes_inseridas'] += 1
                batch = []
                for ordem, it in enumerate(comp['itens']):
                    tipo_item = 'COMPOSICAO' if len(it['codigo_item']) == 6 else 'INSUMO'
                    batch.append((id_comp, tipo_item, it['codigo_item'], it['descricao'], it.get('unidade') or '',
                                  it['coeficiente'], it.get('secao'), it['preco_unitario'], it['custo_parcial'], ordem))
                if batch:
                    db.executemany("""INSERT INTO itens_composicao
                        (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,preco_unitario,custo_parcial,ordem)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""", batch)
                    res['itens_inseridos'] += len(batch)
            db.commit()

        importar_comps(comp_on, 'Onerado', 'ON')
        importar_comps(comp_des, 'Desonerado', 'DES')
        res['mensagem'] = (
            f"GOINFRA/GO {mes_ref}: {res['insumos_inseridos']} insumos novos, "
            f"{res['precos_inseridos']} preços novos, {res['composicoes_inseridas']} composições novas."
        )
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1200:]}), 500
    finally:
        if db:
            try: db.close()
            except Exception: pass
        for p in tmp_paths.values():
            try: os.unlink(p)
            except Exception: pass


# ═══════════════════════════════════════════════════════════════════════════════
# CDHU/SP — IMPORTAÇÃO DE COMPOSIÇÕES ANALÍTICAS + SINTÉTICO COM BDI
# ═══════════════════════════════════════════════════════════════════════════════

def _cdhu_norm(txt):
    import re, unicodedata
    s = '' if txt is None else str(txt)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r'[^A-Za-z0-9]+', ' ', s.upper()).strip()
    return re.sub(r'\s+', ' ', s)

def _cdhu_codigo6(v):
    import re
    s = _seinfra_s(v)
    if not s:
        return ''
    try:
        if re.match(r'^\d+(?:\.0+)?$', s):
            return f'{int(float(s)):06d}'
    except Exception:
        pass
    m = re.search(r'(\d{6})', s)
    return m.group(1) if m else ''

def _cdhu_parse_referencia(texto, fallback_mes=2, fallback_ano=2026):
    import re
    meses = {
        'JANEIRO': 1, 'FEVEREIRO': 2, 'MARCO': 3, 'MARÇO': 3, 'ABRIL': 4,
        'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9,
        'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12
    }
    t = texto or ''
    m = re.search(r'(JANEIRO|FEVEREIRO|MAR[CÇ]O|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s*/\s*(\d{2,4})', t, re.I)
    if m:
        ano = int(m.group(2))
        return meses[_norm_ascii(m.group(1)).upper()], (2000 + ano if ano < 100 else ano)
    m = re.search(r'(20\d{2})[-/.](0?[1-9]|1[0-2])', t)
    if m:
        return int(m.group(2)), int(m.group(1))
    return fallback_mes, fallback_ano

def _cdhu_parse_sintetico(path, bdi_divisor):
    import pandas as pd, re
    xl = pd.ExcelFile(path)
    df = pd.read_excel(path, sheet_name=xl.sheet_names[0], header=None)
    texto_ini = ' '.join(_seinfra_s(v) for v in df.head(8).values.flatten())
    mes, ano = _cdhu_parse_referencia(texto_ini)
    por_codigo = {}
    por_desc_un = {}
    registros = []
    for _, row in df.iterrows():
        cod = _cdhu_codigo6(row.iloc[0] if len(row) > 0 else '')
        desc = _seinfra_s(row.iloc[1] if len(row) > 1 else '')
        un = _seinfra_s(row.iloc[2] if len(row) > 2 else '').upper()[:20]
        preco_bdi = _seinfra_float(row.iloc[3] if len(row) > 3 else None)
        if not cod or not desc or not un or preco_bdi is None:
            continue
        preco_direto = round(float(preco_bdi) / float(bdi_divisor), 6) if bdi_divisor else float(preco_bdi)
        rec = {
            'codigo': cod,
            'descricao': desc,
            'unidade': un,
            'preco_com_bdi': float(preco_bdi),
            'preco_direto': preco_direto,
        }
        por_codigo[cod] = rec
        por_desc_un.setdefault((_cdhu_norm(desc), un), rec)
        registros.append(rec)
    return {'mes': mes, 'ano': ano, 'por_codigo': por_codigo, 'por_desc_un': por_desc_un, 'registros': registros}

def _cdhu_parse_item_line(line):
    import re
    s = _seinfra_s(line)
    unidades = r'(M3|M2|M²|M³|UN|KG|H|L|M|HA|KM|VB|CJ|GL|PC|PÇ|PAR|JG|MES|MÊS)'
    m = re.match(r'^([\d.,]+)\s*' + unidades + r'\s*(.+)$', s, re.I)
    if not m:
        return None
    coef = _seinfra_float(m.group(1))
    if coef is None:
        return None
    unidade = m.group(2).upper().replace('²', '2').replace('³', '3')[:20]
    resto = m.group(3).strip()
    cod_match = re.search(r'([A-Z]\.\d{2}\.\d{3}\.\d{6}|[A-Z]\d{8,})\s*$', resto)
    codigo = cod_match.group(1).upper() if cod_match else ''
    descricao = resto[:cod_match.start()].strip() if cod_match else resto
    return {'coeficiente': coef, 'unidade': unidade, 'descricao': descricao, 'codigo_item': codigo}

def _cdhu_parse_header_line(line):
    import re
    s = _seinfra_s(line)
    if not s or re.match(r'^\d', s):
        return None
    unidades = r'(M3|M2|M²|M³|UN|KG|H|L|M|HA|KM|VB|CJ|GL|PC|PÇ|PAR|JG|MES|MÊS)'
    m = re.match(r'^' + unidades + r'\s*(.+?)(\d{6})$', s, re.I)
    if not m:
        return None
    unidade = m.group(1).upper().replace('²', '2').replace('³', '3')[:20]
    desc = m.group(2).strip()
    codigo = m.group(3)
    if not desc or desc.upper().startswith(('CODIGO', 'CÓDIGO')):
        return None
    return {'codigo': codigo, 'descricao': desc, 'unidade': unidade}

def _cdhu_parse_pdf_analitico(path):
    import re
    reader = _goinfra_pdf_reader(path)
    texto0 = (reader.pages[0].extract_text() or '') if reader.pages else ''
    mes, ano = _cdhu_parse_referencia(texto0)
    comps, comp, header_buf, pending_item = [], None, '', None

    def finalizar_item():
        nonlocal pending_item
        if pending_item and comp and pending_item.get('descricao'):
            comp['itens'].append(pending_item)
        pending_item = None

    def finalizar_comp():
        nonlocal comp
        finalizar_item()
        if comp and comp.get('codigo'):
            comps.append(comp)
        comp = None

    skip = ('Projeto:', 'Data Base:', 'Listagem de Compos', 'Código', 'C�digo',
            'Descri', 'Unidade', 'Coeficiente', 'PADR', 'Padr')
    for page in reader.pages:
        for raw in (page.extract_text() or '').splitlines():
            line = _seinfra_s(raw).replace('\x00', '')
            if not line:
                continue
            if any(line.startswith(p) for p in skip) or re.match(r'^P\S*gina\s+\d+', line, re.I):
                continue
            if re.match(r'^\d{2}/\d{2}/\d{4}', line) or re.match(r'^[A-ZÇ]+/\d{2}$', line, re.I):
                continue

            item = _cdhu_parse_item_line(line)
            if item:
                finalizar_item()
                pending_item = item
                if pending_item.get('codigo_item'):
                    finalizar_item()
                continue

            if pending_item:
                cod = re.search(r'([A-Z]\.\d{2}\.\d{3}\.\d{6}|[A-Z]\d{8,})\s*$', line)
                if cod:
                    prefixo = line[:cod.start()].strip()
                    if prefixo:
                        pending_item['descricao'] = (pending_item['descricao'] + ' ' + prefixo).strip()
                    pending_item['codigo_item'] = cod.group(1).upper()
                    finalizar_item()
                else:
                    pending_item['descricao'] = (pending_item['descricao'] + ' ' + line).strip()
                continue

            header = _cdhu_parse_header_line(line)
            if header:
                finalizar_comp()
                comp = {**header, 'itens': []}
                header_buf = ''
                continue

            if re.match(r'^\d{6}$', line) and header_buf:
                header = _cdhu_parse_header_line(header_buf + line)
                if header:
                    finalizar_comp()
                    comp = {**header, 'itens': []}
                header_buf = ''
                continue

            if comp is None and not header_buf:
                header_buf = line
            elif header_buf:
                header_buf = (header_buf + ' ' + line).strip()
                header = _cdhu_parse_header_line(header_buf)
                if header:
                    finalizar_comp()
                    comp = {**header, 'itens': []}
                    header_buf = ''
    finalizar_comp()
    return {'mes': mes, 'ano': ano, 'composicoes': comps}

def _cdhu_tipo_item_por_codigo(codigo, unidade=''):
    c = (codigo or '').upper()
    u = (unidade or '').upper()
    if c.startswith('B.01') or u in ('H', 'HH'):
        return 'MO'
    if c.startswith(('E', 'F')) or u in ('HP', 'HPR'):
        return 'EQUIPAMENTO'
    if re.match(r'^\d{6}$', c):
        return 'COMPOSICAO'
    return 'INSUMO'

def _cdhu_tipo_insumo(codigo, unidade=''):
    tipo = _cdhu_tipo_item_por_codigo(codigo, unidade)
    if tipo == 'MO':
        return 'Mão de Obra'
    if tipo == 'EQUIPAMENTO':
        return 'Equipamento'
    if tipo == 'COMPOSICAO':
        return 'Serviço Auxiliar'
    return 'Material'

def _cdhu_float_form(v, default=None):
    if v is None or v == '':
        return default
    s = str(v).strip()
    try:
        if ',' in s:
            return float(s.replace('.', '').replace(',', '.'))
        return float(s)
    except Exception:
        return default

@app.route('/api/cdhu/importar', methods=['POST'])
def cdhu_importar():
    import tempfile, os, traceback, re
    obrig = ['arquivo_pdf', 'arquivo_sintetico']
    faltando = [k for k in obrig if k not in request.files]
    if faltando:
        return jsonify({'erro': 'Envie o PDF analítico e o XLS sintético da CDHU/SP.'}), 400

    tmp_paths = {}
    db = None
    try:
        for k in obrig:
            f = request.files[k]
            ext = os.path.splitext(f.filename or '')[1].lower()
            if k == 'arquivo_pdf' and ext != '.pdf':
                return jsonify({'erro': 'O relatório analítico da CDHU deve estar em PDF.'}), 400
            if k == 'arquivo_sintetico' and ext not in ('.xls', '.xlsx', '.xlsm'):
                return jsonify({'erro': 'O relatório sintético da CDHU deve estar em .xls/.xlsx.'}), 400
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            f.save(tmp.name); tmp.close()
            tmp_paths[k] = tmp.name

        bdi_pct = _cdhu_float_form(request.form.get('bdi_percentual'), 20.81)
        divisor = _cdhu_float_form(request.form.get('bdi_divisor'), 1.0 + bdi_pct / 100.0)
        sobrepor = request.form.get('sobrepor', 'true').lower() == 'true'
        sint = _cdhu_parse_sintetico(tmp_paths['arquivo_sintetico'], divisor)
        anal = _cdhu_parse_pdf_analitico(tmp_paths['arquivo_pdf'])
        mes = int(request.form.get('mes') or sint.get('mes') or anal.get('mes') or 2)
        ano = int(request.form.get('ano') or sint.get('ano') or anal.get('ano') or 2026)
        mes_ref = f'{mes:02d}/{ano}'
        comps = anal['composicoes']
        if not comps:
            return jsonify({'erro': 'Nenhuma composição CDHU foi identificada no PDF analítico.'}), 400
        if not sint['registros']:
            return jsonify({'erro': 'Nenhum preço foi identificado no relatório sintético CDHU.'}), 400

        db = get_db()
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [mes, ano]).fetchone()
        id_data_base = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)",
            [mes, ano, f'CDHU/SP {mes_ref}']).lastrowid
        fonte = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='CDHU/SP'").fetchone()
        id_fonte = fonte[0] if fonte else db.execute(
            "INSERT INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia,observacoes) VALUES (?,?,?,?,?)",
            ['CDHU/SP', 'Oficial', 'Companhia de Desenvolvimento Habitacional e Urbano do Estado de São Paulo', 'SP',
             'Fonte estadual CDHU/SP. UF fixa: SP. Custos diretos importados do sintético sem BDI.']).lastrowid
        db.commit()

        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        grp_cache = {r[0]: r[1] for r in db.execute("SELECT nome_grupo,id_grupo_comp FROM grupos_composicoes")}
        cod_ins = {r[0]: r[1] for r in db.execute("SELECT codigo_insumo,id_insumo FROM insumos WHERE origem='CDHU'")}
        comp_cache = {r[0]: r[1] for r in db.execute("SELECT codigo,id_composicao FROM composicoes WHERE fonte='CDHU'")}

        def get_grupo(nome):
            if nome not in grp_cache:
                grp_cache[nome] = db.execute(
                    "INSERT INTO grupos_composicoes (nome_grupo,fonte) VALUES (?,?)",
                    [nome, 'CDHU']).lastrowid
                db.commit()
            return grp_cache[nome]

        def inferir_preco_item(it):
            cod6 = _cdhu_codigo6(it.get('codigo_item'))
            rec = sint['por_codigo'].get(cod6) if cod6 else None
            if not rec:
                rec = sint['por_desc_un'].get((_cdhu_norm(it.get('descricao')), (it.get('unidade') or '').upper()[:20]))
            if rec:
                return rec['preco_direto'], rec['preco_direto'] * float(it.get('coeficiente') or 0), rec
            return None, None, None

        id_grp = get_grupo(f'CDHU/SP {mes_ref}')
        res = {'data_base': mes_ref, 'uf': 'SP', 'bdi_percentual': bdi_pct, 'bdi_divisor': divisor,
               'insumos_inseridos': 0, 'insumos_atualizados': 0, 'precos_inseridos': 0, 'precos_atualizados': 0,
               'composicoes_inseridas': 0, 'composicoes_atualizadas': 0, 'composicoes_sem_preco': 0,
               'itens_inseridos': 0, 'itens_com_preco_inferido': 0}

        for comp in comps:
            rec_comp = sint['por_codigo'].get(comp['codigo']) or sint['por_desc_un'].get((_cdhu_norm(comp['descricao']), comp['unidade'].upper()[:20]))
            custo = rec_comp['preco_direto'] if rec_comp else None
            if custo is None:
                res['composicoes_sem_preco'] += 1
            cod_comp = f"CDHU.{comp['codigo']}"
            if cod_comp in comp_cache:
                id_comp = comp_cache[cod_comp]
                if sobrepor:
                    db.execute("""UPDATE composicoes SET descricao=?,unidade=?,id_grupo_comp=?,mes_referencia=?,
                                  uf_referencia='SP',situacao_ref=?,custo_unitario=?,situacao='Ativo',observacoes=?
                                  WHERE id_composicao=?""",
                               [comp['descricao'], comp['unidade'], id_grp, mes_ref,
                                'COM PREÇO' if custo is not None else 'SEM PREÇO', custo or 0,
                                f'CDHU/SP importado. Preço sintético expurgado por divisor {divisor:.4f}.', id_comp])
                    db.execute("DELETE FROM itens_composicao WHERE id_composicao=?", [id_comp])
                    res['composicoes_atualizadas'] += 1
                else:
                    continue
            else:
                id_comp = db.execute("""INSERT INTO composicoes
                    (codigo,fonte,formato,descricao,unidade,id_grupo_comp,mes_referencia,uf_referencia,situacao_ref,custo_unitario,situacao,observacoes)
                    VALUES (?,'CDHU','UNITARIO',?,?,?,?,?,?,?,'Ativo',?)""",
                    [cod_comp, comp['descricao'], comp['unidade'], id_grp, mes_ref, 'SP',
                     'COM PREÇO' if custo is not None else 'SEM PREÇO', custo or 0,
                     f'CDHU/SP importado. Preço sintético expurgado por divisor {divisor:.4f}.']).lastrowid
                comp_cache[cod_comp] = id_comp
                res['composicoes_inseridas'] += 1

            batch = []
            for ordem, it in enumerate(comp['itens']):
                codigo_item = (it.get('codigo_item') or '').upper()
                preco, parcial, rec_item = inferir_preco_item(it)
                tipo_item = _cdhu_tipo_item_por_codigo(codigo_item, it.get('unidade'))
                if tipo_item != 'COMPOSICAO' and codigo_item:
                    uid = _seinfra_get_unidade(db, unid_cache, it.get('unidade') or 'UN')
                    if codigo_item in cod_ins:
                        id_ins = cod_ins[codigo_item]
                        if sobrepor:
                            db.execute("UPDATE insumos SET descricao=?,tipo_insumo=?,id_unidade=?,observacoes=? WHERE id_insumo=?",
                                       [it.get('descricao') or '', _cdhu_tipo_insumo(codigo_item, it.get('unidade')), uid, 'CDHU/SP', id_ins])
                            res['insumos_atualizados'] += 1
                    else:
                        id_ins = db.execute(
                            "INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao,observacoes) VALUES (?,?,?,?,?,?,?)",
                            [codigo_item, it.get('descricao') or '', _cdhu_tipo_insumo(codigo_item, it.get('unidade')), uid, 'CDHU', 'Ativo', 'CDHU/SP']).lastrowid
                        cod_ins[codigo_item] = id_ins
                        res['insumos_inseridos'] += 1
                    if preco is not None:
                        row = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia='SP'",
                                         [id_ins, id_data_base]).fetchone()
                        if row:
                            if sobrepor:
                                db.execute("UPDATE precos_insumos SET preco_desonerado=?,preco_nao_desonerado=?,preco_referencia=?,observacoes=? WHERE id_preco=?",
                                           [preco, preco, preco, 'Preço inferido do sintético CDHU/SP sem BDI.', row[0]])
                                res['precos_atualizados'] += 1
                        else:
                            db.execute("""INSERT INTO precos_insumos
                                (id_insumo,id_data_base,id_fonte,uf_referencia,preco_desonerado,preco_nao_desonerado,preco_referencia,observacoes)
                                VALUES (?,?,?,?,?,?,?,?)""",
                                [id_ins, id_data_base, id_fonte, 'SP', preco, preco, preco, 'Preço inferido do sintético CDHU/SP sem BDI.'])
                            res['precos_inseridos'] += 1
                        res['itens_com_preco_inferido'] += 1
                batch.append((id_comp, tipo_item, codigo_item, it.get('descricao') or '', it.get('unidade') or '',
                              it.get('coeficiente') or 0, 'PREÇO INFERIDO' if preco is not None else 'SEM PREÇO INFERIDO',
                              preco, parcial, ordem))
            if batch:
                db.executemany("""INSERT INTO itens_composicao
                    (id_composicao,tipo_item,codigo_item,descricao,unidade,coeficiente,situacao_item,preco_unitario,custo_parcial,ordem)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""", batch)
                res['itens_inseridos'] += len(batch)
            db.commit()

        res['mensagem'] = (
            f"CDHU/SP {mes_ref}: {res['composicoes_inseridas']} composições novas, "
            f"{res['composicoes_atualizadas']} atualizadas, {res['itens_inseridos']} itens importados. "
            f"{res['itens_com_preco_inferido']} item(ns) receberam preço inferido pelo sintético sem BDI."
        )
        return jsonify(res)
    except Exception as e:
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1400:]}), 500
    finally:
        if db:
            try: db.close()
            except Exception: pass
        for p in tmp_paths.values():
            try: os.unlink(p)
            except Exception: pass


# ═══════════════════════════════════════════════════════════════════════════════
# SICRO — IMPORTAÇÃO DE COMPOSIÇÕES E INSUMOS
# ═══════════════════════════════════════════════════════════════════════════════

# ── Migração automática: adiciona colunas SICRO à tabela equipamentos_sinapi ──
def _sicro_migrate_db():
    db = get_db()
    # Add columns to equipamentos_sinapi
    cols_equip = [r[1] for r in db.execute("PRAGMA table_info(equipamentos_sinapi)").fetchall()]
    for col, ddl in [
        ('sistema',          "ALTER TABLE equipamentos_sinapi ADD COLUMN sistema TEXT DEFAULT 'SINAPI'"),
        ('custo_produtivo',  "ALTER TABLE equipamentos_sinapi ADD COLUMN custo_produtivo REAL"),
        ('custo_improdutivo',"ALTER TABLE equipamentos_sinapi ADD COLUMN custo_improdutivo REAL"),
    ]:
        if col not in cols_equip:
            try: db.execute(ddl); db.commit()
            except Exception: pass

    # Create composicoes_secoes if not exists
    db.execute("""CREATE TABLE IF NOT EXISTS composicoes_secoes (
        id_secao        INTEGER PRIMARY KEY AUTOINCREMENT,
        id_composicao   INTEGER NOT NULL REFERENCES composicoes(id_composicao) ON DELETE CASCADE,
        letra_secao     TEXT NOT NULL,
        nome_secao      TEXT,
        custo_total_secao REAL DEFAULT 0,
        ordem           INTEGER DEFAULT 0
    )""")

    # Create composicoes_secao_itens if not exists
    db.execute("""CREATE TABLE IF NOT EXISTS composicoes_secao_itens (
        id_item_secao   INTEGER PRIMARY KEY AUTOINCREMENT,
        id_composicao   INTEGER NOT NULL REFERENCES composicoes(id_composicao) ON DELETE CASCADE,
        id_secao        INTEGER REFERENCES composicoes_secoes(id_secao) ON DELETE CASCADE,
        letra_secao     TEXT NOT NULL,
        codigo_item     TEXT,
        descricao       TEXT,
        quantidade      REAL,
        unidade         TEXT,
        util_operativa      REAL,
        util_improdutiva    REAL,
        custo_hp            REAL,
        custo_hi            REAL,
        preco_unitario      REAL,
        custo_total         REAL,
        cod_transporte      TEXT,
        cod_transp_ln   TEXT,
        cod_transp_rp   TEXT,
        cod_transp_p    TEXT,
        fit             REAL,
        dmt             REAL,
        ordem           INTEGER DEFAULT 0
    )""")
    db.commit()
    db.close()

_sicro_migrate_db()

# ── Helpers ───────────────────────────────────────────────────────────────────
_UF_NOME_COD = {
    'Acre':'AC','Alagoas':'AL','Amapá':'AP','Amazonas':'AM','Bahia':'BA','Ceará':'CE',
    'Distrito Federal':'DF','Espírito Santo':'ES','Goiás':'GO','Maranhão':'MA',
    'Mato Grosso':'MT','Mato Grosso do Sul':'MS','Minas Gerais':'MG','Pará':'PA',
    'Paraíba':'PB','Paraná':'PR','Pernambuco':'PE','Piauí':'PI','Rio de Janeiro':'RJ',
    'Rio Grande do Norte':'RN','Rio Grande do Sul':'RS','Rondônia':'RO','Roraima':'RR',
    'Santa Catarina':'SC','São Paulo':'SP','Sergipe':'SE','Tocantins':'TO',
}
_MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
             'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

_sicro_unidade_cache = {}

def _sicro_get_or_create_unidade(db, sigla):
    if not sigla or str(sigla) == 'nan': return None
    sigla = str(sigla).strip().upper()[:20]
    if sigla in _sicro_unidade_cache: return _sicro_unidade_cache[sigla]
    row = db.execute("SELECT id_unidade FROM unidades_medida WHERE sigla=?", [sigla]).fetchone()
    if row:
        _sicro_unidade_cache[sigla] = row[0]; return row[0]
    cur2 = db.execute("INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)",
                      [sigla, sigla, 'Outro'])
    _sicro_unidade_cache[sigla] = cur2.lastrowid
    return cur2.lastrowid


def _pf_sicro(v):
    if v is None: return None
    try:    return float(str(v).replace(',','.').strip())
    except: return None

def _mes_ref_sicro(s):
    """'Janeiro/2026' → '01/2026'"""
    if not s: return ''
    s = str(s).strip()
    for i, m in enumerate(_MESES_PT):
        if s.startswith(m):
            partes = s.split('/')
            if len(partes) == 2:
                return f"{str(i+1).zfill(2)}/{partes[1].strip()}"
    return s

def _parse_sicro_analitico(wb):
    """
    Parseia a planilha analítica SICRO (uma composição por bloco de ~30 linhas).
    Retorna lista de composições com suas seções e itens.
    """
    ws = wb.active
    composicoes = []
    cur = None
    sec_letra = None

    SECTION_LETTERS = {'A -', 'B -', 'C -', 'D -', 'E -', 'F -',
                       'A –', 'B –', 'C –', 'D –', 'E –', 'F –'}

    def _letra(v):
        v = str(v or '').strip()
        for sl in SECTION_LETTERS:
            if v.startswith(sl): return sl[0]
        return None

    def _is_item_code(v, sec=None):
        """Retorna True se v parece código de insumo SICRO (E9xxx, Pxxxx, Mxxxx)."""
        v = str(v or '').strip()
        if not v: return False
        if v[0] in 'EPMGCepmc' and len(v) >= 4 and v[1:].replace('.','').isdigit():
            return True
        # Atividades auxiliares, tempo fixo e momento de transporte referenciam
        # frequentemente codigos numericos de composicoes auxiliares do SICRO.
        return sec in ('D', 'E', 'F') and v.replace('.', '').isdigit() and len(v) >= 5

    for row in ws.iter_rows(values_only=True):
        c = [str(x).strip() if x is not None else '' for x in list(row) + ['']*15]

        v0 = c[0]

        # ── Início de nova composição ─────────────────────────────────────────
        if 'SISTEMA DE CUSTOS REFERENCIAIS' in v0:
            if cur: composicoes.append(cur)
            uf_nome = c[3]
            uf_cod  = _UF_NOME_COD.get(uf_nome, uf_nome[:2].upper() if len(uf_nome) >= 2 else 'XX')
            cur = {
                'uf': uf_cod,
                'fic': _pf_sicro(c[7]),
                'codigo': None, 'descricao': None,
                'mes_referencia': None,
                'producao_equipe': None, 'unidade_producao': None,
                'custo_unitario': None,
                'custo_horario_execucao': None,
                'custo_unitario_execucao': None,
                'custo_fic': None,
                'subtotal_sicro': None,
                'secoes': {},
            }
            sec_letra = None
            continue

        if cur is None:
            continue

        # ── Linha 2: data-base e produção ────────────────────────────────────
        if v0 == 'Custo Unitário de Referência':
            cur['mes_referencia'] = _mes_ref_sicro(c[3])
            cur['producao_equipe'] = _pf_sicro(c[7])
            cur['unidade_producao'] = c[8]
            continue

        # ── Linha 3: código e descrição ──────────────────────────────────────
        if cur['codigo'] is None and v0 and c[1]:
            v0s = v0.replace('.','')
            if v0s.isdigit() and len(v0) >= 5:
                cur['codigo'] = 'SICRO.' + v0
                cur['descricao'] = c[1]
                continue

        # ── Cabeçalho de seção ───────────────────────────────────────────────
        letra = _letra(v0)
        if letra:
            sec_letra = letra
            if letra not in cur['secoes']:
                cur['secoes'][letra] = {'itens': [], 'custo_total_secao': None}
            continue

        if sec_letra is None:
            continue

        sec = cur['secoes'][sec_letra]

        # ── Linhas de totais/resumo (pular como item, capturar total) ────────
        labels_total = ' '.join(str(x or '') for x in c[:9])
        if 'Custo unitário direto total' in labels_total:
            cur['custo_unitario'] = _pf_sicro(c[8])
            continue
        if 'Custo horário total de execução' in labels_total:
            cur['custo_horario_execucao'] = _pf_sicro(c[8])
            continue
        if 'Custo unitário de execução' in labels_total:
            cur['custo_unitario_execucao'] = _pf_sicro(c[8])
            continue
        if 'Custo do FIC' in labels_total:
            cur['custo_fic'] = _pf_sicro(c[8])
            continue
        if any(x in labels_total for x in ['Custo horário total','Custo unitário total',
                                           'Custo total de','Custo horário total de']):
            if sec['custo_total_secao'] is None:
                sec['custo_total_secao'] = _pf_sicro(c[8])
            continue
        if 'Subtotal' in labels_total:
            cur['subtotal_sicro'] = _pf_sicro(c[8])
            if sec['custo_total_secao'] is None:
                sec['custo_total_secao'] = _pf_sicro(c[8])
            continue

        # ── Fim do bloco ─────────────────────────────────────────────────────
        if v0 == 'Obs.':
            sec_letra = None
            continue

        # ── Item real ────────────────────────────────────────────────────────
        if _is_item_code(v0, sec_letra):
            item = {'codigo_item': v0, 'descricao': c[1]}
            if sec_letra == 'A':
                item.update({'quantidade': _pf_sicro(c[2]),
                              'util_operativa':   _pf_sicro(c[3]),
                              'util_improdutiva':  _pf_sicro(c[4]),
                              'custo_hp':  _pf_sicro(c[5]),
                              'custo_hi':  _pf_sicro(c[6]),
                              'custo_total': _pf_sicro(c[8])})
            elif sec_letra == 'B':
                item.update({'quantidade': _pf_sicro(c[2]),
                              'unidade': c[3],
                              'preco_unitario': _pf_sicro(c[5]),
                              'custo_total': _pf_sicro(c[8])})
            elif sec_letra in ('C', 'D'):
                item.update({'quantidade': _pf_sicro(c[2]),
                              'unidade': c[3],
                              'preco_unitario': _pf_sicro(c[5]),
                              'custo_total': _pf_sicro(c[8])})
            elif sec_letra == 'E':
                item.update({'cod_transporte': c[2],
                              'quantidade': _pf_sicro(c[3]),
                              'unidade': c[4],
                              'preco_unitario': _pf_sicro(c[6]),
                              'custo_total': _pf_sicro(c[8])})
            elif sec_letra == 'F':
                item.update({'quantidade': _pf_sicro(c[2]),
                              'unidade': c[3],
                              'cod_transp_ln': c[4],
                              'cod_transp_rp': c[5],
                              'cod_transp_p':  c[6],
                              'custo_total': _pf_sicro(c[8])})
            sec['itens'].append(item)

    if cur:
        composicoes.append(cur)

    return composicoes


# ── Endpoint: analisar arquivo SICRO (pré-importação) ────────────────────────
@app.route('/api/sicro/analisar-composicoes', methods=['POST'])
def sicro_analisar_composicoes():
    import tempfile
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'Arquivo não enviado.'}), 400
    f = request.files['arquivo']
    if not f.filename.lower().endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({'erro': 'Use arquivo .xlsx/.xls/.xlsm.'}), 400

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        f.save(tmp.name); tmp_path = tmp.name

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active

        # Amostra das primeiras composições para detectar UF e data-base
        uf_detectada, mes_ref_detectado = '', ''
        qtd_est = 0

        for row in ws.iter_rows(max_row=5000, values_only=True):
            c = [str(x).strip() if x is not None else '' for x in list(row) + ['']*10]
            if 'SISTEMA DE CUSTOS REFERENCIAIS' in c[0]:
                qtd_est += 1
                if not uf_detectada:
                    uf_nome = c[3]
                    uf_detectada = _UF_NOME_COD.get(uf_nome, uf_nome[:2].upper())
            if not mes_ref_detectado and c[0] == 'Custo Unitário de Referência':
                mes_ref_detectado = _mes_ref_sicro(c[3])

        wb.close()

        # Sobreposição no banco
        db = get_db()
        sob = 0
        if uf_detectada and mes_ref_detectado:
            sob = db.execute(
                "SELECT COUNT(*) FROM composicoes WHERE fonte='SICRO' AND uf_referencia=? AND mes_referencia=?",
                [uf_detectada, mes_ref_detectado]).fetchone()[0]
        db.close()

        return jsonify({
            'uf': uf_detectada,
            'mes_referencia': mes_ref_detectado,
            'qtd_composicoes_estimada': int(qtd_est),
            'sobreposicao': int(sob),
        })
    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-500:]}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


# ── Endpoint: importar composições SICRO ─────────────────────────────────────
@app.route('/api/sicro/importar-composicoes', methods=['POST'])
def sicro_importar_composicoes():
    import tempfile
    db = None
    tmp_path = None
    try:
        if 'arquivo' not in request.files:
            return jsonify({'erro': 'Arquivo não enviado.'}), 400

        # Para SICRO, sobrepor é sempre True — o usuário quer reimportar
        uf_override = request.form.get('uf_override', '').strip().upper()
        f = request.files['arquivo']
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            f.save(tmp.name); tmp_path = tmp.name

        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        composicoes = _parse_sicro_analitico(wb)
        wb.close()

        if not composicoes:
            return jsonify({'erro': 'Nenhuma composição encontrada. Verifique se é o Relatório Analítico de Composições de Custos do SICRO.'}), 400

        db = get_db()

        # Garantir fonte SICRO
        if not db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='SICRO'").fetchone():
            db.execute("INSERT OR IGNORE INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia) VALUES ('SICRO','Oficial','DNIT','Nacional')")
            db.commit()

        for col in (
            "custo_horario_execucao REAL",
            "custo_unitario_execucao REAL",
            "custo_fic REAL",
            "subtotal_sicro REAL",
        ):
            try:
                db.execute(f"ALTER TABLE composicoes ADD COLUMN {col}")
            except Exception:
                pass
        db.commit()

        SEC_NOMES = {'A':'Equipamentos','B':'Mão de Obra','C':'Material',
                     'D':'Atividades Auxiliares','E':'Tempo Fixo','F':'Momento de Transporte'}
        n_ins=0; n_upd=0; n_sec=0; n_item=0

        # Cache de composições existentes
        cod_map = {r[0]: r[1] for r in db.execute(
            "SELECT codigo, id_composicao FROM composicoes WHERE fonte='SICRO'")}

        # Cache de datas-base
        db_cache = {}

        for comp in composicoes:
            uf  = uf_override if uf_override else comp.get('uf','')
            mes = comp.get('mes_referencia','')
            cod = comp.get('codigo','')
            if not cod: continue

            # Data-base
            if mes not in db_cache and mes and len(mes) == 7:
                try:
                    m_num, a_num = int(mes[:2]), int(mes[3:])
                    db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [m_num, a_num]).fetchone()
                    if not db_row:
                        db_cache[mes] = db.execute("INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)",
                                                    [m_num, a_num, f'SICRO {mes}']).lastrowid
                        db.commit()
                    else:
                        db_cache[mes] = db_row[0]
                except: pass

            vals = [comp.get('descricao',''), comp.get('unidade_producao',''), mes, uf,
                    comp.get('fic'), comp.get('producao_equipe'),
                    comp.get('unidade_producao',''), comp.get('custo_unitario'),
                    comp.get('custo_horario_execucao'), comp.get('custo_unitario_execucao'),
                    comp.get('custo_fic'), comp.get('subtotal_sicro'), 'Ativo']

            if cod in cod_map:
                id_comp = cod_map[cod]
                db.execute("UPDATE composicoes SET descricao=?,unidade=?,mes_referencia=?,uf_referencia=?,fic=?,producao_equipe=?,unidade_producao=?,custo_unitario=COALESCE(?, custo_unitario),custo_horario_execucao=?,custo_unitario_execucao=?,custo_fic=?,subtotal_sicro=?,situacao=? WHERE id_composicao=?",
                           vals + [id_comp])
                db.execute("DELETE FROM composicoes_secoes WHERE id_composicao=?", [id_comp])
                db.commit()
                n_upd += 1
            else:
                id_comp = db.execute(
                    "INSERT INTO composicoes (codigo,fonte,formato,descricao,unidade,mes_referencia,uf_referencia,fic,producao_equipe,unidade_producao,custo_unitario,custo_horario_execucao,custo_unitario_execucao,custo_fic,subtotal_sicro,situacao) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    [cod,'SICRO','PRODUCAO_HORARIA'] + vals).lastrowid
                db.commit()
                cod_map[cod] = id_comp
                n_ins += 1

            for ordem_sec, (letra, sec) in enumerate(sorted(comp.get('secoes',{}).items())):
                id_sec = db.execute(
                    "INSERT INTO composicoes_secoes (id_composicao,letra_secao,nome_secao,ordem,custo_total_secao) VALUES (?,?,?,?,?)",
                    [id_comp, letra, SEC_NOMES.get(letra, letra), ordem_sec, sec.get('custo_total_secao')]).lastrowid
                n_sec += 1

                itens = sec.get('itens',[])
                if itens:
                    db.executemany(
                        "INSERT INTO composicoes_secao_itens (id_composicao,id_secao,letra_secao,codigo_item,descricao,quantidade,unidade,util_operativa,util_improdutiva,custo_hp,custo_hi,preco_unitario,custo_total,cod_transporte,cod_transp_ln,cod_transp_rp,cod_transp_p,fit,ordem) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        [[id_comp,id_sec,letra,
                          it.get('codigo_item'), it.get('descricao'), it.get('quantidade'), it.get('unidade'),
                          it.get('util_operativa'), it.get('util_improdutiva'),
                          it.get('custo_hp'), it.get('custo_hi'),
                          it.get('preco_unitario'), it.get('custo_total'),
                          it.get('cod_transporte'), it.get('cod_transp_ln'),
                          it.get('cod_transp_rp'), it.get('cod_transp_p'), it.get('fit'), i]
                         for i, it in enumerate(itens)])
                    n_item += len(itens)
            db.commit()

        return jsonify({
            'composicoes_inseridas':   n_ins,
            'composicoes_atualizadas': n_upd,
            'composicoes_ignoradas':   0,
            'secoes_inseridas':        n_sec,
            'itens_inseridos':         n_item,
            'total_processadas':       len(composicoes),
            'mensagem': f'{n_ins} composições inseridas, {n_upd} atualizadas. {n_sec} seções e {n_item} itens importados.',
        })

    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1000:]}), 500
    finally:
        if db:
            try: db.close()
            except: pass
        if tmp_path:
            try: os.unlink(tmp_path)
            except: pass
@app.route('/api/sicro/importar-insumos', methods=['POST'])
def sicro_importar_insumos():
    import tempfile, re as _re
    db = None
    tmp_paths = {}
    try:
        uf      = request.form.get('uf','').strip().upper()
        mes_ref = request.form.get('mes_ref','').strip()
        sobrepor= request.form.get('sobrepor','false').lower() == 'true'
        if not uf:
            return jsonify({'erro': 'UF é obrigatória.'}), 400
        if not _re.match(r'^\d{2}/\d{4}$', mes_ref):
            return jsonify({'erro': 'Mês de referência inválido. Use MM/AAAA.'}), 400
        for key in ('arq_mo','arq_mat','arq_equip'):
            if key not in request.files:
                return jsonify({'erro': f'Arquivo "{key}" não enviado.'}), 400

        import openpyxl
        for key in ('arq_mo','arq_mat','arq_equip'):
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                request.files[key].save(tmp.name); tmp_paths[key] = tmp.name

        db = get_db()
        m_num = int(mes_ref[:2]); a_num = int(mes_ref[3:])
        db_row = db.execute("SELECT id_data_base FROM datas_base WHERE mes=? AND ano=?", [m_num, a_num]).fetchone()
        id_db  = db_row[0] if db_row else db.execute(
            "INSERT INTO datas_base (mes,ano,descricao) VALUES (?,?,?)", [m_num, a_num, f'SICRO {mes_ref}']).lastrowid
        db.commit()

        fn = db.execute("SELECT id_fonte FROM fontes_referencia WHERE nome_fonte='SICRO'").fetchone()
        if fn:
            id_fonte = fn[0]
        else:
            id_fonte = db.execute("INSERT OR IGNORE INTO fontes_referencia (nome_fonte,tipo_fonte,orgao_responsavel,abrangencia) VALUES ('SICRO','Oficial','DNIT','Nacional')").lastrowid
            db.commit()

        stats = {'ins_insumos':0,'upd_insumos':0,'ins_precos':0,'upd_precos':0,
                 'ins_equip':0,'upd_equip':0,'ins_preco_equip':0,'upd_preco_equip':0}

        # Cache de unidades
        unid_cache = {r[0]: r[1] for r in db.execute("SELECT sigla,id_unidade FROM unidades_medida")}
        def get_unidade(sigla):
            if not sigla: return None
            sigla = str(sigla).strip().upper()[:20]
            if sigla not in unid_cache:
                uid = db.execute("INSERT INTO unidades_medida (sigla,descricao,tipo_unidade) VALUES (?,?,?)",
                                 [sigla, sigla, 'Outro']).lastrowid
                db.commit(); unid_cache[sigla] = uid
            return unid_cache[sigla]

        # Cache de insumos existentes
        ins_cache = {r[0]: r[1] for r in db.execute("SELECT codigo_insumo, id_insumo FROM insumos WHERE origem='SICRO'")}

        def processar_insumos(path, tipo, prefixo):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = [list(r) + [None]*6 for r in wb.active.iter_rows(values_only=True)
                    if r[0] and str(r[0]).strip().upper().startswith(prefixo)]
            wb.close()

            novos_ins=[]; upd_ins=[]; novos_preco=[]; upd_preco=[]
            for c in rows:
                cod  = str(c[0]).strip()
                desc = str(c[1] or '').strip()
                unid = str(c[2] or '').strip()
                preco= _pf_sicro(c[3])
                if preco is None: continue

                id_un = get_unidade(unid)
                if cod in ins_cache:
                    id_ins = ins_cache[cod]
                    if sobrepor:
                        upd_ins.append((desc, id_un, id_ins))
                else:
                    novos_ins.append((cod, desc, tipo, id_un))

            if novos_ins:
                db.executemany("INSERT INTO insumos (codigo_insumo,descricao,tipo_insumo,id_unidade,origem,situacao) VALUES (?,?,?,?,'SICRO','Ativo')",
                               novos_ins)
                db.commit()
                for (cod, desc, tipo_, id_un) in novos_ins:
                    r = db.execute("SELECT id_insumo FROM insumos WHERE codigo_insumo=? AND origem='SICRO'", [cod]).fetchone()
                    if r: ins_cache[cod] = r[0]
                stats['ins_insumos'] += len(novos_ins)
            if upd_ins:
                db.executemany("UPDATE insumos SET descricao=?,id_unidade=? WHERE id_insumo=?", upd_ins)
                db.commit(); stats['upd_insumos'] += len(upd_ins)

            # Preços
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for c in [list(r) + [None]*6 for r in wb.active.iter_rows(values_only=True)
                      if r[0] and str(r[0]).strip().upper().startswith(prefixo)]:
                cod   = str(c[0]).strip()
                preco = _pf_sicro(c[3])
                if preco is None or cod not in ins_cache: continue
                id_ins = ins_cache[cod]
                ep = db.execute("SELECT id_preco FROM precos_insumos WHERE id_insumo=? AND id_data_base=? AND uf_referencia=?",
                                [id_ins, id_db, uf]).fetchone()
                if ep:
                    if sobrepor:
                        db.execute("UPDATE precos_insumos SET preco_referencia=?,preco_nao_desonerado=? WHERE id_preco=?",
                                   [preco, preco, ep[0]]); stats['upd_precos'] += 1
                else:
                    novos_preco.append((id_ins, id_db, id_fonte, uf, preco, preco))
            wb.close()
            if novos_preco:
                db.executemany("INSERT INTO precos_insumos (id_insumo,id_data_base,id_fonte,uf_referencia,preco_referencia,preco_nao_desonerado) VALUES (?,?,?,?,?,?)",
                               novos_preco)
                db.commit(); stats['ins_precos'] += len(novos_preco)

        processar_insumos(tmp_paths['arq_mo'],  'Mão de Obra', 'P')
        processar_insumos(tmp_paths['arq_mat'], 'Material',    'M')

        # ── Equipamentos ──────────────────────────────────────────────────────
        eq_cache = {r[0]: r[1] for r in db.execute(
            "SELECT codigo_chp, id_equip FROM equipamentos_sinapi WHERE sistema='SICRO'")}
        wb = openpyxl.load_workbook(tmp_paths['arq_equip'], read_only=True, data_only=True)
        rows_eq = [list(r)+[None]*12 for r in wb.active.iter_rows(values_only=True)
                   if r[0] and str(r[0]).strip().upper().startswith('E')]
        wb.close()

        novos_eq=[]; upd_eq=[]; novos_peq=[]; upd_peq=[]
        for c in rows_eq:
            cod=str(c[0]).strip(); desc=str(c[1] or '').strip()
            cprod=_pf_sicro(c[9]); cimprod=_pf_sicro(c[10])
            if cprod is None: continue
            val_aq=_pf_sicro(c[2]) or 0; dep=_pf_sicro(c[3]) or 0
            oport=_pf_sicro(c[4]) or 0; seg=_pf_sicro(c[5]) or 0
            manut=_pf_sicro(c[6]) or 0; oper=_pf_sicro(c[7]) or 0; mo_op=_pf_sicro(c[8]) or 0

            if cod in eq_cache:
                id_eq = eq_cache[cod]
                if sobrepor:
                    upd_eq.append((desc, cprod, cimprod, id_eq))
            else:
                novos_eq.append((cod, desc, cprod, cimprod, f'SICRO {mes_ref}'))

        if novos_eq:
            db.executemany("INSERT INTO equipamentos_sinapi (codigo_chp,descricao,sistema,custo_produtivo,custo_improdutivo,situacao,fonte) VALUES (?,?,'SICRO',?,?,'Ativo',?)",
                           novos_eq)
            db.commit()
            for (cod,*_) in novos_eq:
                r = db.execute("SELECT id_equip FROM equipamentos_sinapi WHERE codigo_chp=? AND sistema='SICRO'", [cod]).fetchone()
                if r: eq_cache[cod] = r[0]
            stats['ins_equip'] += len(novos_eq)
        if upd_eq:
            db.executemany("UPDATE equipamentos_sinapi SET descricao=?,custo_produtivo=?,custo_improdutivo=? WHERE id_equip=?", upd_eq)
            db.commit(); stats['upd_equip'] += len(upd_eq)

        # Preços de equipamentos
        for c in rows_eq:
            cod=str(c[0]).strip()
            if cod not in eq_cache: continue
            id_eq=eq_cache[cod]
            cprod=_pf_sicro(c[9]); cimprod=_pf_sicro(c[10]) or 0
            if cprod is None: continue
            val_aq=_pf_sicro(c[2]) or 0; dep=_pf_sicro(c[3]) or 0
            oport=_pf_sicro(c[4]) or 0; seg=_pf_sicro(c[5]) or 0
            manut=_pf_sicro(c[6]) or 0; oper=_pf_sicro(c[7]) or 0; mo_op=_pf_sicro(c[8]) or 0
            ep = db.execute("SELECT id_preco_eq FROM precos_equipamentos WHERE id_equip=? AND id_data_base=? AND uf_referencia=?",
                            [id_eq, id_db, uf]).fetchone()
            if ep:
                if sobrepor:
                    db.execute("UPDATE precos_equipamentos SET preco_aquisicao=?,custo_depreciacao=?,custo_juros=?,custo_imp_seguros=?,custo_manutencao=?,custo_materiais=?,custo_mao_obra=?,chp_calculado=?,chi_calculado=? WHERE id_preco_eq=?",
                               [val_aq,dep,oport,seg,manut,oper,mo_op,cprod,cimprod,ep[0]]); stats['upd_preco_equip']+=1
            else:
                novos_peq.append((id_eq,id_db,id_fonte,uf,val_aq,dep,oport,seg,manut,oper,mo_op,cprod,cimprod))

        if novos_peq:
            db.executemany("INSERT INTO precos_equipamentos (id_equip,id_data_base,id_fonte,uf_referencia,preco_aquisicao,custo_depreciacao,custo_juros,custo_imp_seguros,custo_manutencao,custo_materiais,custo_mao_obra,chp_calculado,chi_calculado) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                           novos_peq)
            db.commit(); stats['ins_preco_equip'] += len(novos_peq)
        if upd_peq:
            db.commit(); stats['upd_preco_equip'] += len(upd_peq)

        return jsonify({**stats, 'uf': uf, 'mes_referencia': mes_ref,
            'mensagem': (f"Insumos: {stats['ins_insumos']} inseridos, {stats['upd_insumos']} atualizados. "
                         f"Preços: {stats['ins_precos']} inseridos, {stats['upd_precos']} atualizados. "
                         f"Equipamentos: {stats['ins_equip']} inseridos, {stats['upd_equip']} atualizados. "
                         f"Custos equip.: {stats['ins_preco_equip']} inseridos.")})

    except Exception as e:
        import traceback
        return jsonify({'erro': str(e), 'detalhe': traceback.format_exc()[-1000:]}), 500
    finally:
        if db:
            try: db.close()
            except: pass
        for p in tmp_paths.values():
            try: os.unlink(p)
            except: pass
if __name__ == '__main__':
    # Forçar UTF-8 no stdout/stderr para evitar erros de encoding no Windows
    import sys, io
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    print(f"\n🚀  Sistema de Orçamentação de Obras iniciado!")
    print(f"    Banco: {DB_PATH}")
    print(f"    Acesse: http://localhost:{PORT}\n")
    try:
        _db_init = get_db()
        ensure_encargos_schema(_db_init)
        ensure_insumos_encargos_schema(_db_init)
        _db_init.commit()
        _db_init.close()
    except Exception as _e:
        print(f"    Aviso: migração inicial de encargos dos insumos não executada: {_e}")
    app.run(host='0.0.0.0', port=PORT, debug=False, threaded=True)
